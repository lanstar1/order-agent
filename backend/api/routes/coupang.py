"""
쿠팡 OPEN API 주문 자동화 라우트
발주서조회 → 상품준비중 → ERP 판매입력 → 송장업로드(발송처리)
+ 상품매핑 (엑셀 기반 쿠팡상품 → ERP품목코드)
"""
import re
import logging
import io
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from pathlib import Path
from typing import List, Optional

from fastapi import APIRouter, Query, Body, UploadFile, File
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel

from config import (
    COUPANG_ACCESS_KEY, COUPANG_SECRET_KEY, COUPANG_VENDOR_ID,
    COUPANG_API_BASE,
    COUPANG_CUST_CODE, COUPANG_EMP_CODE, COUPANG_WH_CODE,
    ERP_COM_CODE, ERP_USER_ID, ERP_ZONE, ERP_API_KEY,
)
from services.coupang_client import (
    CoupangClient, DELIVERY_COMPANIES,
    ORDER_STATUS_ACCEPT, ORDER_STATUS_INSTRUCT,
    ORDER_STATUS_DEPARTURE, ORDER_STATUS_DELIVERING,
)

KST = ZoneInfo("Asia/Seoul")
logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/coupang", tags=["Coupang"])

# 쿠팡 클라이언트 싱글턴
_coupang: CoupangClient | None = None


def _get_coupang() -> CoupangClient:
    global _coupang
    if _coupang is None:
        _coupang = CoupangClient(
            access_key=COUPANG_ACCESS_KEY,
            secret_key=COUPANG_SECRET_KEY,
            vendor_id=COUPANG_VENDOR_ID,
            base_url=COUPANG_API_BASE,
        )
    return _coupang


# ─── 연결 테스트 ───
@router.get("/test-connection")
async def test_connection():
    """쿠팡 API 및 ERP 연결 테스트"""
    results = {}

    # 쿠팡 API 테스트
    cp = _get_coupang()
    cp_result = await cp.test_connection()
    results["coupang"] = cp_result

    # ERP 테스트
    if ERP_COM_CODE and ERP_API_KEY:
        try:
            from services.erp_client import ERPClient
            erp = ERPClient()
            session = await erp.ensure_session()
            results["erp"] = {"ok": bool(session)}
        except Exception as e:
            results["erp"] = {"ok": False, "error": str(e)}
    else:
        results["erp"] = {"ok": False, "error": "ERP 설정 없음"}

    return results


# ─── 발주서 목록 조회 (주문수집) ───
@router.get("/orders")
async def get_orders(
    status: str = Query(ORDER_STATUS_ACCEPT, description="ACCEPT/INSTRUCT/DEPARTURE/DELIVERING"),
    from_date: str = Query("", description="시작일 (YYYY-MM-DD)"),
    to_date: str = Query("", description="종료일 (YYYY-MM-DD)"),
):
    """쿠팡 발주서 목록 조회"""
    cp = _get_coupang()
    if not cp.access_key:
        return JSONResponse(
            status_code=400,
            content={"detail": "쿠팡 API 키가 설정되지 않았습니다. 환경변수를 확인하세요."},
        )

    try:
        orders = await cp.fetch_orders(
            status=status,
            from_date=from_date,
            to_date=to_date,
        )
        return {"orders": orders, "total": len(orders), "status": status}
    except ValueError as e:
        return JSONResponse(status_code=400, content={"detail": str(e)})
    except Exception as e:
        logger.error(f"[쿠팡] 주문조회 오류: {e}", exc_info=True)
        return JSONResponse(status_code=500, content={"detail": str(e)})


# ─── 상품준비중 처리 (발주확인) ───
class ConfirmRequest(BaseModel):
    shipment_box_ids: List[int]


@router.post("/confirm-orders")
async def confirm_orders(req: ConfirmRequest):
    """선택한 발주서를 '상품준비중' 상태로 변경"""
    cp = _get_coupang()
    try:
        # 50개씩 나눠서 처리
        all_results = []
        ids = req.shipment_box_ids
        for i in range(0, len(ids), 50):
            batch = ids[i:i+50]
            result = await cp.confirm_orders(batch)
            all_results.append(result)

        return {"results": all_results, "total": len(ids)}
    except ValueError as e:
        return JSONResponse(status_code=400, content={"detail": str(e)})
    except Exception as e:
        logger.error(f"[쿠팡] 상품준비중 처리 오류: {e}", exc_info=True)
        return JSONResponse(status_code=500, content={"detail": str(e)})


# ─── ERP 판매전표 등록 ───
class ERPSendRequest(BaseModel):
    orders: list  # 주문 데이터 리스트
    io_date: str = ""  # 전표일자 (YYYY-MM-DD)


@router.post("/send-to-erp")
async def send_to_erp(req: ERPSendRequest):
    """주문 데이터를 ERP 판매전표로 등록 (ERPClientSS.save_sale 사용)"""
    if not ERP_COM_CODE or not ERP_API_KEY:
        return JSONResponse(status_code=400, content={"detail": "ERP 설정이 없습니다."})
    if not COUPANG_CUST_CODE:
        return JSONResponse(status_code=400, content={"detail": "COUPANG_CUST_CODE가 설정되지 않았습니다."})

    try:
        from services.erp_client_ss import ERPClientSS
        erp = ERPClientSS()
        await erp.ensure_session()

        # ERP 판매전표 라인 구성
        erp_lines = []
        unmatched_items = []
        for order in req.orders:
            order_items = order.get("orderItems", [])
            for item in order_items:
                prod_cd = _resolve_erp_code(item)
                qty = item.get("shippingCount", 1)
                price = item.get("salesPrice", 0)
                if isinstance(price, dict):
                    price = price.get("units", 0)
                rcv_name = order.get("receiver", {}).get("name", "")

                if not prod_cd:
                    unmatched_items.append({
                        "order_id": order.get("orderId", ""),
                        "item_name": item.get("vendorItemName", ""),
                        "sellerProductId": item.get("sellerProductId", ""),
                        "reason": "매핑 없음 & externalVendorSku 없음",
                    })
                    continue

                erp_lines.append({
                    "prod_cd": prod_cd,
                    "prod_name": item.get("vendorItemName", ""),
                    "qty": qty,
                    "price": round(price / qty, 2) if qty else 0,
                    "rcv_name": rcv_name,
                })

        if not erp_lines:
            return {"success": False, "error": "ERP 전송 대상 없음", "unmatched_items": unmatched_items}

        # ERPClientSS.save_sale() 호출 (스마트스토어와 동일)
        _emp_cd = COUPANG_EMP_CODE or ""
        result = await erp.save_sale(COUPANG_CUST_CODE, erp_lines, COUPANG_WH_CODE, _emp_cd)
        result["lines"] = len(erp_lines)
        result["unmatched_items"] = unmatched_items
        logger.info(f"[쿠팡] ERP 전송 완료: {len(erp_lines)}건, 미매칭: {len(unmatched_items)}건")
        return result

    except Exception as e:
        logger.error(f"[쿠팡] ERP 전송 오류: {e}", exc_info=True)
        return JSONResponse(status_code=500, content={"detail": str(e)})


# ─── 로젠 전송용 엑셀 다운로드 ───
@router.post("/logen-export-excel")
async def logen_export_excel(orders: list = Body(...)):
    """선택한 쿠팡 주문을 로젠 전송용 엑셀로 다운로드

    로젠 구시스템양식 (A타입):
    A:수하인명 B:수하인주소1 C:수하인전화 D:수하인휴대폰
    E:택배수량 F:택배운임 G:운임구분 H:물품명 I:주문번호
    J:제주운임구분 K:배송메세지
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "엑셀파일첫행-제목있음"

    headers = ["수하인명", "수하인주소1", "수하인전화", "수하인휴대폰",
               "택배수량", "택배운임", "운임구분", "물품명", "주문번호",
               None, "배송메세지"]
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF")
    for ci, h in enumerate(headers, 1):
        if h:
            c = ws.cell(1, ci, h)
            c.fill = hdr_fill; c.font = hdr_font; c.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 30
    ws.column_dimensions["I"].width = 22
    ws.column_dimensions["K"].width = 25

    # shipmentBoxId 기준으로 그룹화
    groups: dict = {}
    for o in orders:
        sb_id = str(o.get("shipmentBoxId", ""))
        if not sb_id:
            continue
        if sb_id not in groups:
            groups[sb_id] = o
        # 이미 있으면 첫번째 주문 사용 (같은 shipmentBox)

    row_idx = 2
    for sb_id, o in groups.items():
        receiver = o.get("receiver") or o.get("orderer") or {}
        rcv_name = receiver.get("name", "")
        rcv_tel = receiver.get("safeNumber") or receiver.get("phone") or receiver.get("cellPhone", "")
        rcv_cell = receiver.get("cellPhone") or receiver.get("phone") or rcv_tel
        rcv_addr = (receiver.get("addr1", "") + " " + receiver.get("addr2", "")).strip()

        items = o.get("orderItems", [])
        # 물품명: LS-, LSP-, ZOT-, LSN- 모델코드만 추출
        model_pat = re.compile(r'((?:LS|LSP|ZOT|LSN)-[\w\-]+)')
        goods_parts = []
        for item in items:
            name = item.get("vendorItemName", "")
            qty = item.get("shippingCount", 1)
            match = model_pat.search(name)
            model = match.group(1) if match else name[:20]
            goods_parts.append(f"{model} x{qty}" if qty > 1 else model)
        goods_nm = ", ".join(goods_parts)[:50]

        # 주문번호: shipmentBoxId만 (로젠 반환 시 S열 매칭용)
        order_key = sb_id

        fare_tp = "010"  # 선불

        ws.cell(row_idx, 1, rcv_name)      # A: 수하인명
        ws.cell(row_idx, 2, rcv_addr)      # B: 수하인주소1
        ws.cell(row_idx, 3, rcv_tel)       # C: 수하인전화
        ws.cell(row_idx, 4, rcv_cell)      # D: 수하인휴대폰
        ws.cell(row_idx, 5, 1)             # E: 택배수량 (박스)
        ws.cell(row_idx, 6, 0)             # F: 택배운임
        ws.cell(row_idx, 7, fare_tp)       # G: 운임구분
        ws.cell(row_idx, 8, goods_nm)      # H: 물품명
        ws.cell(row_idx, 9, order_key)     # I: 주문번호 (매칭용)
        jeju = "선착불" if "제주" in rcv_addr else None
        ws.cell(row_idx, 10, jeju)         # J: 제주운임구분
        ws.cell(row_idx, 11, receiver.get("message", ""))  # K: 배송메세지
        row_idx += 1

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    filename = f"coupang_logen_{datetime.now(KST).strftime('%Y%m%d_%H%M%S')}.xlsx"
    logger.info(f"[쿠팡로젠] 엑셀 다운로드: {row_idx - 2}건")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


# ─── 로젠 송장 엑셀 업로드 → 쿠팡 발송처리 ───
@router.post("/logen-dispatch-excel")
async def logen_dispatch_excel(
    file: UploadFile = File(...),
    carrier: str = Query("LOGEN", description="택배사코드"),
):
    """로젠에서 반환된 엑셀 업로드 → 쿠팡 송장업로드 자동 처리

    로젠 반환 파일: D열(index 3)=운송장번호, S열(index 18)=주문번호(shipmentBoxId)
    """
    import openpyxl

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    cp = _get_coupang()
    company_code = DELIVERY_COMPANIES.get(carrier, carrier)

    results = []
    skipped = []

    # 로젠 반환 파일: 1행=타이틀, 2행=헤더, 3행=서브헤더, 4행~=데이터
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not any(row):
            continue
        tracking = str(row[3] or "").strip()    # D열: 운송장번호
        sb_id = str(row[18] or "").strip()      # S열: shipmentBoxId

        if not tracking or not sb_id or tracking == "None" or sb_id == "None":
            skipped.append(sb_id or str(row[0] or ""))
            continue

        # "|" 포함 시 하위호환 (이전 버전 파일)
        if "|" in sb_id:
            sb_id = sb_id.split("|")[0]

        if not sb_id:
            skipped.append(str(row[0] or ""))
            continue

        try:
            invoice_data = [{
                "shipmentBoxId": int(sb_id),
                "deliveryCompanyCode": company_code,
                "invoiceNumber": tracking,
                "splitShipping": False,
                "preSplitShipped": False,
                "estimatedShippingDate": "",
            }]
            result = await cp.upload_invoice(invoice_data)
            results.append({"shipmentBoxId": sb_id, "tracking": tracking, "success": True})
        except Exception as e:
            results.append({"shipmentBoxId": sb_id, "tracking": tracking, "success": False, "error": str(e)})

    success_cnt = sum(1 for r in results if r["success"])
    fail_cnt = sum(1 for r in results if not r["success"])
    logger.info(f"[쿠팡로젠] 송장업로드: 성공={success_cnt}, 실패={fail_cnt}, 건너뜀={len(skipped)}")

    return {
        "success": success_cnt > 0,
        "dispatched_count": success_cnt,
        "fail_count": fail_cnt,
        "skipped_count": len(skipped),
        "results": results,
    }


# ─── 송장업로드 (발송처리) - 수동 ───
class ShipRequest(BaseModel):
    shipments: list  # [{"shipment_box_id": ..., "order_id": ..., "vendor_item_id": ..., "delivery_company": "CJ대한통운", "invoice_number": "..."}, ...]


@router.post("/ship")
async def ship_orders(req: ShipRequest):
    """송장업로드 처리 (발송처리)"""
    cp = _get_coupang()
    results = []
    for s in req.shipments:
        company = s.get("delivery_company", "")
        company_code = DELIVERY_COMPANIES.get(company, s.get("delivery_company_code", ""))
        if not company_code:
            results.append({
                "shipment_box_id": s.get("shipment_box_id"),
                "success": False,
                "error": f"택배사코드 없음: {company}",
            })
            continue

        try:
            invoice_data = [{
                "shipmentBoxId": s["shipment_box_id"],
                "orderId": s["order_id"],
                "vendorItemId": s["vendor_item_id"],
                "deliveryCompanyCode": company_code,
                "invoiceNumber": s["invoice_number"],
                "splitShipping": s.get("split_shipping", False),
                "preSplitShipped": s.get("pre_split_shipped", False),
                "estimatedShippingDate": s.get("estimated_shipping_date", ""),
            }]
            data = await cp.upload_invoice(invoice_data)
            results.append({
                "shipment_box_id": s["shipment_box_id"],
                "success": True,
                "data": data,
            })
        except Exception as e:
            results.append({
                "shipment_box_id": s.get("shipment_box_id"),
                "success": False,
                "error": str(e),
            })

    success = sum(1 for r in results if r["success"])
    return {"results": results, "total": len(results), "success_count": success}


# ─── 택배사 목록 ───
@router.get("/delivery-companies")
async def delivery_companies():
    """사용 가능한 택배사 목록"""
    return {"companies": [{"name": k, "code": v} for k, v in DELIVERY_COMPANIES.items()]}


# ─── 설정 상태 ───
@router.get("/config-status")
async def config_status():
    """현재 설정 상태 반환"""
    return {
        "coupang_configured": bool(COUPANG_ACCESS_KEY and COUPANG_SECRET_KEY),
        "coupang_access_key": COUPANG_ACCESS_KEY[:4] + "****" if COUPANG_ACCESS_KEY else "",
        "vendor_id": COUPANG_VENDOR_ID,
        "erp_configured": bool(ERP_COM_CODE and ERP_API_KEY),
        "coupang_cust_code": COUPANG_CUST_CODE,
        "coupang_wh_code": COUPANG_WH_CODE,
    }


# ═══════════════════════════════════════════
#  쿠팡 상품 → ERP 품목코드 매핑 (엑셀 기반)
# ═══════════════════════════════════════════

# 매핑 딕셔너리: sellerProductId → ERP 품목코드
_coupang_product_map: dict = {}   # sellerProductId → erp_code
_coupang_model_map: dict = {}     # sellerProductId → model_name

# 엑셀 파일 경로
_MAPPING_DIR = Path(__file__).parent.parent.parent / "data"
# v2 파일이 있으면 우선 사용 (실제 ERP 품목코드 매핑)
_MAPPING_FILE_V2 = _MAPPING_DIR / "coupang_product_map_v2.xlsx"
_MAPPING_FILE_V1 = _MAPPING_DIR / "coupang_product_map.xlsx"
_MAPPING_FILE = _MAPPING_FILE_V2 if _MAPPING_FILE_V2.exists() else _MAPPING_FILE_V1


def _load_mapping():
    """엑셀 파일에서 매핑 로드"""
    global _coupang_product_map, _coupang_model_map
    _coupang_product_map.clear()
    _coupang_model_map.clear()

    if not _MAPPING_FILE.exists():
        logger.info("[쿠팡매핑] 매핑 파일 없음, 빈 맵 사용")
        return

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(_MAPPING_FILE), read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            seller_id = str(row[0]).strip()
            erp_code = str(row[1] or "").strip() if len(row) > 1 else ""
            model = str(row[2] or "").strip() if len(row) > 2 else ""
            if seller_id and erp_code:
                _coupang_product_map[seller_id] = erp_code
            if seller_id and model:
                _coupang_model_map[seller_id] = model
        wb.close()
        logger.info(f"[쿠팡매핑] 로드 완료: {len(_coupang_product_map)}건")
    except Exception as e:
        logger.error(f"[쿠팡매핑] 로드 실패: {e}", exc_info=True)


def _save_mapping():
    """매핑을 엑셀 파일로 저장"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    _MAPPING_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "쿠팡매핑"

    # 헤더
    headers = ["쿠팡상품ID(sellerProductId)", "ERP품목코드", "모델명(참고)"]
    header_fill = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 데이터
    all_ids = sorted(set(list(_coupang_product_map.keys()) + list(_coupang_model_map.keys())))
    for row_idx, sid in enumerate(all_ids, 2):
        ws.cell(row_idx, 1, sid)
        ws.cell(row_idx, 2, _coupang_product_map.get(sid, ""))
        ws.cell(row_idx, 3, _coupang_model_map.get(sid, ""))

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 30

    wb.save(str(_MAPPING_FILE))
    logger.info(f"[쿠팡매핑] 저장 완료: {len(all_ids)}건 → {_MAPPING_FILE}")


def _resolve_erp_code(order_item: dict) -> str:
    """주문 아이템에서 ERP 품목코드 결정
    우선순위: 1) 매핑테이블 2) externalVendorSku
    """
    # 1) sellerProductId로 매핑테이블 조회
    seller_id = str(order_item.get("sellerProductId", "")).strip()
    if seller_id and seller_id in _coupang_product_map:
        return _coupang_product_map[seller_id]

    # 2) externalVendorSku 사용
    sku = str(order_item.get("externalVendorSku", "")).strip()
    if sku:
        return sku

    return ""


# 시작 시 매핑 로드
_load_mapping()


# ─── 상품 목록 조회 (쿠팡 API) ───
@router.get("/products")
async def get_products():
    """쿠팡 전체 상품 목록 조회"""
    cp = _get_coupang()
    if not cp.access_key:
        return JSONResponse(status_code=400, content={"detail": "쿠팡 API 키 미설정"})

    try:
        products = await cp.fetch_products()
        # 매핑 상태 추가
        for p in products:
            sid = str(p.get("sellerProductId", ""))
            p["_erp_code"] = _coupang_product_map.get(sid, "")
            p["_model"] = _coupang_model_map.get(sid, "")
            p["_mapped"] = bool(p["_erp_code"])
        return {"products": products, "total": len(products)}
    except Exception as e:
        logger.error(f"[쿠팡] 상품목록 조회 오류: {e}", exc_info=True)
        return JSONResponse(status_code=500, content={"detail": str(e)})


# ─── 매핑 목록 조회 ───
@router.get("/mapping")
async def get_mapping():
    """현재 매핑 목록 반환"""
    items = []
    all_ids = sorted(set(list(_coupang_product_map.keys()) + list(_coupang_model_map.keys())))
    for sid in all_ids:
        items.append({
            "sellerProductId": sid,
            "erpCode": _coupang_product_map.get(sid, ""),
            "model": _coupang_model_map.get(sid, ""),
        })
    return {"mappings": items, "total": len(items)}


# ─── 매핑 추가/수정 ───
@router.post("/mapping")
async def upsert_mapping(entry: dict = Body(...)):
    """매핑 추가/수정"""
    sid = str(entry.get("sellerProductId", "")).strip()
    erp_code = str(entry.get("erpCode", "")).strip()
    model = str(entry.get("model", "")).strip()

    if not sid or not erp_code:
        return JSONResponse(status_code=400, content={"detail": "sellerProductId와 erpCode 필수"})

    action = "수정" if sid in _coupang_product_map else "추가"
    _coupang_product_map[sid] = erp_code
    if model:
        _coupang_model_map[sid] = model
    _save_mapping()

    logger.info(f"[쿠팡매핑] {action}: {sid} → ERP:{erp_code}, 모델:{model}")
    return {"success": True, "action": action, "sellerProductId": sid, "erpCode": erp_code}


# ─── 매핑 삭제 ───
@router.delete("/mapping/{seller_product_id}")
async def delete_mapping(seller_product_id: str):
    """매핑 삭제"""
    if seller_product_id not in _coupang_product_map:
        return JSONResponse(status_code=404, content={"detail": "매핑 없음"})

    erp_code = _coupang_product_map.pop(seller_product_id)
    _coupang_model_map.pop(seller_product_id, None)
    _save_mapping()
    logger.info(f"[쿠팡매핑] 삭제: {seller_product_id} (was {erp_code})")
    return {"success": True, "deleted": seller_product_id}


# ─── 매핑 엑셀 다운로드 ───
@router.get("/mapping/download")
async def download_mapping():
    """매핑 엑셀 파일 다운로드"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "쿠팡매핑"

    headers = ["쿠팡상품ID(sellerProductId)", "ERP품목코드", "모델명(참고)"]
    header_fill = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    all_ids = sorted(set(list(_coupang_product_map.keys()) + list(_coupang_model_map.keys())))
    for row_idx, sid in enumerate(all_ids, 2):
        ws.cell(row_idx, 1, sid)
        ws.cell(row_idx, 2, _coupang_product_map.get(sid, ""))
        ws.cell(row_idx, 3, _coupang_model_map.get(sid, ""))

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"coupang_mapping_{datetime.now(KST).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─── 매핑 엑셀 업로드 ───
@router.post("/mapping/upload")
async def upload_mapping(file: UploadFile = File(...)):
    """매핑 엑셀 파일 업로드 (기존 매핑 교체)"""
    try:
        import openpyxl
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active

        new_map = {}
        new_model = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            sid = str(row[0]).strip()
            erp_code = str(row[1] or "").strip() if len(row) > 1 else ""
            model = str(row[2] or "").strip() if len(row) > 2 else ""
            if sid and erp_code:
                new_map[sid] = erp_code
            if sid and model:
                new_model[sid] = model
        wb.close()

        global _coupang_product_map, _coupang_model_map
        _coupang_product_map = new_map
        _coupang_model_map = new_model
        _save_mapping()

        logger.info(f"[쿠팡매핑] 업로드 완료: {len(new_map)}건")
        return {"success": True, "count": len(new_map)}
    except Exception as e:
        logger.error(f"[쿠팡매핑] 업로드 실패: {e}", exc_info=True)
        return JSONResponse(status_code=500, content={"detail": str(e)})


# ─── 쿠팡 상품 → 미매핑 목록 자동 생성 ───
@router.post("/mapping/auto-generate")
async def auto_generate_mapping():
    """쿠팡 전체 상품을 가져와서 미매핑 상품을 엑셀에 자동 추가"""
    cp = _get_coupang()
    if not cp.access_key:
        return JSONResponse(status_code=400, content={"detail": "쿠팡 API 키 미설정"})

    try:
        products = await cp.fetch_products()
        added = 0
        for p in products:
            sid = str(p.get("sellerProductId", ""))
            if sid and sid not in _coupang_product_map:
                # 미매핑 상품은 상품명을 모델로 저장 (ERP코드는 비워둠)
                _coupang_model_map[sid] = p.get("sellerProductName", "")
                added += 1

        if added > 0:
            _save_mapping()

        mapped = sum(1 for sid in _coupang_product_map if _coupang_product_map.get(sid))
        total = len(set(list(_coupang_product_map.keys()) + list(_coupang_model_map.keys())))
        return {
            "success": True,
            "total_products": len(products),
            "new_added": added,
            "mapped": mapped,
            "unmapped": total - mapped,
        }
    except Exception as e:
        logger.error(f"[쿠팡매핑] 자동생성 오류: {e}", exc_info=True)
        return JSONResponse(status_code=500, content={"detail": str(e)})


# ─── 자동매칭: 쿠팡 모델명 ↔ ERP 품목코드 ───
@router.post("/mapping/auto-match")
async def auto_match_mapping():
    """쿠팡 상품의 모델명/판매자코드에서 ERP 품목코드를 자동으로 매칭

    로직:
    1) ERP 전체 품목 목록 조회
    2) 쿠팡 상품 목록 조회
    3) 쿠팡 상품명/externalVendorSku에서 ERP 품목코드와 일치하는 것을 찾아 매핑
    """
    try:
        # 1) ERP 품목 목록 가져오기
        from services.erp_client_ss import ERPClientSS
        erp = ERPClientSS()
        await erp.ensure_session()

        erp_result = await erp.get_product_list(per_page=5000)
        if not erp_result.get("success"):
            return JSONResponse(status_code=500, content={
                "detail": f"ERP 품목 조회 실패: {erp_result.get('error', '')}"
            })

        # ERP 품목코드 셋 구성
        erp_products = erp_result.get("products", [])
        erp_code_set = set()
        erp_code_to_name = {}
        for p in erp_products:
            cd = str(p.get("PROD_CD", "")).strip()
            if cd:
                erp_code_set.add(cd)
                erp_code_to_name[cd] = p.get("PROD_DES", "")

        logger.info(f"[쿠팡매칭] ERP 품목 수: {len(erp_code_set)}")

        # 페이지네이션으로 전체 ERP 품목 조회 (5000건 이상일 경우)
        total_erp = erp_result.get("total", 0)
        if total_erp > 5000:
            page = 2
            while len(erp_products) < total_erp:
                more = await erp.get_product_list(page=page, per_page=5000)
                if not more.get("success") or not more.get("products"):
                    break
                for p in more["products"]:
                    cd = str(p.get("PROD_CD", "")).strip()
                    if cd:
                        erp_code_set.add(cd)
                        erp_code_to_name[cd] = p.get("PROD_DES", "")
                erp_products.extend(more["products"])
                page += 1
            logger.info(f"[쿠팡매칭] ERP 전체 품목 수: {len(erp_code_set)}")

        # 2) 쿠팡 상품 목록 가져오기
        cp = _get_coupang()
        coupang_products = await cp.fetch_products()
        logger.info(f"[쿠팡매칭] 쿠팡 상품 수: {len(coupang_products)}")

        # 3) 자동 매칭
        matched = 0
        already_mapped = 0
        unmatched = 0

        for p in coupang_products:
            sid = str(p.get("sellerProductId", ""))
            if not sid:
                continue

            # 이미 매핑된 건 스킵
            if sid in _coupang_product_map and _coupang_product_map[sid]:
                already_mapped += 1
                continue

            # 모델명을 _coupang_model_map 또는 상품명에서 가져옴
            model = _coupang_model_map.get(sid, "") or p.get("sellerProductName", "")
            sku = str(p.get("externalVendorSku", "")).strip()

            found_code = ""

            # 방법 1: externalVendorSku가 ERP 품목코드와 직접 일치
            if sku and sku in erp_code_set:
                found_code = sku

            # 방법 2: 모델명/상품명에서 ERP 품목코드와 일치하는 부분 찾기
            if not found_code and model:
                # 모델명에서 하이픈/공백으로 분리한 토큰들을 ERP 코드와 비교
                tokens = re.split(r'[\s\-_/,()]+', model)
                for token in tokens:
                    token = token.strip()
                    if token and token in erp_code_set:
                        found_code = token
                        break

                # 모델명 전체가 ERP 코드에 포함되는지 확인
                if not found_code:
                    for erp_cd in erp_code_set:
                        if erp_cd and len(erp_cd) >= 3 and erp_cd in model:
                            found_code = erp_cd
                            break

            if found_code:
                _coupang_product_map[sid] = found_code
                if not _coupang_model_map.get(sid):
                    _coupang_model_map[sid] = model
                matched += 1
            else:
                if not _coupang_model_map.get(sid):
                    _coupang_model_map[sid] = model
                unmatched += 1

        # 저장
        if matched > 0:
            _save_mapping()

        total_mapped = sum(1 for v in _coupang_product_map.values() if v)
        total_all = len(set(list(_coupang_product_map.keys()) + list(_coupang_model_map.keys())))

        logger.info(f"[쿠팡매칭] 결과: 신규매칭={matched}, 기존매핑={already_mapped}, 미매칭={unmatched}")
        return {
            "success": True,
            "new_matched": matched,
            "already_mapped": already_mapped,
            "unmatched": unmatched,
            "total_mapped": total_mapped,
            "total_products": total_all,
            "erp_product_count": len(erp_code_set),
        }

    except Exception as e:
        logger.error(f"[쿠팡매칭] 자동매칭 오류: {e}", exc_info=True)
        return JSONResponse(status_code=500, content={"detail": str(e)})
