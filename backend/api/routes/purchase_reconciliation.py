"""
매입정산 API 라우터
- POST /api/reconcile/upload-vendor-ledger — 거래처 원장 엑셀 업로드·파싱
- POST /api/reconcile/compare             — 거래처 원장 vs ERP 데이터 비교 (AI 매칭)
- POST /api/reconcile/save-purchase        — 누락 매입전표 ERP 입력
- POST /api/reconcile/validate-purchase    — 매입전표 입력 전 유효성 검사
- GET  /api/reconcile/session/{session_id} — 이전 비교 결과 조회
- GET  /api/reconcile/download-result/{session_id} — 비교 결과를 엑셀로 다운로드
"""
import os
import re
import uuid
import asyncio
import logging
import json
import io
import time
import unicodedata
from pathlib import Path
from fastapi import APIRouter, HTTPException, UploadFile, File, Depends, Query, Form, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional

import sys
sys.path.insert(0, str(Path(__file__).parent.parent.parent))
from security import get_current_user

from config import UPLOAD_DIR, ERP_WH_CD, ERP_EMP_CD
from services.vendor_parser import parse_vendor_ledger
from services.ai_matcher import match_products_ai, check_sales_history, _is_shipping_item, _get_field
from services.erp_client import erp_client
from services.erp_web_scraper import erp_web_scraper
from services.erp_data_parser import parse_erp_purchase, parse_erp_sales
from db.database import get_connection

router = APIRouter(prefix="/api/reconcile", tags=["purchase-reconciliation"])
logger = logging.getLogger(__name__)

# 세션별 데이터 저장 (메모리)
reconcile_sessions: dict = {}

# 거래처 데이터 캐시 (메모리)
_vendor_cache: Optional[list[dict]] = None

# ── 메모리 캐시 (DB 영속 캐시의 인메모리 미러) ──
_purchase_cache: dict = {
    "items": [],
    "total": 0,
    "filename": "",
    "cached": False,
}

_sales_cache: dict = {
    "items": [],
    "total": 0,
    "filename": "",
    "cached": False,
}


# ── DB 영속 캐시 헬퍼 ──
def _save_cache_to_db(cache_key: str, filename: str, total: int, items: list[dict]):
    """메모리 캐시를 DB에 영속 저장"""
    try:
        conn = get_connection()
        data_json = json.dumps(items, ensure_ascii=False)
        # 기존 행 삭제 후 삽입 (SQLite/PG 모두 호환)
        conn.execute("DELETE FROM erp_cache WHERE cache_key = ?", (cache_key,))
        conn.execute(
            """INSERT INTO erp_cache (cache_key, filename, total, data_json, updated_at)
               VALUES (?, ?, ?, ?, datetime('now','localtime'))""",
            (cache_key, filename, total, data_json),
        )
        conn.commit()
        conn.close()
        logger.info(f"[캐시DB] '{cache_key}' 저장 완료: {total}건 ({filename})")
    except Exception as e:
        logger.error(f"[캐시DB] '{cache_key}' 저장 실패: {e}")


def _load_cache_from_db(cache_key: str) -> dict | None:
    """DB에서 캐시 데이터 로드"""
    try:
        conn = get_connection()
        row = conn.execute(
            "SELECT filename, total, data_json FROM erp_cache WHERE cache_key = ?",
            (cache_key,),
        ).fetchone()
        conn.close()
        if row:
            return {
                "filename": row[0] if isinstance(row, (list, tuple)) else row["filename"],
                "total": row[1] if isinstance(row, (list, tuple)) else row["total"],
                "items": json.loads(row[2] if isinstance(row, (list, tuple)) else row["data_json"]),
                "cached": True,
            }
    except Exception as e:
        logger.error(f"[캐시DB] '{cache_key}' 로드 실패: {e}")
    return None


def _delete_cache_from_db(cache_key: str):
    """DB에서 캐시 삭제"""
    try:
        conn = get_connection()
        conn.execute("DELETE FROM erp_cache WHERE cache_key = ?", (cache_key,))
        conn.commit()
        conn.close()
        logger.info(f"[캐시DB] '{cache_key}' 삭제 완료")
    except Exception as e:
        logger.error(f"[캐시DB] '{cache_key}' 삭제 실패: {e}")


def restore_cache_from_db():
    """서버 시작 시 DB에서 캐시 복원 (main.py에서 호출)"""
    global _purchase_cache, _sales_cache

    pc = _load_cache_from_db("purchase")
    if pc:
        _purchase_cache = pc
        logger.info(f"[캐시복원] 구매현황: {pc['total']}건 ({pc['filename']})")

    sc = _load_cache_from_db("sales")
    if sc:
        _sales_cache = sc
        logger.info(f"[캐시복원] 판매현황: {sc['total']}건 ({sc['filename']})")


# ──── 모델 ────
class ERPFetchRequest(BaseModel):
    cust_code: str = ""
    from_date: str = ""         # "20260301"
    to_date: str = ""           # "20260331"


class CompareRequest(BaseModel):
    vendor_items: list[dict]
    erp_purchase_data: list[dict]
    erp_sales_data: list[dict] = []
    use_ai: bool = True


class PurchaseItem(BaseModel):
    io_date: str               # "20260301"
    cust_code: str = ""
    cust_name: str = ""
    wh_cd: str = ""
    prod_cd: str = ""
    prod_name: str = ""
    size_des: str = ""
    qty: int = 1
    price: float = 0
    supply_amt: float = 0
    vat_amt: float = 0
    remarks: str = "매입정산 자동입력"


class PurchaseSaveRequest(BaseModel):
    items: list[PurchaseItem]
    upload_ser_no: str = "1"


# ──── 모델 (병렬 조회용) ────

class ERPFetchBothRequest(BaseModel):
    from_date: str = ""              # "20260401"
    to_date: str = ""                # "20260403"
    purchase_cust_code: str = ""     # 구매현황 거래처 필터
    sales_cust_code: str = ""        # 판매현황 거래처 필터 (보통 빈 문자열)


# ──── 엔드포인트 ────

@router.post("/fetch-erp-data")
async def fetch_erp_data(
    req: ERPFetchBothRequest,
    user: dict = Depends(get_current_user),
):
    """
    구매현황 + 판매현황을 **브라우저 탭 2개로 병렬 조회**.
    기존 순차 호출 대비 약 40~50 % 시간 절약.
    """
    try:
        result = await erp_web_scraper.get_both(
            from_date=req.from_date,
            to_date=req.to_date,
            purchase_cust_code=req.purchase_cust_code,
            sales_cust_code=req.sales_cust_code,
        )

        purchase = result.get("purchase", {})
        sales = result.get("sales", {})

        if not purchase.get("success") and not sales.get("success"):
            raise HTTPException(
                502,
                f"ERP 조회 실패 - 구매: {purchase.get('error', '?')}, 판매: {sales.get('error', '?')}"
            )

        return {
            "success": True,
            "purchase": {
                "success": purchase.get("success", False),
                "items": purchase.get("items", []),
                "total": purchase.get("total", 0),
                "error": purchase.get("error", ""),
            },
            "sales": {
                "success": sales.get("success", False),
                "items": sales.get("items", []),
                "total": sales.get("total", 0),
                "error": sales.get("error", ""),
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"ERP 병렬 조회 오류: {e}", exc_info=True)
        raise HTTPException(500, f"ERP 조회 오류: {str(e)}")


@router.post("/fetch-erp-purchases")
async def fetch_erp_purchases(
    req: ERPFetchRequest,
    user: dict = Depends(get_current_user),
):
    """ERP 웹에서 구매현황 자동 조회 (단독)"""
    try:
        result = await erp_web_scraper.get_purchase_list(
            from_date=req.from_date,
            to_date=req.to_date,
            cust_code=req.cust_code,
        )
        if not result.get("success"):
            raise HTTPException(
                502,
                f"ERP 구매현황 조회 실패: {result.get('error', '알 수 없는 오류')}"
            )
        return {
            "success": True,
            "items": result["items"],
            "total": result["total"],
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"ERP 구매현황 조회 오류: {e}", exc_info=True)
        raise HTTPException(500, f"ERP 조회 오류: {str(e)}")


@router.post("/fetch-erp-sales")
async def fetch_erp_sales(
    req: ERPFetchRequest,
    user: dict = Depends(get_current_user),
):
    """ERP 웹에서 판매현황 자동 조회 (단독)"""
    try:
        result = await erp_web_scraper.get_sales_list(
            from_date=req.from_date,
            to_date=req.to_date,
            cust_code=req.cust_code,
        )
        if not result.get("success"):
            raise HTTPException(
                502,
                f"ERP 판매현황 조회 실패: {result.get('error', '알 수 없는 오류')}"
            )
        return {
            "success": True,
            "items": result["items"],
            "total": result["total"],
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"ERP 판매현황 조회 오류: {e}", exc_info=True)
        raise HTTPException(500, f"ERP 조회 오류: {str(e)}")


@router.post("/upload-vendor-ledger")
async def upload_vendor_ledger(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
):
    """거래처 원장 엑셀 업로드 및 파싱"""
    file_ext = os.path.splitext(file.filename)[1]
    if file_ext.lower() not in (".xlsx", ".xls"):
        raise HTTPException(400, "엑셀 파일(.xlsx)만 업로드 가능합니다")

    os.makedirs(str(UPLOAD_DIR), exist_ok=True)
    saved_name = f"reconcile_{uuid.uuid4().hex[:8]}{file_ext}"
    saved_path = os.path.join(str(UPLOAD_DIR), saved_name)

    with open(saved_path, "wb") as f:
        content = await file.read()
        f.write(content)

    try:
        result = parse_vendor_ledger(saved_path, original_filename=file.filename)
        result["file_id"] = saved_name
        result["original_filename"] = file.filename
        return result
    except Exception as e:
        logger.error(f"거래처 원장 파싱 실패: {e}")
        raise HTTPException(500, f"파싱 실패: {str(e)}")


@router.post("/upload-erp-data")
async def upload_erp_data(
    purchase_file: Optional[UploadFile] = File(None),
    sales_file: Optional[UploadFile] = File(None),
    user: dict = Depends(get_current_user),
):
    """ERP 구매현황/판매현황 엑셀·CSV 업로드 및 파싱

    - purchase_file: 구매현황 엑셀(.xlsx) or CSV
    - sales_file: 판매현황 CSV or 엑셀(.xlsx)
    - 둘 중 하나만 업로드해도 됨
    - 판매현황은 한 번 업로드하면 서버 메모리에 캐시되어, 새 파일을 업로드하거나
      DELETE /clear-sales-cache를 호출하기 전까지 유지됨
    """
    global _sales_cache

    result = {
        "purchase": {"success": False, "items": [], "total": 0, "error": ""},
        "sales": {"success": False, "items": [], "total": 0, "error": ""},
    }

    os.makedirs(str(UPLOAD_DIR), exist_ok=True)

    # 구매현황 파싱
    if purchase_file and purchase_file.filename:
        try:
            ext = os.path.splitext(purchase_file.filename)[1]
            saved = os.path.join(str(UPLOAD_DIR), f"erp_purchase_{uuid.uuid4().hex[:8]}{ext}")
            with open(saved, "wb") as f:
                f.write(await purchase_file.read())
            data = parse_erp_purchase(saved)
            result["purchase"] = {
                "success": True,
                "items": data["items"],
                "total": data["total"],
                "meta": data.get("meta", ""),
                "error": "",
            }
            logger.info(f"구매현황 파싱 완료: {data['total']}건")
        except Exception as e:
            logger.error(f"구매현황 파싱 실패: {e}", exc_info=True)
            result["purchase"]["error"] = str(e)

    # 판매현황 파싱 (새 파일이 있으면 파싱 후 캐시 갱신)
    if sales_file and sales_file.filename:
        try:
            ext = os.path.splitext(sales_file.filename)[1]
            saved = os.path.join(str(UPLOAD_DIR), f"erp_sales_{uuid.uuid4().hex[:8]}{ext}")
            with open(saved, "wb") as f:
                f.write(await sales_file.read())
            data = parse_erp_sales(saved)
            # 캐시 갱신 (메모리 + DB)
            _sales_cache = {
                "items": data["items"],
                "total": data["total"],
                "filename": sales_file.filename,
                "cached": True,
            }
            _save_cache_to_db("sales", sales_file.filename, data["total"], data["items"])
            result["sales"] = {
                "success": True,
                "items": data["items"],
                "total": data["total"],
                "filename": sales_file.filename,
                "error": "",
            }
            logger.info(f"판매현황 파싱 완료 (캐시 갱신): {data['total']}건")
        except Exception as e:
            logger.error(f"판매현황 파싱 실패: {e}", exc_info=True)
            result["sales"]["error"] = str(e)
    elif _sales_cache["cached"]:
        # 새 파일 없지만 캐시가 있으면 캐시 데이터 반환
        result["sales"] = {
            "success": True,
            "items": _sales_cache["items"],
            "total": _sales_cache["total"],
            "filename": _sales_cache["filename"],
            "from_cache": True,
            "error": "",
        }
        logger.info(f"판매현황 캐시 사용: {_sales_cache['total']}건 ({_sales_cache['filename']})")

    return result


@router.get("/erp-cache-status")
async def get_erp_cache_status(
    user: dict = Depends(get_current_user),
):
    """구매현황 + 판매현황 캐시 상태 조회"""
    return {
        "purchase": {
            "cached": _purchase_cache["cached"],
            "total": _purchase_cache["total"],
            "filename": _purchase_cache["filename"],
        },
        "sales": {
            "cached": _sales_cache["cached"],
            "total": _sales_cache["total"],
            "filename": _sales_cache["filename"],
        },
    }


@router.delete("/clear-purchase-cache")
async def clear_purchase_cache(
    user: dict = Depends(get_current_user),
):
    """구매현황 캐시 삭제 (메모리 + DB)"""
    global _purchase_cache
    old = _purchase_cache["filename"]
    _purchase_cache = {"items": [], "total": 0, "filename": "", "cached": False}
    _delete_cache_from_db("purchase")
    logger.info(f"구매현황 캐시 삭제됨 (이전: {old})")
    return {"success": True, "message": f"구매현황 캐시 삭제됨 ({old})"}


@router.delete("/clear-sales-cache")
async def clear_sales_cache(
    user: dict = Depends(get_current_user),
):
    """판매현황 캐시 삭제 (메모리 + DB)"""
    global _sales_cache
    old = _sales_cache["filename"]
    _sales_cache = {"items": [], "total": 0, "filename": "", "cached": False}
    _delete_cache_from_db("sales")
    logger.info(f"판매현황 캐시 삭제됨 (이전: {old})")
    return {"success": True, "message": f"판매현황 캐시 삭제됨 ({old})"}


# ──── 일괄 처리 (거래처원장 여러 개 + 구매현황 + 판매현황 한 번에) ────

def _match_vendor_to_purchase(vendor_name: str, purchase_items: list[dict]) -> list[dict]:
    """거래처명으로 전체 구매현황에서 해당 거래처 항목만 필터링.

    vendor_name(거래처 원장의 시트명/파일명)과 ERP 구매현황의 cust_name을 매칭.
    부분 문자열 포함(예: '유니정보' in '(주)유니정보통신')도 허용.
    """
    if not vendor_name:
        return []

    # macOS NFD→NFC 정규화
    vn = unicodedata.normalize("NFC", vendor_name)
    v_clean = vn.replace("(주)", "").replace("주식회사", "").strip()

    exact = [p for p in purchase_items if unicodedata.normalize("NFC", p.get("cust_name", "")) == vn]
    if exact:
        return exact

    partial = [
        p for p in purchase_items
        if v_clean and (
            v_clean in unicodedata.normalize("NFC", p.get("cust_name", "")).replace("(주)", "").replace("주식회사", "")
            or unicodedata.normalize("NFC", p.get("cust_name", "")).replace("(주)", "").replace("주식회사", "") in v_clean
        )
    ]
    return partial


def _find_vendor_code(vendor_name: str) -> str:
    """vendors.json 캐시에서 거래처코드 찾기"""
    global _vendor_cache
    if not _vendor_cache or not vendor_name:
        return ""
    vn = unicodedata.normalize("NFC", vendor_name)
    v_clean = vn.replace("(주)", "").replace("주식회사", "").strip()
    for v in _vendor_cache:
        name = unicodedata.normalize("NFC", v.get("name", ""))
        if name == vn:
            return v.get("code", "")
        n_clean = name.replace("(주)", "").replace("주식회사", "").strip()
        if v_clean and (v_clean in n_clean or n_clean in v_clean):
            return v.get("code", "")
    return ""


@router.post("/batch-reconcile")
async def batch_reconcile(
    vendor_files: list[UploadFile] = File(...),
    purchase_file: Optional[UploadFile] = File(None),
    sales_file: Optional[UploadFile] = File(None),
    user: dict = Depends(get_current_user),
):
    """거래처원장 여러 개 + 구매현황(전체) + 판매현황을 한 번에 업로드하여 일괄 정산

    - vendor_files: 거래처 원장 엑셀 파일 여러 개 (파일명에 거래처명 포함)
    - purchase_file: 전체 구매현황 엑셀/CSV (거래처 필터 없이 전체)
    - sales_file: 판매현황 CSV/엑셀 (없으면 캐시 사용)
    """
    global _purchase_cache, _sales_cache

    os.makedirs(str(UPLOAD_DIR), exist_ok=True)

    # ── 1. 구매현황 파싱 (새 파일 있으면 캐시 갱신, 없으면 캐시 사용) ──
    all_purchase_items = []
    purchase_info = {"filename": "", "from_cache": False}
    if purchase_file and purchase_file.filename:
        try:
            ext = os.path.splitext(purchase_file.filename)[1]
            saved = os.path.join(str(UPLOAD_DIR), f"erp_purchase_{uuid.uuid4().hex[:8]}{ext}")
            with open(saved, "wb") as f:
                f.write(await purchase_file.read())
            data = parse_erp_purchase(saved)
            all_purchase_items = data["items"]
            _purchase_cache = {
                "items": data["items"],
                "total": data["total"],
                "filename": purchase_file.filename,
                "cached": True,
            }
            _save_cache_to_db("purchase", purchase_file.filename, data["total"], data["items"])
            purchase_info = {"filename": purchase_file.filename, "from_cache": False}
            logger.info(f"[일괄] 구매현황 파싱 (캐시 갱신): {len(all_purchase_items)}건")
        except Exception as e:
            logger.error(f"[일괄] 구매현황 파싱 실패: {e}", exc_info=True)
            raise HTTPException(400, f"구매현황 파싱 실패: {str(e)}")
    elif _purchase_cache["cached"]:
        all_purchase_items = _purchase_cache["items"]
        purchase_info = {"filename": _purchase_cache["filename"], "from_cache": True}
        logger.info(f"[일괄] 구매현황 캐시 사용: {len(all_purchase_items)}건")

    # ── 2. 판매현황 파싱 (새 파일 있으면 캐시 갱신, 없으면 캐시 사용) ──
    all_sales_items = []
    sales_info = {"filename": "", "from_cache": False}
    if sales_file and sales_file.filename:
        try:
            ext = os.path.splitext(sales_file.filename)[1]
            saved = os.path.join(str(UPLOAD_DIR), f"erp_sales_{uuid.uuid4().hex[:8]}{ext}")
            with open(saved, "wb") as f:
                f.write(await sales_file.read())
            data = parse_erp_sales(saved)
            all_sales_items = data["items"]
            _sales_cache = {
                "items": data["items"],
                "total": data["total"],
                "filename": sales_file.filename,
                "cached": True,
            }
            _save_cache_to_db("sales", sales_file.filename, data["total"], data["items"])
            sales_info = {"filename": sales_file.filename, "from_cache": False}
            logger.info(f"[일괄] 판매현황 파싱 (캐시 갱신): {len(all_sales_items)}건")
        except Exception as e:
            logger.error(f"[일괄] 판매현황 파싱 실패: {e}", exc_info=True)
    elif _sales_cache["cached"]:
        all_sales_items = _sales_cache["items"]
        sales_info = {"filename": _sales_cache["filename"], "from_cache": True}
        logger.info(f"[일괄] 판매현황 캐시 사용: {len(all_sales_items)}건")

    # ── 3. 거래처 원장들 파싱 + 매칭 ──
    vendor_results = []

    # vendors.json 미리 로드
    global _vendor_cache
    if _vendor_cache is None:
        try:
            vendor_path = Path(__file__).parent.parent.parent.parent / "data" / "vendors.json"
            if vendor_path.exists():
                with open(vendor_path, "r", encoding="utf-8") as f:
                    _vendor_cache = json.load(f)
        except Exception:
            pass

    for vf in vendor_files:
        if not vf.filename:
            continue

        vf_result = {
            "filename": vf.filename,
            "vendor_name": "",
            "vendor_code": "",
            "error": "",
            "summary": {},
            "matched": [],
            "unmatched": [],
            "sales_check": [],
            "shipping_items": [],
            "amount_mismatches": [],
        }

        try:
            # 파일 저장 및 파싱
            ext = os.path.splitext(vf.filename)[1]
            saved = os.path.join(str(UPLOAD_DIR), f"vendor_{uuid.uuid4().hex[:8]}{ext}")
            with open(saved, "wb") as f:
                f.write(await vf.read())

            ledger = parse_vendor_ledger(saved, original_filename=vf.filename)
            vendor_name = ledger.get("vendor_name", "")

            # 파일명에서 거래처명 추출 시도 (시트명보다 파일명이 더 정확할 수 있음)
            if not vendor_name or vendor_name == "Sheet1" or vendor_name == "Sheet":
                fn_base = os.path.splitext(vf.filename)[0]
                # UUID suffix 제거 (드래그앤드롭 업로드 시 "-79021bc2" 같은 suffix 붙음)
                fn_base = re.sub(r'-[0-9a-f]{6,12}$', '', fn_base)
                # 흔한 패턴: "거래장부내역_유니정보통신", "202603_파워네트", "거래내역_파워네트정보통신" 등
                parts = fn_base.replace("-", "_").split("_")
                # 숫자만인 파트(날짜 등)를 제외하고 마지막 한글/영문 파트를 거래처명으로
                name_parts = [p for p in parts if p and not re.match(r'^\d+$', p)]
                # "거래장부내역", "거래내역", "거래확인서" 같은 공통 접두어 제외
                skip_prefixes = {"거래장부내역", "거래내역", "거래확인서", "거래원장", "원장"}
                name_parts = [p for p in name_parts if p not in skip_prefixes]
                vendor_name = name_parts[-1] if name_parts else fn_base

            vf_result["vendor_name"] = vendor_name
            vf_result["vendor_code"] = _find_vendor_code(vendor_name)

            # 거래처 원장에서 "매출" 항목 = 우리의 매입
            all_vendor_items = ledger.get("sales_items", [])
            if not all_vendor_items:
                # tx_type 없는 경우 전체 사용
                all_vendor_items = ledger.get("transactions", [])
                all_vendor_items = [
                    item for item in all_vendor_items
                    if item.get("tx_type") == "매출" or not item.get("tx_type")
                ]

            if not all_vendor_items:
                vf_result["error"] = "거래처 원장에 매출 항목이 없습니다"
                vendor_results.append(vf_result)
                continue

            # 메모성 항목 제외 (수량=0 & 금액=0인 "출고됨", "직수" 등)
            memo_keywords = {"출고됨", "직수", "이전잔액", "소계", "합계"}
            memo_items = [
                item for item in all_vendor_items
                if (not item.get("qty") and not item.get("amount"))
                or str(item.get("product_name", "")).strip() in memo_keywords
            ]
            real_items = [item for item in all_vendor_items if item not in memo_items]

            # 배송료 분리
            shipping_items = [item for item in real_items if _is_shipping_item(item)]
            regular_items = [item for item in real_items if not _is_shipping_item(item)]

            # 구매현황에서 해당 거래처만 필터
            filtered_purchase = _match_vendor_to_purchase(vendor_name, all_purchase_items)
            logger.info(f"[일괄] '{vendor_name}': 원장 {len(all_vendor_items)}건, 구매현황 {len(filtered_purchase)}건")

            # 1단계: 규칙 기반 매칭 (날짜/금액/수량)
            match_results = await match_products_ai(regular_items, filtered_purchase)

            matched = [r for r in match_results if r["match_type"] != "unmatched"]
            unmatched = [r for r in match_results if r["match_type"] == "unmatched"]

            # 금액 차이 감지
            amount_mismatches = []
            for r in matched:
                v = r.get("vendor_item", {})
                e = r.get("erp_match", {})
                v_amt = float(v.get("amount", 0) or 0)
                try:
                    e_amt = float(str(_get_field(e, "total", "합계", default=0) or 0).replace(",", ""))
                except (ValueError, TypeError):
                    e_amt = 0
                v_qty = int(v.get("qty", 0) or 0)
                try:
                    e_qty = int(float(str(_get_field(e, "qty", "수량", default=0) or 0).replace(",", "")))
                except (ValueError, TypeError):
                    e_qty = 0

                if v_amt and e_amt and abs(v_amt - e_amt) > 1:
                    r["amount_diff"] = v_amt - e_amt
                    r["amount_diff_pct"] = round((v_amt - e_amt) / max(v_amt, 1) * 100, 1)
                    r["vendor_amount"] = v_amt
                    r["erp_amount"] = e_amt
                    amount_mismatches.append(r)
                if v_qty and e_qty and v_qty != e_qty:
                    r["qty_diff"] = v_qty - e_qty

            # 2단계: 누락 건 → 판매이력 확인 (AI)
            sales_check = []
            if unmatched and all_sales_items:
                unmatched_vendor_items = [r["vendor_item"] for r in unmatched]
                sales_check = await check_sales_history(
                    unmatched_vendor_items, all_sales_items
                )

            # 배송료 매칭
            shipping_match_results = []
            if shipping_items and filtered_purchase:
                shipping_match_results = await match_products_ai(
                    shipping_items, filtered_purchase
                )

            vf_result["matched"] = matched
            vf_result["unmatched"] = unmatched
            vf_result["sales_check"] = sales_check
            vf_result["shipping_items"] = [
                {
                    "vendor_item": item,
                    "erp_match": next((r.get("erp_match") for r in shipping_match_results
                                       if r.get("vendor_item") == item and r.get("match_type") != "unmatched"), None),
                    "match_type": next((r.get("match_type") for r in shipping_match_results
                                        if r.get("vendor_item") == item and r.get("match_type") != "unmatched"), "unmatched"),
                }
                for item in shipping_items
            ]
            vf_result["amount_mismatches"] = amount_mismatches
            vf_result["memo_items"] = [item for item in memo_items]
            vf_result["summary"] = {
                "total_vendor_items": len(all_vendor_items),
                "memo_filtered": len(memo_items),
                "matched_count": len(matched),
                "unmatched_count": len(unmatched),
                "with_sales_history": sum(1 for s in sales_check if s.get("has_sales_history")),
                "shipping_count": len(shipping_items),
                "amount_mismatch_count": len(amount_mismatches),
                "purchase_filtered": len(filtered_purchase),
            }

        except Exception as e:
            logger.error(f"[일괄] '{vf.filename}' 처리 실패: {e}", exc_info=True)
            vf_result["error"] = str(e)

        vendor_results.append(vf_result)

    # ── 4. 세션 저장 & 응답 ──
    session_id = uuid.uuid4().hex[:12]
    reconcile_sessions[session_id] = {
        "batch": True,
        "vendor_results": vendor_results,
        "purchase_total": len(all_purchase_items),
        "sales_total": len(all_sales_items),
        "purchase_info": purchase_info,
        "sales_info": sales_info,
    }

    # 전체 요약
    total_summary = {
        "vendor_count": len(vendor_results),
        "total_matched": sum(r["summary"].get("matched_count", 0) for r in vendor_results),
        "total_unmatched": sum(r["summary"].get("unmatched_count", 0) for r in vendor_results),
        "total_shipping": sum(r["summary"].get("shipping_count", 0) for r in vendor_results),
        "total_amount_mismatch": sum(r["summary"].get("amount_mismatch_count", 0) for r in vendor_results),
        "purchase_total": len(all_purchase_items),
        "purchase_from_cache": purchase_info["from_cache"],
        "sales_total": len(all_sales_items),
        "sales_from_cache": sales_info["from_cache"],
        "errors": [r["filename"] for r in vendor_results if r.get("error")],
    }

    return {
        "session_id": session_id,
        "summary": total_summary,
        "vendor_results": vendor_results,
    }


@router.post("/preview-vendors")
async def preview_vendors(
    vendor_files: list[UploadFile] = File(...),
    user: dict = Depends(get_current_user),
):
    """거래처 원장 파일들을 업로드하여 거래처명만 미리 추출 (확인용)

    파일을 서버에 저장하고 거래처명을 추출하여 반환.
    사용자가 확인/수정 후 batch-reconcile-stream에서 사용.
    """
    os.makedirs(str(UPLOAD_DIR), exist_ok=True)
    previews = []

    for vf in vendor_files:
        if not vf.filename:
            continue
        try:
            ext = os.path.splitext(vf.filename)[1]
            file_id = uuid.uuid4().hex[:8]
            saved = os.path.join(str(UPLOAD_DIR), f"vendor_{file_id}{ext}")
            with open(saved, "wb") as f:
                f.write(await vf.read())

            # 거래처명 추출 (원본 파일명 전달)
            ledger = parse_vendor_ledger(saved, original_filename=vf.filename)
            vendor_name = ledger.get("vendor_name", "")

            # 시트명이 generic이면 파일명에서 추출
            if not vendor_name or vendor_name in ("Sheet1", "Sheet", ""):
                fn_base = os.path.splitext(vf.filename)[0]
                fn_base = re.sub(r'-[0-9a-f]{6,12}$', '', fn_base)
                parts = fn_base.replace("-", "_").split("_")
                name_parts = [p for p in parts if p and not re.match(r'^\d+$', p)]
                skip_prefixes = {"거래장부내역", "거래내역", "거래확인서", "거래원장", "원장"}
                name_parts = [p for p in name_parts if p not in skip_prefixes]
                vendor_name = name_parts[-1] if name_parts else fn_base

            item_count = len(ledger.get("sales_items", []) or ledger.get("transactions", []))
            date_from = ledger.get("date_from", "")
            date_to = ledger.get("date_to", "")
            previews.append({
                "file_id": file_id,
                "original_filename": vf.filename,
                "saved_path": saved,
                "vendor_name": vendor_name,
                "item_count": item_count,
                "date_from": date_from,
                "date_to": date_to,
            })
        except Exception as e:
            previews.append({
                "file_id": "",
                "original_filename": vf.filename,
                "saved_path": "",
                "vendor_name": "",
                "item_count": 0,
                "date_from": "",
                "date_to": "",
                "error": str(e),
            })

    return {"previews": previews}


@router.post("/batch-reconcile-stream")
async def batch_reconcile_stream(
    request: Request,
    vendor_files: list[UploadFile] = File(None),
    purchase_file: Optional[UploadFile] = File(None),
    sales_file: Optional[UploadFile] = File(None),
    vendor_names_json: str = Form("[]"),
    vendor_codes_json: str = Form("[]"),
    saved_paths_json: str = Form("[]"),
    user: dict = Depends(get_current_user),
):
    """SSE 스트리밍 일괄 정산 — 실시간 진행 로그를 전송

    vendor_names_json: 사용자가 확인한 거래처명 배열 '["파워네트정보통신(주)","(주)현대모아컴"]'
    vendor_codes_json: 거래처코드 배열 '["1068637925","1068187401"]'
    saved_paths_json: preview-vendors에서 반환된 파일경로 배열
    """
    global _purchase_cache, _sales_cache

    confirmed_names = json.loads(vendor_names_json)
    confirmed_codes = json.loads(vendor_codes_json)
    saved_paths = json.loads(saved_paths_json)

    os.makedirs(str(UPLOAD_DIR), exist_ok=True)

    async def event_stream():
        global _purchase_cache, _sales_cache

        def sse(event: str, data: dict):
            return f"event: {event}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n"

        t0 = time.time()

        # ── 1. 구매현황 ──
        all_purchase_items = []
        purchase_info = {"filename": "", "from_cache": False}
        if purchase_file and purchase_file.filename:
            yield sse("log", {"msg": f"📂 구매현황 파싱 중: {purchase_file.filename}"})
            try:
                ext = os.path.splitext(purchase_file.filename)[1]
                saved = os.path.join(str(UPLOAD_DIR), f"erp_purchase_{uuid.uuid4().hex[:8]}{ext}")
                content = await purchase_file.read()
                with open(saved, "wb") as f:
                    f.write(content)
                data = parse_erp_purchase(saved)
                all_purchase_items = data["items"]
                _purchase_cache = {
                    "items": data["items"], "total": data["total"],
                    "filename": purchase_file.filename, "cached": True,
                }
                _save_cache_to_db("purchase", purchase_file.filename, data["total"], data["items"])
                purchase_info = {"filename": purchase_file.filename, "from_cache": False}
                yield sse("log", {"msg": f"✅ 구매현황 {data['total']}건 로드"})
            except Exception as e:
                yield sse("log", {"msg": f"❌ 구매현황 파싱 실패: {e}", "level": "error"})
        elif _purchase_cache["cached"]:
            all_purchase_items = _purchase_cache["items"]
            purchase_info = {"filename": _purchase_cache["filename"], "from_cache": True}
            yield sse("log", {"msg": f"📦 구매현황 캐시 사용: {_purchase_cache['total']}건 ({_purchase_cache['filename']})"})

        # ── 2. 판매현황 ──
        all_sales_items = []
        sales_info = {"filename": "", "from_cache": False}
        if sales_file and sales_file.filename:
            yield sse("log", {"msg": f"📂 판매현황 파싱 중: {sales_file.filename}"})
            try:
                ext = os.path.splitext(sales_file.filename)[1]
                saved = os.path.join(str(UPLOAD_DIR), f"erp_sales_{uuid.uuid4().hex[:8]}{ext}")
                content = await sales_file.read()
                with open(saved, "wb") as f:
                    f.write(content)
                data = parse_erp_sales(saved)
                all_sales_items = data["items"]
                _sales_cache = {
                    "items": data["items"], "total": data["total"],
                    "filename": sales_file.filename, "cached": True,
                }
                _save_cache_to_db("sales", sales_file.filename, data["total"], data["items"])
                sales_info = {"filename": sales_file.filename, "from_cache": False}
                yield sse("log", {"msg": f"✅ 판매현황 {data['total']}건 로드"})
            except Exception as e:
                yield sse("log", {"msg": f"❌ 판매현황 파싱 실패: {e}", "level": "error"})
        elif _sales_cache["cached"]:
            all_sales_items = _sales_cache["items"]
            sales_info = {"filename": _sales_cache["filename"], "from_cache": True}
            yield sse("log", {"msg": f"📦 판매현황 캐시 사용: {_sales_cache['total']}건 ({_sales_cache['filename']})"})

        # vendors.json 로드
        global _vendor_cache
        if _vendor_cache is None:
            try:
                vendor_path = Path(__file__).parent.parent.parent.parent / "data" / "vendors.json"
                if vendor_path.exists():
                    with open(vendor_path, "r", encoding="utf-8") as f:
                        _vendor_cache = json.load(f)
            except Exception:
                pass

        # ── 3. 거래처 원장 처리 ──
        vendor_results = []

        # 파일 목록 결정: saved_paths가 있으면 서버에 이미 저장된 파일 사용, 없으면 vendor_files 사용
        file_entries = []
        if saved_paths:
            for i, spath in enumerate(saved_paths):
                name = confirmed_names[i] if i < len(confirmed_names) else ""
                code = confirmed_codes[i] if i < len(confirmed_codes) else ""
                file_entries.append({"path": spath, "vendor_name": name, "vendor_code": code, "filename": os.path.basename(spath)})
        elif vendor_files:
            vi_idx = 0
            for i, vf in enumerate(vendor_files):
                if not vf.filename:
                    continue
                ext = os.path.splitext(vf.filename)[1]
                fid = uuid.uuid4().hex[:8]
                spath = os.path.join(str(UPLOAD_DIR), f"vendor_{fid}{ext}")
                content = await vf.read()
                with open(spath, "wb") as f:
                    f.write(content)
                name = confirmed_names[vi_idx] if vi_idx < len(confirmed_names) else ""
                code = confirmed_codes[vi_idx] if vi_idx < len(confirmed_codes) else ""
                file_entries.append({"path": spath, "vendor_name": name, "vendor_code": code, "filename": vf.filename})
                vi_idx += 1

        total_vendors = len(file_entries)
        yield sse("log", {"msg": f"🔍 거래처 {total_vendors}건 정산 시작"})
        yield sse("progress", {"current": 0, "total": total_vendors})

        for vi, entry in enumerate(file_entries):
            saved_path = entry["path"]
            vendor_name = entry["vendor_name"]
            vendor_code = entry.get("vendor_code", "")
            orig_filename = entry["filename"]

            vf_result = {
                "filename": orig_filename,
                "vendor_name": vendor_name,
                "vendor_code": vendor_code,
                "error": "",
                "summary": {},
                "matched": [], "unmatched": [], "sales_check": [],
                "shipping_items": [], "amount_mismatches": [],
            }

            try:
                yield sse("log", {"msg": f"📋 [{vi+1}/{total_vendors}] {vendor_name} 원장 파싱 중..."})

                ledger = parse_vendor_ledger(saved_path, original_filename=orig_filename)

                # 확인된 거래처명이 없으면 파싱 결과 사용
                if not vendor_name:
                    vendor_name = ledger.get("vendor_name", "")
                    if not vendor_name or vendor_name in ("Sheet1", "Sheet"):
                        fn_base = os.path.splitext(orig_filename)[0]
                        fn_base = re.sub(r'-[0-9a-f]{6,12}$', '', fn_base)
                        parts = fn_base.replace("-", "_").split("_")
                        name_parts = [p for p in parts if p and not re.match(r'^\d+$', p)]
                        skip_prefixes = {"거래장부내역", "거래내역", "거래확인서", "거래원장", "원장"}
                        name_parts = [p for p in name_parts if p not in skip_prefixes]
                        vendor_name = name_parts[-1] if name_parts else fn_base

                vf_result["vendor_name"] = vendor_name
                vf_result["vendor_code"] = vendor_code or _find_vendor_code(vendor_name)

                # 거래처 원장 항목 추출
                all_vendor_items = ledger.get("sales_items", [])
                if not all_vendor_items:
                    all_vendor_items = ledger.get("transactions", [])
                    all_vendor_items = [
                        item for item in all_vendor_items
                        if item.get("tx_type") == "매출" or not item.get("tx_type")
                    ]

                if not all_vendor_items:
                    vf_result["error"] = "거래처 원장에 매출 항목이 없습니다"
                    vendor_results.append(vf_result)
                    yield sse("log", {"msg": f"⚠️ {vendor_name}: 매출 항목 없음", "level": "warn"})
                    continue

                # 메모성 항목 제외
                memo_keywords = {"출고됨", "직수", "이전잔액", "소계", "합계"}
                memo_items = [
                    item for item in all_vendor_items
                    if (not item.get("qty") and not item.get("amount"))
                    or str(item.get("product_name", "")).strip() in memo_keywords
                ]
                real_items = [item for item in all_vendor_items if item not in memo_items]
                shipping_items = [item for item in real_items if _is_shipping_item(item)]
                regular_items = [item for item in real_items if not _is_shipping_item(item)]

                yield sse("log", {"msg": f"  → 전체 {len(all_vendor_items)}건 (실거래 {len(regular_items)}, 배송 {len(shipping_items)}, 메모 {len(memo_items)})"})

                # 구매현황에서 해당 거래처 필터
                filtered_purchase = _match_vendor_to_purchase(vendor_name, all_purchase_items)
                yield sse("log", {"msg": f"  → 구매현황 매칭: {len(filtered_purchase)}건 (거래처: {vendor_name})"})

                if not filtered_purchase and all_purchase_items:
                    # 디버깅: 구매현황의 거래처 목록 표시
                    cust_names = list(set(p.get("cust_name", "") for p in all_purchase_items if p.get("cust_name")))
                    yield sse("log", {"msg": f"  ⚠️ 구매현황에서 '{vendor_name}' 못 찾음. 등록된 거래처: {', '.join(cust_names[:10])}{'...' if len(cust_names) > 10 else ''}", "level": "warn"})

                # ── 1단계: 규칙 기반 매칭 ──
                yield sse("log", {"msg": f"  → 1단계: 규칙 기반 매칭 (날짜/금액/수량)..."})
                match_results = await match_products_ai(regular_items, filtered_purchase)

                matched = [r for r in match_results if r["match_type"] != "unmatched"]
                unmatched = [r for r in match_results if r["match_type"] == "unmatched"]

                yield sse("log", {"msg": f"  → 1단계 결과: 매칭 {len(matched)}건, 미매칭 {len(unmatched)}건"})

                # 금액 차이 감지
                amount_mismatches = []
                for r in matched:
                    v = r.get("vendor_item", {})
                    e = r.get("erp_match", {})
                    v_amt = float(v.get("amount", 0) or 0)
                    try:
                        e_amt = float(str(_get_field(e, "total", "합계", default=0) or 0).replace(",", ""))
                    except (ValueError, TypeError):
                        e_amt = 0
                    v_qty = int(v.get("qty", 0) or 0)
                    try:
                        e_qty = int(float(str(_get_field(e, "qty", "수량", default=0) or 0).replace(",", "")))
                    except (ValueError, TypeError):
                        e_qty = 0
                    if v_amt and e_amt and abs(v_amt - e_amt) > 1:
                        r["amount_diff"] = v_amt - e_amt
                        r["amount_diff_pct"] = round((v_amt - e_amt) / max(v_amt, 1) * 100, 1)
                        r["vendor_amount"] = v_amt
                        r["erp_amount"] = e_amt
                        amount_mismatches.append(r)
                    if v_qty and e_qty and v_qty != e_qty:
                        r["qty_diff"] = v_qty - e_qty

                # 거래처별 총액 비교 (거래처원장 매입총액 vs ERP 구매현황 매입총액)
                vendor_ledger_total = sum(
                    float(item.get("amount", 0) or 0) for item in regular_items
                )
                erp_purchase_total = sum(
                    float(str(_get_field(p, "total", "합계", default=0) or 0).replace(",", ""))
                    for p in filtered_purchase
                )
                vendor_total_match = abs(vendor_ledger_total - erp_purchase_total) <= 1

                # ── 할인 흡수 처리 ──
                # 거래처원장: 상품 354,000 + 매출할인 -12,000 = 342,000
                # ERP 구매현황: 상품 342,000 (할인된 가격으로 입력)
                # → 매칭된 항목의 금액차이가 미매칭 할인항목 합계와 일치하면 할인 흡수
                discount_keywords = {"매출할인", "할인", "리베이트", "DC", "할인DC", "에누리"}
                discount_absorbed = []
                still_unmatched = []

                if unmatched:
                    from collections import defaultdict

                    # 미매칭 항목을 할인성/비할인성으로 분류
                    unmatched_discounts_by_date = defaultdict(list)  # 날짜별 할인 항목
                    unmatched_regular = []  # 비할인 미매칭 항목

                    for r in unmatched:
                        v = r.get("vendor_item", {})
                        pname = (v.get("product_name", "") or "").strip()
                        pcat = (v.get("product_category", "") or "").strip()
                        is_discount = any(kw in pname for kw in discount_keywords) or \
                                      any(kw in pcat for kw in discount_keywords)
                        if is_discount:
                            d = v.get("date", "")
                            unmatched_discounts_by_date[d].append(r)
                        else:
                            unmatched_regular.append(r)

                    # 금액불일치 항목에서 할인 흡수 시도
                    absorbed_discount_ids = set()
                    resolved_mismatch_indices = []

                    for mi_idx, r in enumerate(amount_mismatches):
                        v = r.get("vendor_item", {})
                        e = r.get("erp_match", {})
                        v_date = v.get("date", "")
                        v_amt = float(v.get("amount", 0) or 0)
                        e_amt = float(str(_get_field(e, "total", "합계", default=0) or 0).replace(",", ""))
                        diff = v_amt - e_amt  # 양수 = 원장이 더 큼 (할인이 있을 것)

                        # 같은 날짜의 할인 항목 확인
                        date_discounts = unmatched_discounts_by_date.get(v_date, [])
                        if not date_discounts or abs(diff) <= 1:
                            continue

                        # 할인 합계가 금액차이를 설명하는지 확인
                        discount_total = sum(
                            float(dr.get("vendor_item", {}).get("amount", 0) or 0)
                            for dr in date_discounts
                            if id(dr) not in absorbed_discount_ids
                        )
                        # v_amt + discount_total ≈ e_amt → 할인이 단가에 반영됨
                        adjusted = v_amt + discount_total
                        if abs(adjusted - e_amt) <= max(abs(e_amt) * 0.01, 10):
                            # 할인 흡수 성공
                            for dr in date_discounts:
                                if id(dr) not in absorbed_discount_ids:
                                    absorbed_discount_ids.add(id(dr))
                                    dr["match_type"] = "discount_absorbed"
                                    dv = dr.get("vendor_item", {})
                                    dr["reason"] = f"할인({dv.get('product_name', '')}) → {v.get('product_name', '')}에 단가 반영됨"
                                    dr["absorbed_by"] = v.get("product_name", "")
                                    discount_absorbed.append(dr)
                            # 금액불일치도 해소
                            resolved_mismatch_indices.append(mi_idx)

                    # 해소된 금액불일치 제거
                    for idx in sorted(resolved_mismatch_indices, reverse=True):
                        removed = amount_mismatches.pop(idx)
                        # 매칭 데이터에서 diff 관련 키 제거
                        removed.pop("amount_diff", None)
                        removed.pop("amount_diff_pct", None)
                        removed.pop("vendor_amount", None)
                        removed.pop("erp_amount", None)

                    # 흡수되지 않은 할인항목 + 비할인 미매칭 → still_unmatched
                    for d, dlist in unmatched_discounts_by_date.items():
                        for dr in dlist:
                            if id(dr) not in absorbed_discount_ids:
                                still_unmatched.append(dr)
                    still_unmatched.extend(unmatched_regular)

                    if discount_absorbed:
                        yield sse("log", {"msg": f"  → 할인반영 처리: {len(discount_absorbed)}건 (할인이 매입단가에 반영됨)"})
                    unmatched = still_unmatched

                # ── 2단계: 미매칭 건 판매이력 확인 (AI) — 최대 10건만 ──
                sales_check = []
                if unmatched and all_sales_items:
                    ai_limit = min(len(unmatched), 10)  # AI 호출 제한
                    yield sse("log", {"msg": f"  → 2단계: 판매이력 AI 검색 ({ai_limit}건)..."})
                    unmatched_vendor_items = [r["vendor_item"] for r in unmatched[:ai_limit]]
                    for si, uitem in enumerate(unmatched_vendor_items):
                        pname = uitem.get("product_name", "?")
                        yield sse("log", {"msg": f"    AI 검색 [{si+1}/{ai_limit}]: {pname}"})
                        single_result = await check_sales_history([uitem], all_sales_items)
                        sales_check.extend(single_result)
                        await asyncio.sleep(0)  # yield control

                    # AI 제한으로 처리 못한 나머지
                    if len(unmatched) > ai_limit:
                        for r in unmatched[ai_limit:]:
                            sales_check.append({
                                "vendor_item": r["vendor_item"],
                                "has_sales_history": False,
                                "search_range": "미검색",
                                "candidates": [],
                                "best_candidate": None,
                                "recommendation": "AI 검색 제한 초과 — 수동 확인 필요",
                            })
                elif unmatched:
                    yield sse("log", {"msg": f"  → 판매현황 없어 AI 검색 생략"})

                # 배송료 매칭
                shipping_match_results = []
                if shipping_items and filtered_purchase:
                    shipping_match_results = await match_products_ai(shipping_items, filtered_purchase)

                vf_result["matched"] = matched
                vf_result["unmatched"] = unmatched
                vf_result["discount_absorbed"] = discount_absorbed
                vf_result["sales_check"] = sales_check
                vf_result["shipping_items"] = [
                    {
                        "vendor_item": item,
                        "erp_match": next((r.get("erp_match") for r in shipping_match_results
                                           if r.get("vendor_item") == item and r.get("match_type") != "unmatched"), None),
                        "match_type": next((r.get("match_type") for r in shipping_match_results
                                            if r.get("vendor_item") == item and r.get("match_type") != "unmatched"), "unmatched"),
                    }
                    for item in shipping_items
                ]
                vf_result["amount_mismatches"] = amount_mismatches
                vf_result["memo_items"] = [item for item in memo_items]

                # 미매칭 ERP 항목 추적 (거래처원장에 없지만 매입전표에 있는 항목)
                matched_erp_set = set()
                for r in matched:
                    e = r.get("erp_match")
                    if e:
                        # erp_match의 date+prod_cd를 키로 사용
                        ekey = f"{_get_field(e, 'date', '월/일', default='')}|{_get_field(e, 'prod_cd', '품목코드', default='')}"
                        matched_erp_set.add(ekey)
                for si in vf_result["shipping_items"]:
                    e = si.get("erp_match")
                    if e:
                        ekey = f"{_get_field(e, 'date', '월/일', default='')}|{_get_field(e, 'prod_cd', '품목코드', default='')}"
                        matched_erp_set.add(ekey)
                excess_erp = []
                for p in filtered_purchase:
                    pkey = f"{_get_field(p, 'date', '월/일', default='')}|{_get_field(p, 'prod_cd', '품목코드', default='')}"
                    if pkey not in matched_erp_set:
                        excess_erp.append(p)
                vf_result["excess_erp"] = excess_erp

                vf_result["vendor_ledger_total"] = vendor_ledger_total
                vf_result["erp_purchase_total"] = erp_purchase_total
                vf_result["vendor_total_match"] = vendor_total_match
                vf_result["summary"] = {
                    "total_vendor_items": len(all_vendor_items),
                    "memo_filtered": len(memo_items),
                    "matched_count": len(matched) + len(discount_absorbed),
                    "unmatched_count": len(unmatched),
                    "discount_absorbed_count": len(discount_absorbed),
                    "with_sales_history": sum(1 for s in sales_check if s.get("has_sales_history")),
                    "shipping_count": len(shipping_items),
                    "amount_mismatch_count": len(amount_mismatches),
                    "purchase_filtered": len(filtered_purchase),
                    "excess_erp_count": len(excess_erp),
                    "vendor_total_match": vendor_total_match,
                }

                effective_matched = len(matched) + len(discount_absorbed)
                match_rate = round(effective_matched / max(len(regular_items), 1) * 100, 1)
                yield sse("log", {"msg": f"✅ {vendor_name} 완료: 매칭률 {match_rate}% ({effective_matched}/{len(regular_items)})"})
                if discount_absorbed:
                    yield sse("log", {"msg": f"  ↳ 할인반영 {len(discount_absorbed)}건 포함"})

            except Exception as e:
                logger.error(f"[일괄] '{orig_filename}' 처리 실패: {e}", exc_info=True)
                vf_result["error"] = str(e)
                yield sse("log", {"msg": f"❌ {vendor_name} 처리 실패: {e}", "level": "error"})

            vendor_results.append(vf_result)
            yield sse("progress", {"current": vi + 1, "total": total_vendors})

        # ── 세션 저장 & 최종 결과 ──
        session_id = uuid.uuid4().hex[:12]
        reconcile_sessions[session_id] = {
            "batch": True,
            "vendor_results": vendor_results,
            "purchase_total": len(all_purchase_items),
            "sales_total": len(all_sales_items),
            "purchase_info": purchase_info,
            "sales_info": sales_info,
        }

        total_summary = {
            "vendor_count": len(vendor_results),
            "total_matched": sum(r["summary"].get("matched_count", 0) for r in vendor_results),
            "total_unmatched": sum(r["summary"].get("unmatched_count", 0) for r in vendor_results),
            "total_shipping": sum(r["summary"].get("shipping_count", 0) for r in vendor_results),
            "total_amount_mismatch": sum(r["summary"].get("amount_mismatch_count", 0) for r in vendor_results),
            "purchase_total": len(all_purchase_items),
            "purchase_from_cache": purchase_info["from_cache"],
            "sales_total": len(all_sales_items),
            "sales_from_cache": sales_info["from_cache"],
            "errors": [r["filename"] for r in vendor_results if r.get("error")],
        }

        elapsed = round(time.time() - t0, 1)
        yield sse("log", {"msg": f"🎉 전체 정산 완료 ({elapsed}초)"})
        yield sse("result", {
            "session_id": session_id,
            "summary": total_summary,
            "vendor_results": vendor_results,
        })

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@router.post("/compare")
async def compare_ledgers(
    req: CompareRequest,
    user: dict = Depends(get_current_user),
):
    """거래처 원장 vs ERP 데이터 비교 (AI 매칭)"""
    # 거래처 원장에서 "매출" 항목 = 우리의 매입
    all_vendor_items = [
        item for item in req.vendor_items
        if item.get("tx_type") == "매출" or not item.get("tx_type")
    ]

    # 배송료/운송비 항목 분리
    shipping_items = [item for item in all_vendor_items if _is_shipping_item(item)]
    regular_items = [item for item in all_vendor_items if not _is_shipping_item(item)]

    # AI or 규칙 기반 매칭 (일반 품목만)
    match_results = await match_products_ai(
        regular_items, req.erp_purchase_data
    )

    matched = [r for r in match_results if r["match_type"] != "unmatched"]
    unmatched = [r for r in match_results if r["match_type"] == "unmatched"]

    # 금액 차이 감지 (매칭된 항목 중 금액이 다른 건)
    amount_mismatches = []
    for r in matched:
        v = r.get("vendor_item", {})
        e = r.get("erp_match", {})
        v_amt = float(v.get("amount", 0) or 0)
        try:
            e_amt = float(str(_get_field(e, "total", "합계", default=0) or 0).replace(",", ""))
        except (ValueError, TypeError):
            e_amt = 0
        v_qty = int(v.get("qty", 0) or 0)
        try:
            e_qty = int(float(str(_get_field(e, "qty", "수량", default=0) or 0).replace(",", "")))
        except (ValueError, TypeError):
            e_qty = 0

        if v_amt and e_amt and abs(v_amt - e_amt) > 1:
            r["amount_diff"] = v_amt - e_amt
            r["amount_diff_pct"] = round((v_amt - e_amt) / max(v_amt, 1) * 100, 1)
            amount_mismatches.append(r)
        if v_qty and e_qty and v_qty != e_qty:
            r["qty_diff"] = v_qty - e_qty

    # 누락 건 → 판매이력 확인
    sales_check = []
    if unmatched and req.erp_sales_data:
        unmatched_vendor_items = [r["vendor_item"] for r in unmatched]
        sales_check = await check_sales_history(
            unmatched_vendor_items, req.erp_sales_data
        )

    # 배송료 항목도 ERP 구매현황에서 매칭 시도
    shipping_match_results = []
    if shipping_items and req.erp_purchase_data:
        shipping_match_results = await match_products_ai(
            shipping_items, req.erp_purchase_data
        )

    session_id = uuid.uuid4().hex[:12]
    reconcile_sessions[session_id] = {
        "matched": matched,
        "unmatched": unmatched,
        "sales_check": sales_check,
        "shipping_items": [
            {
                "vendor_item": item,
                "erp_match": next((r.get("erp_match") for r in shipping_match_results
                                   if r.get("vendor_item") == item and r.get("match_type") != "unmatched"), None),
                "match_type": next((r.get("match_type") for r in shipping_match_results
                                    if r.get("vendor_item") == item and r.get("match_type") != "unmatched"), "unmatched"),
            }
            for item in shipping_items
        ],
        "amount_mismatches": amount_mismatches,
    }

    return {
        "session_id": session_id,
        "summary": {
            "total_vendor_items": len(all_vendor_items),
            "matched_count": len(matched),
            "unmatched_count": len(unmatched),
            "with_sales_history": sum(
                1 for s in sales_check if s.get("has_sales_history")
            ),
            "shipping_count": len(shipping_items),
            "amount_mismatch_count": len(amount_mismatches),
        },
        "matched": matched,
        "unmatched": unmatched,
        "sales_check": sales_check,
        "shipping_items": reconcile_sessions[session_id]["shipping_items"],
        "amount_mismatches": amount_mismatches,
    }


@router.get("/session/{session_id}")
async def get_session(
    session_id: str,
    user: dict = Depends(get_current_user),
):
    """이전 비교 결과 조회"""
    if session_id not in reconcile_sessions:
        raise HTTPException(404, "세션을 찾을 수 없습니다")
    return reconcile_sessions[session_id]


@router.post("/save-purchase")
async def save_purchase(
    req: PurchaseSaveRequest,
    user: dict = Depends(get_current_user),
):
    """누락 매입전표를 ERP에 입력 (기존 erp_client.save_purchase 활용)"""
    if not req.items:
        raise HTTPException(400, "입력할 항목이 없습니다")

    results = []
    success_count = 0
    fail_count = 0

    # 거래처코드별로 그룹핑
    from collections import defaultdict
    grouped: dict[str, list[PurchaseItem]] = defaultdict(list)
    for item in req.items:
        grouped[item.cust_code].append(item)

    for cust_code, items in grouped.items():
        lines = []
        io_date = items[0].io_date if items else ""

        for item in items:
            lines.append({
                "prod_cd": item.prod_cd,
                "qty": item.qty,
                "unit": "",
                "price": item.price,
            })

        try:
            result = await erp_client.save_purchase(
                cust_code=cust_code,
                lines=lines,
                upload_ser=req.upload_ser_no,
                wh_cd=items[0].wh_cd or ERP_WH_CD,
                emp_cd=ERP_EMP_CD,
                io_date=io_date,
            )

            status = result.get("Status", "")
            data = result.get("Data", {})

            if status == "200":
                s_cnt = int(data.get("SuccessCnt", 0))
                f_cnt = int(data.get("FailCnt", 0))
                success_count += s_cnt
                fail_count += f_cnt
                results.append({
                    "cust_code": cust_code,
                    "status": "success",
                    "success": s_cnt,
                    "failed": f_cnt,
                    "slip_nos": data.get("SlipNos", ""),
                    "details": data.get("ResultDetails", ""),
                })
            else:
                error = result.get("Error", {})
                fail_count += len(items)
                results.append({
                    "cust_code": cust_code,
                    "status": "error",
                    "success": 0,
                    "failed": len(items),
                    "error_message": error.get("Message", "알 수 없는 오류"),
                })

        except Exception as e:
            fail_count += len(items)
            results.append({
                "cust_code": cust_code,
                "status": "error",
                "success": 0,
                "failed": len(items),
                "error_message": str(e),
            })

    return {
        "status": "success" if fail_count == 0 else "partial" if success_count > 0 else "error",
        "total": len(req.items),
        "success": success_count,
        "failed": fail_count,
        "results": results,
    }


@router.post("/validate-purchase")
async def validate_purchase(
    req: PurchaseSaveRequest,
    user: dict = Depends(get_current_user),
):
    """매입전표 입력 전 유효성 검사"""
    errors = []
    for i, item in enumerate(req.items):
        item_errors = []
        if not item.io_date or len(item.io_date) != 8:
            item_errors.append("전표일자(YYYYMMDD) 형식 오류")
        if not item.cust_code and not item.cust_name:
            item_errors.append("거래처코드 또는 거래처명 필요")
        if not item.prod_cd and not item.prod_name:
            item_errors.append("품목코드 또는 품목명 필요")
        if item.qty <= 0:
            item_errors.append("수량은 1 이상이어야 함")
        if item_errors:
            errors.append({"index": i, "errors": item_errors})

    return {
        "valid": len(errors) == 0,
        "error_count": len(errors),
        "errors": errors,
    }


@router.get("/download-result/{session_id}")
async def download_result_excel(
    session_id: str,
    user: dict = Depends(get_current_user),
):
    """비교 결과를 엑셀로 다운로드"""
    if session_id not in reconcile_sessions:
        raise HTTPException(404, "세션을 찾을 수 없습니다")

    data = reconcile_sessions[session_id]

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl이 설치되지 않았습니다")

    wb = openpyxl.Workbook()

    # Style definitions
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    green_fill = PatternFill("solid", fgColor="E2EFDA")
    red_fill = PatternFill("solid", fgColor="FCE4EC")
    yellow_fill = PatternFill("solid", fgColor="FFF9C4")
    blue_fill = PatternFill("solid", fgColor="DBEAFE")
    orange_fill = PatternFill("solid", fgColor="FFF3E0")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def style_header(ws, row, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    def style_data_cell(cell, fill=None):
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if fill:
            cell.fill = fill

    # ── 배치 세션 vs 단일 세션 분기 ──
    is_batch = data.get("batch", False)

    if is_batch:
        # ━━━ 배치 모드: 거래처별 시트 생성 ━━━
        vendor_results = data.get("vendor_results", [])

        # Sheet 1: 전체 요약
        ws1 = wb.active
        ws1.title = "전체요약"
        ws1["A1"] = "매입정산 일괄 비교 결과"
        ws1["A1"].font = Font(bold=True, size=14)
        ws1.merge_cells("A1:E1")

        total_matched = sum(len(vr.get("matched", [])) for vr in vendor_results)
        total_unmatched = sum(len(vr.get("unmatched", [])) for vr in vendor_results)
        total_shipping = sum(len(vr.get("shipping_items", [])) for vr in vendor_results)
        total_sales_check = sum(
            sum(1 for s in vr.get("sales_check", []) if s.get("has_sales_history"))
            for vr in vendor_results
        )
        total_amt_mismatch = sum(len(vr.get("amount_mismatches", [])) for vr in vendor_results)

        summary_headers = ["구분", "건수", "비고"]
        for c, h in enumerate(summary_headers, 1):
            ws1.cell(row=3, column=c, value=h)
        style_header(ws1, 3, 3)

        summary_rows = [
            ["✅ 매칭됨", total_matched, "거래처 원장 = ERP 구매현황"],
            ["🔴 매입전표 누락", total_unmatched, "거래처 원장에 있으나 ERP에 없음"],
            ["📦 판매이력 확인", total_sales_check, "판매이력은 있으나 매입전표 누락"],
            ["🚚 배송료/운송비", total_shipping, "배송 관련 항목"],
            ["⚠️ 금액 불일치", total_amt_mismatch, "매칭되었으나 금액이 다른 항목"],
            ["전체", total_matched + total_unmatched + total_shipping, ""],
        ]
        for r_idx, row_data in enumerate(summary_rows, start=4):
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws1.cell(row=r_idx, column=c_idx, value=val)
                style_data_cell(cell)

        # 거래처별 요약 테이블
        v_start = 4 + len(summary_rows) + 2
        ws1.cell(row=v_start, column=1, value="거래처별 요약").font = Font(bold=True, size=12)
        v_headers = ["거래처명", "매칭", "미매칭", "배송료", "금액차이", "합계"]
        for c, h in enumerate(v_headers, 1):
            ws1.cell(row=v_start + 1, column=c, value=h)
        style_header(ws1, v_start + 1, len(v_headers))

        for vi, vr in enumerate(vendor_results):
            row = v_start + 2 + vi
            m_cnt = len(vr.get("matched", []))
            u_cnt = len(vr.get("unmatched", []))
            s_cnt = len(vr.get("shipping_items", []))
            a_cnt = len(vr.get("amount_mismatches", []))
            vals = [
                vr.get("vendor_name", vr.get("filename", "")),
                m_cnt, u_cnt, s_cnt, a_cnt,
                m_cnt + u_cnt + s_cnt,
            ]
            for c, val in enumerate(vals, 1):
                cell = ws1.cell(row=row, column=c, value=val)
                style_data_cell(cell)

        ws1.column_dimensions["A"].width = 25
        ws1.column_dimensions["B"].width = 12
        ws1.column_dimensions["C"].width = 40

        # 거래처별 상세 시트
        for vr in vendor_results:
            vname = vr.get("vendor_name", "Unknown")[:25]  # 시트명 길이 제한
            if vr.get("error"):
                continue

            matched = vr.get("matched", [])
            unmatched = vr.get("unmatched", [])
            sales_check = vr.get("sales_check", [])
            shipping = vr.get("shipping_items", [])

            # 시트명 중복 방지
            sheet_name = vname
            existing = [ws.title for ws in wb.worksheets]
            idx = 1
            while sheet_name in existing:
                sheet_name = f"{vname[:22]}_{idx}"
                idx += 1

            ws = wb.create_sheet(sheet_name)

            ws["A1"] = f"거래처: {vr.get('vendor_name', '')}"
            ws["A1"].font = Font(bold=True, size=12)
            ws.merge_cells("A1:F1")

            curr_row = 3

            # 매칭됨
            if matched:
                ws.cell(row=curr_row, column=1, value="✅ 매칭됨").font = Font(bold=True, size=11)
                curr_row += 1
                m_headers = ["#", "비교결과", "거래처 날짜", "거래처 품목명", "수량", "금액",
                             "ERP 품목코드", "ERP 품목명", "ERP 수량", "ERP 금액", "금액차이", "신뢰도"]
                for c, h in enumerate(m_headers, 1):
                    ws.cell(row=curr_row, column=c, value=h)
                style_header(ws, curr_row, len(m_headers))
                curr_row += 1

                for i, r in enumerate(matched, 1):
                    v = r.get("vendor_item", {})
                    e = r.get("erp_match", {}) or {}
                    v_amt = float(v.get("amount", 0) or 0)
                    e_amt = float(str(e.get("total", e.get("합계", 0)) or 0).replace(",", ""))
                    diff = v_amt - e_amt if v_amt and e_amt else 0
                    conf = round((r.get("confidence", 0)) * 100)

                    vals = [
                        i,
                        "✅ 일치" if abs(diff) <= 1 else "⚠️ 금액차이",
                        v.get("date", ""),
                        v.get("product_name", ""),
                        v.get("qty", 0),
                        v_amt,
                        e.get("prod_cd", e.get("품목코드", "")),
                        e.get("prod_name", e.get("품명 및 모델", "")),
                        e.get("qty", e.get("수량", "")),
                        e_amt,
                        diff if abs(diff) > 1 else 0,
                        f"{conf}%",
                    ]
                    fill = yellow_fill if abs(diff) > 1 else green_fill
                    for c, val in enumerate(vals, 1):
                        cell = ws.cell(row=curr_row, column=c, value=val)
                        style_data_cell(cell, fill)
                    curr_row += 1
                curr_row += 1

            # 미매칭
            if unmatched:
                ws.cell(row=curr_row, column=1, value="🔴 매입전표 누락").font = Font(bold=True, size=11)
                curr_row += 1
                u_headers = ["#", "날짜", "품목명", "모델명", "수량", "단가", "금액",
                             "판매이력", "추천 품목코드", "추천 품목명", "신뢰도"]
                for c, h in enumerate(u_headers, 1):
                    ws.cell(row=curr_row, column=c, value=h)
                style_header(ws, curr_row, len(u_headers))
                curr_row += 1

                for i, r in enumerate(unmatched, 1):
                    v = r.get("vendor_item", {})
                    sc = None
                    for s in sales_check:
                        sv = s.get("vendor_item", {})
                        if sv.get("product_name") == v.get("product_name") and sv.get("date") == v.get("date"):
                            sc = s
                            break
                    best = sc.get("best_candidate", {}) if sc else {}
                    fill = blue_fill if (sc and sc.get("has_sales_history")) else red_fill

                    vals = [
                        i,
                        v.get("date", ""),
                        v.get("product_name", ""),
                        v.get("model_name", ""),
                        v.get("qty", 0),
                        v.get("unit_price", 0),
                        v.get("amount", 0),
                        "있음" if (sc and sc.get("has_sales_history")) else "없음",
                        best.get("product_code", "") if best else "",
                        best.get("product_name", "") if best else "",
                        f"{round((best.get('confidence', 0)) * 100)}%" if best else "",
                    ]
                    for c, val in enumerate(vals, 1):
                        cell = ws.cell(row=curr_row, column=c, value=val)
                        style_data_cell(cell, fill)
                    curr_row += 1
                curr_row += 1

            # 배송료
            if shipping:
                ws.cell(row=curr_row, column=1, value="🚚 배송료/운송비").font = Font(bold=True, size=11)
                curr_row += 1
                s_headers = ["#", "날짜", "품목명", "수량", "금액", "ERP매칭"]
                for c, h in enumerate(s_headers, 1):
                    ws.cell(row=curr_row, column=c, value=h)
                style_header(ws, curr_row, len(s_headers))
                curr_row += 1

                for i, s_item in enumerate(shipping, 1):
                    sv = s_item.get("vendor_item", {})
                    e_match = s_item.get("erp_match")
                    vals = [
                        i, sv.get("date", ""), sv.get("product_name", ""),
                        sv.get("qty", 0), sv.get("amount", 0),
                        "매칭됨" if e_match else "미매칭",
                    ]
                    fill_s = green_fill if e_match else orange_fill
                    for c, val in enumerate(vals, 1):
                        cell = ws.cell(row=curr_row, column=c, value=val)
                        style_data_cell(cell, fill_s)
                    curr_row += 1

            # 열 너비 설정
            for col_letter in ["A","B","C","D","E","F","G","H","I","J","K","L"]:
                ws.column_dimensions[col_letter].width = 15
            ws.column_dimensions["D"].width = 30
            ws.column_dimensions["H"].width = 30
            ws.column_dimensions["J"].width = 30

    else:
        # ━━━ 단일 거래처 모드 (기존 로직) ━━━
        # Sheet 1: 비교 요약
        ws1 = wb.active
        ws1.title = "비교요약"
        matched = data.get("matched", [])
        unmatched = data.get("unmatched", [])
        sales_check = data.get("sales_check", [])
        shipping = data.get("shipping_items", [])
        amt_mismatch = data.get("amount_mismatches", [])

        ws1["A1"] = "매입정산 비교 결과"
        ws1["A1"].font = Font(bold=True, size=14)
        ws1.merge_cells("A1:D1")

        summary_data = [
            ["구분", "건수", "비고"],
            ["✅ 매칭됨", len(matched), "거래처 원장 = ERP 구매현황"],
            ["🔴 매입전표 누락", len(unmatched), "거래처 원장에 있으나 ERP에 없음"],
            ["📦 판매이력 확인", sum(1 for s in sales_check if s.get("has_sales_history")), "판매이력은 있으나 매입전표 누락"],
            ["🚚 배송료/운송비", len(shipping), "배송 관련 항목"],
            ["⚠️ 금액 불일치", len(amt_mismatch), "매칭되었으나 금액이 다른 항목"],
            ["전체", len(matched) + len(unmatched) + len(shipping), ""],
        ]
        for r_idx, row_data in enumerate(summary_data, start=3):
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws1.cell(row=r_idx, column=c_idx, value=val)
                style_data_cell(cell)
        style_header(ws1, 3, 3)
        ws1.column_dimensions["A"].width = 20
        ws1.column_dimensions["B"].width = 12
        ws1.column_dimensions["C"].width = 40

        # Sheet 2: 매칭 상세
        ws2 = wb.create_sheet("매칭됨")
        headers2 = ["#", "비교결과", "거래처 날짜", "거래처 품목명", "수량", "금액",
                     "ERP 품목코드", "ERP 품목명", "ERP 수량", "ERP 금액", "금액차이", "신뢰도", "비고"]
        for c, h in enumerate(headers2, 1):
            ws2.cell(row=1, column=c, value=h)
        style_header(ws2, 1, len(headers2))

        for i, r in enumerate(matched, 1):
            v = r.get("vendor_item", {})
            e = r.get("erp_match", {}) or {}
            row = i + 1
            v_amt = float(v.get("amount", 0) or 0)
            e_amt = float(str(e.get("total", e.get("합계", 0)) or 0).replace(",", ""))
            diff = v_amt - e_amt if v_amt and e_amt else 0
            conf = round((r.get("confidence", 0)) * 100)

            vals = [
                i,
                "✅ 일치" if abs(diff) <= 1 else "⚠️ 금액차이",
                v.get("date", ""),
                v.get("product_name", ""),
                v.get("qty", 0),
                v_amt,
                e.get("prod_cd", e.get("품목코드", "")),
                e.get("prod_name", e.get("품명 및 모델", "")),
                e.get("qty", e.get("수량", "")),
                e_amt,
                diff if abs(diff) > 1 else 0,
                f"{conf}%",
                r.get("reason", ""),
            ]
            fill = yellow_fill if abs(diff) > 1 else green_fill
            for c, val in enumerate(vals, 1):
                cell = ws2.cell(row=row, column=c, value=val)
                style_data_cell(cell, fill)

        for col_letter in ["A","B","C","D","E","F","G","H","I","J","K","L","M"]:
            ws2.column_dimensions[col_letter].width = 15
        ws2.column_dimensions["D"].width = 30
        ws2.column_dimensions["H"].width = 30
        ws2.column_dimensions["M"].width = 25

        # Sheet 3: 매입전표 누락
        ws3 = wb.create_sheet("매입전표누락")
        headers3 = ["#", "거래처 날짜", "품목명", "모델명", "수량", "단가", "금액", "판매이력", "검색범위", "추천 품목코드", "추천 품목명", "신뢰도", "추천사항"]
        for c, h in enumerate(headers3, 1):
            ws3.cell(row=1, column=c, value=h)
        style_header(ws3, 1, len(headers3))

        row_idx = 2
        for i, r in enumerate(unmatched):
            v = r.get("vendor_item", {})
            sc = None
            for s in sales_check:
                sv = s.get("vendor_item", {})
                if sv.get("product_name") == v.get("product_name") and sv.get("date") == v.get("date"):
                    sc = s
                    break

            best = sc.get("best_candidate", {}) if sc else {}
            fill = blue_fill if (sc and sc.get("has_sales_history")) else red_fill

            vals = [
                i + 1,
                v.get("date", ""),
                v.get("product_name", ""),
                v.get("model_name", ""),
                v.get("qty", 0),
                v.get("unit_price", 0),
                v.get("amount", 0),
                "있음" if (sc and sc.get("has_sales_history")) else "없음",
                sc.get("search_range", "") if sc else "",
                best.get("product_code", "") if best else "",
                best.get("product_name", "") if best else "",
                f"{round((best.get('confidence', 0)) * 100)}%" if best else "",
                sc.get("recommendation", "") if sc else "확인 필요",
            ]
            for c, val in enumerate(vals, 1):
                cell = ws3.cell(row=row_idx, column=c, value=val)
                style_data_cell(cell, fill)
            row_idx += 1

        for col_letter in ["A","B","C","D","E","F","G","H","I","J","K","L","M"]:
            ws3.column_dimensions[col_letter].width = 15
        ws3.column_dimensions["C"].width = 30
        ws3.column_dimensions["K"].width = 30
        ws3.column_dimensions["M"].width = 35

        # Sheet 4: 배송료
        if shipping:
            ws4 = wb.create_sheet("배송료")
            headers4 = ["#", "날짜", "품목명", "수량", "금액", "ERP매칭", "비고"]
            for c, h in enumerate(headers4, 1):
                ws4.cell(row=1, column=c, value=h)
            style_header(ws4, 1, len(headers4))
            for i, s in enumerate(shipping, 1):
                v = s.get("vendor_item", {})
                e_match = s.get("erp_match")
                vals = [
                    i, v.get("date", ""), v.get("product_name", ""),
                    v.get("qty", 0), v.get("amount", 0),
                    "매칭됨" if e_match else "미매칭",
                    s.get("match_type", ""),
                ]
                fill = green_fill if e_match else yellow_fill
                for c, val in enumerate(vals, 1):
                    cell = ws4.cell(row=i+1, column=c, value=val)
                    style_data_cell(cell, fill)
            for col_letter in ["A","B","C","D","E","F","G"]:
                ws4.column_dimensions[col_letter].width = 18
            ws4.column_dimensions["C"].width = 30

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"매입정산_비교결과_{session_id}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


@router.get("/vendor-list")
async def get_vendor_list(
    q: Optional[str] = Query(None),
    user: dict = Depends(get_current_user),
):
    """거래처 목록 조회 및 검색

    - 첫 로드 시 /app/data/vendors.json에서 데이터 로드
    - 메모리에 캐시하여 성능 최적화
    - q 파라미터로 거래처명 또는 코드 검색 (대소문자 무시)
    - 검색 결과는 최대 50개, 검색 없을 시 전체 반환
    """
    global _vendor_cache

    try:
        # 캐시에 없으면 파일에서 로드
        if _vendor_cache is None:
            vendor_path = Path(__file__).parent.parent.parent.parent / "data" / "vendors.json"
            if not vendor_path.exists():
                raise HTTPException(
                    404,
                    f"거래처 데이터 파일을 찾을 수 없습니다: {vendor_path}"
                )

            with open(vendor_path, "r", encoding="utf-8") as f:
                _vendor_cache = json.load(f)

        logger.info(f"[vendor-list] 캐시 로드 완료: {len(_vendor_cache)}건")

        # 검색 쿼리가 있으면 필터링 (macOS NFD→NFC 정규화)
        vendors = _vendor_cache
        if q:
            q_lower = unicodedata.normalize("NFC", q).lower()
            vendors = [
                v for v in _vendor_cache
                if q_lower in v.get("name", "").lower() or
                   q_lower in v.get("code", "").lower()
            ]
            # 검색 결과는 최대 50개
            vendors = vendors[:50]
            logger.info(f"[vendor-list] q='{q}' → {len(vendors)}건 매칭")

        return {
            "vendors": vendors,
            "total": len(vendors),
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"거래처 목록 조회 오류: {e}", exc_info=True)
        raise HTTPException(500, f"거래처 목록 조회 실패: {str(e)}")
