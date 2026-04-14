"""
矛聤陇毛搂聢铆聤赂矛聤陇铆聠聽矛聳麓 矛拢录毛卢赂 矛聻聬毛聫聶铆聶聰 API 毛聺录矛職掳铆聤赂
毛聞陇矛聺麓毛虏聞 矛拢录毛卢赂矛聢聵矛搂聭 芒聠聮 ERP 铆聦聬毛搂陇矛聻聟毛聽楼 芒聠聮 毛隆聹矛聽聽铆聝聺毛掳掳 毛聯卤毛隆聺 芒聠聮 毛掳聹矛聠隆矛虏聵毛娄卢
"""
import re
import json
import asyncio
import logging
from typing import Optional
from pathlib import Path
from datetime import datetime
from zoneinfo import ZoneInfo
from fastapi import APIRouter, Query, HTTPException, Body, UploadFile, File
from pydantic import BaseModel

KST = ZoneInfo("Asia/Seoul")

from config import (
    SMARTSTORE_CUST_CODE, SMARTSTORE_EMP_CODE, SMARTSTORE_WH_CODE,
    SMARTSTORE_PRODUCT_MAP_PATH, SMARTSTORE_MODEL_MAP_PATH,
    SMARTSTORE_OPTION_MAP_PATH, SMARTSTORE_ADDON_MAP_PATH,
    SMARTSTORE_OPTION_TEXT_MAP_PATH, SMARTSTORE_ADDON_TEXT_MAP_PATH,
    SMARTSTORE_CODE_ALIAS_MAP_PATH,
)

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/smartstore", tags=["SmartStore"])

# 毛陋篓毛聧赂矛陆聰毛聯聹 矛聽聲锚路聹矛聥聺
MODEL_CODE_RE = re.compile(r"(LS[PNE]?-[\w\-]+|ZOT-[\w\-]+)", re.IGNORECASE)
EXCLUDE_KEYWORDS = ["铆聴聢毛赂聦毛聻聶", "矛聞聹毛虏聞毛聻聶", "矛潞聬毛鹿聞毛聞路"]

# 矛聥聹铆聤赂1: 毛漏聰矛聺赂矛聝聛铆聮聢 芒聙聰 矛聝聛铆聮聢毛虏聢铆聵赂 芒聠聮 ERP铆聮聢毛陋漏矛陆聰毛聯聹 (矛聵碌矛聟聵 矛聴聠毛聤聰 矛聝聛铆聮聢)
_product_map: dict = {}
# 矛聥聹铆聤赂2: 矛聵碌矛聟聵矛聝聛铆聮聢 芒聙聰 矛聝聛铆聮聢毛虏聢铆聵赂 芒聠聮 ERP铆聮聢毛陋漏矛陆聰毛聯聹 (矛聻聬毛聫聶矛露聰矛露聹 矛聵陇毛虏聞毛聺录矛聺麓毛聯聹)
_option_override_map: dict = {}
# 矛聥聹铆聤赂3: 矛露聰锚掳聙矛聝聛铆聮聢 芒聙聰 矛聝聛铆聮聢毛虏聢铆聵赂 芒聠聮 ERP铆聮聢毛陋漏矛陆聰毛聯聹
_addon_map: dict = {}
# 毛陋篓毛聧赂毛陋聟: 矛聝聛铆聮聢毛虏聢铆聵赂 芒聠聮 毛陋篓毛聧赂毛陋聟 (毛隆聹矛聽聽 矛聠隆矛聻楼矛職漏, 矛聽聞 矛聥聹铆聤赂 锚鲁碌矛職漏)
_model_map: dict = {}
# 矛聴颅毛掳漏铆聳楼: 毛陋篓毛聧赂毛陋聟 芒聠聮 ERP铆聮聢毛陋漏矛陆聰毛聯聹 (矛聵碌矛聟聵 铆聟聧矛聤陇铆聤赂矛聴聬矛聞聹 毛陋篓毛聧赂毛陋聟 矛露聰矛露聹 矛聥聹 ERP矛陆聰毛聯聹毛隆聹 毛鲁聙铆聶聵)
_model_to_erp_map: dict = {}
# 矛聵碌矛聟聵铆聟聧矛聤陇铆聤赂 矛搂聛矛聽聭毛搂陇铆聲聭: "矛聝聛铆聮聢毛虏聢铆聵赂|矛聵碌矛聟聵锚掳聮" 芒聠聮 ERP铆聮聢毛陋漏矛陆聰毛聯聹
_option_text_map: dict = {}
# 矛露聰锚掳聙矛聝聛铆聮聢铆聟聧矛聤陇铆聤赂 矛搂聛矛聽聭毛搂陇铆聲聭: "矛聝聛铆聮聢毛虏聢铆聵赂|矛露聰锚掳聙矛聝聛铆聮聢锚掳聮" 芒聠聮 ERP铆聮聢毛陋漏矛陆聰毛聯聹
_addon_text_map: dict = {}
# 矛露聰矛露聹矛陆聰毛聯聹 毛鲁聞矛鹿颅毛搂碌: HDSVAL-XXX 毛聯卤 芒聠聮 ERP铆聮聢毛陋漏矛陆聰毛聯聹
_code_alias_map: dict = {}


def _load_json(path) -> dict:
    if path.exists():
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    return {}


def _save_json(path, data: dict):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _load_product_map():
    global _product_map, _option_override_map, _addon_map, _model_map, _model_to_erp_map
    global _option_text_map, _addon_text_map, _code_alias_map
    _product_map        = _load_json(SMARTSTORE_PRODUCT_MAP_PATH)
    _option_override_map = _load_json(SMARTSTORE_OPTION_MAP_PATH)
    _addon_map          = _load_json(SMARTSTORE_ADDON_MAP_PATH)
    _model_map          = _load_json(SMARTSTORE_MODEL_MAP_PATH)
    _option_text_map    = _load_json(SMARTSTORE_OPTION_TEXT_MAP_PATH)
    _addon_text_map     = _load_json(SMARTSTORE_ADDON_TEXT_MAP_PATH)
    _code_alias_map     = _load_json(SMARTSTORE_CODE_ALIAS_MAP_PATH)
    # 矛聴颅毛掳漏铆聳楼 毛搂碌 毛鹿聦毛聯聹: 毛陋篓毛聧赂毛陋聟 芒聠聮 ERP铆聮聢毛陋漏矛陆聰毛聯聹 (矛聽聞 矛聥聹铆聤赂 铆聫卢铆聲篓)
    _model_to_erp_map = {}
    for pno, model in _model_map.items():
        if model:
            if pno in _product_map:
                _model_to_erp_map[model] = _product_map[pno]
            elif pno in _option_override_map:
                _model_to_erp_map[model] = _option_override_map[pno]
            elif pno in _addon_map:
                _model_to_erp_map[model] = _addon_map[pno]
    logger.info(
        f"[SS] 毛搂陇铆聲聭 毛隆聹毛聯聹 芒聙聰 毛漏聰矛聺赂:{len(_product_map)} 矛聵碌矛聟聵:{len(_option_override_map)} "
        f"矛露聰锚掳聙:{len(_addon_map)} 毛陋篓毛聧赂矛聴颅毛掳漏铆聳楼:{len(_model_to_erp_map)} "
        f"矛聵碌矛聟聵铆聟聧矛聤陇铆聤赂:{len(_option_text_map)} 矛露聰锚掳聙矛聝聛铆聮聢铆聟聧矛聤陇铆聤赂:{len(_addon_text_map)} 矛陆聰毛聯聹毛鲁聞矛鹿颅:{len(_code_alias_map)}"
    )


def _save_product_map():
    _save_json(SMARTSTORE_PRODUCT_MAP_PATH, _product_map)
    _save_json(SMARTSTORE_OPTION_MAP_PATH,  _option_override_map)
    _save_json(SMARTSTORE_ADDON_MAP_PATH,   _addon_map)
    _save_json(SMARTSTORE_MODEL_MAP_PATH,   _model_map)


_load_product_map()


def _extract_erp_code_from_option(option_text: str) -> Optional[str]:
    """
    矛聵碌矛聟聵 铆聟聧矛聤陇铆聤赂矛聴聬矛聞聹 ERP 铆聮聢毛陋漏矛陆聰毛聯聹 矛露聰矛露聹.
    矛職掳矛聞聽矛聢聹矛聹聞:
      1) 毛搂聢矛搂聙毛搂聣 (矛陆聰毛聯聹) 毛聵聬毛聤聰 [矛陆聰毛聯聹] 芒聙聰 矛聽聲矛聝聛/毛鹿聞矛聽聲矛聝聛 锚麓聞铆聵赂 毛陋篓毛聭聬 矛虏聵毛娄卢
      2) 矛陆聹毛隆聽 毛聮陇 矛陆聰毛聯聹
      3) 矛聵碌矛聟聵 矛聽聞矛虏麓锚掳聙 矛陆聰毛聯聹
      4) LS/LSP/LSN/LST/ZOT 毛隆聹 矛聥聹矛聻聭铆聲聵毛聤聰 矛陆聰毛聯聹锚掳聙 铆聟聧矛聤陇铆聤赂 毛聜麓 铆聫卢铆聲篓
    矛聹聽铆職篓 矛陆聰毛聯聹 锚赂掳矛陇聙: 矛聵聛毛卢赂脗路矛聢芦矛聻聬毛隆聹 矛聥聹矛聻聭, 锚鲁碌毛掳卤 矛聴聠矛聺聦
    """
    if not option_text:
        return None

    text = option_text.strip()

    # '/ 毛掳掳矛聠隆毛掳漏毛虏聲:' 矛聺麓铆聸聞 矛聽聹锚卤掳 (矛聵聢: "LS-750HS / 毛掳掳矛聠隆毛掳漏毛虏聲: 锚虏陆毛聫聶铆聝聺毛掳掳 (矛聸聬铆聲聵毛聤聰 毛掳掳矛聠隆矛搂聙 毛聫聞矛掳漏)")
    if " / " in text:
        text = text.split(" / ")[0].strip()

    def is_valid_code(s: str) -> bool:
        # 矛聵聛毛卢赂脗路矛聢芦矛聻聬 矛聥聹矛聻聭, 锚鲁碌毛掳卤 矛聴聠矛聺聦, 矛碌聹矛聠聦 3矛聻聬 矛聺麓矛聝聛 (毛聥篓矛聺录 毛卢赂矛聻聬 矛聵陇铆聝聬 毛掳漏矛搂聙)
        return bool(s and " " not in s and len(s) >= 3 and re.match(r"[A-Za-z0-9]", s))

    # 铆聦篓铆聞麓 1: 毛搂聢矛搂聙毛搂聣 (矛陆聰毛聯聹) 芒聙聰 毛聥芦铆聻聦 锚麓聞铆聵赂
    m = re.search(r"\(([^()]*(?:\([^)]*\))[^()]*|[^()]+)\)\s*$", text)
    if m:
        candidate = m.group(1).strip()
        if is_valid_code(candidate):
            return candidate

    # 铆聦篓铆聞麓 1-b: 毛聥芦铆聻聢矛搂聙 矛聲聤矛聺聙 锚麓聞铆聵赂 (矛聵聢: (LS-ADOOR(B) )
    m2 = re.search(r"\(([A-Za-z0-9][A-Za-z0-9\-\(\)\.]*)\s*$", text)
    if m2:
        candidate = m2.group(1).strip()
        if is_valid_code(candidate):
            return candidate

    # 铆聦篓铆聞麓 1-c: 毛聦聙锚麓聞铆聵赂 [矛陆聰毛聯聹] (矛聵聢: "0.5M [LS-5UTPD-0.5MG]", "矛聞聹毛虏聞铆聝颅 [HDSVAL-615]")
    m3 = re.search(r"\[([A-Za-z0-9][A-Za-z0-9\-\.]*)\]", text)
    if m3:
        candidate = m3.group(1).strip()
        if is_valid_code(candidate):
            return candidate

    # 铆聦篓铆聞麓 2: 矛陆聹毛隆聽 毛聮陇 矛陆聰毛聯聹
    if ":" in text:
        after_colon = text.rsplit(":", 1)[1].strip()
        if is_valid_code(after_colon):
            return after_colon

    # 铆聦篓铆聞麓 3: 矛聵碌矛聟聵 铆聟聧矛聤陇铆聤赂 矛聽聞矛虏麓锚掳聙 矛陆聰毛聯聹 (矛聵聢: "LS-420HM", "LS-UHS2SR", "LS-WPCOP-C6")
    if is_valid_code(text):
        return text

    # 铆聦篓铆聞麓 4: LS/LSP/LSN/ZOT 毛隆聹 矛聥聹矛聻聭铆聲聵毛聤聰 矛陆聰毛聯聹锚掳聙 铆聟聧矛聤陇铆聤赂 毛聜麓 铆聫卢铆聲篓
    # (矛聵聢: "1. LS-U61MH", "0.5M LS-HF7005", "毛陋篓毛聥聢铆聞掳 4锚掳聹 矛聴掳锚虏掳(LS-UCHD4) 毛娄卢铆聧录矛聽聹铆聮聢")
    # LS- 铆聵聲矛聥聺矛虏聵毛聼录 矛聽聭毛聭聬矛聳麓 毛聥陇矛聺聦矛聴聬 铆聲聵矛聺麓铆聰聢矛聺麓 矛聵陇毛聤聰 锚虏陆矛職掳毛聫聞 铆聫卢铆聲篓
    m4 = re.search(r'\b((?:LS[PNT]?|ZOT)[A-Za-z0-9\-\.]{2,})', text, re.IGNORECASE)
    if m4:
        candidate = m4.group(1).rstrip('-.')
        if is_valid_code(candidate) and len(candidate) >= 4:
            return candidate

    return None


def _match_item_code(order: dict) -> Optional[str]:
    """
    ERP 铆聮聢毛陋漏矛陆聰毛聯聹 锚虏掳矛聽聲 矛職掳矛聞聽矛聢聹矛聹聞:
      1) 矛聥聹铆聤赂2 矛聵陇毛虏聞毛聺录矛聺麓毛聯聹 (矛聵碌矛聟聵矛聝聛铆聮聢, 矛聜卢矛職漏矛聻聬 矛聢聵毛聫聶 矛搂聙矛聽聲)
      2a) 矛聵碌矛聟聵铆聟聧矛聤陇铆聤赂 矛搂聛矛聽聭毛搂陇铆聲聭 ("矛聝聛铆聮聢毛虏聢铆聵赂|矛聵碌矛聟聵锚掳聮" 芒聠聮 ERP矛陆聰毛聯聹)
      2b) 矛聵碌矛聟聵 铆聟聧矛聤陇铆聤赂 矛聻聬毛聫聶 矛露聰矛露聹 芒聠聮 矛陆聰毛聯聹毛鲁聞矛鹿颅毛搂碌 芒聠聮 毛陋篓毛聧赂矛聴颅毛掳漏铆聳楼毛搂碌
      3) 矛聥聹铆聤赂1 (毛漏聰矛聺赂矛聝聛铆聮聢, 矛聝聛铆聮聢毛虏聢铆聵赂 锚赂掳矛陇聙)
      4a) 矛露聰锚掳聙矛聝聛铆聮聢铆聟聧矛聤陇铆聤赂 矛搂聛矛聽聭毛搂陇铆聲聭 ("矛聝聛铆聮聢毛虏聢铆聵赂|矛露聰锚掳聙矛聝聛铆聮聢锚掳聮" 芒聠聮 ERP矛陆聰毛聯聹)
      4b) 矛聥聹铆聤赂3 (矛露聰锚掳聙矛聝聛铆聮聢, 矛聝聛铆聮聢毛虏聢铆聵赂 锚赂掳矛陇聙)
    """
    option_text = (order.get("optionInfo", "") or "").strip()
    addon_text  = (order.get("addProductInfo", "") or "").strip()
    product_no  = str(order.get("productNo", "") or order.get("productId", "") or "")

    if option_text:
        # 1) 矛聥聹铆聤赂2 矛聵陇毛虏聞毛聺录矛聺麓毛聯聹 (矛聝聛铆聮聢毛虏聢铆聵赂 芒聠聮 ERP)
        if product_no and product_no in _option_override_map:
            code = _option_override_map[product_no]
            logger.info(f"[SS] 矛聵碌矛聟聵矛聵陇毛虏聞毛聺录矛聺麓毛聯聹(矛聥聹铆聤赂2): {product_no} 芒聠聮 {code}")
            return code
        # 2a) 矛聵碌矛聟聵铆聟聧矛聤陇铆聤赂 矛搂聛矛聽聭毛搂陇铆聲聭 (矛聝聛铆聮聢毛虏聢铆聵赂|矛聵碌矛聟聵锚掳聮 芒聠聮 ERP)
        opt_key = f"{product_no}|{option_text}"
        if opt_key in _option_text_map:
            code = _option_text_map[opt_key]
            logger.info(f"[SS] 矛聵碌矛聟聵铆聟聧矛聤陇铆聤赂矛搂聛矛聽聭毛搂陇铆聲聭: '{option_text[:40]}' 芒聠聮 {code}")
            return code
        # 2a-2) 矛露聰锚掳聙矛聝聛铆聮聢毛搂碌矛聴聬矛聞聹毛聫聞 optionInfo毛隆聹 锚虏聙矛聝聣 (矛露聰锚掳聙矛聝聛铆聮聢矛聺麓 productOption矛聹录毛隆聹 毛聞聵矛聳麓矛聵陇毛聤聰 锚虏陆矛職掳)
        if product_no and opt_key in _addon_text_map:
            code = _addon_text_map[opt_key]
            logger.info(f"[SS] 矛露聰锚掳聙矛聝聛铆聮聢铆聟聧矛聤陇铆聤赂(optionInfo锚虏陆矛聹聽): '{option_text[:40]}' 芒聠聮 {code}")
            return code
        # 2b) 矛聻聬毛聫聶 矛露聰矛露聹 芒聠聮 矛陆聰毛聯聹毛鲁聞矛鹿颅毛搂碌 芒聠聮 毛陋篓毛聧赂矛聴颅毛掳漏铆聳楼毛搂碌
        code = _extract_erp_code_from_option(option_text)
        if code:
            erp_code = _code_alias_map.get(code) or _model_to_erp_map.get(code, code)
            logger.info(f"[SS] 矛聵碌矛聟聵矛聻聬毛聫聶矛露聰矛露聹: '{option_text[:40]}' 芒聠聮 矛陆聰毛聯聹:{code} 芒聠聮 ERP:{erp_code}")
            return erp_code

    # 3) 矛聥聹铆聤赂1 毛漏聰矛聺赂矛聝聛铆聮聢
    if product_no and product_no in _product_map:
        code = _product_map[product_no]
        logger.info(f"[SS] 毛漏聰矛聺赂矛聝聛铆聮聢(矛聥聹铆聤赂1): {product_no} 芒聠聮 {code}")
        return code

    # 4a) 矛露聰锚掳聙矛聝聛铆聮聢铆聟聧矛聤陇铆聤赂 矛搂聛矛聽聭毛搂陇铆聲聭
    if addon_text and product_no:
        addon_key = f"{product_no}|{addon_text}"
        if addon_key in _addon_text_map:
            code = _addon_text_map[addon_key]
            logger.info(f"[SS] 矛露聰锚掳聙矛聝聛铆聮聢铆聟聧矛聤陇铆聤赂矛搂聛矛聽聭毛搂陇铆聲聭: '{addon_text[:40]}' 芒聠聮 {code}")
            return code
        # 矛露聰锚掳聙矛聝聛铆聮聢毛聫聞 矛陆聰毛聯聹 矛聻聬毛聫聶矛露聰矛露聹 矛聥聹毛聫聞
        code = _extract_erp_code_from_option(addon_text)
        if code:
            erp_code = _code_alias_map.get(code) or _model_to_erp_map.get(code, code)
            logger.info(f"[SS] 矛露聰锚掳聙矛聝聛铆聮聢矛聻聬毛聫聶矛露聰矛露聹: '{addon_text[:40]}' 芒聠聮 矛陆聰毛聯聹:{code} 芒聠聮 ERP:{erp_code}")
            return erp_code

    # 4b) 矛聥聹铆聤赂3 矛露聰锚掳聙矛聝聛铆聮聢 (矛聝聛铆聮聢毛虏聢铆聵赂 锚赂掳矛陇聙)
    if product_no and product_no in _addon_map:
        code = _addon_map[product_no]
        logger.info(f"[SS] 矛露聰锚掳聙矛聝聛铆聮聢(矛聥聹铆聤赂3): {product_no} 芒聠聮 {code}")
        return code

    seller_code = (order.get("sellerProductCode", "") or "").strip()
    logger.warning(
        f"[SS] 毛搂陇矛鹿颅 矛聥陇铆聦篓: productNo={product_no}, option='{option_text[:30]}', "
        f"sellerCode={seller_code}, name={order.get('productName','')[:40]}"
    )
    return None


def _is_excluded(order: dict) -> bool:
    # 矛陇聭矛虏漏 锚碌卢矛隆掳(_rawOrders: {productOrder: {...}})矛聶聙 铆聫聣铆聝聞铆聶聰 锚碌卢矛隆掳(锚路赂毛拢鹿 毛聜麓毛露聙 dict) 毛陋篓毛聭聬 矛搂聙矛聸聬
    po = order.get("productOrder") or {}
    name   = po.get("productName", "")   or order.get("productName", "")   or ""
    option = po.get("productOption", "") or po.get("optionInfo", "") or \
             order.get("optionInfo", "") or order.get("productOption", "") or ""
    combined = (name + " " + option).lower()
    return any(kw in combined for kw in EXCLUDE_KEYWORDS)


def _build_goods_nm(orders_in_group: list[dict]) -> str:
    model_qty = {}
    for o in orders_in_group:
        code = _match_item_code(o) or "UNKNOWN"
        qty = int(o.get("quantity", 1) or 1)
        model_qty[code] = model_qty.get(code, 0) + qty
    return ", ".join(f"{c}({q})" for c, q in model_qty.items())


@router.get("/token-test")
async def token_test():
    import httpx
    # 矛聞聹毛虏聞 outbound IP 铆聶聲矛聺赂 (毛聞陇矛聺麓毛虏聞 IP 铆聶聰矛聺麓铆聤赂毛娄卢矛聤陇铆聤赂 毛聯卤毛隆聺矛職漏)
    server_ip = None
    try:
        async with httpx.AsyncClient(timeout=5) as c:
            r = await c.get("https://api.ipify.org?format=json")
            server_ip = r.json().get("ip")
    except Exception:
        pass

    try:
        from services.naver_client import naver_client
        token = await naver_client.get_token()
        return {"success": True, "token_prefix": token[:20] + "..." if token else None, "server_ip": server_ip}
    except Exception as e:
        return {"success": False, "error": str(e), "server_ip": server_ip}


@router.get("/orders")
async def fetch_orders(
    date_from: Optional[str] = Query(None, description="矛聥聹矛聻聭矛聺录 YYYY-MM-DD"),
    date_to: Optional[str] = Query(None, description="矛垄聟毛拢聦矛聺录 YYYY-MM-DD"),
    order_type: str = Query("NEW_BEFORE", description="NEW_BEFORE|NEW_AFTER|DELIVERING"),
):
    try:
        from services.naver_client import naver_client
        orders = await naver_client.fetch_orders(
            date_from=date_from, date_to=date_to, order_type=order_type,
        )
        return {"success": True, "orders": orders, "count": len(orders)}
    except Exception as e:
        logger.error(f"[SS] 矛拢录毛卢赂矛聢聵矛搂聭 矛聵陇毛楼聵: {e}", exc_info=True)
        return {"success": False, "error": str(e), "orders": []}


class SendErpRequest(BaseModel):
    orders: list[dict]
    emp_cd: str = ""
    wh_cd: str = ""  # "10"=矛職漏矛聜掳, "30"=铆聠碌矛搂聞, ""=锚赂掳毛鲁赂锚掳聮

@router.post("/send-erp")
async def send_erp_only(
    req: SendErpRequest,
):
    """ERP 铆聦聬毛搂陇矛聽聞铆聭聹毛搂聦 矛聽聞矛聠隆 (毛隆聹矛聽聽 毛炉赂铆聫卢铆聲篓)"""
    selected_orders = req.orders
    _emp_cd = req.emp_cd or SMARTSTORE_EMP_CODE
    _wh_cd = req.wh_cd if req.wh_cd else SMARTSTORE_WH_CODE
    from services.erp_client_ss import ERPClientSS

    if not selected_orders:
        return {"success": True, "message": "矛聞聽铆聝聺毛聬聹 矛拢录毛卢赂矛聺麓 矛聴聠矛聤碌毛聥聢毛聥陇.", "lines": 0}

    order_groups = {}
    unmatched_items = []

    order_shipping: dict = {}   # orderId 芒聠聮 毛掳掳矛聠隆毛鹿聞 锚赂聢矛聲隆
    for o in selected_orders:
        od = o.get("order", {})
        po = o.get("productOrder", {})
        oid = od.get("orderId", "")
        poid = po.get("productOrderId", "")
        if not oid or not poid:
            continue
        if oid not in order_groups:
            order_groups[oid] = []
            # 毛掳掳矛聠隆毛鹿聞: productOrder.deliveryFeeAmount
            fee = float(po.get("deliveryFeeAmount", 0) or 0)
            order_shipping[oid] = fee
        product_id = str(po.get("productId", "") or po.get("productNo", "") or "")
        seller_code = po.get("sellerProductCode", "") or ""
        order_groups[oid].append({
            "orderId": oid, "productOrderId": poid,
            "productName": po.get("productName", ""),
            "productNo": product_id, "productId": product_id,
            "sellerProductCode": seller_code,
            "optionInfo": po.get("productOption", "") or seller_code,
            "quantity": po.get("quantity", 1),
            "settlementAmount": po.get("expectedSettlementAmount", 0) or po.get("totalPaymentAmount", 0),
            "rcvName": po.get("shippingAddress", {}).get("name", ""),
        })

    DELIVERY_PROD_CD = "DEL-毛搂陇矛露聹毛掳掳002"
    erp_lines = []

    for oid, group in order_groups.items():
        for o in group:
            # 矛搂聛矛聽聭 矛聞聽铆聝聺铆聲麓矛聞聹 ERP 矛聽聞矛聠隆铆聲聵毛聤聰 锚虏陆矛職掳 矛聽聹矛聶赂 铆聜陇矛聸聦毛聯聹 毛卢麓矛聥聹 (矛聜卢矛職漏矛聻聬 毛陋聟矛聥聹矛聽聛 矛聞聽铆聝聺 矛職掳矛聞聽)
            code = _match_item_code(o)
            qty = int(o.get("quantity", 1) or 1)
            settle = float(o.get("settlementAmount", 0) or 0)
            if code:
                erp_lines.append({"prod_cd": code, "prod_name": o.get("productName", ""), "qty": qty, "price": round(settle / qty, 2) if qty else 0, "rcv_name": o.get("rcvName", "")})
            else:
                unmatched_items.append({
                    "orderId": oid,
                    "productOrderId": o.get("productOrderId", ""),
                    "productNo": o.get("productNo", "") or o.get("productId", ""),
                    "productName": o.get("productName", ""),
                    "optionInfo": o.get("optionInfo", ""),
                    "quantity": qty, "settlementAmount": settle,
                })

    # 毛掳掳矛聠隆毛鹿聞: 锚赂聢矛聲隆毛鲁聞毛隆聹 毛卢露矛聳麓矛聞聹 (锚掳聶矛聺聙 锚赂聢矛聲隆 芒聠聮 矛聢聵毛聼聣 铆聲漏矛聜掳, 毛聥陇毛楼赂 锚赂聢矛聲隆 芒聠聮 毛鲁聞毛聫聞 毛聺录矛聺赂)
    from collections import defaultdict
    delivery_by_fee: dict = defaultdict(int)
    for oid in order_groups:
        fee = order_shipping.get(oid, 0)
        delivery_by_fee[fee] += 1
    for fee_amount, count in delivery_by_fee.items():
        erp_lines.append({"prod_cd": DELIVERY_PROD_CD, "qty": count, "price": int(fee_amount), "rcv_name": order_groups[oid][0].get("rcvName", "") if order_groups.get(oid) else ""})

    if not erp_lines:
        return {"success": False, "error": "ERP 矛聽聞矛聠隆 毛聦聙矛聝聛 矛聴聠矛聺聦", "unmatched_items": unmatched_items}

    if not SMARTSTORE_CUST_CODE:
        return {"success": False, "error": "SMARTSTORE_CUST_CODE 毛炉赂矛聞陇矛聽聲"}

    # 毛掳聹矛拢录铆聶聲矛聺赂 毛聦聙矛聝聛 productOrderId 矛聢聵矛搂聭
    all_po_ids = []
    for o in selected_orders:
        po = o.get("productOrder", {})
        poid = po.get("productOrderId", "")
        if poid:
            all_po_ids.append(poid)

    try:
        erp = ERPClientSS()
        await erp.ensure_session()
        r = await erp.save_sale(SMARTSTORE_CUST_CODE, erp_lines, _wh_cd, _emp_cd)
        delivery_count = len(delivery_by_fee)
        r["lines"] = len(erp_lines)
        r["erp_matched"] = len(erp_lines) - delivery_count
        r["erp_unmatched"] = len(unmatched_items)
        r["unmatched_items"] = unmatched_items

        # ERP 矛聽聞矛聠隆 矛聞卤锚鲁碌 矛聥聹 毛聞陇矛聺麓毛虏聞 毛掳聹矛拢录铆聶聲矛聺赂 矛虏聵毛娄卢 芒聠聮 "矛聥聽锚路聹矛拢录毛卢赂(毛掳聹矛拢录 铆聸聞)"毛隆聹 矛聺麓毛聫聶
        if r.get("success") and all_po_ids:
            from services.naver_client import naver_client
            confirm_result = await naver_client.confirm_orders(all_po_ids)
            r["confirm"] = confirm_result
            logger.info(f"[SS] 毛掳聹矛拢录铆聶聲矛聺赂: {confirm_result.get('confirmed', 0)}锚卤麓")

        return r
    except Exception as e:
        logger.error(f"[SS] ERP 矛聽聞矛聠隆 矛聵陇毛楼聵: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


class ExcludedSendErpRequest(BaseModel):
    orders: list[dict]
    emp_cd: str = ""

@router.post("/excluded-send-erp")
async def excluded_send_erp(
    req: ExcludedSendErpRequest,
):
    """矛聽聹矛聶赂 铆聜陇矛聸聦毛聯聹 矛拢录毛卢赂 芒聠聮 ERP 铆聦聬毛搂陇矛聽聞铆聭聹 矛聽聞矛聠隆 (毛鹿聞锚鲁聽矛聜卢铆聲颅矛聴聬 锚虏陆毛聫聶铆聝聺毛掳掳矛聞聽毛露聢/矛掳漏毛露聢 矛聻聬毛聫聶 锚赂掳矛聻聟)"""
    selected_orders = req.orders
    _emp_cd = req.emp_cd or SMARTSTORE_EMP_CODE
    from services.erp_client_ss import ERPClientSS
    from collections import defaultdict

    if not selected_orders:
        return {"success": False, "error": "矛聞聽铆聝聺毛聬聹 矛拢录毛卢赂矛聺麓 矛聴聠矛聤碌毛聥聢毛聥陇."}

    # 矛聽聹矛聶赂 铆聜陇矛聸聦毛聯聹 铆聫卢铆聲篓 矛拢录毛卢赂毛搂聦 铆聲聞铆聞掳
    excluded_orders = []
    for o in selected_orders:
        if _is_excluded(o):
            excluded_orders.append(o)

    if not excluded_orders:
        return {"success": False, "error": "矛聽聹矛聶赂 铆聜陇矛聸聦毛聯聹(铆聴聢毛赂聦毛聻聶/矛聞聹毛虏聞毛聻聶/矛潞聬毛鹿聞毛聞路) 矛拢录毛卢赂矛聺麓 矛聴聠矛聤碌毛聥聢毛聥陇."}

    # orderId 锚赂掳矛陇聙 锚路赂毛拢鹿铆聶聰
    order_groups: dict = {}
    order_shipping: dict = {}
    order_feetype: dict = {}
    order_addr: dict = {}    # orderId 芒聠聮 毛掳掳矛聠隆矛搂聙 矛聽聲毛鲁麓 (毛鹿聞锚鲁聽矛聜卢铆聲颅 铆聲漏矛聜掳矛職漏)
    for o in excluded_orders:
        od = o.get("order") or {}
        po = o.get("productOrder") or {}
        oid = od.get("orderId", "") or po.get("orderId", "")
        poid = po.get("productOrderId", "")
        if not oid or not poid:
            continue
        if oid not in order_groups:
            order_groups[oid] = []
            fee = float(po.get("deliveryFeeAmount", 0) or 0)
            order_shipping[oid] = fee
            fee_type = str(po.get("shippingFeeType") or od.get("shippingFeeType") or "")
            order_feetype[oid] = fee_type
            # 毛掳掳矛聠隆矛搂聙 矛聽聲毛鲁麓 矛聢聵矛搂聭
            addr = po.get("shippingAddress") or {}
            rcv       = addr.get("name", "") or ""
            tel       = addr.get("tel2", "") or addr.get("tel1", "") or ""
            full_addr = ((addr.get("baseAddress", "") or "") + " " + (addr.get("detailedAddress", "") or "")).strip()
            cust_msg  = (po.get("shippingMemo") or po.get("deliveryMemo") or
                         po.get("deliveryMessage") or od.get("shippingMemo") or
                         od.get("deliveryMemo") or "")
            order_addr[oid] = {"rcv": rcv, "tel": tel, "addr": full_addr, "msg": cust_msg}
        product_id = str(po.get("productId", "") or po.get("productNo", "") or "")
        seller_code = po.get("sellerProductCode", "") or ""
        order_groups[oid].append({
            "orderId": oid, "productOrderId": poid,
            "productName": po.get("productName", ""),
            "productNo": product_id, "productId": product_id,
            "sellerProductCode": seller_code,
            "optionInfo": po.get("productOption", "") or seller_code,
            "quantity": po.get("quantity", 1),
            "settlementAmount": po.get("expectedSettlementAmount", 0) or po.get("totalPaymentAmount", 0),
        })

    DELIVERY_PROD_CD = "DEL-毛搂陇矛露聹毛掳掳002"
    erp_lines = []
    unmatched_items = []

    for oid, group in order_groups.items():
        # 矛聞聽毛露聢/矛掳漏毛露聢 铆聦聬毛聥篓
        fee_type = order_feetype.get(oid, "")
        if "矛掳漏毛露聢" in fee_type or fee_type.upper() in ("COLLECT", "COD"):
            delivery_type = "锚虏陆毛聫聶铆聝聺毛掳掳矛掳漏毛露聢"
        else:
            delivery_type = "锚虏陆毛聫聶铆聝聺毛掳掳矛聞聽毛露聢"

        # 毛鹿聞锚鲁聽矛聜卢铆聲颅: 锚虏陆毛聫聶铆聝聺毛掳掳矛聞聽毛露聢/矛掳漏毛露聢 / 矛聽聞铆聭聹矛聽聹矛聶赂 / 矛聢聵毛聽鹿矛聺赂 / 矛聴掳毛聺陆矛虏聵 / 矛拢录矛聠聦 / 毛掳掳矛聠隆毛漏聰矛聞赂矛搂聙
        ai = order_addr.get(oid, {})
        parts = [f"{delivery_type} / 矛聽聞铆聭聹矛聽聹矛聶赂"]
        if ai.get("rcv"):   parts.append(ai["rcv"])
        if ai.get("tel"):   parts.append(ai["tel"])
        if ai.get("addr"):  parts.append(ai["addr"])
        if ai.get("msg"):   parts.append(ai["msg"])
        remark = " / ".join(parts)

        for o in group:
            code = _match_item_code(o)
            qty = int(o.get("quantity", 1) or 1)
            settle = float(o.get("settlementAmount", 0) or 0)
            if code:
                erp_lines.append({"prod_cd": code, "qty": qty,
                                   "price": round(settle / qty, 2) if qty else 0,
                                   "remark": remark})
            else:
                unmatched_items.append({
                    "orderId": oid,
                    "productOrderId": o.get("productOrderId", ""),
                    "productNo": o.get("productNo", "") or o.get("productId", ""),
                    "productName": o.get("productName", ""),
                    "optionInfo": o.get("optionInfo", ""),
                    "quantity": qty, "settlementAmount": settle,
                })

    # 毛掳掳矛聠隆毛鹿聞 毛聺录矛聺赂 (锚虏陆毛聫聶铆聝聺毛掳掳 毛掳掳矛聠隆毛鹿聞) 芒聙聰 毛鹿聞锚鲁聽矛聜卢铆聲颅 毛聫聶矛聺录铆聲聵锚虏聦 铆聫卢铆聲篓
    # (毛陋篓毛聯聽 毛聺录矛聺赂矛聴聬 CHAR5锚掳聙 矛聻聢矛聳麓矛聲录 Ecount锚掳聙 毛搂聢矛搂聙毛搂聣 毛聺录矛聺赂矛聹录毛隆聹 毛聧庐矛聳麓矛聯掳矛搂聙 矛聲聤矛聺聦)
    first_remark = erp_lines[0]["remark"] if erp_lines and erp_lines[0].get("remark") else ""
    delivery_by_fee: dict = defaultdict(int)
    for oid in order_groups:
        fee = order_shipping.get(oid, 0)
        delivery_by_fee[fee] += 1
    for fee_amount, count in delivery_by_fee.items():
        erp_lines.append({"prod_cd": DELIVERY_PROD_CD, "qty": count, "price": int(fee_amount),
                           "remark": first_remark})

    if not erp_lines:
        return {"success": False, "error": "ERP 矛聽聞矛聠隆 毛聦聙矛聝聛 矛聴聠矛聺聦", "unmatched_items": unmatched_items}

    if not SMARTSTORE_CUST_CODE:
        return {"success": False, "error": "SMARTSTORE_CUST_CODE 毛炉赂矛聞陇矛聽聲"}

    # 毛掳聹矛拢录铆聶聲矛聺赂 毛聦聙矛聝聛 productOrderId 矛聢聵矛搂聭
    all_po_ids = []
    for o in excluded_orders:
        po = o.get("productOrder") or {}
        poid = po.get("productOrderId", "")
        if poid:
            all_po_ids.append(poid)

    try:
        erp = ERPClientSS()
        await erp.ensure_session()
        r = await erp.save_sale(SMARTSTORE_CUST_CODE, erp_lines, SMARTSTORE_WH_CODE, _emp_cd)
        r["lines"] = len(erp_lines)
        r["erp_matched"] = len(erp_lines)
        r["erp_unmatched"] = len(unmatched_items)
        r["unmatched_items"] = unmatched_items
        logger.info(f"[SS] 锚虏陆毛聫聶铆聝聺毛掳掳 ERP 矛聽聞矛聠隆: {len(erp_lines)}锚卤麓, 毛炉赂毛搂陇矛鹿颅: {len(unmatched_items)}锚卤麓")

        # ERP 矛聽聞矛聠隆 矛聞卤锚鲁碌 矛聥聹 毛聞陇矛聺麓毛虏聞 毛掳聹矛拢录铆聶聲矛聺赂 矛虏聵毛娄卢 芒聠聮 "矛聥聽锚路聹矛拢录毛卢赂(毛掳聹矛拢录 铆聸聞)"毛隆聹 矛聺麓毛聫聶
        if r.get("success") and all_po_ids:
            from services.naver_client import naver_client
            confirm_result = await naver_client.confirm_orders(all_po_ids)
            r["confirm"] = confirm_result
            logger.info(f"[SS] 锚虏陆毛聫聶 毛掳聹矛拢录铆聶聲矛聺赂: {confirm_result.get('confirmed', 0)}锚卤麓")

        return r
    except Exception as e:
        logger.error(f"[SS] 锚虏陆毛聫聶铆聝聺毛掳掳 ERP 矛聽聞矛聠隆 矛聵陇毛楼聵: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


@router.post("/excluded-export-excel")
async def excluded_export_excel(
    selected_orders: list[dict] = Body(...),
):
    """矛聽聹矛聶赂 铆聜陇矛聸聦毛聯聹 矛拢录毛卢赂 芒聠聮 锚虏陆毛聫聶铆聝聺毛掳掳 矛聽聞铆聭聹矛職漏 矛聴聭矛聟聙 毛聥陇矛職麓毛隆聹毛聯聹"""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    # 矛聽聹矛聶赂 铆聜陇矛聸聦毛聯聹 铆聫卢铆聲篓 矛拢录毛卢赂毛搂聦 铆聲聞铆聞掳 (orderId 锚赂掳矛陇聙 锚路赂毛拢鹿)
    excluded_oids: set = set()
    for o in selected_orders:
        po = o.get("productOrder") or {}
        combined = ((po.get("productName", "") or "") + " " + (po.get("productOption", "") or "")).lower()
        if any(kw in combined for kw in EXCLUDE_KEYWORDS):
            od = o.get("order") or {}
            oid = od.get("orderId", "") or po.get("orderId", "")
            if oid:
                excluded_oids.add(oid)
    logger.info(f"[SS] excluded-export-excel: 矛聽聹矛聶赂 矛拢录毛卢赂 {len(excluded_oids)}锚掳聹 orderId={excluded_oids}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "锚虏陆毛聫聶铆聝聺毛掳掳矛聽聞铆聭聹"

    headers = ["毛鹿聞锚鲁聽矛聜卢铆聲颅", "矛聢聵毛聽鹿矛聺赂", "矛聴掳毛聺陆矛虏聵", "矛拢录矛聠聦", "毛掳掳矛聠隆毛漏聰矛聞赂矛搂聙"]
    hdr_fill = PatternFill("solid", fgColor="C00000")
    hdr_font = Font(bold=True, color="FFFFFF")
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 25

    seen: set = set()
    row = 2
    for o in selected_orders:
        try:
            od = o.get("order") or {}
            po = o.get("productOrder") or {}
            oid = od.get("orderId", "") or po.get("orderId", "")
            if not oid or oid not in excluded_oids or oid in seen:
                continue
            seen.add(oid)
            addr      = po.get("shippingAddress") or {}
            rcv       = addr.get("name", "") or ""
            tel       = addr.get("tel2", "") or addr.get("tel1", "") or ""
            full_addr = ((addr.get("baseAddress", "") or "") + " " + (addr.get("detailedAddress", "") or "")).strip()
            cust_msg  = (po.get("shippingMemo") or po.get("deliveryMemo") or
                         po.get("deliveryMessage") or od.get("shippingMemo") or
                         od.get("deliveryMemo") or "")

            fee_type = str(po.get("shippingFeeType") or od.get("shippingFeeType") or "")
            if "矛掳漏毛露聢" in fee_type or fee_type.upper() in ("COLLECT", "COD"):
                delivery_type = "锚虏陆毛聫聶铆聝聺毛掳掳矛掳漏毛露聢"
            else:
                delivery_type = "锚虏陆毛聫聶铆聝聺毛掳掳矛聞聽毛露聢"

            ws.cell(row, 1, f"{delivery_type} / 矛聽聞铆聭聹矛聽聹矛聶赂")
            ws.cell(row, 2, rcv)
            ws.cell(row, 3, tel)
            ws.cell(row, 4, full_addr)
            ws.cell(row, 5, cust_msg)
            row += 1
        except Exception as ex:
            logger.error(f"[SS] excluded-export-excel 铆聳聣 矛虏聵毛娄卢 矛聵陇毛楼聵: {ex}", exc_info=True)
            continue

    if row == 2:
        raise HTTPException(status_code=404, detail="矛聽聹矛聶赂 铆聜陇矛聸聦毛聯聹(铆聴聢毛赂聦毛聻聶/矛聞聹毛虏聞毛聻聶/矛潞聬毛鹿聞毛聞路) 矛拢录毛卢赂矛聺麓 矛聴聠矛聤碌毛聥聢毛聥陇.")

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    from urllib.parse import quote
    filename = f"锚虏陆毛聫聶铆聝聺毛掳掳矛聽聞铆聭聹_{datetime.now(KST).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    )


@router.post("/logen-export-excel")
async def logen_export_excel(
    selected_orders: list[dict] = Body(...),
):
    """矛聞聽铆聝聺 矛拢录毛卢赂矛聺聞 毛隆聹矛聽聽 矛聽聞矛聠隆矛職漏 矛聴聭矛聟聙毛隆聹 毛聥陇矛職麓毛隆聹毛聯聹.
    矛禄卢毛聼录: 矛拢录毛卢赂毛虏聢铆聵赂 | 矛聝聛铆聮聢矛拢录毛卢赂毛虏聢铆聵赂 | 矛聢聵毛聽鹿矛聺赂 | 矛聴掳毛聺陆矛虏聵 | 矛拢录矛聠聦 | 矛聝聛铆聮聢毛陋聟 | 矛聢聵毛聼聣
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "矛聴聭矛聟聙铆聦聦矛聺录矛虏芦铆聳聣-矛聽聹毛陋漏矛聻聢矛聺聦"

    # 毛隆聹矛聽聽 锚碌卢矛聥聹矛聤陇铆聟聹矛聳聭矛聥聺 (A铆聝聙矛聻聟) - 矛聜卢矛職漏矛聻聬 矛聞陇矛聽聲 矛禄卢毛聼录 矛聢聹矛聞聹
    # A:矛聢聵铆聲聵矛聺赂毛陋聟 B:矛聢聵铆聲聵矛聺赂矛拢录矛聠聦1 C:矛聢聵铆聲聵矛聺赂矛聽聞铆聶聰 D:矛聢聵铆聲聵矛聺赂铆聹麓毛聦聙铆聫掳
    # E:铆聝聺毛掳掳矛聢聵毛聼聣 F:铆聝聺毛掳掳矛職麓矛聻聞 G:矛職麓矛聻聞锚碌卢毛露聞 H:毛卢录铆聮聢毛陋聟 I:矛拢录毛卢赂毛虏聢铆聵赂(芒聠聮毛掳聵铆聶聵铆聦聦矛聺录 S矛聴麓 毛搂陇矛鹿颅矛職漏)
    # J:矛聽聹矛拢录矛職麓矛聻聞锚碌卢毛露聞 K:毛掳掳矛聠隆毛漏聰矛聞赂矛搂聙
    headers = ["矛聢聵铆聲聵矛聺赂毛陋聟", "矛聢聵铆聲聵矛聺赂矛拢录矛聠聦1", "矛聢聵铆聲聵矛聺赂矛聽聞铆聶聰", "矛聢聵铆聲聵矛聺赂铆聹麓毛聦聙铆聫掳",
               "铆聝聺毛掳掳矛聢聵毛聼聣", "铆聝聺毛掳掳矛職麓矛聻聞", "矛職麓矛聻聞锚碌卢毛露聞", "毛卢录铆聮聢毛陋聟", "矛拢录毛卢赂毛虏聢铆聵赂",
               None, "毛掳掳矛聠隆毛漏聰矛聞赂矛搂聙"]
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF")
    for ci, h in enumerate(headers, 1):
        if h:
            c = ws.cell(1, ci, h)
            c.fill = hdr_fill; c.font = hdr_font; c.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 30
    ws.column_dimensions["I"].width = 22
    ws.column_dimensions["K"].width = 25

    # orderId 锚赂掳矛陇聙矛聹录毛隆聹 锚路赂毛拢鹿铆聶聰
    groups: dict = {}
    for o in selected_orders:
        od = o.get("order", {})
        po = o.get("productOrder", {})
        oid = od.get("orderId", "")
        if not oid:
            continue
        if oid not in groups:
            groups[oid] = {"od": od, "po": po, "items": []}
        groups[oid]["items"].append(po)

    row = 2
    for oid, g in groups.items():
        od = g["od"]
        po = g["po"]   # 矛虏芦 毛虏聢矛搂赂 productOrder (矛聢聵毛聽鹿矛聺赂 矛聽聲毛鲁麓矛職漏)
        addr     = po.get("shippingAddress", {})
        rcv      = addr.get("name", "")
        tel_home = addr.get("tel1", "")
        tel_cell = addr.get("tel2", "") or addr.get("tel1", "")
        full_addr = ((addr.get("baseAddress","") or "") + " " + (addr.get("detailedAddress","") or "")).strip()
        ship_fee = int(float(po.get("deliveryFeeAmount", 0) or 0))
        fare_tp  = "010"

        # 铆聮聢毛陋漏毛陋聟: 毛陋篓毛聧赂毛陋聟(or ERP矛陆聰毛聯聹) + 矛聢聵毛聼聣 矛職聰矛聲陆
        model_qty: dict = {}
        total_qty = 0
        first_poid = ""
        cust_msg = ""
        for item_po in g["items"]:
            if not first_poid:
                first_poid = item_po.get("productOrderId", "")
                cust_msg = (item_po.get("shippingMemo", "") or
                            item_po.get("deliveryMemo", "") or
                            item_po.get("deliveryMessage", "") or
                            od.get("shippingMemo", "") or
                            od.get("deliveryMemo", "") or "")
            product_id = str(item_po.get("productId", "") or item_po.get("productNo", "") or "")
            # 毛陋篓毛聧赂毛陋聟 矛職掳矛聞聽, 矛聴聠矛聹录毛漏麓 ERP矛陆聰毛聯聹, 矛聴聠矛聹录毛漏麓 矛聝聛铆聮聢毛陋聟
            model = _model_map.get(product_id, "")
            if not model:
                model = _product_map.get(product_id, "") or item_po.get("productName", "")[:20]
            qty = int(item_po.get("quantity", 1) or 1)
            model_qty[model] = model_qty.get(model, 0) + qty
            total_qty += qty

        goods = ", ".join(f"{m} x{q}" for m, q in model_qty.items())[:50]

        ws.cell(row, 1,  rcv)         # A: 矛聢聵铆聲聵矛聺赂毛陋聟
        ws.cell(row, 2,  full_addr)   # B: 矛聢聵铆聲聵矛聺赂矛拢录矛聠聦1
        ws.cell(row, 3,  tel_home)    # C: 矛聢聵铆聲聵矛聺赂矛聽聞铆聶聰
        ws.cell(row, 4,  tel_cell)    # D: 矛聢聵铆聲聵矛聺赂铆聹麓毛聦聙铆聫掳
        ws.cell(row, 5,  1)            # E: 铆聝聺毛掳掳矛聢聵毛聼聣 (毛掳聲矛聤陇 矛聢聵毛聼聣, 铆聲颅矛聝聛 1)
        ws.cell(row, 6,  ship_fee)    # F: 铆聝聺毛掳掳矛職麓矛聻聞
        ws.cell(row, 7,  fare_tp)     # G: 矛職麓矛聻聞锚碌卢毛露聞
        ws.cell(row, 8,  goods)       # H: 毛卢录铆聮聢毛陋聟 (毛陋篓毛聧赂毛陋聟+矛聢聵毛聼聣)
        ws.cell(row, 9,  first_poid)  # I: 矛拢录毛卢赂毛虏聢铆聵赂 芒聠聮 毛掳聵铆聶聵铆聦聦矛聺录 S矛聴麓(index 18)毛隆聹 毛搂陇矛鹿颅
        jeju = "矛聞聽矛掳漏毛露聢" if "矛聽聹矛拢录" in full_addr else None
        ws.cell(row, 10, jeju)        # J: 矛聽聹矛拢录矛職麓矛聻聞锚碌卢毛露聞 (矛聽聹矛拢录 矛拢录矛聠聦毛漏麓 矛聞聽矛掳漏毛露聢 矛聻聬毛聫聶)
        ws.cell(row, 11, cust_msg)    # K: 毛掳掳矛聠隆毛漏聰矛聞赂矛搂聙 (锚鲁聽锚掳聺 矛職聰矛虏颅矛聜卢铆聲颅)
        row += 1

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    filename = f"logen_{datetime.now(KST).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


@router.post("/logen-dispatch-excel")
async def logen_dispatch_excel(
    file: UploadFile = File(...),
    carrier: str = Form("LOGEN"),
):
    """矛聠隆矛聻楼毛虏聢铆聵赂 锚赂掳矛聻聟毛聬聹 矛聴聭矛聟聙 矛聴聟毛隆聹毛聯聹 芒聠聮 毛聞陇矛聺麓毛虏聞 毛掳聹矛聠隆矛虏聵毛娄卢.
    H矛聴麓(8毛虏聢矛搂赂)矛聴聬 矛聠隆矛聻楼毛虏聢铆聵赂, A矛聴麓(1毛虏聢矛搂赂)矛聴聬 矛拢录毛卢赂毛虏聢铆聵赂, B矛聴麓(2毛虏聢矛搂赂)矛聴聬 矛聝聛铆聮聢矛拢录毛卢赂毛虏聢铆聵赂
    """
    import io
    import openpyxl
    from services.naver_client import naver_client

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    dispatch_list = []
    skipped = []
    # 毛隆聹矛聽聽 毛掳聵铆聶聵 铆聦聦矛聺录 锚碌卢矛隆掳:
    #   1铆聳聣: 铆聝聙矛聺麓铆聥聙, 2铆聳聣: 铆聴陇毛聧聰, 3铆聳聣: 矛聞聹毛赂聦铆聴陇毛聧聰, 4铆聳聣~: 毛聧掳矛聺麓铆聞掳
    #   D矛聴麓(index 3): 矛職麓矛聠隆矛聻楼毛虏聢铆聵赂
    #   S矛聴麓(index 18): 矛拢录毛卢赂毛虏聢铆聵赂 芒聠聮 毛聥陇矛職麓毛隆聹毛聯聹 矛聥聹 I矛聴麓矛聴聬 矛聜陆矛聻聟铆聲聹 productOrderId
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not any(row):
            continue
        tracking = str(row[3]  or "").strip()   # D矛聴麓: 矛職麓矛聠隆矛聻楼毛虏聢铆聵赂
        poid     = str(row[18] or "").strip()   # S矛聴麓: 矛拢录毛卢赂毛虏聢铆聵赂(=productOrderId)
        if not tracking or not poid or tracking == "None" or poid == "None":
            skipped.append(poid or str(row[6] or ""))
            continue
        dispatch_list.append({
            "productOrderId": poid,
            "deliveryCompanyCode": carrier,
            "trackingNumber": tracking,
        })

    if not dispatch_list:
        return {"success": False, "error": f"矛聠隆矛聻楼毛虏聢铆聵赂锚掳聙 矛聻聟毛聽楼毛聬聹 铆聳聣矛聺麓 矛聴聠矛聤碌毛聥聢毛聥陇. (毛鹿聢 铆聳聣: {len(skipped)}锚掳聹)"}

    try:
        result = await naver_client.dispatch_orders(dispatch_list)
        result["dispatched_count"] = len(dispatch_list)
        result["skipped_count"] = len(skipped)
        logger.info(f"[SS] 矛聴聭矛聟聙毛掳聹矛聠隆矛虏聵毛娄卢: {len(dispatch_list)}锚卤麓, 锚虏掳锚鲁录={result}")
        return {"success": True, **result}
    except Exception as e:
        logger.error(f"[SS] 矛聴聭矛聟聙毛掳聹矛聠隆矛虏聵毛娄卢 矛聵陇毛楼聵: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


@router.post("/dispatch-manual")
async def dispatch_manual(
    body: dict = Body(...),
):
    """矛聢聵锚赂掳 矛聠隆矛聻楼毛虏聢铆聵赂毛隆聹 毛聞陇矛聺麓毛虏聞 毛掳聹矛聠隆矛虏聵毛娄卢.
    body: { "items": [{"productOrderId": "...", "trackingNumber": "..."}, ...] }
    """
    from services.naver_client import naver_client
    items = body.get("items", [])
    if not items:
        return {"success": False, "error": "矛聠隆矛聻楼 毛聧掳矛聺麓铆聞掳锚掳聙 矛聴聠矛聤碌毛聥聢毛聥陇."}

    dispatch_list = [
        {"productOrderId": it["productOrderId"], "deliveryCompanyCode": "LOGEN", "trackingNumber": it["trackingNumber"]}
        for it in items if it.get("productOrderId") and it.get("trackingNumber")
    ]
    if not dispatch_list:
        return {"success": False, "error": "矛聹聽铆職篓铆聲聹 矛聠隆矛聻楼毛虏聢铆聵赂锚掳聙 矛聴聠矛聤碌毛聥聢毛聥陇."}

    try:
        result = await naver_client.dispatch_orders(dispatch_list)
        result["dispatched_count"] = len(dispatch_list)
        logger.info(f"[SS] 矛聢聵锚赂掳毛掳聹矛聠隆矛虏聵毛娄卢: {len(dispatch_list)}锚卤麓, 锚虏掳锚鲁录={result}")
        return {"success": True, **result}
    except Exception as e:
        logger.error(f"[SS] 矛聢聵锚赂掳毛掳聹矛聠隆矛虏聵毛娄卢 矛聵陇毛楼聵: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


@router.post("/register-logen")
async def register_logen_only(
    warehouse: str = Query(..., pattern="^(gimpo|yongsan)$"),
    selected_orders: list[dict] = Body(...),
):
    """毛隆聹矛聽聽铆聝聺毛掳掳 毛聯卤毛隆聺 + 毛掳聹矛拢录铆聶聲矛聺赂 + 毛掳聹矛聠隆矛虏聵毛娄卢 (ERP 毛炉赂铆聫卢铆聲篓)"""
    from services.naver_client import naver_client
    from services.ilogen_client import register_orders, get_sender

    if not selected_orders:
        return {"success": True, "message": "矛聞聽铆聝聺毛聬聹 矛拢录毛卢赂矛聺麓 矛聴聠矛聤碌毛聥聢毛聥陇."}

    order_groups = {}
    all_po_ids = []

    for o in selected_orders:
        od = o.get("order", {})
        po = o.get("productOrder", {})
        oid = od.get("orderId", "")
        poid = po.get("productOrderId", "")
        if not oid or not poid:
            continue
        all_po_ids.append(poid)
        if oid not in order_groups:
            order_groups[oid] = []
        product_id = str(po.get("productId", "") or po.get("productNo", "") or "")
        seller_code = po.get("sellerProductCode", "") or ""
        order_groups[oid].append({
            "orderId": oid, "productOrderId": poid,
            "productName": po.get("productName", ""),
            "productNo": product_id, "productId": product_id,
            "sellerProductCode": seller_code,
            "optionInfo": po.get("productOption", "") or seller_code,
            "quantity": po.get("quantity", 1),
            "deliveryFeeType": po.get("shippingFeeType", ""),
            "rcvName": po.get("shippingAddress", {}).get("name", ""),
            "rcvTel": po.get("shippingAddress", {}).get("tel1", ""),
            "rcvAddr": (po.get("shippingAddress", {}).get("baseAddress", "") + " " + po.get("shippingAddress", {}).get("detailedAddress", "")).strip(),
        })

    sender = get_sender(warehouse)
    ilogen_orders = []
    oid_to_idx = {}

    for oid, group in order_groups.items():
        first = group[0]
        fare_code = "020" if "矛掳漏毛露聢" in str(first.get("deliveryFeeType", "")) else "030"
        ilogen_orders.append({
            "snd_name": sender["name"], "snd_tel": sender["tel"], "snd_addr": sender["addr"],
            "rcv_name": first["rcvName"], "rcv_tel": first["rcvTel"], "rcv_addr": first["rcvAddr"],
            "fare_code": fare_code, "goods_nm": _build_goods_nm(group),
        })
        oid_to_idx[oid] = len(ilogen_orders) - 1

    try:
        logen_res = await register_orders(warehouse, ilogen_orders)
        tns = logen_res.get("tracking_numbers", [])
        logen_ok = logen_res.get("success", False) and len(tns) > 0

        confirm_result = {"confirmed": 0, "message": "毛隆聹矛聽聽 毛聯卤毛隆聺 矛聥陇铆聦篓毛隆聹 毛鲁麓毛楼聵"}
        dispatch_result = {"dispatched": 0, "message": "毛隆聹矛聽聽 毛聯卤毛隆聺 矛聥陇铆聦篓毛隆聹 毛鲁麓毛楼聵"}

        if logen_ok:
            confirm_result = await naver_client.confirm_orders(all_po_ids)
            if confirm_result.get("confirmed", 0) > 0:
                oid_slip = {}
                for tn in tns:
                    for oid, idx in oid_to_idx.items():
                        if idx == tn["index"]:
                            oid_slip[oid] = tn["slip_no"]
                            break
                dispatch_list = []
                for oid, group in order_groups.items():
                    slip = oid_slip.get(oid)
                    if not slip:
                        continue
                    for o in group:
                        dispatch_list.append({"productOrderId": o["productOrderId"], "deliveryCompanyCode": "LOGEN", "trackingNumber": slip})
                dispatch_result = await naver_client.dispatch_orders(dispatch_list) if dispatch_list else {"success": True, "message": "毛聦聙矛聝聛 矛聴聠矛聺聦"}

        return {
            "success": logen_ok,
            "logen": logen_res,
            "confirm": confirm_result,
            "dispatch": dispatch_result,
            "tracking_count": len(tns),
            "total_orders": len(all_po_ids),
        }
    except Exception as e:
        logger.error(f"[SS] 毛隆聹矛聽聽毛聯卤毛隆聺 矛聵陇毛楼聵: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


class AutoRegisterLogenRequest(BaseModel):
    orders: list[dict]
    emp_cd: str = ""

@router.post("/auto-register-logen")
async def auto_register_logen(
    warehouse: str = Query(..., pattern="^(gimpo|yongsan)$"),
    req: AutoRegisterLogenRequest = Body(...),
):
    selected_orders = req.orders
    _emp_cd = req.emp_cd or SMARTSTORE_EMP_CODE
    from services.naver_client import naver_client
    from services.ilogen_client import register_orders, get_sender
    from services.erp_client_ss import ERPClientSS

    result = {"step1_confirm": None, "step2_erp": None, "step2_logen": None, "step3_dispatch": None, "summary": {}}

    try:
        if not selected_orders:
            return {"success": True, "message": "矛聞聽铆聝聺毛聬聹 矛拢录毛卢赂矛聺麓 矛聴聠矛聤碌毛聥聢毛聥陇.", **result}

        order_groups = {}
        all_po_ids = []

        order_shipping: dict = {}   # orderId 芒聠聮 毛掳掳矛聠隆毛鹿聞 锚赂聢矛聲隆
        for o in selected_orders:
            od = o.get("order", {})
            po = o.get("productOrder", {})
            oid = od.get("orderId", "")
            poid = po.get("productOrderId", "")
            if not oid or not poid:
                continue
            all_po_ids.append(poid)
            if oid not in order_groups:
                order_groups[oid] = []
                fee = float(po.get("deliveryFeeAmount", 0) or 0)
                order_shipping[oid] = fee
            product_id = str(po.get("productId", "") or po.get("productNo", "") or "")
            seller_code = po.get("sellerProductCode", "") or ""
            order_groups[oid].append({
                "orderId": oid, "productOrderId": poid,
                "productName": po.get("productName", ""),
                "productNo": product_id, "productId": product_id,
                "sellerProductCode": seller_code,
                "optionInfo": po.get("productOption", "") or seller_code,
                "quantity": po.get("quantity", 1),
                "settlementAmount": po.get("expectedSettlementAmount", 0) or po.get("totalPaymentAmount", 0),
                "deliveryFeeType": po.get("shippingFeeType", ""),
                "rcvName": po.get("shippingAddress", {}).get("name", ""),
                "rcvTel": po.get("shippingAddress", {}).get("tel1", ""),
                "rcvAddr": (po.get("shippingAddress", {}).get("baseAddress", "") + " " + po.get("shippingAddress", {}).get("detailedAddress", "")).strip(),
            })

        logger.info(f"[SS] 矛聞聽铆聝聺 矛拢录毛卢赂: {len(all_po_ids)}锚卤麓, {len(order_groups)}锚路赂毛拢鹿")

        # ERP 毛聺录矛聺赂 锚碌卢矛聞卤
        DELIVERY_PROD_CD = "DEL-毛搂陇矛露聹毛掳掳002"
        erp_lines = []
        unmatched_items = []

        for oid, group in order_groups.items():
            for o in group:
                if _is_excluded(o):
                    logger.info(f"[SS] 矛聽聹矛聶赂 铆聜陇矛聸聦毛聯聹 铆聲聞铆聞掳: {o.get('productName','')[:40]}")
                    continue
                code = _match_item_code(o)
                qty = int(o.get("quantity", 1) or 1)
                settle = float(o.get("settlementAmount", 0) or 0)
                if code:
                    erp_lines.append({"prod_cd": code, "qty": qty, "price": round(settle / qty, 2) if qty else 0})
                else:
                    unmatched_items.append({
                        "orderId": oid,
                        "productOrderId": o.get("productOrderId", ""),
                        "productNo": o.get("productNo", "") or o.get("productId", ""),
                        "productName": o.get("productName", ""),
                        "optionInfo": o.get("optionInfo", ""),
                        "quantity": qty, "settlementAmount": settle,
                        "deliveryMethod": o.get("deliveryFeeType", ""),
                        "rcvName": o.get("rcvName", ""),
                    })

        # 毛掳掳矛聠隆毛鹿聞: 锚赂聢矛聲隆毛鲁聞毛隆聹 毛卢露矛聳麓矛聞聹 (锚掳聶矛聺聙 锚赂聢矛聲隆 芒聠聮 矛聢聵毛聼聣 铆聲漏矛聜掳, 毛聥陇毛楼赂 锚赂聢矛聲隆 芒聠聮 毛鲁聞毛聫聞 毛聺录矛聺赂)
        from collections import defaultdict
        delivery_by_fee: dict = defaultdict(int)
        for oid in order_groups:
            fee = order_shipping.get(oid, 0)
            delivery_by_fee[fee] += 1
        for fee_amount, count in delivery_by_fee.items():
            erp_lines.append({"prod_cd": DELIVERY_PROD_CD, "qty": count, "price": int(fee_amount)})

        sender = get_sender(warehouse)
        ilogen_orders = []
        oid_to_idx = {}

        for oid, group in order_groups.items():
            first = group[0]
            fare_code = "020" if "矛掳漏毛露聢" in str(first.get("deliveryFeeType", "")) else "030"
            ilogen_orders.append({
                "snd_name": sender["name"], "snd_tel": sender["tel"], "snd_addr": sender["addr"],
                "rcv_name": first["rcvName"], "rcv_tel": first["rcvTel"], "rcv_addr": first["rcvAddr"],
                "fare_code": fare_code, "goods_nm": _build_goods_nm(group),
            })
            oid_to_idx[oid] = len(ilogen_orders) - 1

        async def _do_erp():
            if not erp_lines:
                return {"success": True, "lines": 0, "message": "ERP 矛聻聟毛聽楼 毛聦聙矛聝聛 矛聴聠矛聺聦"}
            if not SMARTSTORE_CUST_CODE:
                return {"success": False, "lines": len(erp_lines), "error": "SMARTSTORE_CUST_CODE 毛炉赂矛聞陇矛聽聲"}
            erp = ERPClientSS()
            await erp.ensure_session()
            r = await erp.save_sale(SMARTSTORE_CUST_CODE, erp_lines, SMARTSTORE_WH_CODE, _emp_cd)
            r["lines"] = len(erp_lines)
            r["sent_prod_codes"] = [l["prod_cd"] for l in erp_lines]
            return r

        async def _do_logen():
            if not ilogen_orders:
                return {"success": True, "tracking_numbers": []}
            return await register_orders(warehouse, ilogen_orders)

        erp_res, logen_res = await asyncio.gather(_do_erp(), _do_logen())
        result["step1_erp"] = erp_res
        result["step1_logen"] = logen_res

        tns = logen_res.get("tracking_numbers", [])
        erp_ok = erp_res.get("success", False)
        logen_ok = logen_res.get("success", False) and len(tns) > 0

        if erp_ok and logen_ok:
            confirm_result = await naver_client.confirm_orders(all_po_ids)
            result["step2_confirm"] = confirm_result
        else:
            reasons = []
            if not erp_ok: reasons.append("ERP 铆聦聬毛搂陇矛聻聟毛聽楼 矛聥陇铆聦篓")
            if not logen_ok: reasons.append("毛隆聹矛聽聽 矛聠隆矛聻楼毛掳聹锚赂聣 矛聥陇铆聦篓")
            result["step2_confirm"] = {"confirmed": 0, "message": f"毛掳聹矛拢录铆聶聲矛聺赂 毛鲁麓毛楼聵 ({', '.join(reasons)})"}

        if tns and erp_ok and result["step2_confirm"].get("confirmed", 0) > 0:
            oid_slip = {}
            for tn in tns:
                for oid, idx in oid_to_idx.items():
                    if idx == tn["index"]:
                        oid_slip[oid] = tn["slip_no"]
                        break
            dispatch_list = []
            for oid, group in order_groups.items():
                slip = oid_slip.get(oid)
                if not slip: continue
                for o in group:
                    dispatch_list.append({"productOrderId": o["productOrderId"], "deliveryCompanyCode": "LOGEN", "trackingNumber": slip})
            result["step3_dispatch"] = await naver_client.dispatch_orders(dispatch_list) if dispatch_list else {"success": True, "message": "毛聦聙矛聝聛 矛聴聠矛聺聦"}
        else:
            skip_reason = "ERP/毛隆聹矛聽聽 毛炉赂矛聶聞毛拢聦" if not (erp_ok and logen_ok) else "矛職麓矛聠隆矛聻楼 矛聴聠矛聺聦"
            result["step3_dispatch"] = {"dispatched": 0, "message": f"毛掳聹矛聠隆矛虏聵毛娄卢 毛鲁麓毛楼聵 ({skip_reason})"}

        result["unmatched_items"] = unmatched_items
        result["summary"] = {
            "total_orders": len(all_po_ids), "total_groups": len(order_groups),
            "erp_matched": len(erp_lines) - (1 if delivery_count > 0 else 0),
            "erp_unmatched": len(unmatched_items),
            "erp_delivery_count": delivery_count,
            "erp_lines": len(erp_lines),
            "logen_registered": len(ilogen_orders),
            "tracking_numbers": len(tns),
        }
        result["success"] = True
        return result

    except Exception as e:
        logger.error(f"[SS] 矛聻聬毛聫聶毛聯卤毛隆聺 矛聵陇毛楼聵: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# 芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬
# 矛聻卢锚鲁聽铆聵聞铆聶漏 矛隆掳铆職聦
# 芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬

@router.post("/inventory")
async def get_inventory(body: dict = Body(...)):
    """
    ERP 矛聻卢锚鲁聽铆聵聞铆聶漏 矛隆掳铆職聦 (矛掳陆锚鲁聽毛鲁聞).
    body: { "orders": [...] }
    矛職漏矛聜掳(10), 铆聠碌矛搂聞(30) 矛掳陆锚鲁聽 矛聻卢锚鲁聽毛楼录 锚掳聛锚掳聛 矛隆掳铆職聦铆聲聵矛聴卢 毛掳聵铆聶聵.
    """
    from services.erp_client_ss import ERPClientSS

    try:
        orders = body.get("orders", [])
        if not orders:
            return {"success": True, "inventory": {}, "message": "矛拢录毛卢赂 矛聴聠矛聺聦"}

        prod_codes = set()
        order_erp_map = {}
        for o in orders:
            po = o.get("productOrder", {})
            poid = po.get("productOrderId", "")
            product_id = str(po.get("productId", "") or po.get("productNo", "") or "")
            seller_code = po.get("sellerProductCode", "") or ""
            option_info = po.get("productOption", "") or seller_code

            item = {
                "productNo": product_id,
                "productId": product_id,
                "sellerProductCode": seller_code,
                "optionInfo": option_info,
                "productName": po.get("productName", ""),
            }
            code = _match_item_code(item)
            if code:
                prod_codes.add(code)
                if poid:
                    order_erp_map[poid] = code

        if not prod_codes:
            return {"success": True, "inventory": {"yongsan": {}, "tongjin": {}}, "order_erp_map": {}, "message": "毛搂陇矛鹿颅毛聬聹 铆聮聢毛陋漏矛陆聰毛聯聹 矛聴聠矛聺聦"}

        erp = ERPClientSS()
        await erp.ensure_session()
        # ECOUNT API PROD_CD 铆聲聞铆聞掳锚掳聙 矛陆陇毛搂聢 锚碌卢毛露聞矛聺聞 矛搂聙矛聸聬铆聲聵矛搂聙 矛聲聤矛聹录毛炉聙毛隆聹
        # 矛聽聞矛虏麓 矛聻卢锚鲁聽毛楼录 矛隆掳铆職聦铆聲聹 铆聸聞 Python矛聴聬矛聞聹 铆聲聞铆聞掳毛搂聛
        result = await erp.get_inventory_by_warehouses(prod_codes=None)

        # 铆聲聞矛職聰铆聲聹 铆聮聢毛陋漏矛陆聰毛聯聹毛搂聦 铆聲聞铆聞掳毛搂聛
        if result.get("success") and result.get("inventory"):
            inv = result["inventory"]
            filtered_yongsan = {k: v for k, v in inv.get("yongsan", {}).items() if k in prod_codes}
            filtered_tongjin = {k: v for k, v in inv.get("tongjin", {}).items() if k in prod_codes}
            result["inventory"] = {"yongsan": filtered_yongsan, "tongjin": filtered_tongjin}

        result["order_erp_map"] = order_erp_map

        return result
    except Exception as e:
        logger.error(f"[SS] 矛聻卢锚鲁聽矛隆掳铆職聦 矛聵陇毛楼聵: {e}", exc_info=True)
        return {"success": False, "error": str(e), "inventory": {}}



@router.get("/inventory-debug")
async def inventory_debug():
    """毛聰聰毛虏聞锚路赂: PROD_CD 矛聴聠矛聺麓 矛聻卢锚鲁聽 矛聽聞矛虏麓 矛隆掳铆職聦 (矛碌聹毛聦聙 毛陋聡锚卤麓 毛掳聵铆聶聵毛聬聵毛聤聰矛搂聙 铆聶聲矛聺赂)"""
    from services.erp_client_ss import ERPClientSS
    try:
        erp = ERPClientSS()
        await erp.ensure_session()
        r10 = await erp.get_inventory_balance(prod_codes=None, wh_cd="10")
        r30 = await erp.get_inventory_balance(prod_codes=None, wh_cd="30")
        return {
            "success": True,
            "yongsan_count": r10.get("total", 0),
            "yongsan_sample": dict(list(r10.get("inventory", {}).items())[:5]),
            "tongjin_count": r30.get("total", 0),
            "tongjin_sample": dict(list(r30.get("inventory", {}).items())[:5]),
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.post("/reload-product-map")
async def reload_product_map():
    _load_product_map()
    return {"success": True, "count": len(_product_map)}


# 芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬
# 毛搂陇铆聲聭 锚麓聙毛娄卢 API
# 芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬

@router.get("/product-map")
async def get_product_map(
    search: Optional[str] = Query(None),
):
    items = []
    for prod_no, erp_code in _product_map.items():
        model = _model_map.get(prod_no, "")
        if search:
            q = search.lower()
            if q not in prod_no.lower() and q not in erp_code.lower() and q not in model.lower():
                continue
        items.append({"productNo": prod_no, "erpCode": erp_code, "model": model})
    return {"success": True, "items": items, "total": len(_product_map), "filtered": len(items)}


@router.post("/product-map")
async def add_product_map(entry: dict = Body(...)):
    prod_no = str(entry.get("productNo", "")).strip()
    erp_code = str(entry.get("erpCode", "")).strip()
    model = str(entry.get("model", "")).strip()

    if not prod_no or not erp_code:
        return {"success": False, "error": "矛聝聛铆聮聢毛虏聢铆聵赂矛聶聙 铆聮聢毛陋漏矛陆聰毛聯聹毛聤聰 铆聲聞矛聢聵矛聻聟毛聥聢毛聥陇."}

    is_new = prod_no not in _product_map
    _product_map[prod_no] = erp_code
    if model:
        _model_map[prod_no] = model
    _save_product_map()

    action = "矛露聰锚掳聙" if is_new else "矛聢聵矛聽聲"
    logger.info(f"[SS] 毛搂陇铆聲聭 {action}: {prod_no} 芒聠聮 ERP:{erp_code}, 毛陋篓毛聧赂:{model}")
    return {"success": True, "action": action, "productNo": prod_no, "erpCode": erp_code, "model": model,
            "total": len(_product_map)}


@router.delete("/product-map/{product_no}")
async def delete_product_map(product_no: str):
    if product_no not in _product_map:
        return {"success": False, "error": f"矛聝聛铆聮聢毛虏聢铆聵赂 {product_no} 毛搂陇铆聲聭矛聺麓 矛聴聠矛聤碌毛聥聢毛聥陇."}
    erp_code = _product_map.pop(product_no)
    _model_map.pop(product_no, None)
    _save_product_map()
    logger.info(f"[SS] 毛搂陇铆聲聭 矛聜颅矛聽聹: {product_no} (was {erp_code})")
    return {"success": True, "deleted": product_no, "total": len(_product_map)}


# 芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬
# Excel 矛聴聟毛隆聹毛聯聹/毛聥陇矛職麓毛隆聹毛聯聹 API
# 芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬芒聲聬

def _make_header(ws, headers: list, fill_color: str):
    """锚鲁碌铆聠碌 铆聴陇毛聧聰 矛聤陇铆聝聙矛聺录 矛聽聛矛職漏"""
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def _find_data_start(ws) -> int:
    """铆聴陇毛聧聰铆聳聣(矛聝聛铆聮聢毛虏聢铆聵赂 铆聫卢铆聲篓) 毛聥陇矛聺聦 铆聳聣 毛掳聵铆聶聵"""
    for r in range(1, min(5, ws.max_row + 1)):
        if "矛聝聛铆聮聢毛虏聢铆聵赂" in str(ws.cell(r, 1).value or ""):
            return r + 1
    return 2  # 铆聴陇毛聧聰 矛聴聠矛聹录毛漏麓 2铆聳聣毛露聙铆聞掳


def _read_sheet_2col(ws) -> tuple[dict, dict]:
    """矛聝聛铆聮聢毛虏聢铆聵赂|ERP铆聮聢毛陋漏矛陆聰毛聯聹|毛陋篓毛聧赂毛陋聟 矛聥聹铆聤赂 矛聺陆锚赂掳 芒聠聮 (map, model_map)"""
    data_start = _find_data_start(ws)
    prod_map, model_map = {}, {}
    for r in range(data_start, ws.max_row + 1):
        pno  = str(ws.cell(r, 1).value or "").strip()
        code = str(ws.cell(r, 2).value or "").strip()
        mdl  = str(ws.cell(r, 3).value or "").strip()
        if pno and code and pno != "None" and code != "None":
            prod_map[pno] = code
            if mdl and mdl != "None":
                model_map[pno] = mdl
    return prod_map, model_map


@router.get("/product-map/export-excel")
async def export_product_map_excel():
    """铆聵聞矛聻卢 毛搂陇铆聲聭 矛聽聞矛虏麓毛楼录 3矛聥聹铆聤赂 Excel毛隆聹 毛聥陇矛職麓毛隆聹毛聯聹
       矛聥聹铆聤赂1: 毛漏聰矛聺赂矛聝聛铆聮聢 / 矛聥聹铆聤赂2: 矛聵碌矛聟聵矛聝聛铆聮聢(矛聵陇毛虏聞毛聺录矛聺麓毛聯聹) / 矛聥聹铆聤赂3: 矛露聰锚掳聙矛聝聛铆聮聢
    """
    import io
    from fastapi.responses import StreamingResponse
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 毛炉赂矛聞陇矛鹿聵")

    wb = openpyxl.Workbook()

    # 芒聰聙芒聰聙 矛聥聹铆聤赂1: 毛漏聰矛聺赂矛聝聛铆聮聢 芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙
    ws1 = wb.active
    ws1.title = "1_毛漏聰矛聺赂矛聝聛铆聮聢"
    _make_header(ws1, ["矛聝聛铆聮聢毛虏聢铆聵赂", "ERP铆聮聢毛陋漏矛陆聰毛聯聹", "毛陋篓毛聧赂毛陋聟(毛隆聹矛聽聽矛聠隆矛聻楼矛職漏)"], "1F4E79")
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 30
    ws1.column_dimensions["C"].width = 30
    for i, (pno, code) in enumerate(_product_map.items(), start=2):
        ws1.cell(i, 1, pno); ws1.cell(i, 2, code); ws1.cell(i, 3, _model_map.get(pno, ""))

    # 芒聰聙芒聰聙 矛聥聹铆聤赂2: 矛聵碌矛聟聵矛聝聛铆聮聢 芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙
    ws2 = wb.create_sheet("2_矛聵碌矛聟聵矛聝聛铆聮聢")
    _make_header(ws2, ["矛聝聛铆聮聢毛虏聢铆聵赂", "矛聵碌矛聟聵铆聟聧矛聤陇铆聤赂(矛掳赂锚鲁聽矛職漏)", "矛聻聬毛聫聶矛露聰矛露聹矛陆聰毛聯聹(矛掳赂锚鲁聽矛職漏)", "ERP铆聮聢毛陋漏矛陆聰毛聯聹(毛鹿聞矛職掳毛漏麓矛聻聬毛聫聶)", "毛陋篓毛聧赂毛陋聟(毛隆聹矛聽聽矛聠隆矛聻楼矛職漏)"], "375623")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 45
    ws2.column_dimensions["C"].width = 25
    ws2.column_dimensions["D"].width = 30
    ws2.column_dimensions["E"].width = 30
    for i, (pno, code) in enumerate(_option_override_map.items(), start=2):
        ws2.cell(i, 1, pno); ws2.cell(i, 4, code); ws2.cell(i, 5, _model_map.get(pno, ""))

    # 芒聰聙芒聰聙 矛聥聹铆聤赂3: 矛露聰锚掳聙矛聝聛铆聮聢 芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙芒聰聙
    ws3 = wb.create_sheet("3_矛露聰锚掳聙矛聝聛铆聮聢")
    _make_header(ws3, ["矛聝聛铆聮聢毛虏聢铆聵赂", "ERP铆聮聢毛陋漏矛陆聰毛聯聹", "毛陋篓毛聧赂毛陋聟(毛隆聹矛聽聽矛聠隆矛聻楼矛職漏)"], "7B3F00")
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 30
    ws3.column_dimensions["C"].width = 30
    for i, (pno, code) in enumerate(_addon_map.items(), start=2):
        ws3.cell(i, 1, pno); ws3.cell(i, 2, code); ws3.cell(i, 3, _model_map.get(pno, ""))

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    logger.info(f"[SS] Excel 毛聜麓毛鲁麓毛聜麓锚赂掳 芒聙聰 毛漏聰矛聺赂:{len(_product_map)} 矛聵碌矛聟聵:{len(_option_override_map)} 矛露聰锚掳聙:{len(_addon_map)}")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=smartstore_product_map.xlsx"},
    )


@router.post("/product-map/import-excel")
async def import_product_map_excel(file: bytes = Body(..., media_type="application/octet-stream")):
    """3矛聥聹铆聤赂 Excel 矛聴聟毛隆聹毛聯聹 芒聠聮 矛聽聞矛虏麓 毛搂陇铆聲聭 锚掳卤矛聥聽
       矛聥聹铆聤赂1: 毛漏聰矛聺赂矛聝聛铆聮聢 / 矛聥聹铆聤赂2: 矛聵碌矛聟聵矛聝聛铆聮聢 / 矛聥聹铆聤赂3: 矛露聰锚掳聙矛聝聛铆聮聢
    """
    import io
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 毛炉赂矛聞陇矛鹿聵")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file), data_only=True)

        new_product_map, new_model_map, new_option_map, new_addon_map = {}, {}, {}, {}

        # 矛聥聹铆聤赂1: 毛漏聰矛聺赂矛聝聛铆聮聢 (矛禄卢毛聼录: 矛聝聛铆聮聢毛虏聢铆聵赂|ERP铆聮聢毛陋漏矛陆聰毛聯聹|毛陋篓毛聧赂毛陋聟)
        if len(wb.sheetnames) >= 1:
            m, mdl = _read_sheet_2col(wb.worksheets[0])
            new_product_map.update(m); new_model_map.update(mdl)

        # 矛聥聹铆聤赂2: 矛聵碌矛聟聵矛聝聛铆聮聢 (矛禄卢毛聼录: 矛聝聛铆聮聢毛虏聢铆聵赂|矛聵碌矛聟聵铆聟聧矛聤陇铆聤赂|矛聻聬毛聫聶矛露聰矛露聹矛陆聰毛聯聹|ERP铆聮聢毛陋漏矛陆聰毛聯聹|毛陋篓毛聧赂毛陋聟)
        if len(wb.sheetnames) >= 2:
            ws2 = wb.worksheets[1]
            data_start = _find_data_start(ws2)
            for r in range(data_start, ws2.max_row + 1):
                pno  = str(ws2.cell(r, 1).value or "").strip()
                code = str(ws2.cell(r, 4).value or "").strip()  # D矛聴麓: ERP铆聮聢毛陋漏矛陆聰毛聯聹
                mdl  = str(ws2.cell(r, 5).value or "").strip()  # E矛聴麓: 毛陋篓毛聧赂毛陋聟
                if pno and code and pno != "None" and code != "None":
                    new_option_map[pno] = code
                    if mdl and mdl != "None":
                        new_model_map[pno] = mdl

        # 矛聥聹铆聤赂3: 矛露聰锚掳聙矛聝聛铆聮聢 (矛禄卢毛聼录: 矛聝聛铆聮聢毛虏聢铆聵赂|ERP铆聮聢毛陋漏矛陆聰毛聯聹|毛陋篓毛聧赂毛陋聟)
        if len(wb.sheetnames) >= 3:
            m, mdl = _read_sheet_2col(wb.worksheets[2])
            new_addon_map.update(m); new_model_map.update(mdl)

        total = len(new_product_map) + len(new_option_map) + len(new_addon_map)
        if total == 0:
            return {"success": False, "error": "矛聹聽铆職篓铆聲聹 毛聧掳矛聺麓铆聞掳锚掳聙 矛聴聠矛聤碌毛聥聢毛聥陇."}

        global _product_map, _option_override_map, _addon_map, _model_map, _model_to_erp_map
        _product_map         = new_product_map
        _option_override_map = new_option_map
        _addon_map           = new_addon_map
        _model_map           = new_model_map
        _save_product_map()
        _load_product_map()   # 矛聴颅毛掳漏铆聳楼 毛搂碌 矛聻卢毛鹿聦毛聯聹

        logger.info(f"[SS] Excel 锚掳聙矛聽赂矛聵陇锚赂掳 芒聙聰 毛漏聰矛聺赂:{len(_product_map)} 矛聵碌矛聟聵:{len(_option_override_map)} 矛露聰锚掳聙:{len(_addon_map)}")
        return {
            "success": True,
            "sheet1_main": len(_product_map),
            "sheet2_option": len(_option_override_map),
            "sheet3_addon": len(_addon_map),
            "message": f"毛漏聰矛聺赂 {len(_product_map)}锚卤麓 / 矛聵碌矛聟聵 {len(_option_override_map)}锚卤麓 / 矛露聰锚掳聙矛聝聛铆聮聢 {len(_addon_map)}锚卤麓 锚掳卤矛聥聽 矛聶聞毛拢聦",
        }

    except Exception as e:
        logger.error(f"[SS] Excel 锚掳聙矛聽赂矛聵陇锚赂掳 矛聵陇毛楼聵: {e}", exc_info=True)
        raise HTTPException(status_code=400, detail=f"铆聦聦矛聺录 铆聦聦矛聥卤 矛聵陇毛楼聵: {e}")


@router.post("/product-map/fetch-options-excel")
async def fetch_options_excel(body: dict = Body(...)):
    """矛聝聛铆聮聢毛虏聢铆聵赂 毛陋漏毛隆聺矛聹录毛隆聹 毛聞陇矛聺麓毛虏聞 API 矛隆掳铆職聦 芒聠聮 矛聵碌矛聟聵/矛露聰锚掳聙矛聝聛铆聮聢 毛搂陇铆聲聭 矛聻聭矛聴聟矛職漏 Excel 毛聥陇矛職麓毛隆聹毛聯聹"""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse
    from services.naver_client import naver_client

    product_nos = body.get("productNos", [])
    if not product_nos:
        raise HTTPException(status_code=400, detail="productNos 铆聲聞矛職聰")

    items = await naver_client.fetch_products_with_options(product_nos)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "矛聵碌矛聟聵_矛露聰锚掳聙矛聝聛铆聮聢_毛搂陇铆聲聭"

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    yel_fill = PatternFill("solid", fgColor="FFF2CC")
    data_font = Font(name="Arial", size=10)

    cols = ["锚碌卢毛露聞", "矛聝聛铆聮聢毛虏聢铆聵赂", "矛聝聛铆聮聢毛陋聟", "矛聵碌矛聟聵铆聟聧矛聤陇铆聤赂", "铆聦聬毛搂陇矛聻聬矛陆聰毛聯聹(矛掳赂锚鲁聽)", "矛聻卢锚鲁聽", "ERP铆聮聢毛陋漏矛陆聰毛聯聹(矛聻聟毛聽楼)", "毛陋篓毛聧赂毛陋聟(矛聻聟毛聽楼)"]
    for ci, col in enumerate(cols, 1):
        c = ws.cell(1, ci, col)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center")

    widths = [10, 18, 40, 40, 20, 8, 25, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for ri, item in enumerate(items, 2):
        # 铆聵聞矛聻卢 毛搂陇铆聲聭 矛聴卢毛露聙 铆聶聲矛聺赂
        pno = item["productNo"]
        existing_erp = _product_map.get(pno, "") or _option_override_map.get(pno, "")
        existing_model = _model_map.get(pno, "")

        vals = [
            item["type"], pno, item["productName"], item["optionText"],
            item["sellerCode"], item["stock"],
            existing_erp,   # 锚赂掳矛隆麓 毛搂陇铆聲聭 矛聻聢矛聹录毛漏麓 毛炉赂毛娄卢 矛卤聞矛聸聦矛陇聦
            existing_model,
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(ri, ci, val)
            c.font = data_font
            if ci in (7, 8):
                c.fill = yel_fill

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    logger.info(f"[SS] 矛聵碌矛聟聵矛隆掳铆職聦 Excel: {len(items)}锚掳聹 铆聲颅毛陋漏")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=smartstore_options.xlsx"},
    )


@router.get("/product-map/debug-option/{product_no}")
async def debug_channel_product(product_no: str, naver_no: str = ""):
    """矛卤聞毛聞聬矛聝聛铆聮聢 API 矛聺聭毛聥碌 锚碌卢矛隆掳 铆聶聲矛聺赂矛職漏 (毛聰聰毛虏聞锚路赂)
    product_no = 矛聝聛铆聮聢毛虏聢铆聵赂(矛聤陇毛搂聢铆聤赂矛聤陇铆聠聽矛聳麓) = 矛聸聬矛聝聛铆聮聢毛虏聢铆聵赂
    naver_no   = 毛聞陇矛聺麓毛虏聞矛聡录铆聲聭矛聝聛铆聮聢毛虏聢铆聵赂    = 矛卤聞毛聞聬矛聝聛铆聮聢毛虏聢铆聵赂
    """
    from services.naver_client import naver_client
    headers = await naver_client._headers()
    import httpx
    from config import NAVER_COMMERCE_URL

    endpoints = {
        # 矛聸聬矛聝聛铆聮聢 矛隆掳铆職聦 (v2) 芒聙聰 矛聝聛铆聮聢毛虏聢铆聵赂(矛聤陇毛搂聢铆聤赂矛聤陇铆聠聽矛聳麓) 矛聜卢矛職漏
        "v2_矛聸聬矛聝聛铆聮聢": f"{NAVER_COMMERCE_URL}/external/v2/products/origin-products/{product_no}",
        # 矛卤聞毛聞聬矛聝聛铆聮聢 矛隆掳铆職聦 (v2) 芒聙聰 矛聝聛铆聮聢毛虏聢铆聵赂毛隆聹 矛聥聹毛聫聞
        "v2_矛卤聞毛聞聬_矛聸聬矛聝聛铆聮聢毛虏聢铆聵赂": f"{NAVER_COMMERCE_URL}/external/v2/channel-products/{product_no}",
    }
    if naver_no:
        # 矛卤聞毛聞聬矛聝聛铆聮聢 矛隆掳铆職聦 (v2) 芒聙聰 毛聞陇矛聺麓毛虏聞矛聡录铆聲聭矛聝聛铆聮聢毛虏聢铆聵赂毛隆聹 矛聥聹毛聫聞
        endpoints["v2_矛卤聞毛聞聬_矛聡录铆聲聭毛虏聢铆聵赂"] = f"{NAVER_COMMERCE_URL}/external/v2/channel-products/{naver_no}"

    results = {}
    async with httpx.AsyncClient(timeout=15) as client:
        for key, url in endpoints.items():
            try:
                r = await client.get(url, headers=headers)
                results[key] = {"status": r.status_code, "url": url, "body": r.json()}
            except Exception as e:
                results[key] = {"error": str(e), "url": url}
    return results
