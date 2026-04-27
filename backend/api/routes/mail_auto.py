"""
메일 자동화 API 라우터
- GET  /api/mail-auto/dashboard    대시보드 데이터
- GET  /api/mail-auto/logs         처리 로그
- POST /api/mail-auto/trigger      수동 실행
- POST /api/mail-auto/process-file 업로드 Excel HS코드 처리 (테스트용)
- GET  /api/mail-auto/exchange-rate 환율 조회
- POST /api/mail-auto/exchange-rate 환율 수동 설정
- GET  /api/mail-auto/oem-products  OEM 미매핑 제품
- POST /api/mail-auto/oem-products  OEM 품목코드 매핑
- POST /api/mail-auto/auth          비밀번호 인증
"""

import logging
import io
import json
from datetime import datetime, timedelta, timezone
from pathlib import Path
from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Query, Request
from fastapi.responses import StreamingResponse
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))

from db.database import get_connection
from services.mail_auto_service import (
    fetch_bor_emails, process_excel_hs_code, create_purchase_slip,
    fetch_exchange_rate, run_mail_automation_pipeline,
    get_auto_state, start_mail_auto_scheduler, stop_mail_auto_scheduler,
    MAIL_AUTO_PASSWORD, ERP_SUPPLIER_CODE,
)

router = APIRouter(prefix="/api/mail-auto", tags=["mail-auto"])
logger = logging.getLogger(__name__)
KST = timezone(timedelta(hours=9))

# 메모리 캐시 (세션 비밀번호 검증 대용 - 간단 구현)
_exchange_rate_cache = {"rate": None, "updated": None}


def _ensure_tables():
    """테이블 생성 (없으면)"""
    conn = get_connection()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS mail_processing_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            message_id TEXT UNIQUE,
            subject TEXT,
            sender TEXT,
            received_at TEXT,
            attachment_count INTEGER DEFAULT 0,
            status TEXT DEFAULT 'pending',
            hs_code_count INTEGER DEFAULT 0,
            erp_slip_id TEXT DEFAULT '',
            reply_sent INTEGER DEFAULT 0,
            error_message TEXT DEFAULT '',
            result_json TEXT DEFAULT '{}',
            processed_at TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS oem_product_mapping (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            description TEXT NOT NULL,
            item_code TEXT DEFAULT '',
            category TEXT DEFAULT '',
            mapped_by TEXT DEFAULT 'manual',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS mail_auto_exchange_rate (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rate REAL NOT NULL,
            source TEXT DEFAULT 'auto',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS product_code_mapping (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            model_name TEXT NOT NULL,
            prod_cd TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS mail_attachment_processed (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            log_id INTEGER DEFAULT 0,
            message_id TEXT DEFAULT '',
            bor_number TEXT DEFAULT '',
            filename TEXT DEFAULT '',
            hs_filled INTEGER DEFAULT 0,
            hs_unknown INTEGER DEFAULT 0,
            erp_success INTEGER DEFAULT 0,
            erp_lines_count INTEGER DEFAULT 0,
            erp_error TEXT DEFAULT '',
            erp_failure_kind TEXT DEFAULT '',
            file_b64 TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
    """)
    conn.commit()

    # 매핑 데이터 초기 로드 (비어있으면)
    cnt = conn.execute("SELECT COUNT(*) FROM product_code_mapping").fetchone()[0]
    if cnt == 0:
        _load_product_mapping_from_file(conn)


def _load_product_mapping_from_file(conn):
    """data/product_code_mapping.xlsx에서 품목코드 매핑 로드"""
    import openpyxl
    import os
    
    # 여러 경로 시도
    candidates = [
        Path(__file__).parent.parent.parent.parent / "data" / "product_code_mapping.xlsx",
        Path(__file__).parent.parent.parent / "data" / "product_code_mapping.xlsx",
        Path(os.getcwd()).parent / "data" / "product_code_mapping.xlsx",
        Path(os.getcwd()) / "data" / "product_code_mapping.xlsx",
        Path(os.getcwd()) / ".." / "data" / "product_code_mapping.xlsx",
    ]
    
    mapping_path = None
    for p in candidates:
        if p.exists():
            mapping_path = p
            break
    
    if not mapping_path:
        logger.warning(f"[매핑] 파일 없음. 시도한 경로: {[str(p) for p in candidates]}")
        return 0
    
    wb = openpyxl.load_workbook(mapping_path, data_only=True)
    ws = wb.active
    count = 0
    for r in range(2, ws.max_row + 1):
        prod_cd = ws.cell(row=r, column=1).value
        model = ws.cell(row=r, column=2).value
        if prod_cd and model:
            model_str = str(model).strip()
            try:
                conn.execute(
                    "INSERT INTO product_code_mapping (model_name, prod_cd) VALUES (?, ?)",
                    (model_str, str(prod_cd).strip())
                )
                count += 1
            except Exception:
                pass
    conn.commit()
    wb.close()
    logger.info(f"[매핑] {count}건 로드 완료")
    return count


def _lookup_prod_cd(model_name: str, conn=None) -> str:
    """모델명 → 품목코드 조회. 정확→전방→역방→클린 순서.

    `conn`을 외부에서 주입하면 그것을 그대로 쓰고 close하지 않는다.
    내부에서 새로 연 경우에만 close를 보장(try/finally) — PG 풀 누수 방지.
    """
    own_conn = False
    if conn is None:
        conn = get_connection()
        own_conn = True
    try:
        # 1. 정확 매칭
        row = conn.execute(
            "SELECT prod_cd FROM product_code_mapping WHERE model_name = ?",
            (model_name,)
        ).fetchone()
        if row:
            return row[0]

        # 2. 매핑 model_name이 검색어로 시작 (매핑에 추가 텍스트 있는 경우)
        #    예: 매핑 "LSP-GIC-FJM, STP" → 검색 "LSP-GIC-FJM"
        row = conn.execute(
            "SELECT prod_cd FROM product_code_mapping WHERE model_name LIKE ? ORDER BY LENGTH(model_name) LIMIT 1",
            (model_name + "%",)
        ).fetchone()
        if row:
            return row[0]

        # 3. 검색어가 매핑 model_name으로 시작 (색상 접미사 등)
        #    예: 검색 "LS-5STPD-2MG" → 매핑 "LS-5STPD-2M" (가장 긴 매칭 우선)
        prefix = model_name[:-1] if len(model_name) > 3 else model_name
        rows = conn.execute(
            "SELECT model_name, prod_cd FROM product_code_mapping WHERE model_name LIKE ?",
            (prefix + "%",)
        ).fetchall()
        best = ("", "")
        for mname, pcd in rows:
            clean = mname.split(",")[0].strip()
            if model_name.startswith(clean) and len(clean) > len(best[0]):
                best = (clean, pcd)
        if best[1]:
            return best[1]

        return ""
    finally:
        if own_conn:
            try:
                conn.close()
            except Exception:
                pass


_ensure_tables()


# ─── 인증 ────────────────────────────────────────────────

@router.post("/auth")
async def mail_auto_auth(request: Request):
    """비밀번호 인증"""
    body = await request.json()
    pw = body.get("password", "")
    if pw == MAIL_AUTO_PASSWORD:
        return {"success": True}
    raise HTTPException(status_code=401, detail="비밀번호가 올바르지 않습니다")


# ─── 대시보드 ─────────────────────────────────────────────

@router.get("/dashboard")
async def get_dashboard():
    """대시보드 요약 데이터"""
    conn = get_connection()
    
    # 최근 처리 통계
    today = datetime.now(KST).strftime("%Y-%m-%d")
    week_ago = (datetime.now(KST) - timedelta(days=7)).strftime("%Y-%m-%d")
    
    stats = {
        "today_count": 0, "week_count": 0, "total_count": 0,
        "today_hs": 0, "total_hs": 0,
        "pending_oem": 0,
    }
    
    try:
        r = conn.execute("SELECT COUNT(*) FROM mail_processing_log WHERE processed_at LIKE ?", (f"{today}%",))
        stats["today_count"] = r.fetchone()[0]
        
        r = conn.execute("SELECT COUNT(*) FROM mail_processing_log WHERE processed_at >= ?", (week_ago,))
        stats["week_count"] = r.fetchone()[0]
        
        r = conn.execute("SELECT COUNT(*), COALESCE(SUM(hs_code_count),0) FROM mail_processing_log")
        row = r.fetchone()
        stats["total_count"] = row[0]
        stats["total_hs"] = row[1]
        
        r = conn.execute("SELECT COALESCE(SUM(hs_code_count),0) FROM mail_processing_log WHERE processed_at LIKE ?", (f"{today}%",))
        stats["today_hs"] = r.fetchone()[0]
        
        r = conn.execute("SELECT COUNT(*) FROM oem_product_mapping WHERE item_code='' OR item_code IS NULL")
        stats["pending_oem"] = r.fetchone()[0]
    except Exception as e:
        logger.warning(f"[대시보드] 통계 조회 오류: {e}")
    
    # 환율
    rate = _exchange_rate_cache.get("rate")
    rate_updated = _exchange_rate_cache.get("updated")
    
    return {
        "stats": stats,
        "exchange_rate": rate,
        "exchange_rate_updated": rate_updated,
        "auto": get_auto_state(),
    }


# ─── 자동 실행 스케줄러 제어 ───────────────────────────────

@router.post("/scheduler/start")
async def scheduler_start(request: Request):
    """자동 실행 시작"""
    body = await request.json() if request.headers.get("content-type") == "application/json" else {}
    interval = body.get("interval_min", 10)
    auto_reply = body.get("auto_reply", False)
    reply_template = body.get("reply_template", "")
    
    from services.mail_auto_service import _auto_state
    _auto_state["auto_reply"] = auto_reply
    _auto_state["reply_template"] = reply_template
    
    start_mail_auto_scheduler(interval_min=interval)
    return {"success": True, "state": get_auto_state()}


@router.post("/scheduler/stop")
async def scheduler_stop():
    """자동 실행 중지"""
    stop_mail_auto_scheduler()
    return {"success": True, "state": get_auto_state()}


@router.get("/scheduler/status")
async def scheduler_status():
    """자동 실행 상태"""
    return get_auto_state()


# ─── 처리 로그 ────────────────────────────────────────────

@router.get("/logs")
async def get_logs(limit: int = 20, offset: int = 0):
    """처리 로그 목록"""
    conn = get_connection()
    rows = conn.execute(
        """SELECT id, message_id, subject, sender, received_at, 
                  attachment_count, status, hs_code_count, reply_sent,
                  error_message, processed_at, result_json
           FROM mail_processing_log 
           ORDER BY id DESC LIMIT ? OFFSET ?""",
        (limit, offset)
    ).fetchall()
    
    total = conn.execute("SELECT COUNT(*) FROM mail_processing_log").fetchone()[0]
    
    return {
        "total": total,
        "logs": [
            {
                "id": r[0], "message_id": r[1], "subject": r[2],
                "sender": r[3], "received_at": r[4],
                "attachment_count": r[5], "status": r[6],
                "hs_code_count": r[7], "reply_sent": bool(r[8]),
                "error_message": r[9], "processed_at": r[10],
                "result": json.loads(r[11]) if r[11] else {},
            }
            for r in rows
        ],
    }


# ─── 수동 실행 ────────────────────────────────────────────

@router.post("/trigger")
async def trigger_pipeline(request: Request):
    """메일 자동화 파이프라인 수동 실행"""
    body = await request.json() if request.headers.get("content-type") == "application/json" else {}
    days_back = body.get("days_back", 30)
    auto_reply = body.get("auto_reply", False)
    auto_erp = body.get("auto_erp", True)
    custom_rate = body.get("exchange_rate", None)
    reply_template = body.get("reply_template", "")
    
    exchange_rate = custom_rate or _exchange_rate_cache.get("rate")
    conn = get_connection()
    
    result = await run_mail_automation_pipeline(
        days_back=days_back,
        exchange_rate=exchange_rate,
        auto_reply=auto_reply,
        auto_erp=auto_erp,
        db_conn=conn,
        reply_template=reply_template,
    )
    
    # 텔레그램 알림 (신규 처리 건이 있을 때)
    if result.get("new_processed", 0) > 0:
        from services.mail_auto_service import _send_telegram_notification
        await _send_telegram_notification(result)
    
    # 환율 캐시 갱신
    if result.get("exchange_rate"):
        _exchange_rate_cache["rate"] = result["exchange_rate"]
        _exchange_rate_cache["updated"] = datetime.now(KST).isoformat()
    
    return result


# ─── Excel 테스트 처리 ────────────────────────────────────

@router.post("/process-file")
async def process_uploaded_file(file: UploadFile = File(...)):
    """업로드된 Excel 파일에 HS코드 적용 (테스트용)"""
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "xlsx 파일만 지원합니다")
    
    data = await file.read()
    result = process_excel_hs_code(data, file.filename)
    
    if not result["success"]:
        raise HTTPException(400, result.get("error", "처리 실패"))
    
    # 처리된 Excel 반환
    return {
        "filename": result["filename"],
        "stats": result["stats"],
        "items": result["items"][:50],  # 최대 50개
        "erp_lines": result["erp_lines"][:50],
        "oem_items": result["oem_items"],
    }


@router.post("/process-file/download")
async def process_and_download(file: UploadFile = File(...)):
    """업로드 Excel → HS코드 입력 → 다운로드"""
    data = await file.read()
    result = process_excel_hs_code(data, file.filename)

    if not result["success"]:
        raise HTTPException(400, result.get("error"))

    return StreamingResponse(
        io.BytesIO(result["output_data"]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=HS_{file.filename}"},
    )


@router.get("/attachment/{att_id}/download")
async def download_processed_attachment(att_id: int):
    """자동 파이프라인이 처리한 첨부 Excel(HS코드 입력 완료) 다운로드"""
    import base64 as _b64
    from urllib.parse import quote as _urlquote

    conn = get_connection()
    try:
        row = conn.execute(
            "SELECT filename, file_b64 FROM mail_attachment_processed WHERE id = ?",
            (att_id,)
        ).fetchone()
    finally:
        try:
            conn.close()
        except Exception:
            pass

    if not row:
        raise HTTPException(404, "첨부를 찾을 수 없습니다")

    # sqlite3.Row, dict, tuple 모두 호환
    try:
        filename = row["filename"]
        file_b64 = row["file_b64"]
    except Exception:
        filename, file_b64 = row[0], row[1]

    if not file_b64:
        raise HTTPException(410, "처리된 파일 데이터가 없습니다 (이전 버전에서 처리된 메일은 재실행 필요)")

    try:
        file_bytes = _b64.b64decode(file_b64)
    except Exception as e:
        raise HTTPException(500, f"파일 디코딩 실패: {e}")

    safe_name = _urlquote(f"HS_{filename}")
    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{safe_name}"
        },
    )


# ─── ERP 구매전표 테스트 ──────────────────────────────────

@router.post("/test-erp")
async def test_erp_purchase(file: UploadFile = File(...), exchange_rate: float = Form(0)):
    """업로드 Excel → ERP 구매전표 미리보기 (실제 전송 안함)"""
    try:
        data = await file.read()
        result = process_excel_hs_code(data, file.filename)

        if not result.get("success"):
            raise HTTPException(400, result.get("error", "Excel 처리 실패"))

        # 환율 조회 (실패해도 계속 진행)
        rate = exchange_rate
        if rate <= 0:
            try:
                rate_info = await fetch_exchange_rate()
                rate = float(rate_info.get("rate") or 0)
            except Exception as e:
                logger.warning(f"[ERP 미리보기] 환율 조회 실패, fallback 사용: {e}")
                rate = 0
            if rate <= 0:
                # 마지막 fallback - DB의 가장 최근 환율
                try:
                    conn = get_connection()
                    row = conn.execute(
                        "SELECT rate FROM mail_auto_exchange_rate ORDER BY id DESC LIMIT 1"
                    ).fetchone()
                    rate = float(row[0]) if row and row[0] else 1400.0
                except Exception:
                    rate = 1400.0

        # ERP 라인 미리보기 생성 (모델명 → 품목코드 변환)
        # 한 conn으로 N개 모델 조회 (풀 고갈 방지)
        erp_preview = []
        unmapped_models = []
        lookup_conn = get_connection()
        try:
            for item in result.get("erp_lines") or []:
                try:
                    price_usd = float(item.get("price_usd") or 0)
                    tax_rate = float(item.get("tax_rate") or 1.18)
                    qty = float(item.get("qty") or 0)
                    price_krw = round(price_usd * tax_rate * rate)
                    model = item.get("prod_cd") or ""
                    prod_cd = _lookup_prod_cd(model, conn=lookup_conn) if model else ""

                    erp_preview.append({
                        "model_name": model,
                        "prod_cd": prod_cd,
                        "qty": qty,
                        "price_usd": price_usd,
                        "tax_rate": tax_rate,
                        "price_krw": price_krw,
                        "supply_amt": round(price_krw * qty),
                        "description": item.get("description", ""),
                    })
                    if not prod_cd:
                        unmapped_models.append(model)
                except Exception as e:
                    logger.error(f"[ERP 미리보기] 라인 변환 실패 item={item}: {e}")
                    continue
        finally:
            try:
                lookup_conn.close()
            except Exception:
                pass

        # 전표일자 (오늘 +8일)
        io_date = (datetime.now(KST) + timedelta(days=8)).strftime("%Y-%m-%d")

        return {
            "success": True,
            "erp_lines": erp_preview,
            "oem_items": result.get("oem_items") or [],
            "unmapped_models": unmapped_models,
            "total_lines": len(erp_preview),
            "total_amount": sum(l["supply_amt"] for l in erp_preview),
            "exchange_rate": rate,
            "io_date": io_date,
            "cust_code": ERP_SUPPLIER_CODE,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"[ERP 미리보기] 처리 중 예외: {e}")
        raise HTTPException(500, f"ERP 미리보기 처리 실패: {e}")


@router.post("/submit-erp")
async def submit_erp_purchase(request: Request):
    """ERP 구매전표 실제 전송 (미리보기 데이터 기반)"""
    from services.erp_client import erp_client
    
    body = await request.json()
    erp_lines = body.get("erp_lines", [])
    io_date_str = body.get("io_date", "")
    
    if not erp_lines:
        raise HTTPException(400, "전표 항목이 없습니다")
    
    io_date = io_date_str.replace("-", "") if io_date_str else (
        datetime.now(KST) + timedelta(days=8)
    ).strftime("%Y%m%d")
    
    # 미리보기에서 이미 KRW 단가가 계산됨 — 품목코드 있는 것만 전송
    lines = []
    skipped = []
    for item in erp_lines:
        prod_cd = item.get("prod_cd", "")
        if not prod_cd:
            skipped.append(item.get("model_name", ""))
            continue
        lines.append({
            "prod_cd": prod_cd,
            "qty": item["qty"],
            "unit": "EA",
            "price": item["price_krw"],
        })
    
    if not lines:
        return {"success": False, "error": "전송 가능한 품목이 없습니다 (모두 미매핑)"}
    
    try:
        result = await erp_client.save_purchase(
            cust_code=ERP_SUPPLIER_CODE,
            lines=lines,
            io_date=io_date,
        )
        if skipped:
            result["skipped_models"] = skipped
            result["skipped_count"] = len(skipped)
        return result
    except Exception as e:
        return {"success": False, "error": str(e)}


# ─── 환율 ─────────────────────────────────────────────────

@router.get("/exchange-rate")
async def get_exchange_rate():
    """환율 조회 (전신환매도율 = 기준율 + 스프레드)"""
    cached = _exchange_rate_cache.get("rate")
    if cached and _exchange_rate_cache.get("base_rate"):
        return {
            "rate": cached,
            "base_rate": _exchange_rate_cache.get("base_rate"),
            "spread": _exchange_rate_cache.get("spread", 1.75),
            "source": _exchange_rate_cache.get("source", "cache"),
            "updated": _exchange_rate_cache.get("updated"),
        }
    
    rate_info = await fetch_exchange_rate()
    _exchange_rate_cache["rate"] = rate_info["rate"]
    _exchange_rate_cache["base_rate"] = rate_info.get("base_rate")
    _exchange_rate_cache["spread"] = rate_info.get("spread", 1.75)
    _exchange_rate_cache["source"] = rate_info.get("source", "")
    _exchange_rate_cache["updated"] = datetime.now(KST).isoformat()
    
    return {
        "rate": rate_info["rate"],
        "base_rate": rate_info.get("base_rate"),
        "spread": rate_info.get("spread", 1.75),
        "source": rate_info.get("source", ""),
        "updated": _exchange_rate_cache["updated"],
    }


@router.post("/exchange-rate")
async def set_exchange_rate(request: Request):
    """환율 수동 설정"""
    body = await request.json()
    rate = body.get("rate")
    if not rate or float(rate) <= 0:
        raise HTTPException(400, "유효한 환율을 입력하세요")
    
    _exchange_rate_cache["rate"] = float(rate)
    _exchange_rate_cache["updated"] = datetime.now(KST).isoformat()
    
    # DB 저장
    conn = get_connection()
    conn.execute(
        "INSERT INTO mail_auto_exchange_rate (rate, source) VALUES (?, 'manual')",
        (float(rate),)
    )
    conn.commit()
    
    return {"rate": float(rate), "source": "manual", "updated": _exchange_rate_cache["updated"]}


# ─── OEM 제품 매핑 ─────────────────────────────────────────

@router.get("/oem-products")
async def get_oem_products():
    """OEM 미매핑 제품 목록"""
    conn = get_connection()
    rows = conn.execute(
        """SELECT id, description, item_code, category, mapped_by, created_at 
           FROM oem_product_mapping ORDER BY id DESC"""
    ).fetchall()
    
    return {
        "products": [
            {
                "id": r[0], "description": r[1], "item_code": r[2],
                "category": r[3], "mapped_by": r[4], "created_at": r[5],
            }
            for r in rows
        ]
    }


@router.post("/oem-products")
async def map_oem_product(request: Request):
    """OEM 제품 품목코드 매핑"""
    body = await request.json()
    product_id = body.get("id")
    item_code = body.get("item_code", "").strip()
    
    if not item_code:
        raise HTTPException(400, "품목코드를 입력하세요")
    
    conn = get_connection()
    if product_id:
        conn.execute(
            "UPDATE oem_product_mapping SET item_code=?, updated_at=datetime('now','localtime') WHERE id=?",
            (item_code, product_id)
        )
    else:
        desc = body.get("description", "")
        category = body.get("category", "")
        conn.execute(
            "INSERT INTO oem_product_mapping (description, item_code, category) VALUES (?,?,?)",
            (desc, item_code, category)
        )
    conn.commit()
    
    return {"success": True}


# ─── 메일 미리보기 (IMAP 검색만) ──────────────────────────

@router.get("/preview-emails")
async def preview_emails(days_back: int = 30):
    """처리 대상 메일 미리보기 (실제 처리 없이 검색만)"""
    emails = fetch_bor_emails(days_back=days_back)
    
    conn = get_connection()
    processed_ids = set()
    try:
        cursor = conn.execute("SELECT message_id FROM mail_processing_log")
        processed_ids = {row[0] for row in cursor.fetchall()}
    except Exception:
        pass
    
    return {
        "total": len(emails),
        "emails": [
            {
                "subject": m["subject"],
                "date": str(m.get("date_kst", "")),
                "message_id": m["message_id"],
                "attachments": [
                    {"filename": a["filename"], "bor_number": a.get("bor_number", "")}
                    for a in m["attachments"]
                ],
                "already_processed": m["message_id"] in processed_ids,
            }
            for m in emails
        ],
    }


# ─── 품목코드 매핑 관리 ────────────────────────────────────

@router.get("/product-mapping/count")
async def get_mapping_count():
    """품목코드 매핑 건수"""
    conn = get_connection()
    cnt = conn.execute("SELECT COUNT(*) FROM product_code_mapping").fetchone()[0]
    return {"count": cnt}


@router.post("/product-mapping/reload")
async def reload_mapping():
    """data/product_code_mapping.xlsx에서 매핑 새로고침"""
    conn = get_connection()
    conn.execute("DELETE FROM product_code_mapping")
    conn.commit()
    count = _load_product_mapping_from_file(conn)
    return {"success": True, "loaded": count}


@router.post("/product-mapping/upload")
async def upload_mapping(file: UploadFile = File(...)):
    """품목코드 매핑 Excel 업로드"""
    import openpyxl
    try:
        data = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        
        conn = get_connection()
        
        # 테이블 재생성 (기존 UNIQUE 제약 제거)
        conn.execute("DROP TABLE IF EXISTS product_code_mapping")
        conn.commit()
        conn.execute("""
            CREATE TABLE IF NOT EXISTS product_code_mapping (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                model_name TEXT NOT NULL,
                prod_cd TEXT NOT NULL
            )
        """)
        conn.commit()
        
        count = 0
        for r in range(2, ws.max_row + 1):
            prod_cd = ws.cell(row=r, column=1).value
            model = ws.cell(row=r, column=2).value
            if prod_cd and model:
                # 모델명 정제: 콤마/★ 이후 제거
                clean_model = str(model).strip()
                for sep in [",", "★", "  "]:
                    clean_model = clean_model.split(sep)[0].strip()
                conn.execute(
                    "INSERT INTO product_code_mapping (model_name, prod_cd) VALUES (?, ?)",
                    (clean_model, str(prod_cd).strip())
                )
                count += 1
        conn.commit()
        wb.close()
        return {"success": True, "loaded": count}
    except Exception as e:
        import traceback
        logger.error(f"[매핑 업로드] 오류: {traceback.format_exc()}")
        return {"success": False, "error": str(e)}


@router.post("/product-mapping/add")
async def add_mapping(request: Request):
    """단건 품목코드 매핑 추가"""
    body = await request.json()
    model = body.get("model_name", "").strip()
    prod_cd = body.get("prod_cd", "").strip()
    if not model or not prod_cd:
        raise HTTPException(400, "모델명과 품목코드를 입력하세요")
    
    conn = get_connection()
    # 기존 매핑 삭제 후 추가 (업데이트)
    conn.execute("DELETE FROM product_code_mapping WHERE model_name = ?", (model,))
    conn.execute("INSERT INTO product_code_mapping (model_name, prod_cd) VALUES (?, ?)", (model, prod_cd))
    conn.commit()
    return {"success": True, "model_name": model, "prod_cd": prod_cd}


@router.get("/product-mapping/search")
async def search_mapping(q: str = Query("")):
    """품목코드 매핑 검색"""
    if not q.strip():
        return {"results": []}
    conn = get_connection()
    rows = conn.execute(
        "SELECT model_name, prod_cd FROM product_code_mapping WHERE model_name LIKE ? ORDER BY model_name LIMIT 20",
        (f"%{q.strip()}%",)
    ).fetchall()
    return {"results": [{"model_name": r[0], "prod_cd": r[1]} for r in rows]}


# ─── 승인 기반 처리 ────────────────────────────────────────

@router.get("/pending")
async def get_pending_mails():
    """승인 대기 중인 메일 목록"""
    conn = get_connection()
    rows = conn.execute("""
        SELECT id, message_id, subject, received_at, attachment_count, status
        FROM mail_processing_log WHERE status = 'pending'
        ORDER BY received_at DESC
    """).fetchall()
    return {"pending": [
        {"id": r[0], "message_id": r[1], "subject": r[2], 
         "received_at": r[3], "attachment_count": r[4], "status": r[5]}
        for r in rows
    ]}


@router.post("/approve")
async def approve_pending(request: Request):
    """pending 메일 승인 → 파이프라인 실행"""
    body = await request.json()
    message_ids = body.get("message_ids", [])
    auto_reply = body.get("auto_reply", False)
    reply_template = body.get("reply_template", "")
    
    if not message_ids:
        raise HTTPException(400, "승인할 메일을 선택하세요")
    
    conn = get_connection()
    
    # pending → processing
    for mid in message_ids:
        conn.execute(
            "UPDATE mail_processing_log SET status = 'processing' WHERE message_id = ?",
            (mid,)
        )
    conn.commit()
    
    # 파이프라인 실행
    result = await run_mail_automation_pipeline(
        days_back=30,
        auto_reply=auto_reply,
        auto_erp=True,
        db_conn=conn,
        reply_template=reply_template,
    )
    
    # 텔레그램 결과 알림
    if result.get("new_processed", 0) > 0:
        from services.mail_auto_service import _send_telegram_notification
        await _send_telegram_notification(result)
    
    return result


@router.post("/reject")
async def reject_pending(request: Request):
    """pending 메일 거부 (스킵)"""
    body = await request.json()
    message_ids = body.get("message_ids", [])
    
    conn = get_connection()
    for mid in message_ids:
        conn.execute(
            "UPDATE mail_processing_log SET status = 'skipped' WHERE message_id = ?",
            (mid,)
        )
    conn.commit()
    return {"success": True, "skipped": len(message_ids)}


@router.post("/scan-now")
async def scan_now():
    """즉시 메일 스캔 (승인 대기 생성)"""
    from services.mail_auto_service import _auto_check_and_process
    await _auto_check_and_process()
    
    conn = get_connection()
    rows = conn.execute("""
        SELECT id, message_id, subject, received_at, attachment_count
        FROM mail_processing_log WHERE status = 'pending'
        ORDER BY received_at DESC
    """).fetchall()
    
    return {
        "scanned": True,
        "pending_count": len(rows),
        "pending": [
            {"id": r[0], "message_id": r[1], "subject": r[2],
             "received_at": r[3], "attachment_count": r[4]}
            for r in rows
        ],
    }
