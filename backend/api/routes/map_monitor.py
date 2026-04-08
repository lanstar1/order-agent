"""
MAP Monitor API Routes - 지도가 감시 시스템
"""
from fastapi import APIRouter, HTTPException, UploadFile, File, Query
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timedelta
import json, csv, io, logging

from db.database import get_connection

logger = logging.getLogger("map_monitor")
router = APIRouter(prefix="/api/map", tags=["MAP Monitor"])


# ── Pydantic Models ──────────────────────────────────

class SettingsUpdate(BaseModel):
    min_price: Optional[int] = None
    tolerance_pct: Optional[float] = None
    schedules: Optional[List[str]] = None
    watch_interval_hours: Optional[int] = None
    platforms: Optional[List[str]] = None
    alert_email: Optional[bool] = None
    alert_kakao: Optional[bool] = None
    alert_nateon: Optional[bool] = None
    alert_email_address: Optional[str] = None
    admin_password: Optional[str] = None  # 관리자 인증용

import os
MAP_ADMIN_PASSWORD = os.getenv("MAP_ADMIN_PASSWORD", "admin")

class ProductCreate(BaseModel):
    model_name: str
    product_name: str
    brand: str = "LANstar"
    features: str = ""
    map_price: int
    tolerance_pct: Optional[float] = None
    search_keywords: str = ""

class ProductUpdate(BaseModel):
    product_name: Optional[str] = None
    brand: Optional[str] = None
    features: Optional[str] = None
    map_price: Optional[int] = None
    tolerance_pct: Optional[float] = None
    search_keywords: Optional[str] = None
    is_active: Optional[bool] = None
    is_watched: Optional[bool] = None
    watch_interval_hours: Optional[int] = None

class ViolationResolve(BaseModel):
    resolution_note: str = ""


def _days_ago(days: int) -> str:
    """N일 전 날짜 문자열 반환 (YYYY-MM-DD HH:MM:SS)"""
    return (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")


# ═══════════════════════════════════════════════════════
# 1. 설정 API
# ═══════════════════════════════════════════════════════

@router.get("/settings")
async def get_settings():
    conn = get_connection()
    row = conn.execute("SELECT * FROM map_settings WHERE id = 1").fetchone()
    conn.close()
    if not row:
        return {}
    result = dict(row)
    for f in ['schedules', 'platforms']:
        if result.get(f) and isinstance(result[f], str):
            try: result[f] = json.loads(result[f])
            except: pass
    return result

class PasswordCheck(BaseModel):
    password: str

@router.post("/settings/verify")
async def verify_admin_password(data: PasswordCheck):
    """관리자 비밀번호 확인"""
    if data.password == MAP_ADMIN_PASSWORD:
        return {"ok": True}
    raise HTTPException(403, "비밀번호가 올바르지 않습니다")

@router.put("/settings")
async def update_settings(data: SettingsUpdate):
    # 관리자 비밀번호 확인
    if data.admin_password != MAP_ADMIN_PASSWORD:
        raise HTTPException(403, "관리자 비밀번호가 올바르지 않습니다")
    conn = get_connection()
    sets, vals = [], []
    for k, v in data.dict(exclude_unset=True).items():
        if k == 'admin_password': continue  # 비밀번호는 DB에 저장하지 않음
        if k in ['schedules', 'platforms'] and isinstance(v, list):
            sets.append(f"{k} = ?"); vals.append(json.dumps(v, ensure_ascii=False))
        elif isinstance(v, bool):
            sets.append(f"{k} = ?"); vals.append(1 if v else 0)
        else:
            sets.append(f"{k} = ?"); vals.append(v)
    if not sets:
        conn.close(); raise HTTPException(400, "변경할 항목 없음")
    sets.append("updated_at = datetime('now','localtime')")
    conn.execute(f"UPDATE map_settings SET {', '.join(sets)} WHERE id = 1", vals)
    conn.commit()
    # 스케줄러 재로드 (스케줄 또는 간격 변경 시)
    if any(k in data.dict(exclude_unset=True) for k in ['schedules', 'watch_interval_hours']):
        try:
            from services.map_scheduler import reload_map_schedule
            reload_map_schedule()
        except Exception as e:
            logger.warning(f"스케줄러 재로드 실패: {e}")
    row = conn.execute("SELECT * FROM map_settings WHERE id = 1").fetchone()
    conn.close()
    result = dict(row)
    for f in ['schedules', 'platforms']:
        if result.get(f) and isinstance(result[f], str):
            try: result[f] = json.loads(result[f])
            except: pass
    return result


# ═══════════════════════════════════════════════════════
# 2. 제품 API
# ═══════════════════════════════════════════════════════

@router.get("/products")
async def list_products(search: str = "", active_only: bool = True, watched_only: bool = False):
    conn = get_connection()
    cutoff = _days_ago(1)
    # 단일 쿼리로 최저가 + 위반 건수 함께 조회 (N+1 쿼리 제거)
    sql = """SELECT p.*,
        COALESCE(pr.min_price, 0) as current_min_price,
        COALESCE(vc.cnt, 0) as active_violations
        FROM map_products p
        LEFT JOIN (SELECT product_id, MIN(effective_price) as min_price
            FROM map_price_records WHERE collected_at > ? GROUP BY product_id) pr ON p.id = pr.product_id
        LEFT JOIN (SELECT product_id, COUNT(*) as cnt
            FROM map_violations WHERE is_resolved = 0 GROUP BY product_id) vc ON p.id = vc.product_id
        WHERE 1=1"""
    params = [cutoff]
    if active_only: sql += " AND p.is_active = 1"
    if watched_only: sql += " AND p.is_watched = 1"
    if search:
        sql += " AND (p.model_name LIKE ? OR p.product_name LIKE ?)"
        params += [f"%{search}%", f"%{search}%"]
    sql += " ORDER BY p.model_name"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@router.post("/products")
async def create_product(data: ProductCreate):
    conn = get_connection()
    if not data.search_keywords:
        data.search_keywords = f"{data.brand} {data.model_name} {data.product_name}"
    try:
        cur = conn.execute(
            """INSERT INTO map_products (model_name, product_name, brand, features, map_price, tolerance_pct, search_keywords)
               VALUES (?,?,?,?,?,?,?)""",
            (data.model_name, data.product_name, data.brand, data.features,
             data.map_price, data.tolerance_pct, data.search_keywords))
        conn.commit()
        pid = cur.lastrowid
        conn.close()
        return {"id": pid, "message": "제품 등록 완료"}
    except Exception as e:
        conn.close()
        if "unique" in str(e).lower() or "UNIQUE" in str(e):
            raise HTTPException(409, f"모델명 '{data.model_name}'이 이미 등록됨")
        raise HTTPException(500, str(e))

@router.put("/products/{product_id}")
async def update_product(product_id: int, data: ProductUpdate):
    conn = get_connection()
    ud = data.dict(exclude_unset=True)
    if not ud: conn.close(); raise HTTPException(400, "변경 항목 없음")
    # bool → int 변환 (SQLite)
    for k in ['is_active', 'is_watched']:
        if k in ud and isinstance(ud[k], bool):
            ud[k] = 1 if ud[k] else 0
    sets = [f"{k} = ?" for k in ud]; vals = list(ud.values())
    sets.append("updated_at = datetime('now','localtime')")
    conn.execute(f"UPDATE map_products SET {', '.join(sets)} WHERE id = ?", vals + [product_id])
    conn.commit(); conn.close()
    return {"message": "수정 완료"}

@router.delete("/products/{product_id}")
async def delete_product(product_id: int):
    conn = get_connection()
    conn.execute("DELETE FROM map_violations WHERE product_id = ?", (product_id,))
    conn.execute("DELETE FROM map_price_records WHERE product_id = ?", (product_id,))
    conn.execute("DELETE FROM map_products WHERE id = ?", (product_id,))
    conn.commit(); conn.close()
    return {"message": "삭제 완료"}

class BulkUpdate(BaseModel):
    product_ids: List[int]
    is_active: Optional[bool] = None
    is_watched: Optional[bool] = None

@router.put("/products/bulk")
async def bulk_update_products(data: BulkUpdate):
    """일괄 감시/상시감시 ON/OFF"""
    if not data.product_ids:
        raise HTTPException(400, "제품 ID가 없습니다")
    conn = get_connection()
    sets = []
    vals = []
    if data.is_active is not None:
        sets.append("is_active = ?"); vals.append(1 if data.is_active else 0)
    if data.is_watched is not None:
        sets.append("is_watched = ?"); vals.append(1 if data.is_watched else 0)
        if not data.is_watched:
            sets.append("watch_interval_hours = NULL")
    if not sets:
        conn.close(); raise HTTPException(400, "변경 항목 없음")
    sets.append("updated_at = datetime('now','localtime')")
    placeholders = ",".join(["?"] * len(data.product_ids))
    conn.execute(f"UPDATE map_products SET {', '.join(sets)} WHERE id IN ({placeholders})", vals + data.product_ids)
    conn.commit(); conn.close()
    return {"message": f"{len(data.product_ids)}개 제품 업데이트 완료"}

@router.post("/products/upload")
async def upload_products_excel(file: UploadFile = File(...)):
    """엑셀/CSV로 제품 일괄 등록. 컬럼: 모델명|품명|브랜드|제품특징|지도가"""
    conn = get_connection()
    content = await file.read()
    added = updated = 0; errors = []
    try:
        if file.filename.endswith('.csv'):
            rows = list(csv.DictReader(io.StringIO(content.decode('utf-8-sig'))))
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = [c.value for c in ws[1] if c.value]
            rows = []
            for r in ws.iter_rows(min_row=2, values_only=True):
                if r[0]:
                    rows.append(dict(zip(headers, r)))
        col_map = {'모델명':'mn','model':'mn','model_name':'mn',
                   '품명':'pn','품목명':'pn','제품명':'pn','product_name':'pn','name':'pn',
                   '브랜드':'br','brand':'br',
                   '제품특징':'ft','특징':'ft','features':'ft',
                   '지도가':'mp','가격':'mp','price':'mp','map_price':'mp'}
        for i, row in enumerate(rows, 2):
            try:
                n = {}
                for k, v in row.items():
                    if k and k.strip() in col_map: n[col_map[k.strip()]] = v
                mn = str(n.get('mn','')).strip()
                pn = str(n.get('pn','')).strip()
                br = str(n.get('br','LANstar')).strip()
                ft = str(n.get('ft','')).strip()
                mp = int(float(str(n.get('mp',0)).replace(',','')))
                if not mn or not pn: errors.append(f"행{i}: 모델명/품명 누락"); continue
                sk = f"{br} {mn} {pn} {ft}"
                existing = conn.execute("SELECT id FROM map_products WHERE model_name=?", (mn,)).fetchone()
                if existing:
                    conn.execute("""UPDATE map_products SET product_name=?, brand=?, features=?, map_price=?,
                        search_keywords=?, updated_at=datetime('now','localtime') WHERE model_name=?""",
                        (pn, br, ft, mp, sk, mn))
                    updated += 1
                else:
                    conn.execute("""INSERT INTO map_products (model_name, product_name, brand, features, map_price, search_keywords)
                        VALUES (?,?,?,?,?,?)""", (mn, pn, br, ft, mp, sk))
                    added += 1
            except Exception as e: errors.append(f"행{i}: {e}")
        conn.commit()
    except Exception as e:
        conn.rollback(); conn.close(); raise HTTPException(500, f"파일 처리 오류: {e}")
    conn.close()
    return {"message": f"신규 {added}개, 업데이트 {updated}개", "added": added, "updated": updated, "errors": errors[:20]}

@router.put("/products/{product_id}/watch")
async def toggle_watch(product_id: int, watched: bool = True, interval_hours: int = 2):
    conn = get_connection()
    conn.execute("""UPDATE map_products SET is_watched=?, watch_interval_hours=?,
        updated_at=datetime('now','localtime') WHERE id=?""",
        (1 if watched else 0, interval_hours if watched else None, product_id))
    conn.commit(); conn.close()
    return {"message": f"상시감시 {'ON' if watched else 'OFF'}"}

@router.put("/products/{product_id}/map-price")
async def update_map_price(product_id: int, map_price: int = Query(...), reason: str = Query("")):
    conn = get_connection()
    # 기존 가격 조회
    old = conn.execute("SELECT map_price FROM map_products WHERE id=?", (product_id,)).fetchone()
    old_price = old["map_price"] if old else 0
    # 가격 업데이트
    conn.execute("UPDATE map_products SET map_price=?, updated_at=datetime('now','localtime') WHERE id=?",
                 (map_price, product_id))
    # 변경 이력 기록
    if old_price != map_price:
        conn.execute("""INSERT INTO map_price_history (product_id, old_price, new_price, reason)
            VALUES (?,?,?,?)""", (product_id, old_price, map_price, reason))
    conn.commit(); conn.close()
    return {"message": f"지도가 {old_price:,}원 → {map_price:,}원"}

@router.put("/products/batch")
async def batch_update_products(action: str = Query(...), ids: str = Query(...)):
    """일괄 업데이트. action: monitor_on/monitor_off/watch_on/watch_off, ids: 쉼표 구분 ID"""
    conn = get_connection()
    id_list = [int(i) for i in ids.split(",") if i.strip()]
    if not id_list:
        conn.close(); raise HTTPException(400, "대상 ID가 없습니다")
    placeholders = ",".join(["?"] * len(id_list))
    if action == "monitor_on":
        conn.execute(f"UPDATE map_products SET is_active=1, updated_at=datetime('now','localtime') WHERE id IN ({placeholders})", id_list)
    elif action == "monitor_off":
        conn.execute(f"UPDATE map_products SET is_active=0, updated_at=datetime('now','localtime') WHERE id IN ({placeholders})", id_list)
    elif action == "watch_on":
        conn.execute(f"UPDATE map_products SET is_watched=1, watch_interval_hours=2, updated_at=datetime('now','localtime') WHERE id IN ({placeholders})", id_list)
    elif action == "watch_off":
        conn.execute(f"UPDATE map_products SET is_watched=0, watch_interval_hours=NULL, updated_at=datetime('now','localtime') WHERE id IN ({placeholders})", id_list)
    else:
        conn.close(); raise HTTPException(400, f"알 수 없는 action: {action}")
    conn.commit(); conn.close()
    labels = {"monitor_on":"감시 ON","monitor_off":"감시 OFF","watch_on":"상시감시 ON","watch_off":"상시감시 OFF"}
    return {"message": f"{len(id_list)}개 제품 {labels.get(action,action)} 완료"}


# ═══════════════════════════════════════════════════════
# 3. 위반 API
# ═══════════════════════════════════════════════════════

@router.get("/violations")
async def list_violations(
    severity: str = "", platform: str = "", resolved: bool = False,
    days: int = 7, limit: int = 200, offset: int = 0, search: str = "",
    sort: str = "detected_at", order: str = "desc"
):
    conn = get_connection()
    cutoff = _days_ago(days)
    sql = """SELECT v.*, p.model_name, p.product_name, p.brand
        FROM map_violations v JOIN map_products p ON v.product_id = p.id
        WHERE v.detected_at > ?"""
    params = [cutoff]
    if not resolved: sql += " AND v.is_resolved = 0"
    if severity: sql += " AND v.severity = ?"; params.append(severity)
    if platform: sql += " AND v.platform = ?"; params.append(platform)
    if search:
        sql += " AND (p.model_name LIKE ? OR p.product_name LIKE ?)"
        params += [f"%{search}%", f"%{search}%"]

    count_sql = sql.replace("SELECT v.*, p.model_name, p.product_name, p.brand", "SELECT COUNT(*) as cnt")
    total = conn.execute(count_sql, params).fetchone()["cnt"]

    # 정렬 (허용 컬럼만)
    allowed_sorts = {"detected_at": "v.detected_at", "severity": "v.severity", "deviation_pct": "v.deviation_pct",
                     "map_price": "v.map_price", "violated_price": "v.violated_price", "seller_name": "v.seller_name",
                     "model_name": "p.model_name"}
    sort_col = allowed_sorts.get(sort, "v.detected_at")
    sort_dir = "ASC" if order == "asc" else "DESC"
    sql += f" ORDER BY {sort_col} {sort_dir} LIMIT ? OFFSET ?"
    params += [limit, offset]
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return {"violations": [dict(r) for r in rows], "total": total}


@router.get("/violations/grouped")
async def list_violations_grouped(severity: str = "", days: int = 7, search: str = ""):
    """제품별 그룹핑 위반 현황"""
    conn = get_connection()
    cutoff = _days_ago(days)
    sql = """SELECT p.id as product_id, p.model_name, p.product_name, p.map_price,
        COUNT(v.id) as violation_count,
        COUNT(DISTINCT v.seller_name) as seller_count,
        MIN(v.violated_price) as min_price,
        MAX(v.deviation_pct) as max_deviation,
        MAX(v.severity) as worst_severity,
        MAX(v.detected_at) as last_detected,
        STRING_AGG(DISTINCT v.seller_name, ',') as sellers
        FROM map_violations v JOIN map_products p ON v.product_id = p.id
        WHERE v.detected_at > ? AND v.is_resolved = 0"""
    params = [cutoff]
    if severity: sql += " AND v.severity = ?"; params.append(severity)
    if search:
        sql += " AND (p.model_name LIKE ? OR p.product_name LIKE ?)"
        params += [f"%{search}%", f"%{search}%"]
    sql += " GROUP BY p.id, p.model_name, p.product_name, p.map_price ORDER BY violation_count DESC"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    results = []
    for r in rows:
        d = dict(r)
        d["sellers"] = d.get("sellers", "").split(",") if d.get("sellers") else []
        results.append(d)
    return {"products": results, "total_products": len(results),
            "total_violations": sum(r["violation_count"] for r in results)}


@router.get("/violations/export")
async def export_violations_excel(severity: str = "", days: int = 7, search: str = ""):
    """위반 데이터 엑셀 다운로드"""
    from fastapi.responses import StreamingResponse
    import io, openpyxl
    conn = get_connection()
    cutoff = _days_ago(days)
    sql = """SELECT p.model_name, p.product_name, v.seller_name, v.platform,
        v.map_price, v.violated_price, v.deviation_pct, v.severity, v.violation_type,
        v.detected_at, v.evidence_url
        FROM map_violations v JOIN map_products p ON v.product_id = p.id
        WHERE v.detected_at > ? AND v.is_resolved = 0"""
    params = [cutoff]
    if severity: sql += " AND v.severity = ?"; params.append(severity)
    if search:
        sql += " AND (p.model_name LIKE ? OR p.product_name LIKE ?)"
        params += [f"%{search}%", f"%{search}%"]
    sql += " ORDER BY v.detected_at DESC"
    rows = conn.execute(sql, params).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "지도가 위반 현황"
    headers = ["모델명", "제품명", "셀러", "플랫폼", "지도가", "판매가", "편차(%)", "심각도", "유형", "탐지일시", "판매URL"]
    ws.append(headers)
    for r in rows:
        ws.append([r["model_name"], r["product_name"], r["seller_name"], r["platform"],
                   r["map_price"], r["violated_price"], r["deviation_pct"], r["severity"],
                   r["violation_type"], r["detected_at"], r["evidence_url"]])
    # 열폭 자동
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or "")) for c in col[:5]) + 4

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=MAP_violations_{days}d.xlsx"})

@router.put("/violations/{violation_id}/resolve")
async def resolve_violation(violation_id: int, data: ViolationResolve):
    conn = get_connection()
    conn.execute("""UPDATE map_violations SET is_resolved=1,
        resolved_at=datetime('now','localtime'), resolution_note=? WHERE id=?""",
        (data.resolution_note, violation_id))
    conn.commit(); conn.close()
    return {"message": "해결 처리 완료"}

@router.get("/violations/by-product/{product_id}")
async def violations_by_product(product_id: int, days: int = 7):
    """특정 제품의 개별 위반 목록"""
    conn = get_connection()
    cutoff = _days_ago(days)
    rows = conn.execute("""SELECT v.*, p.model_name, p.product_name FROM map_violations v
        JOIN map_products p ON v.product_id=p.id
        WHERE v.product_id=? AND v.detected_at>? AND v.is_resolved=0
        ORDER BY v.violated_price ASC""", (product_id, cutoff)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@router.post("/violations/cleanup")
async def cleanup_duplicate_violations():
    """기존 누적된 중복 위반 정리 — 같은 제품+셀러에 대해 최신 1건만 남기고 삭제"""
    conn = get_connection()
    try:
        before = conn.execute("SELECT COUNT(*) as c FROM map_violations WHERE is_resolved=0").fetchone()["c"]
        # 1단계: 유지할 ID 조회
        keep_rows = conn.execute("""SELECT MAX(id) as keep_id FROM map_violations
            WHERE is_resolved = 0 GROUP BY product_id, seller_name""").fetchall()
        keep_ids = [r["keep_id"] for r in keep_rows if r["keep_id"]]
        if keep_ids:
            placeholders = ",".join(["?"] * len(keep_ids))
            # 2단계: 삭제 대상의 경고 메일 먼저 삭제 (외래키 참조)
            conn.execute(f"""DELETE FROM map_warning_emails WHERE violation_id IS NOT NULL
                AND violation_id NOT IN ({placeholders})""", keep_ids)
            # 3단계: 중복 위반 삭제
            conn.execute(f"DELETE FROM map_violations WHERE is_resolved = 0 AND id NOT IN ({placeholders})", keep_ids)
        conn.commit()
        after = conn.execute("SELECT COUNT(*) as c FROM map_violations WHERE is_resolved=0").fetchone()["c"]
        conn.close()
        return {"message": f"중복 위반 {before - after}건 정리 완료. {before}건 → {after}건"}
    except Exception as e:
        conn.close()
        logger.error(f"cleanup 오류: {e}")
        raise HTTPException(500, f"정리 오류: {e}")


# ═══════════════════════════════════════════════════════
# 4. 대시보드 API
# ═══════════════════════════════════════════════════════

@router.get("/dashboard")
async def get_dashboard():
    conn = get_connection()
    # 설정
    s = conn.execute("SELECT * FROM map_settings WHERE id=1").fetchone()
    s = dict(s) if s else {}
    mp = s.get('min_price', 5000)

    # 제품 통계
    monitored = conn.execute("SELECT COUNT(*) as c FROM map_products WHERE is_active=1 AND map_price>=?", (mp,)).fetchone()["c"]
    watched = conn.execute("SELECT COUNT(*) as c FROM map_products WHERE is_watched=1").fetchone()["c"]
    total_p = conn.execute("SELECT COUNT(*) as c FROM map_products WHERE is_active=1").fetchone()["c"]

    # 위반 통계 (7일)
    cutoff7 = _days_ago(7)
    vs = conn.execute("""SELECT COUNT(*) as total,
        SUM(CASE WHEN severity='CRITICAL' THEN 1 ELSE 0 END) as critical,
        SUM(CASE WHEN severity='HIGH' THEN 1 ELSE 0 END) as high,
        SUM(CASE WHEN severity='MEDIUM' THEN 1 ELSE 0 END) as medium,
        COUNT(DISTINCT product_id) as product_count
        FROM map_violations WHERE detected_at>? AND is_resolved=0""", (cutoff7,)).fetchone()
    vs = dict(vs) if vs else {"total":0,"critical":0,"high":0,"medium":0,"product_count":0}

    # Top 셀러 (30일)
    cutoff30 = _days_ago(30)
    top_rows = conn.execute("""SELECT seller_name, platform, COUNT(*) as cnt,
        MAX(detected_at) as last_at FROM map_violations WHERE detected_at>?
        GROUP BY seller_name, platform ORDER BY cnt DESC LIMIT 10""", (cutoff30,)).fetchall()
    top_sellers = [dict(r) for r in top_rows]

    # 최근 수집
    lc = conn.execute("SELECT * FROM map_collection_logs ORDER BY started_at DESC LIMIT 1").fetchone()

    # 최근 위반 5건
    rv = conn.execute("""SELECT v.*, p.model_name, p.product_name FROM map_violations v
        JOIN map_products p ON v.product_id=p.id WHERE v.is_resolved=0
        ORDER BY v.detected_at DESC LIMIT 5""").fetchall()

    conn.close()
    schedules = json.loads(s['schedules']) if isinstance(s.get('schedules'), str) else s.get('schedules', [])
    platforms = json.loads(s['platforms']) if isinstance(s.get('platforms'), str) else s.get('platforms', [])
    return {
        "total_products": total_p, "monitored_count": monitored, "watch_count": watched,
        "violation_stats": vs, "top_sellers": top_sellers,
        "recent_violations": [dict(r) for r in rv],
        "last_collection": dict(lc) if lc else None,
        "settings_summary": {"min_price": mp, "schedules": schedules,
                             "platforms": platforms, "watch_interval_hours": s.get('watch_interval_hours', 2)}
    }


# ═══════════════════════════════════════════════════════
# 5. 셀러 API
# ═══════════════════════════════════════════════════════

@router.get("/sellers")
async def list_sellers(risk: str = "", platform: str = ""):
    conn = get_connection()
    cutoff30 = _days_ago(30)
    sql = """SELECT s.*, (SELECT COUNT(*) FROM map_violations v
        WHERE v.seller_name=s.seller_name AND v.platform=s.platform AND v.detected_at>?) as recent_violations
        FROM map_sellers s WHERE 1=1"""
    params = [cutoff30]
    if risk: sql += " AND s.risk_level=?"; params.append(risk)
    if platform: sql += " AND s.platform=?"; params.append(platform)
    sql += " ORDER BY s.total_violations DESC"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ═══════════════════════════════════════════════════════
# 6. 가격 이력 API
# ═══════════════════════════════════════════════════════

@router.get("/price-history/{product_id}")
async def get_price_history(product_id: int, days: int = 7, platform: str = ""):
    conn = get_connection()
    cutoff = _days_ago(days)
    sql = """SELECT seller_name, platform, effective_price, display_price,
        coupon_price, is_violation, collected_at FROM map_price_records
        WHERE product_id=? AND collected_at>?"""
    params = [product_id, cutoff]
    if platform: sql += " AND platform=?"; params.append(platform)
    sql += " ORDER BY collected_at ASC"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ═══════════════════════════════════════════════════════
# 7. 수집 실행 API (백그라운드 + 진행률)
# ═══════════════════════════════════════════════════════

@router.post("/collect/run")
async def run_collection_now():
    """즉시 가격 수집 - 백그라운드 실행"""
    from services.map_collector_service import start_collection_background, collection_progress
    if collection_progress.get("running"):
        return {"message": "이미 수집 중입니다", "status": "running", "progress": collection_progress}
    import asyncio
    asyncio.create_task(start_collection_background())
    return {"message": "수집 시작됨", "status": "started"}

@router.get("/collect/progress")
async def get_collection_progress():
    """수집 진행률 조회"""
    from services.map_collector_service import collection_progress
    return collection_progress

@router.get("/collect/status")
async def get_collection_status():
    conn = get_connection()
    rows = conn.execute("SELECT * FROM map_collection_logs ORDER BY started_at DESC LIMIT 10").fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ═══════════════════════════════════════════════════════
# 8. 통계 API
# ═══════════════════════════════════════════════════════

@router.get("/stats/violations-by-day")
async def violations_by_day(days: int = 30):
    conn = get_connection()
    cutoff = _days_ago(days)
    rows = conn.execute("""SELECT DATE(detected_at) as date, COUNT(*) as total,
        SUM(CASE WHEN severity='CRITICAL' THEN 1 ELSE 0 END) as critical,
        SUM(CASE WHEN severity='HIGH' THEN 1 ELSE 0 END) as high
        FROM map_violations WHERE detected_at>? GROUP BY DATE(detected_at) ORDER BY date""", (cutoff,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@router.get("/stats/violations-by-platform")
async def violations_by_platform(days: int = 30):
    conn = get_connection()
    cutoff = _days_ago(days)
    rows = conn.execute("""SELECT platform, COUNT(*) as count,
        SUM(CASE WHEN severity IN ('CRITICAL','HIGH') THEN 1 ELSE 0 END) as severe
        FROM map_violations WHERE detected_at>? GROUP BY platform ORDER BY count DESC""", (cutoff,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

# ═══════════════════════════════════════════════════════
# 9. 스케줄러 상태 API
# ═══════════════════════════════════════════════════════

@router.get("/scheduler/status")
async def scheduler_status():
    try:
        from services.map_scheduler import get_scheduler_status
        return get_scheduler_status()
    except Exception as e:
        return {"running": False, "error": str(e), "jobs": []}

# ═══════════════════════════════════════════════════════
# 10. 지도가 변경 이력 API
# ═══════════════════════════════════════════════════════

@router.get("/price-change-history/{product_id}")
async def get_price_change_history(product_id: int):
    """제품별 지도가 변경 이력"""
    conn = get_connection()
    rows = conn.execute("""SELECT h.*, p.model_name, p.product_name
        FROM map_price_history h JOIN map_products p ON h.product_id=p.id
        WHERE h.product_id=? ORDER BY h.changed_at DESC LIMIT 50""", (product_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@router.get("/price-change-history")
async def get_all_price_changes(days: int = 30):
    """전체 지도가 변경 이력"""
    conn = get_connection()
    cutoff = _days_ago(days)
    rows = conn.execute("""SELECT h.*, p.model_name, p.product_name
        FROM map_price_history h JOIN map_products p ON h.product_id=p.id
        WHERE h.changed_at>? ORDER BY h.changed_at DESC LIMIT 100""", (cutoff,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ═══════════════════════════════════════════════════════
# 11. 경고 메일 템플릿 API
# ═══════════════════════════════════════════════════════

@router.post("/warning-email/generate/{violation_id}")
async def generate_warning_email(violation_id: int):
    """위반 건에 대한 경고 메일 템플릿 생성"""
    conn = get_connection()
    v = conn.execute("""SELECT v.*, p.model_name, p.product_name, p.brand, p.map_price as current_map
        FROM map_violations v JOIN map_products p ON v.product_id=p.id WHERE v.id=?""", (violation_id,)).fetchone()
    if not v:
        conn.close(); raise HTTPException(404, "위반 건 없음")
    v = dict(v)

    subject = f"[LANstar] 지도가 위반 경고 - {v['product_name']} ({v['model_name']})"
    body = f"""안녕하세요, {v['seller_name']} 담당자님.

귀사에서 판매 중인 LANstar 제품의 지도가(MAP) 위반이 확인되어 경고드립니다.

■ 위반 상세
  - 제품명: {v['product_name']}
  - 모델명: {v['model_name']}
  - 판매 플랫폼: {v['platform']}
  - 지도가: {v['map_price']:,}원
  - 판매가: {v['violated_price']:,}원
  - 편차: -{v['deviation_pct']}%
  - 위반 유형: {v['violation_type']}
  - 탐지 일시: {v.get('detected_at','')}

■ 요청 사항
  지도가 정책에 따라 즉시 판매가를 {v['map_price']:,}원 이상으로 조정해 주시기 바랍니다.
  지속적인 위반 시 거래 조건 변경 등 불이익이 있을 수 있습니다.

■ 참고
  본 메일은 LANstar 지도가 감시 시스템에서 자동 생성되었습니다.
  문의사항은 영업 담당자에게 연락 부탁드립니다.

감사합니다.
LANstar Co., Ltd."""

    # DB에 저장
    cur = conn.execute("""INSERT INTO map_warning_emails
        (violation_id, seller_name, platform, product_name, model_name, email_subject, email_body, status)
        VALUES (?,?,?,?,?,?,?,?)""",
        (violation_id, v['seller_name'], v['platform'], v['product_name'], v['model_name'],
         subject, body, 'draft'))
    email_id = cur.lastrowid
    conn.commit(); conn.close()

    return {"id": email_id, "subject": subject, "body": body, "seller": v['seller_name'], "platform": v['platform']}


@router.get("/warning-emails")
async def list_warning_emails(status: str = ""):
    """경고 메일 목록"""
    conn = get_connection()
    sql = "SELECT * FROM map_warning_emails WHERE 1=1"
    params = []
    if status: sql += " AND status=?"; params.append(status)
    sql += " ORDER BY created_at DESC LIMIT 50"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@router.put("/warning-email/{email_id}/send")
async def send_warning_email(email_id: int, email_to: str = Query(...)):
    """경고 메일 발송 (SMTP)"""
    import smtplib, os
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    conn = get_connection()
    em = conn.execute("SELECT * FROM map_warning_emails WHERE id=?", (email_id,)).fetchone()
    if not em:
        conn.close(); raise HTTPException(404, "메일 없음")
    em = dict(em)

    smtp_host = os.getenv("SMTP_HOST", "")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USER", "")
    smtp_pass = os.getenv("SMTP_PASS", "")

    if not smtp_host or not smtp_user:
        # SMTP 미설정 시 발송 없이 상태만 업데이트
        conn.execute("UPDATE map_warning_emails SET email_to=?, status='ready' WHERE id=?", (email_to, email_id))
        conn.commit(); conn.close()
        return {"message": "SMTP 미설정. 메일 내용이 저장되었습니다. (발송하려면 SMTP 환경변수 설정 필요)", "status": "ready"}

    try:
        msg = MIMEMultipart()
        msg['From'] = smtp_user
        msg['To'] = email_to
        msg['Subject'] = em['email_subject']
        msg.attach(MIMEText(em['email_body'], 'plain', 'utf-8'))

        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, email_to, msg.as_string())

        from datetime import datetime, timezone, timedelta
        KST = timezone(timedelta(hours=9))
        now_kst = datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S")
        conn.execute("UPDATE map_warning_emails SET email_to=?, status='sent', sent_at=? WHERE id=?",
                     (email_to, now_kst, email_id))
        conn.commit(); conn.close()
        return {"message": f"{email_to}로 경고 메일 발송 완료", "status": "sent"}
    except Exception as e:
        conn.execute("UPDATE map_warning_emails SET email_to=?, status='failed' WHERE id=?", (email_to, email_id))
        conn.commit(); conn.close()
        raise HTTPException(500, f"메일 발송 오류: {e}")


# ═══════════════════════════════════════════════════════
# 12. 위반 증거 스크린샷 API
# ═══════════════════════════════════════════════════════

@router.post("/screenshot/{violation_id}")
async def capture_screenshot(violation_id: int):
    """위반 페이지 스크린샷 캡처 (Playwright)"""
    conn = get_connection()
    v = conn.execute("""SELECT v.*, p.model_name FROM map_violations v
        JOIN map_products p ON v.product_id=p.id WHERE v.id=?""", (violation_id,)).fetchone()
    if not v:
        conn.close(); raise HTTPException(404, "위반 건 없음")
    v = dict(v)

    url = v.get("evidence_url", "")
    if not url:
        conn.close(); raise HTTPException(400, "판매 URL이 없어 스크린샷 불가")

    try:
        from playwright.async_api import async_playwright
        import os, base64
        from datetime import datetime, timezone, timedelta
        KST = timezone(timedelta(hours=9))
        timestamp = datetime.now(KST).strftime("%Y%m%d_%H%M%S")

        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            page = await browser.new_page(viewport={"width": 1280, "height": 900})
            await page.goto(url, wait_until="domcontentloaded", timeout=30000)
            await page.wait_for_timeout(3000)  # JS 렌더링 대기

            # 스크린샷 바이너리
            screenshot_bytes = await page.screenshot(full_page=False)
            await browser.close()

        # base64로 반환 (DB에 경로 저장 대신 즉시 반환)
        b64 = base64.b64encode(screenshot_bytes).decode()

        # DB에 스크린샷 경로 기록
        fname = f"vio_{violation_id}_{timestamp}.png"
        conn.execute("UPDATE map_violations SET evidence_screenshot=? WHERE id=?", (fname, violation_id))
        conn.commit(); conn.close()

        return {
            "message": "스크린샷 캡처 완료",
            "filename": fname,
            "image_base64": b64,
            "url": url,
            "model_name": v["model_name"],
        }
    except ImportError:
        conn.close()
        raise HTTPException(500, "Playwright가 설치되지 않았습니다")
    except Exception as e:
        conn.close()
        logger.error(f"스크린샷 오류: {e}")
        raise HTTPException(500, f"스크린샷 캡처 오류: {e}")
