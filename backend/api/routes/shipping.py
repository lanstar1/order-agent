"""
택배 발송 조회 API
- 발송 내역 등록 (수동 / 엑셀 업로드)
- 받는사람 이름 검색
- 날짜별 조회
- 운송장번호 화물추적 (추적 결과 자동 DB 저장)
- 대량 운송장 동기화 (운송장번호 목록 → API 추적 → DB 저장)
"""
import logging
import io
from datetime import datetime
from fastapi import APIRouter, Depends, Query, UploadFile, File
from pydantic import BaseModel
from typing import Optional
from security import get_current_user
from db.database import get_connection

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/shipping", tags=["shipping"])


# ─── Request / Response 모델 ───────────────────
class ShipmentCreate(BaseModel):
    warehouse: str = "용산"          # 용산 / 김포
    slip_no: str                     # 운송장번호
    rcv_name: str                    # 받는사람
    rcv_tel: str = ""
    rcv_cell: str = ""
    rcv_addr1: str = ""
    rcv_addr2: str = ""
    rcv_zip: str = ""
    snd_name: str = ""               # 보내는사람
    snd_tel: str = ""
    snd_addr: str = ""
    goods_nm: str = ""               # 물품명
    qty: int = 1
    take_dt: str = ""                # 접수일자 YYYYMMDD
    memo: str = ""


class TrackRequest(BaseModel):
    slip_nos: list[str]              # 운송장번호 목록


class SyncRequest(BaseModel):
    slip_nos: list[str]              # 동기화할 운송장번호 목록


class ShipmentBulk(BaseModel):
    items: list[ShipmentCreate]


def _save_tracking_to_db(tracking_items: list[dict]):
    """
    추적 결과를 shipments 테이블에 자동 저장.
    이미 존재하는 운송장번호는 상태만 업데이트.
    """
    if not tracking_items:
        return 0

    conn = get_connection()
    saved = 0
    try:
        for item in tracking_items:
            slip_no = item.get("slipNo", "").strip()
            if not slip_no:
                continue

            warehouse = item.get("_warehouse", "")
            result_cd = item.get("resultCd", "")

            # 추적 성공한 건만 저장
            if result_cd != "TRUE":
                continue

            # data1에서 최신 상태, 받는사람 정보 추출
            scans = item.get("data1", [])
            status = "접수"
            rcv_name = item.get("rcvNm", "") or ""
            rcv_addr = item.get("rcvAddr", "") or ""
            snd_name = item.get("sndNm", "") or ""
            take_dt = ""
            goods_nm = item.get("goodsNm", "") or ""

            if scans:
                # 최신 상태 (마지막 스캔)
                last = scans[-1] if scans else {}
                status = last.get("statNm", "접수")
                # 접수일자: 첫 번째 스캔 날짜
                first = scans[0]
                take_dt = first.get("scanDt", "").replace("-", "")[:8]

            if not take_dt:
                take_dt = datetime.now().strftime("%Y%m%d")

            if not rcv_name:
                rcv_name = "미확인"

            # UPSERT: 있으면 상태 업데이트, 없으면 삽입
            existing = conn.execute(
                "SELECT id FROM shipments WHERE slip_no = ? AND warehouse = ?",
                (slip_no, warehouse)
            ).fetchone()

            if existing:
                conn.execute(
                    "UPDATE shipments SET status = ? WHERE slip_no = ? AND warehouse = ?",
                    (status, slip_no, warehouse)
                )
            else:
                conn.execute(
                    """INSERT OR IGNORE INTO shipments
                       (warehouse, slip_no, rcv_name, rcv_addr1, snd_name,
                        goods_nm, take_dt, status)
                       VALUES(?,?,?,?,?,?,?,?)""",
                    (warehouse, slip_no, rcv_name, rcv_addr, snd_name,
                     goods_nm, take_dt, status)
                )
            saved += 1

        conn.commit()
    except Exception as e:
        logger.error(f"[Shipping] 추적결과 DB 저장 오류: {e}")
    finally:
        conn.close()

    return saved


# ─── 발송 내역 등록 (단건) ────────────────────────
@router.post("/register")
async def register_shipment(req: ShipmentCreate, user: dict = Depends(get_current_user)):
    conn = get_connection()
    take_dt = req.take_dt or datetime.now().strftime("%Y%m%d")
    try:
        conn.execute(
            """INSERT OR IGNORE INTO shipments
               (warehouse, slip_no, rcv_name, rcv_tel, rcv_cell,
                rcv_addr1, rcv_addr2, rcv_zip,
                snd_name, snd_tel, snd_addr,
                goods_nm, qty, take_dt, memo)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (req.warehouse, req.slip_no, req.rcv_name, req.rcv_tel, req.rcv_cell,
             req.rcv_addr1, req.rcv_addr2, req.rcv_zip,
             req.snd_name, req.snd_tel, req.snd_addr,
             req.goods_nm, req.qty, take_dt, req.memo)
        )
        conn.commit()
        return {"success": True, "message": "등록 완료"}
    except Exception as e:
        logger.error(f"[Shipping] 등록 오류: {e}")
        return {"success": False, "message": str(e)}
    finally:
        conn.close()


# ─── 발송 내역 등록 (대량) ────────────────────────
@router.post("/register-bulk")
async def register_bulk(req: ShipmentBulk, user: dict = Depends(get_current_user)):
    conn = get_connection()
    inserted = 0
    try:
        for item in req.items:
            take_dt = item.take_dt or datetime.now().strftime("%Y%m%d")
            conn.execute(
                """INSERT OR IGNORE INTO shipments
                   (warehouse, slip_no, rcv_name, rcv_tel, rcv_cell,
                    rcv_addr1, rcv_addr2, rcv_zip,
                    snd_name, snd_tel, snd_addr,
                    goods_nm, qty, take_dt, memo)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (item.warehouse, item.slip_no, item.rcv_name, item.rcv_tel, item.rcv_cell,
                 item.rcv_addr1, item.rcv_addr2, item.rcv_zip,
                 item.snd_name, item.snd_tel, item.snd_addr,
                 item.goods_nm, item.qty, take_dt, item.memo)
            )
            inserted += 1
        conn.commit()
        return {"success": True, "inserted": inserted}
    except Exception as e:
        logger.error(f"[Shipping] 대량 등록 오류: {e}")
        return {"success": False, "message": str(e)}
    finally:
        conn.close()


# ─── 엑셀 업로드로 대량 등록 ──────────────────────
@router.post("/upload-excel")
async def upload_excel(
    file: UploadFile = File(...),
    warehouse: str = Query("용산"),
    user: dict = Depends(get_current_user),
):
    """
    엑셀 파일 업로드로 발송내역 대량 등록
    필수 컬럼: 운송장번호, 받는분, 접수일자
    선택 컬럼: 받는분전화, 받는분휴대폰, 받는분주소, 물품명, 수량, 비고
    """
    import openpyxl

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb[wb.sheetnames[0]]

        # 헤더 매핑
        headers = []
        col_map = {}
        for i, row_data in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c or "").strip() for c in row_data]
                # 유연한 컬럼 매핑
                for idx, h in enumerate(headers):
                    h_lower = h.replace(" ", "")
                    if "운송장" in h_lower or "송장번호" in h_lower or "슬립" in h_lower:
                        col_map["slip_no"] = idx
                    elif "받는" in h_lower and ("이름" in h_lower or "분" in h_lower or "사람" in h_lower or "명" in h_lower):
                        col_map["rcv_name"] = idx
                    elif "받는" in h_lower and "전화" in h_lower:
                        col_map["rcv_tel"] = idx
                    elif "받는" in h_lower and ("휴대" in h_lower or "핸드" in h_lower or "셀" in h_lower):
                        col_map["rcv_cell"] = idx
                    elif "받는" in h_lower and "주소" in h_lower:
                        col_map["rcv_addr1"] = idx
                    elif "접수일" in h_lower or "발송일" in h_lower or "일자" in h_lower:
                        col_map["take_dt"] = idx
                    elif "물품" in h_lower or "상품" in h_lower:
                        col_map["goods_nm"] = idx
                    elif "수량" in h_lower:
                        col_map["qty"] = idx
                    elif "비고" in h_lower or "메모" in h_lower:
                        col_map["memo"] = idx
                    elif "보내는" in h_lower and ("이름" in h_lower or "분" in h_lower or "명" in h_lower):
                        col_map["snd_name"] = idx
                    elif "수하인" in h_lower or ("수취" in h_lower and "인" in h_lower):
                        col_map["rcv_name"] = idx
                continue

            # 데이터 행
            slip_no = str(row_data[col_map.get("slip_no", 0)] or "").strip()
            rcv_name = str(row_data[col_map.get("rcv_name", 1)] or "").strip()
            if not slip_no or not rcv_name:
                continue

            take_dt_raw = row_data[col_map["take_dt"]] if "take_dt" in col_map else None
            if take_dt_raw:
                take_dt = str(take_dt_raw).strip().replace("-", "").replace("/", "")[:8]
            else:
                take_dt = datetime.now().strftime("%Y%m%d")

            # DB 저장은 아래에서 일괄 처리
            break  # 헤더만 먼저 파싱

        # 실제 데이터 삽입
        conn = get_connection()
        inserted = 0
        skipped = 0
        for i, row_data in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            slip_no = str(row_data[col_map.get("slip_no", 0)] or "").strip()
            rcv_name = str(row_data[col_map.get("rcv_name", 1)] or "").strip()
            if not slip_no or not rcv_name:
                skipped += 1
                continue

            take_dt_raw = row_data[col_map["take_dt"]] if "take_dt" in col_map else None
            if take_dt_raw:
                take_dt = str(take_dt_raw).strip().replace("-", "").replace("/", "")[:8]
            else:
                take_dt = datetime.now().strftime("%Y%m%d")

            rcv_tel = str(row_data[col_map["rcv_tel"]] or "").strip() if "rcv_tel" in col_map else ""
            rcv_cell = str(row_data[col_map["rcv_cell"]] or "").strip() if "rcv_cell" in col_map else ""
            rcv_addr1 = str(row_data[col_map["rcv_addr1"]] or "").strip() if "rcv_addr1" in col_map else ""
            goods_nm = str(row_data[col_map["goods_nm"]] or "").strip() if "goods_nm" in col_map else ""
            qty_raw = row_data[col_map["qty"]] if "qty" in col_map else 1
            qty = int(qty_raw) if qty_raw else 1
            memo = str(row_data[col_map["memo"]] or "").strip() if "memo" in col_map else ""
            snd_name = str(row_data[col_map["snd_name"]] or "").strip() if "snd_name" in col_map else ""

            conn.execute(
                """INSERT OR IGNORE INTO shipments
                   (warehouse, slip_no, rcv_name, rcv_tel, rcv_cell,
                    rcv_addr1, snd_name, goods_nm, qty, take_dt, memo)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                (warehouse, slip_no, rcv_name, rcv_tel, rcv_cell,
                 rcv_addr1, snd_name, goods_nm, qty, take_dt, memo)
            )
            inserted += 1

        conn.commit()
        conn.close()
        wb.close()

        return {
            "success": True,
            "inserted": inserted,
            "skipped": skipped,
            "columns_found": list(col_map.keys()),
        }
    except Exception as e:
        logger.error(f"[Shipping] 엑셀 업로드 오류: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


# ─── 받는사람 이름 검색 ──────────────────────────
@router.get("/search")
async def search_shipments(
    q: str = Query("", description="받는사람 이름"),
    date: str = Query("", description="접수일자 YYYYMMDD"),
    warehouse: str = Query("", description="창고 필터 (용산/김포/전체)"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    user: dict = Depends(get_current_user),
):
    conn = get_connection()
    where_parts = ["1=1"]
    params = []

    if q:
        where_parts.append("rcv_name LIKE ?")
        params.append(f"%{q}%")
    if date:
        where_parts.append("take_dt = ?")
        params.append(date.replace("-", ""))
    if warehouse and warehouse != "전체":
        where_parts.append("warehouse = ?")
        params.append(warehouse)

    where = " AND ".join(where_parts)

    total = conn.execute(
        f"SELECT COUNT(*) as cnt FROM shipments WHERE {where}", params
    ).fetchone()["cnt"]

    offset = (page - 1) * page_size
    rows = conn.execute(
        f"""SELECT * FROM shipments WHERE {where}
            ORDER BY take_dt DESC, created_at DESC
            LIMIT ? OFFSET ?""",
        params + [page_size, offset]
    ).fetchall()

    conn.close()
    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size if total > 0 else 0,
        "items": [dict(r) for r in rows],
    }


# ─── 날짜별 발송내역 조회 ────────────────────────
@router.get("/daily")
async def daily_shipments(
    date: str = Query(..., description="조회 날짜 YYYYMMDD 또는 YYYY-MM-DD"),
    warehouse: str = Query("", description="창고 필터"),
    user: dict = Depends(get_current_user),
):
    clean_date = date.replace("-", "")
    conn = get_connection()

    where_parts = ["take_dt = ?"]
    params = [clean_date]

    if warehouse and warehouse != "전체":
        where_parts.append("warehouse = ?")
        params.append(warehouse)

    where = " AND ".join(where_parts)

    rows = conn.execute(
        f"""SELECT * FROM shipments WHERE {where}
            ORDER BY warehouse, created_at DESC""",
        params
    ).fetchall()

    # 창고별 요약
    summary = conn.execute(
        f"""SELECT warehouse, COUNT(*) as cnt
            FROM shipments WHERE {where}
            GROUP BY warehouse""",
        params
    ).fetchall()

    conn.close()
    return {
        "date": clean_date,
        "total": len(rows),
        "summary": [dict(r) for r in summary],
        "items": [dict(r) for r in rows],
    }


# ─── 운송장번호로 화물추적 (결과 자동 DB 저장) ──────────
@router.post("/track")
async def track_shipment(req: TrackRequest, user: dict = Depends(get_current_user)):
    import re
    from services.logen_client import track_shipments_both

    if not req.slip_nos:
        return {"success": False, "message": "운송장번호를 입력하세요"}

    # 운송장번호 정규화
    clean_slips = [re.sub(r'\D', '', s.strip()) for s in req.slip_nos if s.strip()]
    clean_slips = [s for s in clean_slips if len(s) >= 10]

    if not clean_slips:
        return {"success": False, "message": "유효한 운송장번호가 없습니다"}

    # 1) 로젠 OpenAPI로 추적 시도
    results = await track_shipments_both(clean_slips)

    # 추적 성공한 운송장번호 집합
    tracked_slips = set()
    for item in results:
        if item.get("resultCd") == "TRUE" and item.get("data1"):
            tracked_slips.add(item.get("slipNo", ""))

    # 2) API 추적 실패한 건은 DB 상태 정보로 보충
    failed_slips = [s for s in clean_slips if s not in tracked_slips]
    if failed_slips:
        conn = get_connection()
        try:
            for slip in failed_slips:
                row = conn.execute(
                    "SELECT * FROM shipments WHERE slip_no = ?", (slip,)
                ).fetchone()
                if row:
                    r = dict(row)
                    # DB 정보를 Logen API 형식으로 변환하여 결과에 추가
                    status = r.get("status", "접수")
                    take_dt = r.get("take_dt", "")
                    db_item = {
                        "slipNo": slip,
                        "resultCd": "DB",
                        "resultMsg": "DB 저장 정보",
                        "_warehouse": r.get("warehouse", ""),
                        "_db_info": {
                            "rcv_name": r.get("rcv_name", ""),
                            "rcv_addr1": r.get("rcv_addr1", ""),
                            "goods_nm": r.get("goods_nm", ""),
                            "status": status,
                            "take_dt": take_dt,
                            "snd_dt": r.get("snd_dt", ""),
                            "dlv_dt": r.get("dlv_dt", ""),
                        },
                        "data1": [],
                    }
                    # DB 상태 정보를 스캔 이력으로 변환
                    if take_dt:
                        db_item["data1"].append({
                            "scanDt": take_dt,
                            "scanTm": "000000",
                            "statNm": "접수",
                            "branNm": r.get("warehouse", ""),
                            "salesNm": "",
                        })
                    if status and status != "접수":
                        snd_dt = r.get("snd_dt", "") or take_dt
                        db_item["data1"].append({
                            "scanDt": snd_dt,
                            "scanTm": "000000",
                            "statNm": status,
                            "branNm": r.get("warehouse", ""),
                            "salesNm": "",
                        })

                    # 이미 동일 슬립 있으면 교체, 없으면 추가
                    existing_idx = None
                    for idx, existing in enumerate(results):
                        if existing.get("slipNo") == slip:
                            existing_idx = idx
                            break
                    if existing_idx is not None:
                        results[existing_idx] = db_item
                    else:
                        results.append(db_item)
        finally:
            conn.close()

    # 추적 결과를 자동으로 DB에 저장
    saved = _save_tracking_to_db(results)

    return {
        "success": True,
        "total": len(results),
        "saved_to_db": saved,
        "items": results,
    }


# ─── 대량 운송장번호 동기화 (API 추적 → DB 저장) ────────
@router.post("/sync")
async def sync_shipments(req: SyncRequest, user: dict = Depends(get_current_user)):
    """
    운송장번호 목록을 받아서 로젠 API로 추적 후 결과를 DB에 자동 저장.
    한번에 최대 50건씩 처리.
    """
    from services.logen_client import track_shipments_both

    if not req.slip_nos:
        return {"success": False, "message": "운송장번호를 입력하세요"}

    all_slip_nos = [s.strip() for s in req.slip_nos if s.strip()]
    total_saved = 0
    total_tracked = 0
    errors = []

    # 50건씩 배치 처리
    batch_size = 50
    for i in range(0, len(all_slip_nos), batch_size):
        batch = all_slip_nos[i:i + batch_size]
        try:
            results = await track_shipments_both(batch)
            total_tracked += len(results)
            saved = _save_tracking_to_db(results)
            total_saved += saved
        except Exception as e:
            logger.error(f"[Shipping] 동기화 배치 오류: {e}")
            errors.append(str(e))

    return {
        "success": True,
        "total_requested": len(all_slip_nos),
        "total_tracked": total_tracked,
        "total_saved": total_saved,
        "errors": errors,
    }


# ─── 단건 발송 + 로젠 API 등록 ───────────────────
@router.post("/register-and-send")
async def register_and_send(req: ShipmentCreate, user: dict = Depends(get_current_user)):
    """자체 DB 저장 + 로젠 API로 송장 등록"""
    from services.logen_client import register_slip

    take_dt = req.take_dt or datetime.now().strftime("%Y%m%d")

    # 1) 자체 DB 저장
    conn = get_connection()
    conn.execute(
        """INSERT OR IGNORE INTO shipments
           (warehouse, slip_no, rcv_name, rcv_tel, rcv_cell,
            rcv_addr1, rcv_addr2, rcv_zip,
            snd_name, snd_tel, snd_addr,
            goods_nm, qty, take_dt, memo)
           VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (req.warehouse, req.slip_no, req.rcv_name, req.rcv_tel, req.rcv_cell,
         req.rcv_addr1, req.rcv_addr2, req.rcv_zip,
         req.snd_name, req.snd_tel, req.snd_addr,
         req.goods_nm, req.qty, take_dt, req.memo)
    )
    conn.commit()
    conn.close()

    # 2) 로젠 API 호출
    slip_data = {
        "printYn": "Y",
        "slipNo": req.slip_no,
        "slipTy": "100",
        "custCd": "",
        "sndCustNm": req.snd_name,
        "sndTelNo": req.snd_tel,
        "sndCustAddr1": req.snd_addr,
        "sndCustAddr2": "",
        "rcvCustNm": req.rcv_name,
        "rcvTelNo": req.rcv_tel,
        "rcvCellNo": req.rcv_cell,
        "rcvCustAddr1": req.rcv_addr1,
        "rcvCustAddr2": req.rcv_addr2,
        "fareTy": req.goods_nm if req.goods_nm else "020",
        "qty": req.qty,
        "goodsNm": req.goods_nm,
        "dlvFare": 0,
        "extraFare": 0,
        "goodsAmt": 0,
        "takeDt": take_dt,
        "remarks": req.memo,
    }

    api_result = await register_slip(req.warehouse, slip_data)
    return {
        "success": True,
        "db_saved": True,
        "api_result": api_result,
    }


# ─── 엑셀에서 운송장번호 추출 → 추적 → DB 저장 ─────────
@router.post("/sync-excel")
async def sync_from_excel(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
):
    """엑셀 파일에서 운송장번호를 추출하여 로젠 API 추적 후 DB에 저장"""
    from services.logen_client import track_shipments_both
    import openpyxl
    import re

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb[wb.sheetnames[0]]

        slip_nos = set()
        slip_col_idx = None

        for i, row_data in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                # 헤더에서 운송장번호 컬럼 찾기
                for idx, cell in enumerate(row_data):
                    h = str(cell or "").replace(" ", "")
                    if "운송장" in h or "송장번호" in h or "슬립" in h or "slipNo" in h.lower():
                        slip_col_idx = idx
                        break
                if slip_col_idx is None:
                    # 헤더 없으면 첫 번째 컬럼 시도
                    slip_col_idx = 0
                continue

            # 데이터 행
            val = str(row_data[slip_col_idx] if slip_col_idx < len(row_data) else "").strip()
            # 10~15자리 숫자만 유효한 운송장번호로 간주
            cleaned = re.sub(r'\D', '', val)
            if 10 <= len(cleaned) <= 15:
                slip_nos.add(cleaned)

        wb.close()

        if not slip_nos:
            return {"success": False, "message": "유효한 운송장번호를 찾지 못했습니다."}

        slip_list = list(slip_nos)
        total_saved = 0
        total_tracked = 0

        # 50건씩 배치 처리
        batch_size = 50
        for i in range(0, len(slip_list), batch_size):
            batch = slip_list[i:i + batch_size]
            try:
                results = await track_shipments_both(batch)
                total_tracked += len(results)
                saved = _save_tracking_to_db(results)
                total_saved += saved
            except Exception as e:
                logger.error(f"[Shipping] 엑셀 동기화 배치 오류: {e}")

        return {
            "success": True,
            "total_extracted": len(slip_list),
            "total_tracked": total_tracked,
            "total_saved": total_saved,
        }
    except Exception as e:
        logger.error(f"[Shipping] 엑셀 동기화 오류: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


# ─── SmartLogen 자동 가져오기 ─────────────────────
@router.post("/auto-fetch")
async def auto_fetch_shipments(
    warehouse: str = Query("", description="창고 (용산/김포/전체)"),
    from_date: str = Query("", description="시작일 YYYYMMDD"),
    to_date: str = Query("", description="종료일 YYYYMMDD"),
    days: int = Query(7, description="조회 기간 (일)"),
    user: dict = Depends(get_current_user),
):
    """
    SmartLogen 포털에서 발송 실적을 자동으로 가져와 DB에 저장.
    로그인 → PickSndRecordSelect 호출 → SEED 복호화 → DB 저장.
    """
    from services.smart_logen_client import fetch_shipments, save_fetched_to_db

    try:
        # SmartLogen에서 발송 내역 조회
        wh = warehouse if warehouse and warehouse != "전체" else ""
        records = await fetch_shipments(
            warehouse=wh,
            from_date=from_date,
            to_date=to_date,
            days=days,
        )

        if not records:
            return {
                "success": True,
                "message": "조회된 발송 내역이 없습니다.",
                "fetched": 0,
                "saved": 0,
            }

        # DB에 저장
        conn = get_connection()
        try:
            saved = save_fetched_to_db(records, conn)
        finally:
            conn.close()

        return {
            "success": True,
            "fetched": len(records),
            "saved": saved,
            "message": f"SmartLogen에서 {len(records)}건 조회, {saved}건 저장 완료",
        }

    except Exception as e:
        logger.error(f"[Shipping] SmartLogen 자동 가져오기 오류: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


# ─── 통계 ────────────────────────────────────
@router.get("/stats")
async def shipping_stats(user: dict = Depends(get_current_user)):
    """택배 발송 통계"""
    import os
    use_pg = bool(os.getenv("DATABASE_URL", ""))

    conn = get_connection()

    total = conn.execute("SELECT COUNT(*) as cnt FROM shipments").fetchone()["cnt"]

    # 창고별 통계
    by_warehouse = conn.execute(
        "SELECT warehouse, COUNT(*) as cnt FROM shipments GROUP BY warehouse"
    ).fetchall()

    # 최근 7일 일별 통계
    if use_pg:
        recent = conn.execute("""
            SELECT take_dt, COUNT(*) as cnt
            FROM shipments
            WHERE take_dt >= to_char(NOW() - INTERVAL '7 days', 'YYYYMMDD')
            GROUP BY take_dt ORDER BY take_dt DESC
        """).fetchall()
    else:
        recent = conn.execute("""
            SELECT take_dt, COUNT(*) as cnt
            FROM shipments
            WHERE take_dt >= strftime('%Y%m%d', 'now', '-7 days')
            GROUP BY take_dt ORDER BY take_dt DESC
        """).fetchall()

    conn.close()
    return {
        "total": total,
        "by_warehouse": [dict(r) for r in by_warehouse],
        "recent_daily": [dict(r) for r in recent],
    }


# ─── 자동 동기화 스케줄러 상태 ───────────────────
@router.get("/scheduler/status")
async def scheduler_status(user: dict = Depends(get_current_user)):
    """자동 동기화 스케줄러 상태 조회"""
    from services.scheduler_service import get_scheduler_status
    return get_scheduler_status()


@router.post("/scheduler/run-now")
async def scheduler_run_now(
    days: int = Query(3, description="조회 기간 (일)"),
    user: dict = Depends(get_current_user),
):
    """자동 동기화 즉시 실행"""
    from services.scheduler_service import run_auto_fetch
    result = await run_auto_fetch(days=days)
    return result


# ─── 삭제 ────────────────────────────────────
@router.delete("/{shipment_id}")
async def delete_shipment(shipment_id: int, user: dict = Depends(get_current_user)):
    conn = get_connection()
    conn.execute("DELETE FROM shipments WHERE id = ?", (shipment_id,))
    conn.commit()
    conn.close()
    return {"success": True}
