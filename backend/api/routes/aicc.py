"""
AICC REST API 라우터
"""
import json
import os
import random
import uuid
import base64
import io
import tempfile
import httpx
import xml.etree.ElementTree as ET
from datetime import date, timedelta
from typing import Dict, Any
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from services.aicc_data_loader import data_loader
from services.aicc_session_manager import session_manager
from services import aicc_db
from security import get_current_user

router = APIRouter()


# ── 리뷰 기반 랜덤 예시 질문 생성 ─────────────────────────
_EXAMPLE_TEMPLATES = [
    "{model} 추천해주세요",
    "{model} 어떤 제품인가요?",
    "{category} 추천 부탁드려요",
    "{category} 어떤 게 좋을까요?",
    "{model} 사용 후기 궁금해요",
    "{model} 호환되는 제품 있나요?",
    "{category} 인기 제품 알려주세요",
]

_review_samples: list[dict] = []  # [{model, category}, ...]


def _init_review_samples():
    """리뷰 데이터에서 예시 질문용 샘플 추출 (서버 시작 시 1회)"""
    global _review_samples
    if _review_samples:
        return
    try:
        # 제품 데이터에서 카테고리 매핑
        samples = []
        for model_name, prod in data_loader.product_data.items():
            cat = prod.get("카테고리", "")
            if cat and model_name.startswith(("LS-", "LSP-", "ZOT-")):
                samples.append({"model": model_name, "category": cat})
        if samples:
            _review_samples = samples
            print(f"[AICC] 예시 질문 샘플 {len(samples)}개 로드 완료")
    except Exception as e:
        print(f"[AICC] 예시 질문 샘플 로드 실패: {e}")


@router.get("/placeholder-examples")
async def get_placeholder_examples(count: int = 5):
    """리뷰/제품 데이터 기반 랜덤 예시 질문 반환"""
    _init_review_samples()
    if not _review_samples:
        return {"examples": ["HDMI 케이블 추천해주세요", "USB 허브 추천 부탁드려요"]}

    examples = []
    picked = random.sample(_review_samples, min(count * 2, len(_review_samples)))
    for s in picked:
        tmpl = random.choice(_EXAMPLE_TEMPLATES)
        ex = tmpl.format(model=s["model"], category=s["category"])
        if ex not in examples:
            examples.append(ex)
        if len(examples) >= count:
            break
    return {"examples": examples}


def _clean_qna_question(raw: str, model: str) -> str:
    """
    고객 QnA 원문을 짧은 질문형 placeholder로 정제.
    - 인사말/감사 제거
    - 첫 번째 실질 질문 문장만 추출
    - 60자 이내로 자르기
    """
    import re as _re

    # 줄바꿈 → 공백
    text = raw.replace("\n", " ").strip()

    # 인사/감사/주문/부탁 등 비질문 패턴 제거
    noise = [
        r"안녕하세요[.\s,]*",
        r"감사합니다[.\s!~]*",
        r"네\s+감사합니다[.\s!~]*",
        r"부탁\s*드립니다[.\s!~]*",
        r"꼭\s*좀\s*부탁[^\s]*",
        r"랜스타\s*(입니다)?[.\s,]*",
        r"주문\s*했습니다[.\s,]*",
        r"어제\s*저녁에[^\s]*",
        r"행사\s*때문에[^,]*,?",
        r"주중으로는[^.]*\.",
    ]
    for pat in noise:
        text = _re.sub(pat, " ", text, flags=_re.IGNORECASE)

    text = _re.sub(r'\s+', ' ', text).strip()
    if not text:
        return ""

    # 문장 분리: 물음표, 마침표, 쉼표+공백 기준
    sentences = _re.split(r'(?<=[?.!])\s+|(?<=요)\s+|(?<=지)\s+|,\s+', text)
    question_sent = ""
    for sent in sentences:
        sent = sent.strip()
        if len(sent) < 5:
            continue
        # 질문스러운 문장 우선
        if _re.search(r'[?]|인가요|까요|나요|는지|런가요|을까요|ㄹ까요|궁금|알\s*수|가능한|되나요|호환|사용|연결|설치|차이|어떻게|몇|무게|크기|규격|지원|방법', sent):
            question_sent = sent
            break
    if not question_sent:
        # 질문 마커 없으면 첫 번째 의미있는 문장
        for sent in sentences:
            sent = sent.strip()
            if len(sent) >= 8:
                question_sent = sent
                break
        if not question_sent:
            question_sent = text

    # 55자 초과 시 자르기
    if len(question_sent) > 55:
        # 물음표/~요/~지 기준 앞부분만
        cut = _re.search(r'^(.{15,50}[?요지])', question_sent)
        if cut:
            question_sent = cut.group(1)
        else:
            question_sent = question_sent[:52] + "..."

    # 끝에 물음표 없으면 추가 (질문형이어야 하므로)
    if question_sent and not question_sent.endswith("?") and not question_sent.endswith("요") and not question_sent.endswith("요."):
        if _re.search(r'인가|까요|나요|는지|런가|을까|ㄹ까|궁금|알\s*수|가능한|되나|호환|차이|어떻게|몇', question_sent):
            question_sent = question_sent.rstrip(".!~ ") + "?"

    return question_sent.strip()


@router.get("/placeholder-examples/tech")
async def get_tech_placeholder_examples(model: str = "", count: int = 5):
    """
    기술문의용: 모델명 기반 QnA 예시 질문 반환.
    technical_qna + product_data(QnA) 에서 실제 질문을 추출·정제.
    """
    if not model or len(model.strip()) < 3:
        return {"examples": [], "model": model}

    model = model.strip()
    model_upper = model.upper()
    # 베이스 모델명 (길이 접미사 제거)
    base_model = data_loader._extract_model_base(model).upper()

    raw_questions = []

    # 1. technical_qna에서 검색 (정확 매칭 + 베이스 모델 매칭)
    for product in data_loader.technical_qna:
        pm = product.get("model", "")
        pm_upper = pm.upper()
        pm_base = data_loader._extract_model_base(pm).upper()

        if pm_upper == model_upper or pm_base == base_model:
            for qna in product.get("qna", []):
                q = qna.get("question", "").strip()
                if q and len(q) > 10:
                    raw_questions.append(q)

    # 2. product_data(01_제품별_통합데이터.json) QnA 검색
    for pmodel, pdata in data_loader.product_data.items():
        pm_upper = pmodel.upper()
        pm_base = data_loader._extract_model_base(pmodel).upper()

        if pm_upper == model_upper or pm_base == base_model:
            for qna in pdata.get("QnA", []):
                q = qna.get("문의", "").strip()
                if q and len(q) > 10:
                    raw_questions.append(q)

    if not raw_questions:
        # 카테고리 기반 폴백: 같은 카테고리 제품의 질문 사용
        target_cat = ""
        prod = data_loader.product_data.get(model, {})
        if prod:
            target_cat = prod.get("카테고리", "")
        if not target_cat:
            for product in data_loader.technical_qna:
                if product.get("model", "").upper() == model_upper:
                    target_cat = product.get("category", "")
                    break

        if target_cat:
            for product in data_loader.technical_qna:
                if product.get("category", "") == target_cat:
                    for qna in product.get("qna", []):
                        q = qna.get("question", "").strip()
                        if q and len(q) > 10:
                            raw_questions.append(q)
                    if len(raw_questions) >= 20:
                        break

    if not raw_questions:
        return {"examples": [], "model": model}

    # 정제
    cleaned = []
    seen = set()
    # 비기술 질문 필터 (배송/주문/교환/환불 관련은 제외)
    _skip_patterns = re.compile(r"배송|택배|주문|교환|환불|반품|쿠폰|적립|영수증|송장|결제|입금|계좌|카드")
    random.shuffle(raw_questions)
    for raw in raw_questions:
        q = _clean_qna_question(raw, model)
        if q and len(q) >= 8 and q not in seen and not _skip_patterns.search(q):
            seen.add(q)
            cleaned.append(q)
        if len(cleaned) >= count:
            break

    return {"examples": cleaned, "model": model}


class AdminMessageBody(BaseModel):
    content: str


class ImageUploadBody(BaseModel):
    session_id: str
    image: str          # data:image/jpeg;base64,...
    file_name: str = "image.jpg"


ALLOWED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/gif", "image/webp"}


@router.post("/upload")
async def upload_image(body: ImageUploadBody):
    """고객 이미지 업로드 — base64로 세션에 저장"""
    # 세션 확인
    s = None
    for sess in session_manager.sessions.values():
        if body.session_id in (sess.get("session_id", ""),):
            s = sess
            break
    # session_id가 프론트에서 보낸 client-side ID일 수 있으므로 유연하게 처리
    if not s:
        # 모든 세션을 확인하여 customer_ws가 연결된 세션 찾기
        for sess in session_manager.sessions.values():
            if sess.get("status") == "active":
                s = sess
                break
    if not s:
        return {"ok": False, "message": "세션을 찾을 수 없습니다."}

    # data URL 파싱: data:image/jpeg;base64,/9j/4AAQ...
    if not body.image.startswith("data:image/"):
        return {"ok": False, "message": "잘못된 이미지 형식입니다."}

    try:
        header, b64_data = body.image.split(",", 1)
        media_type = header.split(":")[1].split(";")[0]  # image/jpeg
    except (ValueError, IndexError):
        return {"ok": False, "message": "이미지 데이터 파싱 실패"}

    if media_type not in ALLOWED_IMAGE_TYPES:
        return {"ok": False, "message": f"지원하지 않는 이미지 형식: {media_type}"}

    # 크기 확인 (base64 디코딩 전 대략 체크: base64는 원본의 ~1.33배)
    if len(b64_data) > 7 * 1024 * 1024:  # ~5MB 원본
        return {"ok": False, "message": "이미지 크기가 너무 큽니다. (최대 5MB)"}

    image_id = str(uuid.uuid4())[:8]
    s["images"][image_id] = {
        "media_type": media_type,
        "base64_data": b64_data,
        "file_name": body.file_name,
    }

    return {"ok": True, "image_id": image_id, "session_id": s["session_id"]}


@router.get("/models")
async def get_models(q: str = ""):
    """드롭다운 자동완성용 모델 목록"""
    if q and len(q.strip()) >= 2:
        return data_loader.search_models(q.strip(), limit=15)
    # q 없으면 전체 반환 (초기 로드용)
    return data_loader.dropdown_models


@router.get("/sessions")
async def get_sessions(menu: str = "", current_user=Depends(get_current_user)):
    """관리자: 세션 목록 (인메모리 + DB 병합, 재배포 후에도 유지)"""
    # 1. 인메모리 세션
    memory_sessions = session_manager.all_serialized()
    memory_ids = {s["session_id"] for s in memory_sessions}

    # 2. DB 세션 (인메모리에 없는 것만 추가)
    db_sessions = aicc_db.get_all_sessions(limit=200)
    for ds in db_sessions:
        if ds["id"] not in memory_ids:
            memory_sessions.append({
                "session_id": ds["id"],
                "customer_name": ds.get("customer_name", ""),
                "selected_model": ds.get("selected_model", ""),
                "erp_code": ds.get("erp_code", ""),
                "selected_menu": ds.get("selected_menu", ""),
                "status": ds.get("status", "closed"),
                "is_admin_intervened": False,
                "messages": [],
                "created_at": ds.get("created_at", ""),
                "updated_at": ds.get("updated_at", ""),
                "from_db": True,
            })

    # 3. 메뉴 필터
    if menu:
        memory_sessions = [s for s in memory_sessions if s.get("selected_menu", "") == menu]

    # 최신순 정렬
    memory_sessions.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return {"sessions": memory_sessions}


@router.get("/sessions/{session_id}")
async def get_session(session_id: str, current_user=Depends(get_current_user)):
    """관리자: 세션 상세 (DB 메시지 우선, 인메모리 보완)"""
    s = session_manager.get(session_id)

    # DB에서 메시지 조회 (영구 저장된 완전한 기록)
    try:
        db_messages = aicc_db.get_session_messages(session_id)
    except Exception as e:
        print(f"[AICC] DB 메시지 조회 오류: {e}")
        db_messages = []

    print(f"[AICC] 세션 조회: {session_id[:8]}… 인메모리={'있음' if s else '없음'}, DB메시지={len(db_messages)}건")

    if s:
        result = session_manager.serialize(s)
        # DB 메시지가 있으면 DB 우선 (더 완전한 기록)
        if db_messages:
            result["messages"] = [
                {"role": m["role"], "content": m["content"], "timestamp": m.get("created_at", "")}
                for m in db_messages
            ]
        elif not result.get("messages"):
            # DB에도 인메모리에도 메시지 없으면 인메모리 대화이력에서 추출
            mem_msgs = []
            for msg in s.get("messages", []):
                if isinstance(msg, dict):
                    role = msg.get("role", "user")
                    content = msg.get("content", "")
                    if isinstance(content, list):
                        # Claude API 형식: [{"type":"text","text":"..."}]
                        content = " ".join(
                            c.get("text", "") for c in content if isinstance(c, dict) and c.get("type") == "text"
                        )
                    if content:
                        mem_msgs.append({"role": role, "content": content, "timestamp": ""})
            result["messages"] = mem_msgs
        return result

    # 인메모리에 없으면 DB에서 세션 정보도 조회
    db_sessions = aicc_db.get_all_sessions(limit=500)
    db_session = next((ds for ds in db_sessions if ds["id"] == session_id), None)
    if not db_session:
        raise HTTPException(404, "세션 없음")

    return {
        "session_id": db_session["id"],
        "customer_name": db_session.get("customer_name", ""),
        "selected_model": db_session.get("selected_model", ""),
        "erp_code": db_session.get("erp_code", ""),
        "selected_menu": db_session.get("selected_menu", ""),
        "status": db_session.get("status", "closed"),
        "is_admin_intervened": False,
        "messages": [{"role": m["role"], "content": m["content"], "timestamp": m.get("created_at", "")} for m in db_messages],
        "created_at": db_session.get("created_at", ""),
        "updated_at": db_session.get("updated_at", ""),
        "from_db": True,
    }


@router.post("/sessions/{session_id}/intervene")
async def intervene(session_id: str, current_user=Depends(get_current_user)):
    """관리자 개입"""
    session_manager.intervene(session_id)
    await session_manager.send_customer(session_id, {
        "type": "admin_joined",
        "content": "담당자가 연결되었습니다. 직접 안내해 드리겠습니다."
    })
    return {"ok": True}


@router.post("/sessions/{session_id}/close")
async def close_session(session_id: str, current_user=Depends(get_current_user)):
    """세션 종료"""
    session_manager.close(session_id)
    await session_manager.send_customer(session_id, {
        "type": "session_closed",
        "content": "상담이 종료되었습니다. 감사합니다."
    })
    return {"ok": True}


@router.post("/sessions/{session_id}/admin-message")
async def admin_message(
    session_id: str,
    body: AdminMessageBody,
    current_user=Depends(get_current_user)
):
    """관리자 메시지 (WebSocket 미연결 시 REST 폴백)"""
    session_manager.add_message(session_id, "admin", body.content)
    await session_manager.send_customer(session_id, {
        "type": "admin_message",
        "content": body.content
    })
    return {"ok": True}


@router.get("/inventory/{model_name}")
async def get_inventory(model_name: str):
    """ERP 재고조회 — check_inventory() 호출 후 창고별 재고 반환"""
    erp_code = data_loader.get_erp_code(model_name)
    if not erp_code:
        return {"ok": False, "message": "해당 제품의 ERP 코드를 찾을 수 없습니다."}
    try:
        from services.erp_client import ERPClient
        erp = ERPClient()
        # check_inventory: 전체 창고 조회 (wh_cd="" → 모든 창고)
        result = await erp.check_inventory(prod_cd=erp_code, wh_cd="")
        if not result.get("success"):
            return {"ok": False, "message": result.get("error", "재고 조회 실패")}

        # 창고코드로 용산·김포 분류 (용산=10, 김포=30)
        yongsan = 0
        gimpo = 0
        other = 0
        for item in result.get("data", []):
            qty = int(float(item.get("qty", 0)))
            wh_cd = str(item.get("wh_cd", "")).strip()
            if wh_cd == "10":
                yongsan += qty
            elif wh_cd == "30":
                gimpo += qty
            else:
                other += qty

        total = yongsan + gimpo + other
        return {
            "ok": True,
            "model_name": model_name,
            "erp_code": erp_code,
            "yongsan": yongsan,
            "gimpo": gimpo,
            "total": total,
        }
    except Exception as e:
        import traceback
        print(f"[AICC Inventory] 오류: {e}\n{traceback.format_exc()}")
        return {"ok": False, "message": "재고 조회 중 오류가 발생했습니다. 전화(02-717-3386) 문의 바랍니다."}


# ── 주문상태코드 → 한글 ──────────────────────────────────────

ORDER_STATUS = {
    "o1": "입금대기",  "o2": "결제완료",  "o3": "상품준비중",
    "g1": "구매발주",  "g2": "구매발주",  "g3": "상품입고",
    "d1": "배송중",    "d2": "배송완료",  "p1": "구매확정",
    "r1": "환불접수",  "r2": "환불완료",
    "b1": "반품접수",  "b2": "반품완료",
    "e1": "교환접수",  "e2": "교환완료",
    "c1": "취소요청",  "c2": "취소완료",
}


def _xml_text(el, tag: str) -> str:
    """XML 엘리먼트에서 태그 텍스트 추출 (None-safe)"""
    child = el.find(tag)
    return (child.text or "").strip() if child is not None else ""


@router.get("/orders")
async def get_customer_orders(memNo: str = "", phone: str = ""):
    """
    고도몰 Open API 주문조회
    - memNo(회원번호) 또는 phone(휴대폰번호)으로 필터링
    - API 제한: 최대 30일 단위 조회
    - 응답: XML → 파싱 후 JSON 반환
    """
    if not memNo and (not phone or len(phone.replace("-", "").strip()) < 9):
        return {"orders": [], "message": "회원 정보가 필요합니다. 로그인 후 이용해 주세요."}

    partner_key = os.getenv("GODOMALL_PARTNER_KEY", os.getenv("GODOMALL_MALL_ID", ""))
    user_key = os.getenv("GODOMALL_USER_KEY", os.getenv("GODOMALL_API_KEY", ""))

    end_date = date.today().strftime("%Y-%m-%d")
    start_date = (date.today() - timedelta(days=30)).strftime("%Y-%m-%d")

    payload = {
        "partner_key": partner_key,
        "key": user_key,
        "dateType": "order",
        "startDate": start_date,
        "endDate": end_date,
    }

    try:
        async with httpx.AsyncClient(timeout=20.0) as http_client:
            res = await http_client.post(
                "https://openhub.godo.co.kr/godomall5/order/Order_Search.php",
                data=payload,
            )
        root = ET.fromstring(res.text)
    except Exception as e:
        print(f"[AICC Orders] API 오류: {e}")
        return {"orders": [], "message": f"주문 조회 중 오류가 발생했습니다."}

    code = _xml_text(root, ".//code")
    if code != "000":
        msg = _xml_text(root, ".//msg")
        return {"orders": [], "message": f"조회 실패: {msg}"}

    # ── 전체 주문 데이터 파싱 + memNo/phone 필터링 ──
    phone_clean = phone.replace("-", "").strip() if phone else ""
    all_order_els = root.findall(".//order_data")
    orders = []

    for order_el in all_order_els:
        # memNo 필터
        if memNo:
            order_memNo = _xml_text(order_el, "memNo")
            if order_memNo != memNo:
                continue

        # phone 필터 (memNo 없을 때)
        if not memNo and phone_clean:
            info_el = order_el.find("orderInfoData")
            if info_el is not None:
                order_phone = _xml_text(info_el, "orderCellPhone").replace("-", "")
                receiver_phone = _xml_text(info_el, "receiverCellPhone").replace("-", "")
                if phone_clean not in (order_phone, receiver_phone):
                    continue
            else:
                continue

        order_no = _xml_text(order_el, "orderNo")
        order_date = _xml_text(order_el, "orderDate")[:10]
        settle_price = _xml_text(order_el, "settlePrice")
        order_status_top = _xml_text(order_el, "orderStatus")

        # orderGoodsData 파싱
        goods_els = order_el.findall("orderGoodsData")
        goods_items = []
        primary_status = ""
        primary_status_text = ""

        for g in goods_els:
            inv_no = _xml_text(g, "invoiceNo")
            status = _xml_text(g, "orderStatus")
            status_text = ORDER_STATUS.get(status, status)
            if not primary_status:
                primary_status = status
                primary_status_text = status_text
            goods_items.append({
                "goods_name": _xml_text(g, "goodsNm"),
                "goods_image": _xml_text(g, "listImageData"),
                "order_status": status,
                "order_status_text": status_text,
                "invoice_company": _xml_text(g, "invoiceCompany"),
                "invoice_no": inv_no,
                "delivery_dt": _xml_text(g, "deliveryDt")[:10],
                "delivery_complete_dt": _xml_text(g, "deliveryCompleteDt")[:10],
                "tracking_url": f"https://www.ilogen.com/m/personal/trace/{inv_no}" if inv_no else "",
            })

        if not primary_status:
            primary_status = order_status_top
            primary_status_text = ORDER_STATUS.get(order_status_top, order_status_top)

        first_name = goods_items[0]["goods_name"] if goods_items else _xml_text(order_el, "orderGoodsNm")
        summary = first_name if len(goods_items) <= 1 else f"{first_name} 외 {len(goods_items)-1}건"

        orders.append({
            "order_no": order_no,
            "order_date": order_date,
            "settle_price": settle_price,
            "goods_summary": summary,
            "order_status": primary_status,
            "order_status_text": primary_status_text,
            "goods": goods_items,
        })

        if len(orders) >= 10:
            break

    orders.sort(key=lambda x: x["order_date"], reverse=True)
    return {"orders": orders, "total": len(orders)}


# ── 제품 지식 DB 관리 API ────────────────────────────────

class ProductKnowledgeBody(BaseModel):
    model_name: str
    category: str = ""
    data: Dict[str, Any]


class BulkProductKnowledgeBody(BaseModel):
    products: Dict[str, Dict[str, Any]]


@router.post("/knowledge")
async def upsert_knowledge(
    body: ProductKnowledgeBody,
    current_user=Depends(get_current_user)
):
    """제품 지식 단건 등록/수정"""
    aicc_db.upsert_product_knowledge(body.model_name, body.category, body.data)
    return {"ok": True, "model_name": body.model_name}


@router.post("/knowledge/bulk")
async def bulk_upload_knowledge(
    body: BulkProductKnowledgeBody,
    current_user=Depends(get_current_user),
):
    """
    제품 지식 일괄 등록 — JSON body
    형식: { "products": { "LS-ANDOOR-S": { "카테고리": "도어락", ... }, ... } }
    """
    count = aicc_db.bulk_upsert_product_knowledge(body.products)
    return {"ok": True, "count": count, "message": f"{count}개 제품 지식 등록 완료"}


@router.post("/knowledge/upload")
async def upload_knowledge_file(
    file: UploadFile = File(...),
    current_user=Depends(get_current_user),
):
    """
    제품 지식 JSON 파일 업로드
    파일 형식: { "LS-ANDOOR-S": { "카테고리": "도어락", ... }, ... }
    """
    content = await file.read()
    try:
        products = json.loads(content.decode("utf-8"))
    except Exception:
        raise HTTPException(400, "JSON 파일 파싱 실패")

    if not isinstance(products, dict):
        raise HTTPException(400, "JSON 최상위는 객체여야 합니다 (모델명: {데이터})")

    count = aicc_db.bulk_upsert_product_knowledge(products)
    return {"ok": True, "count": count, "message": f"{count}개 제품 지식 등록 완료"}


@router.get("/knowledge")
async def list_knowledge(current_user=Depends(get_current_user)):
    """전체 제품 지식 목록"""
    items = aicc_db.get_all_product_knowledge()
    return {"items": items, "total": len(items)}


@router.get("/knowledge-count")
async def knowledge_count():
    """제품 지식 DB 총 개수 (인증 불필요, 디버깅용)"""
    items = aicc_db.get_all_product_knowledge()
    return {"total": len(items)}


@router.get("/knowledge/{model_name}")
async def get_knowledge(model_name: str, current_user=Depends(get_current_user)):
    """특정 제품 지식 상세"""
    item = aicc_db.get_product_knowledge(model_name)
    if not item:
        raise HTTPException(404, "해당 제품 지식이 없습니다")
    return item


@router.get("/knowledge-check/{model_name}")
async def check_knowledge(model_name: str):
    """제품 지식 존재 여부 확인 (인증 불필요)"""
    item = aicc_db.get_product_knowledge(model_name)
    if not item:
        return {"exists": False, "model_name": model_name}
    return {"exists": True, "model_name": model_name, "keys": list(item.get("data", {}).keys())}


@router.delete("/knowledge/{model_name}")
async def delete_knowledge(model_name: str, current_user=Depends(get_current_user)):
    """제품 지식 삭제"""
    aicc_db.delete_product_knowledge(model_name)
    return {"ok": True}


# ── 채팅 이력 조회 API ───────────────────────────────────

@router.get("/history/sessions")
async def get_chat_history_sessions(current_user=Depends(get_current_user)):
    """저장된 전체 AICC 채팅 세션 목록"""
    sessions = aicc_db.get_all_sessions(limit=200)
    return {"sessions": sessions, "total": len(sessions)}


@router.get("/history/sessions/{session_id}")
async def get_chat_history_messages(
    session_id: str,
    current_user=Depends(get_current_user)
):
    """특정 세션의 전체 채팅 메시지"""
    messages = aicc_db.get_session_messages(session_id)
    return {"session_id": session_id, "messages": messages, "total": len(messages)}


# ── 미답변 관리 API ───────────────────────────────────

@router.get("/unanswered")
async def get_unanswered_list(
    resolved: bool = False,
    current_user=Depends(get_current_user)
):
    """미답변 목록 조회"""
    items = aicc_db.get_unanswered(resolved=resolved)
    return {"items": items, "total": len(items)}


@router.get("/unanswered/count")
async def get_unanswered_count(current_user=Depends(get_current_user)):
    """미해결 미답변 수"""
    return {"count": aicc_db.count_unanswered()}


@router.post("/unanswered/{item_id}/resolve")
async def resolve_unanswered(
    item_id: int,
    current_user=Depends(get_current_user)
):
    """미답변 해결 처리"""
    aicc_db.resolve_unanswered(item_id)
    return {"ok": True}


class AddKnowledgeFromUnanswered(BaseModel):
    model_name: str
    key: str       # 예: "손잡이_들뜸_해결"
    value: str     # 예: "사각봉을 35mm용으로 교체하세요"


@router.post("/unanswered/{item_id}/add-knowledge")
async def add_knowledge_from_unanswered(
    item_id: int,
    body: AddKnowledgeFromUnanswered,
    current_user=Depends(get_current_user)
):
    """미답변에서 직접 제품 지식 DB에 항목 추가"""
    existing = aicc_db.get_product_knowledge(body.model_name)
    if existing:
        data = existing["data"]
    else:
        data = {"카테고리": ""}

    # 기존 데이터에 새 키 추가
    if body.key in data and isinstance(data[body.key], list):
        data[body.key].append(body.value)
    elif body.key in data and isinstance(data[body.key], str):
        data[body.key] = [data[body.key], body.value]
    else:
        data[body.key] = body.value

    aicc_db.upsert_product_knowledge(body.model_name, data.get("카테고리", ""), data)
    aicc_db.resolve_unanswered(item_id, admin_note=f"DB 추가: {body.key}")
    return {"ok": True, "model_name": body.model_name, "key": body.key}


class DirectKnowledgeAdd(BaseModel):
    model_name: str
    key: str       # 예: "FAQ", "비밀번호등록"
    value: str     # 단일 텍스트 또는 줄바꿈 구분 리스트


@router.post("/knowledge/add-direct")
async def add_knowledge_direct(
    body: DirectKnowledgeAdd,
    current_user=Depends(get_current_user)
):
    """관리자가 직접 제품 지식 DB에 항목 추가 (미답변 연동 없음)"""
    # 모델명이 실제 존재하는지 확인
    found = data_loader.search_models(body.model_name, limit=1)
    if not found or found[0]["model_name"] != body.model_name:
        # 정확히 일치하지 않아도 DB에는 추가 가능 (이미 등록된 제품일 수 있음)
        pass

    existing = aicc_db.get_product_knowledge(body.model_name)
    if existing:
        data = existing["data"]
        category = existing.get("category", data.get("카테고리", ""))
    else:
        data = {"카테고리": ""}
        category = ""

    # 줄바꿈이 있으면 리스트로 변환
    lines = [l.strip() for l in body.value.split("\n") if l.strip()]
    new_value = lines if len(lines) > 1 else body.value.strip()

    # 기존 데이터에 추가/병합
    if body.key in data and isinstance(data[body.key], list):
        if isinstance(new_value, list):
            data[body.key].extend(new_value)
        else:
            data[body.key].append(new_value)
    elif body.key in data and isinstance(data[body.key], str) and data[body.key]:
        if isinstance(new_value, list):
            data[body.key] = [data[body.key]] + new_value
        else:
            data[body.key] = [data[body.key], new_value]
    else:
        data[body.key] = new_value

    aicc_db.upsert_product_knowledge(body.model_name, category, data)
    return {"ok": True, "model_name": body.model_name, "key": body.key}


# ── 상품 엑셀 다운로드 ────────────────────────────────────────
@router.get("/goods-excel")
async def download_goods_excel(
    level: int = Query(0, description="회원등급 (0=일반, 1=LV1, 2=LV2사업자, 3=LV3업체)"),
    category: str = Query("", description="카테고리 코드 (빈값=전체)"),
):
    """
    고도몰 DB에서 상품 데이터를 조회하여 엑셀 파일로 반환.
    - SSH → PHP → MySQL(localhost) 경로로 조회 (SELECT only)
    - 회원등급에 따라 가격 컬럼 필터링
    - LV2(사업자) 이상만 호출 가능
    """
    if level < 2:
        raise HTTPException(status_code=403, detail="엑셀 다운로드는 사업자회원(LV2) 이상만 가능합니다.")

    import paramiko

    # ── PHP 쿼리 생성 (SELECT only) ──
    where_clause = "open=1"
    if category:
        where_clause += f" AND goodsno IN (SELECT goodsno FROM gd_goods_link WHERE category LIKE '{category}%')"

    php_code = f"""<?php
$conn = new mysqli("localhost", "lanmartDq", "Sbxxe97EB^L", "lanmart_godo_co_kr");
if ($conn->connect_error) {{ die(json_encode(["error" => $conn->connect_error])); }}
$conn->set_charset("utf8");

// 카테고리 조회용
function getCateName($conn, $goodsno, $len) {{
    $res = $conn->query("SELECT catnm FROM gd_category WHERE category IN (SELECT category FROM gd_goods_link WHERE goodsno='$goodsno' AND LEFT(category,3)!='013' AND LENGTH(category)=$len) LIMIT 1");
    if ($res && $row = $res->fetch_assoc()) return $row['catnm'];
    return '';
}}

$res = $conn->query("SELECT goodsno, goodscd, model_name, goodsnm, maker, origin, totstock, addstock, goods_consumer, goods_prices0, goods_prices1, goods_prices2, goods_prices3, img_l, img_m, img_s, img_i, longdesc, shortdesc, launchdt, runout FROM gd_goods WHERE {where_clause} ORDER BY goodsnm");
if (!$res) {{ die(json_encode(["error" => $conn->error])); }}

$rows = [];
while ($data = $res->fetch_assoc()) {{
    $data['cate1'] = getCateName($conn, $data['goodsno'], 3);
    $data['cate2'] = getCateName($conn, $data['goodsno'], 6);
    $data['cate3'] = getCateName($conn, $data['goodsno'], 9);
    $data['cate4'] = getCateName($conn, $data['goodsno'], 12);
    // 이미지 URL 처리
    $imgs = explode('|', $data['img_l']);
    $data['img_l'] = isset($imgs[0]) ? $imgs[0] : '';
    $rows[] = $data;
}}
echo json_encode($rows);
$conn->close();
?>"""

    try:
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(
            '182.162.22.143', port=22,
            username='lanmart', password='LINEUP5303**',
            timeout=15
        )

        sftp = ssh.open_sftp()
        remote_path = '/tmp/_goods_export.php'
        with sftp.open(remote_path, 'w') as f:
            f.write(php_code)
        sftp.close()

        stdin, stdout, stderr = ssh.exec_command(
            f'php {remote_path} && rm {remote_path}',
            timeout=60
        )
        stdout.channel.settimeout(60)
        raw = stdout.read().decode('utf-8', errors='replace')
        ssh.close()

        goods_list = json.loads(raw)
        if isinstance(goods_list, dict) and "error" in goods_list:
            raise HTTPException(status_code=500, detail=f"DB 오류: {goods_list['error']}")

    except paramiko.SSHException as e:
        raise HTTPException(status_code=500, detail=f"서버 연결 실패: {str(e)}")
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="데이터 파싱 실패")

    # ── 엑셀 생성 ──
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "상품목록"

    # 헤더 정의 (등급별 가격 컬럼 필터링)
    headers = [
        ("상품코드", "goodsno", 12),
        ("모델명(상품코드)", "goodscd", 18),
        ("1차카테고리", "cate1", 14),
        ("2차카테고리", "cate2", 14),
        ("3차카테고리", "cate3", 14),
        ("4차카테고리", "cate4", 14),
        ("품목코드(모델명)", "model_name", 20),
        ("상품명", "goodsnm", 40),
        ("제조사", "maker", 16),
        ("원산지", "origin", 10),
        ("총재고", "_totalstock", 8),
        ("용산창고", "totstock", 8),
        ("김포창고", "addstock", 8),
        ("소비자가", "goods_prices0", 12),
    ]

    # 등급별 가격 추가
    if level >= 1:
        headers.append(("오픈마켓 등록가", "goods_prices1", 16))
    if level >= 2:
        headers.append(("온라인 노출가", "goods_prices2", 16))
    if level >= 3:
        headers.append(("딜러가", "goods_prices3", 12))

    headers += [
        ("상품이미지", "img_l", 40),
        ("짧은설명", "shortdesc", 30),
        ("등록일", "launchdt", 12),
        ("품절여부", "_runout", 8),
    ]

    # 스타일
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="1a1a2e")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )
    blue_font = Font(color="0000FF", size=10)
    red_font = Font(color="FF0000", size=10)
    price_fmt = '#,##0'

    # 헤더 행
    for col_idx, (label, key, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    # 데이터 행
    for row_idx, item in enumerate(goods_list, 2):
        for col_idx, (label, key, width) in enumerate(headers, 1):
            if key == "_totalstock":
                tot = int(item.get("totstock") or 0) + int(item.get("addstock") or 0)
                value = tot if tot > 0 else "품절"
            elif key == "_runout":
                value = "품절" if item.get("runout") == "1" else "판매중"
            elif key in ("goods_prices0", "goods_prices1", "goods_prices2", "goods_prices3", "goods_consumer"):
                value = int(float(item.get(key) or 0))
            elif key in ("totstock", "addstock"):
                raw_val = int(item.get(key) or 0)
                value = raw_val if raw_val > 0 else "품절"
            else:
                value = item.get(key, "")

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            # 가격 컬럼 서식
            if key in ("goods_prices0", "goods_prices1", "goods_prices2", "goods_prices3") and isinstance(value, (int, float)):
                cell.number_format = price_fmt
            # 용산재고/오픈마켓 등록가 파란색
            if key in ("totstock", "goods_prices1"):
                cell.font = blue_font
            # 김포재고/온라인 노출가 빨간색
            if key in ("addstock", "goods_prices2"):
                cell.font = red_font

    # 엑셀 파일을 메모리에 저장
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    today = date.today().strftime("%Y%m%d")
    filename = f"lanstar_goods_{today}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
