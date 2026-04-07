"""
발주서 처리 강화 (Few-Shot Learning) 서비스

핵심 개념:
- 원본 발주서(raw PO)와 정확한 판매전표(Excel)를 매칭하여 학습 데이터로 저장
- 새 발주서 처리 시, 같은 거래처의 과거 매칭 데이터를 few-shot 예시로 활용
- 품목코드(item_code)가 핵심: ERP 전송 시 품목코드로 정확한 품목명/모델명 자동 매칭

흐름:
1. 사용자가 원본 발주서 텍스트 + 매칭된 판매전표 엑셀을 업로드
2. 엑셀 파싱 → JSON 구조화 → DB 저장
3. 새 발주서 처리 시 → 같은 거래처의 과거 매칭 조회 → few-shot prompt 생성
"""
import json
import logging
import io
from typing import List, Optional
from pathlib import Path
from datetime import datetime
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))
from db.database import get_connection, column_exists

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────
#  DB 테이블 마이그레이션
# ─────────────────────────────────────────
def _ensure_training_tables():
    """학습 데이터 테이블 생성 (없으면)"""
    conn = get_connection()
    conn.executescript("""
        -- 발주서-판매전표 매칭 쌍 (학습 세트)
        CREATE TABLE IF NOT EXISTS po_training_pairs (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            cust_code        TEXT NOT NULL,
            cust_name        TEXT NOT NULL,
            raw_po_text      TEXT DEFAULT '',
            raw_po_image     BLOB DEFAULT NULL,
            raw_po_image_type TEXT DEFAULT '',
            order_id         TEXT DEFAULT '',
            memo             TEXT DEFAULT '',
            created_at       TEXT DEFAULT (datetime('now','localtime'))
        );

        -- 매칭된 판매전표 개별 항목 (품목코드가 핵심)
        CREATE TABLE IF NOT EXISTS po_training_items (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            pair_id        INTEGER NOT NULL,
            raw_line_text  TEXT DEFAULT '',
            item_code      TEXT NOT NULL,
            product_name   TEXT DEFAULT '',
            model_name     TEXT DEFAULT '',
            spec           TEXT DEFAULT '',
            qty            REAL DEFAULT 0,
            unit           TEXT DEFAULT 'EA',
            unit_price     REAL DEFAULT 0,
            supply_price   REAL DEFAULT 0,
            tax            REAL DEFAULT 0,
            FOREIGN KEY (pair_id) REFERENCES po_training_pairs(id) ON DELETE CASCADE
        );

        CREATE INDEX IF NOT EXISTS idx_training_pairs_cust
            ON po_training_pairs(cust_code);
        CREATE INDEX IF NOT EXISTS idx_training_items_pair
            ON po_training_items(pair_id);
        CREATE INDEX IF NOT EXISTS idx_training_items_code
            ON po_training_items(item_code);
    """)

    # 기존 테이블에 이미지 컬럼이 없으면 추가 (마이그레이션, SQLite/PG 모두 지원)
    from db.database import safe_add_column
    safe_add_column(conn, 'po_training_pairs', 'raw_po_image', "BLOB DEFAULT NULL")
    safe_add_column(conn, 'po_training_pairs', 'raw_po_image_type', "TEXT DEFAULT ''")

    # ── 대량 학습 테이블 ──
    conn.executescript("""
        -- 대량 학습 세션
        CREATE TABLE IF NOT EXISTS bulk_training_sessions (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id  TEXT UNIQUE NOT NULL,
            cust_code   TEXT NOT NULL,
            cust_name   TEXT NOT NULL,
            excel_data  TEXT DEFAULT '',
            status      TEXT DEFAULT 'uploading',
            created_at  TEXT DEFAULT (datetime('now','localtime')),
            updated_at  TEXT DEFAULT (datetime('now','localtime'))
        );

        -- 발주서별 AI 추출 결과
        CREATE TABLE IF NOT EXISTS bulk_training_extractions (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id      TEXT NOT NULL,
            po_filename     TEXT NOT NULL,
            po_image        BLOB DEFAULT NULL,
            po_image_type   TEXT DEFAULT '',
            order_date      TEXT DEFAULT '',
            extracted_lines TEXT DEFAULT '[]',
            raw_text        TEXT DEFAULT '',
            status          TEXT DEFAULT 'pending',
            created_at      TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE INDEX IF NOT EXISTS idx_bulk_ext_session
            ON bulk_training_extractions(session_id);
    """)

    conn.close()


# 모듈 로드 시 테이블 보장
_ensure_training_tables()


# ─────────────────────────────────────────
#  엑셀 파싱 (판매전표)
# ─────────────────────────────────────────
def parse_sales_slip_excel(file_bytes: bytes, filename: str = "") -> dict:
    """
    ECOUNT 판매전표 엑셀을 파싱하여 구조화된 데이터 반환

    Returns: {
        "vendor": "거래처명",
        "items": [
            {
                "date": "01/06",
                "item_code": "LS-HPSC-06040L",
                "product_name": "고급형 스피드 CAT.6 UTP...",
                "model_name": "LS-HPSC-06040L",
                "spec": "규격...",
                "qty": 10,
                "unit": "EA",
                "unit_price": 5500,
                "supply_price": 55000,
                "tax": 5500,
            }
        ],
        "total_items": 10
    }
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # 헤더 행 탐색: '품목코드' 또는 '품명' 포함 행 찾기
    header_row = None
    header_map = {}

    # 헤더 키워드 매핑 (엑셀 컬럼명 → 내부 필드명)
    HEADER_KEYWORDS = {
        "item_code": ["품목코드", "품목 코드", "PROD_CD", "상품코드"],
        "product_name": ["품명", "품명 및 규격", "품목명", "상품명", "제품명"],
        "model_name": ["모델명", "모델", "MODEL"],
        "spec": ["규격", "사양", "SPEC"],
        "qty": ["수량", "QTY", "수 량"],
        "unit": ["단위", "UNIT"],
        "unit_price": ["단가", "매입단가", "출고단가"],
        "supply_price": ["공급가액", "공급가", "금액"],
        "tax": ["세액", "부가세", "VAX"],
        "date": ["월/일", "날짜", "일자", "전표일자"],
        "vendor": ["판매처명", "거래처명", "거래처", "판1매처명", "판 매처명"],
    }

    for row_idx in range(1, min(ws.max_row + 1, 20)):  # 상위 20행에서 헤더 탐색
        row_values = [str(ws.cell(row=row_idx, column=c).value or "").strip()
                      for c in range(1, ws.max_column + 1)]
        row_text = " ".join(row_values).lower()

        if "품목코드" in row_text or "품목 코드" in row_text or "품명" in row_text:
            header_row = row_idx
            # 컬럼 인덱스 매핑
            for col_idx, val in enumerate(row_values):
                val_clean = val.strip()
                for field, keywords in HEADER_KEYWORDS.items():
                    for kw in keywords:
                        if kw in val_clean:
                            if field not in header_map:
                                header_map[field] = col_idx
                            break
            break

    if header_row is None:
        logger.warning(f"[Training] 엑셀에서 헤더 행을 찾을 수 없음: {filename}")
        # 폴백: 첫 행을 헤더로 간주
        header_row = 1
        row_values = [str(ws.cell(row=1, column=c).value or "").strip()
                      for c in range(1, ws.max_column + 1)]
        for col_idx, val in enumerate(row_values):
            for field, keywords in HEADER_KEYWORDS.items():
                for kw in keywords:
                    if kw in val.strip():
                        if field not in header_map:
                            header_map[field] = col_idx
                        break

    logger.info(f"[Training] 헤더행={header_row}, 매핑: {header_map}")

    # 데이터 행 파싱
    items = []
    vendor_name = ""

    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]

        # 빈 행 건너뛰기
        if not any(v for v in row_values if v is not None and str(v).strip()):
            continue

        def get_val(field: str, default=""):
            idx = header_map.get(field)
            if idx is not None and idx < len(row_values):
                v = row_values[idx]
                return str(v).strip() if v is not None else default
            return default

        def get_num(field: str, default=0):
            idx = header_map.get(field)
            if idx is not None and idx < len(row_values):
                v = row_values[idx]
                if v is None:
                    return default
                if isinstance(v, (int, float)):
                    return v
                try:
                    return float(str(v).replace(",", "").strip())
                except (ValueError, TypeError):
                    return default
            return default

        item_code = get_val("item_code")
        if not item_code:
            continue  # 품목코드 없는 행은 스킵

        # 거래처명 추출 (첫 번째 유효한 값 사용)
        if not vendor_name:
            vendor_name = get_val("vendor")

        item = {
            "date": get_val("date"),
            "item_code": item_code,
            "product_name": get_val("product_name"),
            "model_name": get_val("model_name"),
            "spec": get_val("spec"),
            "qty": get_num("qty"),
            "unit": get_val("unit", "EA"),
            "unit_price": get_num("unit_price"),
            "supply_price": get_num("supply_price"),
            "tax": get_num("tax"),
        }
        items.append(item)

    wb.close()

    return {
        "vendor": vendor_name,
        "items": items,
        "total_items": len(items),
    }


# ─────────────────────────────────────────
#  학습 데이터 저장
# ─────────────────────────────────────────
def save_training_pair(
    cust_code: str,
    cust_name: str,
    raw_po_text: str,
    items: List[dict],
    order_id: str = "",
    memo: str = "",
    raw_po_image: bytes = None,
    raw_po_image_type: str = "",
) -> dict:
    """
    발주서-판매전표 매칭 쌍을 DB에 저장

    Args:
        cust_code: 거래처코드
        cust_name: 거래처명
        raw_po_text: 원본 발주서 텍스트
        items: 판매전표 항목 리스트 [{item_code, product_name, model_name, qty, ...}]
        order_id: 연결된 발주서 ID (선택)
        memo: 메모 (선택)
        raw_po_image: 원본 발주서 이미지 바이너리 (선택)
        raw_po_image_type: 이미지 MIME 타입 (선택)

    Returns: {"success": True, "pair_id": 123, "item_count": 10}
    """
    conn = get_connection()
    try:
        cur = conn.execute(
            """INSERT INTO po_training_pairs(cust_code, cust_name, raw_po_text, raw_po_image, raw_po_image_type, order_id, memo)
               VALUES(?, ?, ?, ?, ?, ?, ?)""",
            (cust_code, cust_name, raw_po_text, raw_po_image, raw_po_image_type, order_id, memo)
        )
        pair_id = cur.lastrowid

        item_count = 0
        for item in items:
            if not item.get("item_code"):
                continue
            conn.execute(
                """INSERT INTO po_training_items(
                    pair_id, raw_line_text, item_code, product_name, model_name,
                    spec, qty, unit, unit_price, supply_price, tax
                ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    pair_id,
                    item.get("raw_line_text", ""),
                    item["item_code"],
                    item.get("product_name", ""),
                    item.get("model_name", ""),
                    item.get("spec", ""),
                    item.get("qty", 0),
                    item.get("unit", "EA"),
                    item.get("unit_price", 0),
                    item.get("supply_price", 0),
                    item.get("tax", 0),
                )
            )
            item_count += 1

        conn.commit()
        logger.info(f"[Training] 학습쌍 저장: pair_id={pair_id}, 거래처={cust_name}, 항목={item_count}건")
        return {"success": True, "pair_id": pair_id, "item_count": item_count}

    except Exception as e:
        conn.rollback()
        logger.error(f"[Training] 저장 실패: {e}")
        return {"success": False, "error": str(e)}
    finally:
        conn.close()


# ─────────────────────────────────────────
#  학습 데이터 조회
# ─────────────────────────────────────────
def get_training_pairs(cust_code: str = "", limit: int = 50) -> list:
    """
    학습 데이터 목록 조회
    cust_code가 있으면 해당 거래처만, 없으면 전체
    """
    conn = get_connection()
    # 목록에서는 BLOB 제외, has_image 플래그만 반환
    select_cols = """p.id, p.cust_code, p.cust_name, p.raw_po_text, p.raw_po_image_type,
                     p.order_id, p.memo, p.created_at,
                     CASE WHEN p.raw_po_image IS NOT NULL AND length(p.raw_po_image) > 0 THEN 1 ELSE 0 END as has_image,
                     COUNT(i.id) as item_count"""
    if cust_code:
        rows = conn.execute(
            f"""SELECT {select_cols}
               FROM po_training_pairs p
               LEFT JOIN po_training_items i ON i.pair_id = p.id
               WHERE p.cust_code = ?
               GROUP BY p.id
               ORDER BY p.created_at DESC
               LIMIT ?""",
            (cust_code, limit)
        ).fetchall()
    else:
        rows = conn.execute(
            f"""SELECT {select_cols}
               FROM po_training_pairs p
               LEFT JOIN po_training_items i ON i.pair_id = p.id
               GROUP BY p.id
               ORDER BY p.created_at DESC
               LIMIT ?""",
            (limit,)
        ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_training_pair_detail(pair_id: int) -> dict:
    """학습 데이터 상세 조회 (항목 포함, BLOB 제외)"""
    conn = get_connection()
    pair = conn.execute(
        """SELECT id, cust_code, cust_name, raw_po_text, raw_po_image_type,
                  order_id, memo, created_at,
                  CASE WHEN raw_po_image IS NOT NULL AND length(raw_po_image) > 0 THEN 1 ELSE 0 END as has_image
           FROM po_training_pairs WHERE id = ?""", (pair_id,)
    ).fetchone()
    if not pair:
        conn.close()
        return {}

    items = conn.execute(
        "SELECT * FROM po_training_items WHERE pair_id = ? ORDER BY id",
        (pair_id,)
    ).fetchall()
    conn.close()

    result = dict(pair)
    result["items"] = [dict(i) for i in items]
    return result


def get_training_pair_image(pair_id: int) -> tuple:
    """학습 데이터의 발주서 이미지 반환 (bytes, mime_type) or (None, None)"""
    conn = get_connection()
    row = conn.execute(
        "SELECT raw_po_image, raw_po_image_type FROM po_training_pairs WHERE id = ?",
        (pair_id,)
    ).fetchone()
    conn.close()
    if row and row["raw_po_image"]:
        return row["raw_po_image"], row["raw_po_image_type"] or "image/png"
    return None, None


def get_training_image_base64(pair_id: int) -> str:
    """학습 데이터의 발주서 이미지를 base64 문자열로 반환 (few-shot용)"""
    import base64
    img_bytes, mime_type = get_training_pair_image(pair_id)
    if img_bytes:
        return base64.b64encode(img_bytes).decode("utf-8")
    return ""


def delete_training_pair(pair_id: int) -> dict:
    """학습 데이터 삭제"""
    conn = get_connection()
    try:
        conn.execute("DELETE FROM po_training_items WHERE pair_id = ?", (pair_id,))
        conn.execute("DELETE FROM po_training_pairs WHERE id = ?", (pair_id,))
        conn.commit()
        return {"success": True}
    except Exception as e:
        conn.rollback()
        return {"success": False, "error": str(e)}
    finally:
        conn.close()


# ─────────────────────────────────────────
#  Few-Shot 프롬프트 생성 (핵심 기능)
# ─────────────────────────────────────────
def get_fewshot_examples(cust_code: str, max_examples: int = 5) -> str:
    """
    거래처의 과거 매칭 데이터를 few-shot 예시 텍스트로 생성

    발주서 처리(extraction + resolution) 시 이 텍스트를 시스템 프롬프트에 추가하여
    AI의 매칭 정확도를 높입니다.

    Returns: few-shot 예시 텍스트 (빈 문자열이면 학습 데이터 없음)
    """
    conn = get_connection()

    # 최근 학습 데이터에서 항목 조회 (거래처별)
    pairs = conn.execute(
        """SELECT id, raw_po_text, cust_name
           FROM po_training_pairs
           WHERE cust_code = ?
           ORDER BY created_at DESC
           LIMIT ?""",
        (cust_code, max_examples)
    ).fetchall()

    if not pairs:
        conn.close()
        return ""

    examples = []
    for pair in pairs:
        items = conn.execute(
            """SELECT item_code, product_name, model_name, qty, unit, unit_price
               FROM po_training_items
               WHERE pair_id = ?
               ORDER BY id""",
            (pair["id"],)
        ).fetchall()

        if not items:
            continue

        # 예시 포맷
        items_text = "\n".join(
            f"  - 품목코드: {i['item_code']}, 품명: {i['product_name']}, "
            f"모델: {i['model_name']}, 수량: {i['qty']}{i['unit']}"
            + (f", 단가: {int(i['unit_price']):,}" if i['unit_price'] else "")
            for i in items
        )

        raw_po = pair["raw_po_text"].strip()
        if raw_po:
            example = f"""[과거 매칭 예시 - {pair['cust_name']}]
발주서 원문:
{raw_po[:500]}

매칭 결과 (판매전표):
{items_text}
"""
        else:
            example = f"""[과거 매칭 예시 - {pair['cust_name']}]
매칭된 품목 목록:
{items_text}
"""
        examples.append(example)

    conn.close()

    if not examples:
        return ""

    return (
        "\n\n## 과거 매칭 참고 데이터\n"
        "아래는 이 거래처에서 과거에 발주한 내역과 실제 매칭된 품목코드입니다.\n"
        "동일하거나 유사한 상품이 있다면 이 매핑을 참고하세요.\n\n"
        + "\n".join(examples)
    )


def get_fewshot_item_map(cust_code: str) -> dict:
    """
    거래처의 과거 매칭에서 빠른 조회용 딕셔너리 생성

    Returns: {
        "모델명/품명 키워드": {"item_code": "...", "product_name": "...", "model_name": "..."},
        ...
    }
    모델명, 품명을 키로 사용하여 빠른 매칭 가능
    """
    conn = get_connection()
    items = conn.execute(
        """SELECT DISTINCT i.item_code, i.product_name, i.model_name
           FROM po_training_items i
           JOIN po_training_pairs p ON p.id = i.pair_id
           WHERE p.cust_code = ?""",
        (cust_code,)
    ).fetchall()
    conn.close()

    item_map = {}
    for i in items:
        # 여러 키로 매핑 (모델명, 품명, 품목코드 자체)
        keys = set()
        if i["model_name"]:
            keys.add(i["model_name"].strip().lower())
        if i["product_name"]:
            keys.add(i["product_name"].strip().lower())
        if i["item_code"]:
            keys.add(i["item_code"].strip().lower())

        entry = {
            "item_code": i["item_code"],
            "product_name": i["product_name"],
            "model_name": i["model_name"],
        }
        for k in keys:
            if k:
                item_map[k] = entry

    return item_map


def get_training_stats() -> dict:
    """학습 데이터 통계"""
    conn = get_connection()
    total_pairs = conn.execute("SELECT COUNT(*) as cnt FROM po_training_pairs").fetchone()["cnt"]
    total_items = conn.execute("SELECT COUNT(*) as cnt FROM po_training_items").fetchone()["cnt"]
    vendors = conn.execute(
        """SELECT cust_code, cust_name, COUNT(*) as pair_count
           FROM po_training_pairs
           GROUP BY cust_code
           ORDER BY pair_count DESC"""
    ).fetchall()
    conn.close()

    return {
        "total_pairs": total_pairs,
        "total_items": total_items,
        "vendors": [dict(v) for v in vendors],
    }
