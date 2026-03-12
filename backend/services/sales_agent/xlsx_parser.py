"""
판매 데이터 xlsx 파싱 및 전처리
- ECOUNT ERP 판매현황 내보내기 포맷 우선 지원
- 다양한 컬럼명 자동 매핑
- 거래처/품목 마스터 자동 추출
- SalesData 구조로 변환
"""
import logging
import re
from datetime import datetime, date
from pathlib import Path
from typing import Optional
import openpyxl

logger = logging.getLogger(__name__)

# ── 컬럼명 매핑 (다양한 한글/영문 컬럼명 지원) ──
COLUMN_ALIASES = {
    "transaction_date": ["거래일", "판매일", "일자", "날짜", "date", "거래일자", "판매일자", "전표일자", "매출일", "매출일자", "월/일"],
    "customer_code": ["거래처코드", "고객코드", "업체코드", "cust_code", "거래처 코드", "거래처CD"],
    "customer_name": ["거래처명", "고객명", "업체명", "거래처", "고객", "cust_name", "거래처 명", "판매처명", "판1매처명"],
    "product_code": ["품목코드", "상품코드", "제품코드", "item_code", "품목 코드", "품목CD"],
    "product_name": ["품목명", "상품명", "제품명", "품목", "item_name", "품명", "품목 명", "품명 및 규격"],
    "model_name": ["모델명", "모델", "model"],
    "category": ["카테고리", "분류", "대분류", "품목분류", "구분", "품목군", "품목그룹1명"],
    "quantity": ["수량", "판매수량", "qty", "수 량", "판매 수량"],
    "unit_price": ["단가", "판매단가", "price", "판매가", "매출단가"],
    "supply_price": ["공급가액", "공급가", "원가", "매입가", "cost"],
    "cost_price": ["입고단가", "매입단가"],
    "total_amount": ["합계금액", "합계", "판매금액", "금액", "amount", "합계 금액", "매출액", "합 계"],
    "vat": ["부가세", "세액", "vat", "부가가치세"],
    "grand_total": ["총액", "총합계", "합산금액", "총 금액"],
    "sales_rep": ["담당영업", "담당자", "영업담당", "담당", "영업사원", "영업담당자", "전표출력"],
    "warehouse": ["창고", "출고창고", "warehouse"],
    "notes": ["비고", "메모", "참고", "note", "비 고", "발송수단 및 비고사항"],
    "safety_stock": ["안전재고수량", "안전재고"],
    "display_code": ["진열코드", "진열"],
    "cust_group": ["거래처그룹1명", "거래처그룹"],
    # 거래처 마스터 추가 필드
    "industry": ["업종", "업종분류", "산업"],
    "company_size": ["규모", "기업규모", "회사규모"],
    "region": ["지역", "소재지", "지역구분"],
    "contact_name": ["담당자명", "구매담당", "구매담당자"],
    "contact_phone": ["연락처", "전화번호", "전화"],
    "contact_email": ["이메일", "email", "메일"],
    "contract_type": ["계약유형", "거래유형", "계약형태"],
    "first_deal_date": ["거래시작일", "최초거래일", "첫거래일"],
    # 품목 마스터 추가 필드
    "sub_category": ["소분류", "서브카테고리", "세분류"],
    "brand": ["브랜드", "제조사", "메이커", "벤더"],
    "standard_cost": ["표준원가", "기준원가", "최신매입가"],
    "current_stock": ["현재고", "재고수량", "재고"],
    "lead_time_days": ["리드타임", "입고일수", "발주리드타임"],
}


def _find_column(header_row: list, target_key: str) -> Optional[int]:
    """헤더 행에서 대상 키에 매핑되는 컬럼 인덱스를 찾는다.
    1) 정확 매칭 우선, 2) 부분 매칭은 alias가 cell에 포함될 때만 (역방향 제외)"""
    aliases = COLUMN_ALIASES.get(target_key, [target_key])

    # Pass 1: 정확 매칭
    for idx, cell_val in enumerate(header_row):
        if cell_val is None:
            continue
        cell_str = str(cell_val).strip()
        if cell_str in aliases or cell_str.lower() == target_key:
            return idx

    # Pass 2: 부분 매칭 (alias가 cell_str에 포함되는 경우만 — 더 구체적인 alias 우선)
    # "품목" alias가 "품목코드" cell에 매칭되는 것을 방지하기 위해
    # alias 길이가 긴 것 우선 매칭
    sorted_aliases = sorted(aliases, key=len, reverse=True)
    for idx, cell_val in enumerate(header_row):
        if cell_val is None:
            continue
        cell_str = str(cell_val).strip()
        for alias in sorted_aliases:
            # alias가 cell에 포함되되, cell이 다른 key의 정확매칭이면 스킵
            if alias in cell_str and len(alias) >= 2:
                # "품목"이 "품목코드"에 매칭되는 것 방지: cell이 다른 alias 세트의 정확매칭이면 스킵
                is_exact_other = False
                for other_key, other_aliases in COLUMN_ALIASES.items():
                    if other_key != target_key and cell_str in other_aliases:
                        is_exact_other = True
                        break
                if not is_exact_other:
                    return idx
    return None


def _parse_date(val) -> Optional[str]:
    """다양한 날짜 형식을 YYYY-MM-DD 문자열로 변환"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d", "%m/%d/%Y", "%d/%m/%Y"]:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def _parse_int(val) -> int:
    """숫자값을 int로 변환 (쉼표, 원 기호 등 제거)"""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip().replace(",", "").replace("원", "").replace("₩", "").replace(" ", "")
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


def _safe_load_workbook(file_path: str, read_only: bool = True):
    """스타일시트가 손상된 xlsx 파일도 안전하게 로드"""
    fp = Path(file_path)
    try:
        return openpyxl.load_workbook(str(fp), data_only=True, read_only=read_only)
    except Exception as e1:
        logger.warning(f"[xlsx_parser] 기본 로드 실패, 스타일시트 복구 시도: {e1}")
        import warnings
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
        from zipfile import ZipFile
        import shutil, tempfile

        tmp_dir = tempfile.mkdtemp()
        try:
            tmp_path = Path(tmp_dir) / fp.name
            # 100개의 xf 엔트리를 가진 넉넉한 스타일시트로 교체
            xf_entries = '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>' * 100
            replacement = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
                '<fonts count="1"><font><sz val="11"/></font></fonts>'
                '<fills count="2"><fill><patternFill patternType="none"/></fill>'
                '<fill><patternFill patternType="gray125"/></fill></fills>'
                '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
                '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
                f'<cellXfs count="100">{xf_entries}</cellXfs></styleSheet>'
            ).encode()

            with ZipFile(str(fp), 'r') as zin:
                with ZipFile(str(tmp_path), 'w') as zout:
                    for item in zin.infolist():
                        data = zin.read(item.filename)
                        if item.filename == 'xl/styles.xml':
                            data = replacement
                        zout.writestr(item, data)

            wb = openpyxl.load_workbook(str(tmp_path), data_only=True, read_only=False)
            logger.info("[xlsx_parser] 손상된 스타일시트 복구 후 파싱 진행")
            return wb
        except Exception as e2:
            raise ValueError(f"xlsx 파일을 열 수 없습니다 (스타일시트 복구 실패): {e2}")
        finally:
            shutil.rmtree(tmp_dir, ignore_errors=True)


def _detect_ecount_format(rows: list) -> dict:
    """
    ECOUNT ERP 판매현황 내보내기 형식인지 감지.
    Row 1: '회사명 : XXX / 거래처명 / 기간' 형태
    Row 2: 헤더 ('월/일', '품목코드', '판매처명' 등)

    Returns: {"is_ecount": bool, "company": str, "customer": str,
              "period_start": str, "period_end": str, "header_row_idx": int}
    """
    if len(rows) < 3:
        return {"is_ecount": False}

    row1_val = str(rows[0][0] or "").strip() if rows[0] and rows[0][0] else ""
    row2_vals = [str(c or "").strip() for c in rows[1]] if rows[1] else []

    # ECOUNT 판매현황 감지: 1행에 "회사명" 또는 "/" 구분 패턴, 2행에 "월/일" 또는 "품목코드"
    is_ecount = False
    company = ""
    customer = ""
    period_start = ""
    period_end = ""

    if ("/" in row1_val or "회사명" in row1_val) and any(
        k in " ".join(row2_vals) for k in ["월/일", "품목코드", "판매처명", "판1매처명", "품명 및 규격"]
    ):
        is_ecount = True
        # 파싱: "회사명 : 라인업시스템(주) / (주)컴퓨존 / 2026/01/01  ~ 2026/03/12"
        # 또는 "라인업시스템(주) / (주)컴퓨존 / 2026/01/01  ~ 2026/03/12"
        cleaned = row1_val.replace("회사명", "").replace(":", "").strip()
        parts = [p.strip() for p in cleaned.split("/") if p.strip()]

        # 기간 파싱 (마지막 부분에서 ~ 패턴 찾기)
        full_text = cleaned
        period_match = re.search(r'(\d{4}/\d{2}/\d{2})\s*~\s*(\d{4}/\d{2}/\d{2})', full_text)
        if period_match:
            period_start = period_match.group(1).replace("/", "-")
            period_end = period_match.group(2).replace("/", "-")

        # 회사명과 거래처명 추출
        # 패턴: "회사명 / 거래처명 / 시작일자 ~ 종료일자"
        # parts에서 날짜가 아닌 것들 추출
        non_date_parts = []
        for p in parts:
            # 날짜 패턴이면 스킵
            if re.match(r'^\d{4}$', p) or re.match(r'^\d{2}$', p):
                continue
            if "~" in p:
                continue
            non_date_parts.append(p)

        if len(non_date_parts) >= 2:
            company = non_date_parts[0]
            customer = non_date_parts[1]
        elif len(non_date_parts) == 1:
            company = non_date_parts[0]

        logger.info(f"[xlsx_parser] ECOUNT 형식 감지: 회사={company}, 거래처={customer}, 기간={period_start}~{period_end}")

    return {
        "is_ecount": is_ecount,
        "company": company,
        "customer": customer,
        "period_start": period_start,
        "period_end": period_end,
        "header_row_idx": 1 if is_ecount else 0,
    }


def _parse_ecount_date(val: str, year_hint: str = "") -> Optional[str]:
    """
    ECOUNT 월/일 형식 파싱: 'MM/DD-전표번호' → 'YYYY-MM-DD'
    예: '01/02-27' → '2026-01-02', '03 계' → None (소계행)
    """
    if not val:
        return None
    s = str(val).strip()

    # 소계/합계행 스킵
    if "계" in s or "합계" in s:
        return None

    # MM/DD-번호 패턴
    m = re.match(r'^(\d{2})/(\d{2})', s)
    if m:
        month, day = m.group(1), m.group(2)
        year = year_hint[:4] if year_hint else str(datetime.now().year)
        try:
            return f"{year}-{month}-{day}"
        except Exception:
            return None

    # 일반 날짜 형식도 시도
    return _parse_date(val)


def parse_xlsx(file_path: str) -> dict:
    """
    xlsx 파일을 파싱하여 판매 데이터 구조로 변환한다.
    ECOUNT ERP 판매현황 내보내기 형식을 우선 지원.

    Returns:
        {
            "transactions": [...],
            "customers": [...],
            "products": [...],
            "period_start": "YYYY-MM-DD",
            "period_end": "YYYY-MM-DD",
            "file_name": "...",
            "summary": {"total_rows": N, "total_customers": N, "total_products": N, ...}
        }
    """
    fp = Path(file_path)
    if not fp.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")

    wb = _safe_load_workbook(str(fp))
    result = {
        "transactions": [],
        "customers": {},
        "products": {},
        "file_name": fp.name,
    }

    # ── 시트 탐색 ──
    tx_sheet = None
    cust_sheet = None
    prod_sheet = None

    tx_keywords = ["판매", "매출", "트랜잭션", "거래", "sales", "transaction"]
    cust_keywords = ["거래처", "고객", "customer", "업체"]
    prod_keywords = ["품목", "상품", "제품", "product", "item"]

    for sn in wb.sheetnames:
        sn_lower = sn.lower().strip()
        if any(k in sn_lower for k in cust_keywords) and "판매" not in sn_lower:
            cust_sheet = wb[sn]
        elif any(k in sn_lower for k in prod_keywords) and "판매" not in sn_lower:
            prod_sheet = wb[sn]
        elif tx_sheet is None:
            if any(k in sn_lower for k in tx_keywords) or tx_sheet is None:
                tx_sheet = wb[sn]

    if tx_sheet is None:
        tx_sheet = wb[wb.sheetnames[0]]

    # ── 행 읽기 ──
    rows = list(tx_sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        wb.close()
        raise ValueError("데이터가 부족합니다 (최소 헤더 + 1행 필요)")

    # ── ECOUNT 형식 감지 ──
    ecount = _detect_ecount_format(rows)
    header_idx = ecount["header_row_idx"]
    header = [str(c).strip() if c else "" for c in rows[header_idx]]

    # ECOUNT에서 추출한 메타 정보
    meta_customer = ecount.get("customer", "")
    meta_period_start = ecount.get("period_start", "")
    meta_period_end = ecount.get("period_end", "")
    year_hint = meta_period_start[:4] if meta_period_start else ""

    # ── 컬럼 매핑 ──
    col_map = {}
    for key in ["transaction_date", "customer_code", "customer_name", "product_code",
                 "product_name", "model_name", "category", "quantity", "unit_price",
                 "supply_price", "cost_price", "total_amount", "vat", "grand_total",
                 "sales_rep", "warehouse", "notes", "safety_stock", "cust_group"]:
        idx = _find_column(header, key)
        if idx is not None:
            col_map[key] = idx

    logger.info(f"[xlsx_parser] 컬럼 매핑: {list(col_map.keys())} ({len(col_map)}개), ECOUNT={ecount['is_ecount']}")

    # 필수 컬럼 체크 (ECOUNT: product_name만 있으면 OK, 거래처는 메타에서)
    if not ecount["is_ecount"]:
        required = ["customer_name", "product_name"]
        missing = [k for k in required if k not in col_map]
        if missing:
            wb.close()
            raise ValueError(f"필수 컬럼을 찾을 수 없습니다: {missing}. 헤더: {header}")
    else:
        if "product_name" not in col_map and "product_code" not in col_map:
            wb.close()
            raise ValueError(f"품목 관련 컬럼을 찾을 수 없습니다. 헤더: {header}")

    # ── 데이터 행 파싱 ──
    dates = []
    data_start = header_idx + 1

    for row_idx, row in enumerate(rows[data_start:], start=data_start + 1):
        if all(c is None for c in row):
            continue

        # A열 값으로 소계/합계/타임스탬프 행 스킵
        a_val = str(row[0] or "").strip() if row[0] is not None else ""
        if not a_val:
            continue
        if "계" in a_val or "합계" in a_val:
            continue
        # 날짜 타임스탬프 행 스킵 (예: "2026/03/12 (목) 오후 12:35:36")
        if re.match(r'^\d{4}/\d{2}/\d{2}\s*\(', a_val):
            continue

        def _get(key):
            idx = col_map.get(key)
            return row[idx] if idx is not None and idx < len(row) else None

        # 날짜 파싱
        if ecount["is_ecount"]:
            tx_date = _parse_ecount_date(a_val, year_hint)
        else:
            tx_date = _parse_date(_get("transaction_date"))

        if not tx_date and ecount["is_ecount"]:
            # ECOUNT에서 날짜 파싱 못하면 데이터행이 아닐 가능성
            continue

        # 거래처
        cust_name = str(_get("customer_name") or "").strip()
        cust_code = str(_get("customer_code") or "").strip()
        if not cust_name and meta_customer:
            cust_name = meta_customer

        # 품목
        prod_code = str(_get("product_code") or "").strip()
        prod_name = str(_get("product_name") or "").strip()
        model_name = str(_get("model_name") or "").strip()
        category = str(_get("category") or "").strip()

        quantity = _parse_int(_get("quantity")) or 1
        unit_price = _parse_int(_get("unit_price"))
        supply_price = _parse_int(_get("supply_price"))
        cost_price = _parse_int(_get("cost_price"))
        total_amount = _parse_int(_get("total_amount"))
        vat = _parse_int(_get("vat"))
        sales_rep = str(_get("sales_rep") or "").strip()
        warehouse = str(_get("warehouse") or "").strip()
        notes = str(_get("notes") or "").strip()

        if not cust_name and not prod_name:
            continue

        # ECOUNT에서는 supply_price(공급가액, H열)가 메인 금액
        # total_amount(합 계, J열)는 공급가액+부가세
        if ecount["is_ecount"]:
            # supply_price가 실제 공급가액 (메인 매출액)
            if supply_price != 0 and total_amount == 0:
                total_amount = supply_price
            elif supply_price == 0 and total_amount != 0:
                supply_price = total_amount
        else:
            # 자동 계산 보완
            if total_amount == 0 and unit_price > 0:
                total_amount = unit_price * quantity
            if unit_price == 0 and total_amount > 0 and quantity > 0:
                unit_price = total_amount // quantity

        # 거래처코드 자동 생성
        if not cust_code and cust_name:
            cust_code = f"C{hash(cust_name) % 10000:04d}"
        # 품목코드 자동 생성
        if not prod_code and prod_name:
            prod_code = f"P{hash(prod_name) % 10000:04d}"

        tx = {
            "transaction_date": tx_date or "",
            "customer_code": cust_code,
            "customer_name": cust_name,
            "product_code": prod_code,
            "product_name": prod_name,
            "model_name": model_name,
            "category": category,
            "quantity": quantity,
            "unit_price": unit_price,
            "supply_price": supply_price,
            "cost_price": cost_price,
            "total_amount": total_amount,
            "vat": vat,
            "sales_rep": sales_rep,
            "warehouse": warehouse,
            "notes": notes,
        }
        result["transactions"].append(tx)

        if tx_date:
            dates.append(tx_date)

        # 거래처 자동 수집
        if cust_code and cust_code not in result["customers"]:
            result["customers"][cust_code] = {
                "customer_code": cust_code,
                "customer_name": cust_name,
                "industry": "",
                "company_size": "",
                "region": "",
                "contract_type": "",
                "cust_group": str(_get("cust_group") or "").strip(),
            }

        # 품목 자동 수집
        if prod_code and prod_code not in result["products"]:
            result["products"][prod_code] = {
                "product_code": prod_code,
                "product_name": prod_name,
                "model_name": model_name,
                "category": category,
                "brand": "",
                "standard_cost": cost_price or supply_price,
            }

    # ── 거래처 마스터 시트 파싱 (있으면) ──
    if cust_sheet:
        try:
            c_rows = list(cust_sheet.iter_rows(values_only=True))
            if len(c_rows) >= 2:
                c_header = [str(c).strip() if c else "" for c in c_rows[0]]
                c_map = {}
                for key in ["customer_code", "customer_name", "industry", "company_size",
                            "region", "contact_name", "contact_phone", "contact_email",
                            "contract_type", "first_deal_date"]:
                    idx = _find_column(c_header, key)
                    if idx is not None:
                        c_map[key] = idx

                for crow in c_rows[1:]:
                    if all(c is None for c in crow):
                        continue
                    cc = str(crow[c_map["customer_code"]] if "customer_code" in c_map else "").strip()
                    if cc and cc in result["customers"]:
                        for k, cidx in c_map.items():
                            if k != "customer_code" and cidx < len(crow) and crow[cidx]:
                                result["customers"][cc][k] = str(crow[cidx]).strip()
        except Exception as e:
            logger.warning(f"[xlsx_parser] 거래처 마스터 파싱 실패: {e}")

    # ── 품목 마스터 시트 파싱 (있으면) ──
    if prod_sheet:
        try:
            p_rows = list(prod_sheet.iter_rows(values_only=True))
            if len(p_rows) >= 2:
                p_header = [str(c).strip() if c else "" for c in p_rows[0]]
                p_map = {}
                for key in ["product_code", "product_name", "category", "sub_category",
                            "brand", "standard_cost", "current_stock", "lead_time_days"]:
                    idx = _find_column(p_header, key)
                    if idx is not None:
                        p_map[key] = idx

                for prow in p_rows[1:]:
                    if all(c is None for c in prow):
                        continue
                    pc = str(prow[p_map["product_code"]] if "product_code" in p_map else "").strip()
                    if pc and pc in result["products"]:
                        for k, pidx in p_map.items():
                            if k != "product_code" and pidx < len(prow) and prow[pidx]:
                                result["products"][pc][k] = str(prow[pidx]).strip()
        except Exception as e:
            logger.warning(f"[xlsx_parser] 품목 마스터 파싱 실패: {e}")

    wb.close()

    # ── 결과 정리 ──
    result["customers"] = list(result["customers"].values())
    result["products"] = list(result["products"].values())

    # 기간: ECOUNT 메타 우선, 없으면 데이터에서 추출
    if meta_period_start and meta_period_end:
        result["period_start"] = meta_period_start
        result["period_end"] = meta_period_end
    elif dates:
        dates_sorted = sorted(dates)
        result["period_start"] = dates_sorted[0]
        result["period_end"] = dates_sorted[-1]
    else:
        result["period_start"] = ""
        result["period_end"] = ""

    result["summary"] = {
        "total_rows": len(result["transactions"]),
        "total_customers": len(result["customers"]),
        "total_products": len(result["products"]),
        "total_amount": sum(t["total_amount"] for t in result["transactions"]),
        "period": f"{result['period_start']} ~ {result['period_end']}",
    }

    logger.info(
        f"[xlsx_parser] 파싱 완료: {result['summary']['total_rows']}건, "
        f"거래처 {result['summary']['total_customers']}개, "
        f"품목 {result['summary']['total_products']}개, "
        f"총액 {result['summary']['total_amount']:,}원"
    )

    return result
