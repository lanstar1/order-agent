"""
메일 자동화 서비스
- IMAP 접속 → guzhiyi@bor-cable.com 선적 메일 검색
- BOR Excel 첨부파일 → HS코드 자동 입력
- ERP 구매전표 자동 생성
- SMTP 회신 (HS코드 입력된 Excel 첨부)
"""

import imaplib
import email
import smtplib
from email.header import decode_header
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import logging
import io
import os
import re
import json
import httpx
from datetime import datetime, timedelta, timezone
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))

from services.hs_code_engine import HSCodeEngine
from config import (
    MAIL_IMAP_SERVER, MAIL_IMAP_PORT, MAIL_USER, MAIL_PASSWORD,
)

logger = logging.getLogger(__name__)
KST = timezone(timedelta(hours=9))
hs_engine = HSCodeEngine()

TARGET_SENDER = os.getenv("MAIL_TARGET_SENDER", "guzhiyi@bor-cable.com")
MAIL_SMTP_HOST = os.getenv("MAIL_SMTP_HOST", "wsmtp.ecount.com")
MAIL_SMTP_PORT = int(os.getenv("MAIL_SMTP_PORT", "587"))
ERP_SUPPLIER_CODE = os.getenv("ERP_SUPPLIER_CODE", "1111122222")
MAIL_AUTO_PASSWORD = os.getenv("MAIL_AUTO_PASSWORD", "lanstar2026")


# ─── 유틸리티 ────────────────────────────────────────────

def _decode_header_value(value):
    if not value:
        return ""
    decoded = decode_header(value)
    parts = []
    for part, charset in decoded:
        if isinstance(part, bytes):
            parts.append(part.decode(charset or "utf-8", errors="replace"))
        else:
            parts.append(str(part))
    return " ".join(parts)


def _parse_email_date(date_str):
    if not date_str:
        return None
    try:
        from email.utils import parsedate_to_datetime
        dt = parsedate_to_datetime(date_str)
        return dt.astimezone(KST)
    except Exception:
        return None


# ─── 환율 조회 (전신환매도율 근사) ─────────────────────────
# 무료 API 기준율 + 은행 스프레드(약 1.75%) = 전신환매도율 근사
# 스프레드는 프론트에서 조정 가능

TT_SELL_SPREAD = float(os.getenv("TT_SELL_SPREAD", "1.75"))  # 기본 1.75%


async def fetch_exchange_rate() -> dict:
    """
    USD/KRW 전신환매도율 근사 조회
    = 매매기준율(무료API) × (1 + 스프레드%)
    
    Returns: {"rate": float, "base_rate": float, "spread": float, "source": str}
    """
    base_rate = None
    source = ""

    # 1차: exchangerate-api (무료, 매매기준율 근사)
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            r = await client.get("https://open.er-api.com/v6/latest/USD")
            if r.status_code == 200:
                data = r.json()
                rate = data.get("rates", {}).get("KRW")
                if rate:
                    base_rate = float(rate)
                    source = "er-api"
    except Exception as e:
        logger.warning(f"[환율] er-api 실패: {e}")

    # 2차: 폴백 API
    if not base_rate:
        try:
            async with httpx.AsyncClient(timeout=10) as client:
                r = await client.get("https://api.exchangerate.host/latest?base=USD&symbols=KRW")
                if r.status_code == 200:
                    data = r.json()
                    rate = data.get("rates", {}).get("KRW")
                    if rate:
                        base_rate = float(rate)
                        source = "exchangerate.host"
        except Exception as e:
            logger.warning(f"[환율] 폴백 API 실패: {e}")

    # 3차: 고정값
    if not base_rate:
        base_rate = 1450.0
        source = "기본값"
        logger.warning("[환율] 모든 API 실패, 기본값 사용")

    # 전신환매도율 = 기준율 × (1 + 스프레드%)
    spread = TT_SELL_SPREAD
    tt_sell_rate = round(base_rate * (1 + spread / 100), 2)

    logger.info(f"[환율] 기준율={base_rate}, 스프레드={spread}%, 매도율={tt_sell_rate} ({source})")

    return {
        "rate": tt_sell_rate,         # 전신환매도율 (실제 적용 환율)
        "base_rate": base_rate,       # 매매기준율
        "spread": spread,             # 스프레드 %
        "source": source,
    }


# ─── IMAP 메일 수신 ──────────────────────────────────────

def fetch_bor_emails(days_back: int = 30) -> list:
    """
    IMAP에서 guzhiyi@bor-cable.com 발신, BOR Excel 첨부 메일 검색
    
    Returns: [{
        "message_id": str,
        "uid": str,
        "subject": str,
        "date": str,
        "date_kst": datetime,
        "attachments": [{"filename": str, "data": bytes, "bor_number": str}],
    }]
    """
    if not MAIL_USER or not MAIL_PASSWORD:
        logger.error("[메일자동화] 메일 계정 미설정")
        return []

    results = []
    mail = None
    try:
        mail = imaplib.IMAP4_SSL(MAIL_IMAP_SERVER, MAIL_IMAP_PORT)
        mail.login(MAIL_USER, MAIL_PASSWORD)
        mail.select("INBOX", readonly=True)

        since = (datetime.now(KST) - timedelta(days=days_back)).strftime("%d-%b-%Y")
        
        # UID 기반 검색
        try:
            status, data = mail.uid("search", None, f'(FROM "{TARGET_SENDER}" SINCE {since})')
        except Exception:
            status, data = mail.search(None, f'(FROM "{TARGET_SENDER}" SINCE {since})')
        
        if status != "OK" or not data[0]:
            logger.info("[메일자동화] 새 메일 없음")
            return []
        
        uids = data[0].split()
        logger.info(f"[메일자동화] {TARGET_SENDER} 발신 메일 {len(uids)}건 발견")

        for uid in uids:
            try:
                status, msg_data = mail.uid("fetch", uid, "(RFC822)")
                if status != "OK":
                    continue
                
                raw_body = None
                for part in msg_data:
                    if isinstance(part, tuple) and len(part) >= 2:
                        raw_body = part[1]
                        break
                if not raw_body:
                    continue

                msg = email.message_from_bytes(raw_body)
                subject = _decode_header_value(msg.get("Subject", ""))
                date_str = msg.get("Date", "")
                date_kst = _parse_email_date(date_str)
                message_id = msg.get("Message-ID", "")

                # 첨부파일 추출
                attachments = []
                for part in msg.walk():
                    if part.get_content_disposition() != "attachment":
                        continue
                    filename = _decode_header_value(part.get_filename() or "")
                    if not filename.lower().endswith(".xlsx"):
                        continue
                    
                    file_data = part.get_payload(decode=True)
                    if not file_data:
                        continue
                    
                    # BOR 번호 추출
                    bor_match = re.search(r'(BOR-\d{7})', filename, re.IGNORECASE)
                    bor_number = bor_match.group(1).upper() if bor_match else ""
                    
                    attachments.append({
                        "filename": filename,
                        "data": file_data,
                        "bor_number": bor_number,
                    })

                if attachments:
                    results.append({
                        "message_id": message_id,
                        "uid": uid.decode() if isinstance(uid, bytes) else str(uid),
                        "subject": subject,
                        "date": date_str,
                        "date_kst": date_kst,
                        "attachments": attachments,
                    })

            except Exception as e:
                logger.error(f"[메일자동화] 메일 처리 오류 (UID {uid}): {e}")
                continue

    except Exception as e:
        logger.error(f"[메일자동화] IMAP 접속 오류: {e}")
    finally:
        if mail:
            try:
                mail.close()
                mail.logout()
            except Exception:
                pass

    return results


# ─── Excel HS코드 처리 ───────────────────────────────────

def process_excel_hs_code(file_data: bytes, filename: str) -> dict:
    """
    BOR Invoice Excel에 HS코드를 자동 입력
    
    Returns: {
        "success": bool,
        "output_data": bytes (수정된 Excel),
        "items": [{"model": str, "category": str, "hs_code": str, "rule": str, ...}],
        "erp_lines": [{"prod_cd": str, "qty": float, "price_usd": float}],
        "oem_items": [{"description": str, "category": str}],
        "stats": {"total": int, "hs_filled": int, "skipped": int, "unknown": int},
    }
    """
    import openpyxl
    from openpyxl.styles import Font
    
    hs_font = Font(bold=True, color="FF0000")  # 빨간색 + 볼드
    
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_data))
    except Exception as e:
        return {"success": False, "error": f"Excel 열기 실패: {e}"}
    
    if "Invoice" not in wb.sheetnames:
        return {"success": False, "error": "Invoice 시트가 없습니다"}
    
    ws = wb["Invoice"]
    
    items = []
    erp_lines = []
    oem_items = []
    current_category = ""
    stats = {"total": 0, "hs_filled": 0, "skipped": 0, "unknown": 0}
    
    def _safe_float(val):
        """수식('=1.62*0.97') 등 변환 불가 셀 안전 처리"""
        if val is None:
            return 0.0
        try:
            return float(val)
        except (ValueError, TypeError):
            # 수식 문자열이면 eval 시도
            if isinstance(val, str) and val.startswith("="):
                try:
                    return float(eval(val[1:].replace(",", "")))
                except Exception:
                    pass
            return 0.0
    
    # I열(9번째) 헤더 확인/추가
    header_row = 13
    if ws.cell(row=header_row, column=9).value not in ("HS CODE", "HS code", "HS Code"):
        ws.cell(row=header_row, column=9).value = "HS CODE"
    
    for row_idx in range(14, ws.max_row + 1):
        a_val = ws.cell(row=row_idx, column=1).value
        b_val = ws.cell(row=row_idx, column=2).value
        c_val = ws.cell(row=row_idx, column=3).value  # Quantity
        f_val = ws.cell(row=row_idx, column=6).value   # Unit Price
        i_val = ws.cell(row=row_idx, column=9).value   # HS CODE (기존)
        
        # BOR 주문번호 행 → 건너뜀
        if a_val and isinstance(a_val, str) and re.match(r'^BOR-\d', a_val):
            continue
        if a_val and isinstance(a_val, str) and re.match(r'^PAR-\d', a_val):
            continue
        
        # 카테고리 행 (B열에 텍스트, A열에 모델명 없음)
        if b_val and isinstance(b_val, str) and len(b_val.strip()) > 2:
            if not a_val or (isinstance(a_val, str) and (
                a_val.startswith("BOR-") or a_val.startswith("PAR-")
            )):
                current_category = b_val.strip()
                continue
            
            # B열에 설명이 있고 A열에 모델명이 있는 경우 (LS-68R | Crimping Tool)
            if a_val and isinstance(a_val, str):
                model = hs_engine.extract_model_name(str(a_val))
                if model:
                    combined_desc = f"{a_val}, {b_val}"
                    result = hs_engine.match(current_category, combined_desc)
                    stats["total"] += 1
                    
                    # HS코드 입력
                    if result.hs_code and not i_val:
                        cell = ws.cell(row=row_idx, column=9); cell.value = result.hs_code; cell.font = hs_font
                        stats["hs_filled"] += 1
                    elif result.confidence == "skip":
                        stats["skipped"] += 1
                    elif result.confidence == "unknown":
                        stats["unknown"] += 1
                    else:
                        stats["skipped"] += 1
                    
                    items.append({
                        "row": row_idx, "model": model,
                        "category": current_category,
                        "hs_code": result.hs_code,
                        "rule": result.rule_name,
                        "confidence": result.confidence,
                    })
                    
                    # ERP 라인
                    if hs_engine.is_erp_target(model):
                        erp_lines.append({
                            "prod_cd": model,
                            "qty": _safe_float(c_val),
                            "price_usd": _safe_float(f_val),
                            "description": combined_desc[:100],
                        })
                    else:
                        oem_items.append({
                            "description": combined_desc[:100],
                            "category": current_category,
                        })
                    continue
        
        # 일반 품목행 (A열에 모델명)
        if a_val and isinstance(a_val, str):
            model = hs_engine.extract_model_name(str(a_val))
            if model:
                result = hs_engine.match(current_category, str(a_val))
                stats["total"] += 1
                
                if result.hs_code and not i_val:
                    cell = ws.cell(row=row_idx, column=9); cell.value = result.hs_code; cell.font = hs_font
                    stats["hs_filled"] += 1
                elif result.confidence == "skip":
                    stats["skipped"] += 1
                elif result.confidence == "unknown":
                    stats["unknown"] += 1
                else:
                    stats["skipped"] += 1
                
                items.append({
                    "row": row_idx, "model": model,
                    "category": current_category,
                    "hs_code": result.hs_code,
                    "rule": result.rule_name,
                    "confidence": result.confidence,
                })
                
                if hs_engine.is_erp_target(model):
                    erp_lines.append({
                        "prod_cd": model,
                        "qty": _safe_float(c_val),
                        "price_usd": _safe_float(f_val),
                        "description": str(a_val)[:100],
                    })
                elif not model.startswith(("LS-", "LSP-", "LSN-", "ZOT-")):
                    oem_items.append({
                        "description": str(a_val)[:100],
                        "category": current_category,
                    })
    
    # 수정된 Excel 저장
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return {
        "success": True,
        "output_data": output.read(),
        "filename": filename,
        "items": items,
        "erp_lines": erp_lines,
        "oem_items": oem_items,
        "stats": stats,
    }


# ─── ERP 구매전표 생성 ────────────────────────────────────

async def create_purchase_slip(
    erp_lines: list,
    exchange_rate: float,
    email_date: datetime = None,
) -> dict:
    """
    ERP 구매전표 자동 생성
    
    단가 = USD × 1.2 × 환율
    전표일자 = 메일수신일 + 8일
    """
    from services.erp_client import erp_client
    
    if not erp_lines:
        return {"success": False, "error": "입력할 품목이 없습니다"}
    
    # 전표일자: 메일수신일 +8일
    if email_date:
        io_date = (email_date + timedelta(days=8)).strftime("%Y%m%d")
    else:
        io_date = (datetime.now(KST) + timedelta(days=8)).strftime("%Y%m%d")
    
    # 라인 변환
    lines = []
    for item in erp_lines:
        price_usd = item.get("price_usd", 0)
        # 단가: USD × 1.2(관세/부가세) × 환율
        price_krw = round(price_usd * 1.2 * exchange_rate)
        
        lines.append({
            "prod_cd": item["prod_cd"],
            "qty": item["qty"],
            "unit": "EA",
            "price": price_krw,
        })
    
    logger.info(f"[구매전표] {len(lines)}개 품목, 환율={exchange_rate}, 일자={io_date}")
    
    try:
        result = await erp_client.save_purchase(
            cust_code=ERP_SUPPLIER_CODE,
            lines=lines,
            io_date=io_date,
        )
        return result
    except Exception as e:
        logger.error(f"[구매전표] ERP 전송 실패: {e}")
        return {"success": False, "error": str(e)}


# ─── SMTP 회신 ────────────────────────────────────────────

def send_reply_email(
    to_address: str,
    subject: str,
    body: str,
    attachment_data: bytes = None,
    attachment_filename: str = None,
    in_reply_to: str = None,
) -> bool:
    """HS코드 입력된 Excel을 첨부하여 회신"""
    if not MAIL_USER or not MAIL_PASSWORD:
        logger.error("[SMTP] 계정 미설정")
        return False
    
    try:
        msg = MIMEMultipart()
        msg["From"] = MAIL_USER
        msg["To"] = to_address
        msg["Subject"] = f"Re: {subject}" if not subject.startswith("Re:") else subject
        if in_reply_to:
            msg["In-Reply-To"] = in_reply_to
            msg["References"] = in_reply_to
        
        msg.attach(MIMEText(body, "plain", "utf-8"))
        
        if attachment_data and attachment_filename:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment_data)
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename={attachment_filename}")
            msg.attach(part)
        
        with smtplib.SMTP(MAIL_SMTP_HOST, MAIL_SMTP_PORT) as server:
            server.starttls()
            server.login(MAIL_USER, MAIL_PASSWORD)
            server.send_message(msg)
        
        logger.info(f"[SMTP] 회신 발송 완료 → {to_address}")
        return True
    
    except Exception as e:
        logger.error(f"[SMTP] 회신 발송 실패: {e}")
        return False


# ─── 전체 파이프라인 ──────────────────────────────────────

async def run_mail_automation_pipeline(
    days_back: int = 30,
    exchange_rate: float = None,
    auto_reply: bool = False,
    auto_erp: bool = True,
    db_conn=None,
    reply_template: str = "",
) -> dict:
    """
    전체 메일 자동화 파이프라인 실행
    """
    # 환율 조회
    rate_info = {}
    if not exchange_rate:
        rate_info = await fetch_exchange_rate()
        exchange_rate = rate_info["rate"]
    else:
        rate_info = {"rate": exchange_rate, "base_rate": exchange_rate, "spread": 0, "source": "manual"}
    
    # 처리된 메일 ID 조회 (중복 방지)
    processed_ids = set()
    if db_conn:
        try:
            cursor = db_conn.execute(
                "SELECT message_id FROM mail_processing_log WHERE status='completed'"
            )
            processed_ids = {row[0] for row in cursor.fetchall()}
        except Exception:
            pass
    
    # 메일 수신
    emails = fetch_bor_emails(days_back=days_back)
    
    pipeline_results = []
    
    for mail_info in emails:
        msg_id = mail_info["message_id"]
        
        # 이미 처리된 메일 건너뜀
        if msg_id in processed_ids:
            continue
        
        mail_result = {
            "message_id": msg_id,
            "subject": mail_info["subject"],
            "date": str(mail_info.get("date_kst", "")),
            "attachments_processed": [],
            "erp_result": None,
            "reply_sent": False,
            "status": "processing",
        }
        
        all_erp_lines = []
        all_oem_items = []
        processed_excels = []
        
        for att in mail_info["attachments"]:
            # HS코드 처리
            excel_result = process_excel_hs_code(att["data"], att["filename"])
            
            if excel_result["success"]:
                mail_result["attachments_processed"].append({
                    "filename": att["filename"],
                    "bor_number": att.get("bor_number", ""),
                    "stats": excel_result["stats"],
                    "items_count": len(excel_result["items"]),
                    "hs_items": [i for i in excel_result["items"] if i["hs_code"]],
                })
                all_erp_lines.extend(excel_result["erp_lines"])
                all_oem_items.extend(excel_result["oem_items"])
                processed_excels.append(excel_result)
        
        # ERP 구매전표
        if auto_erp and all_erp_lines:
            erp_result = await create_purchase_slip(
                erp_lines=all_erp_lines,
                exchange_rate=exchange_rate,
                email_date=mail_info.get("date_kst"),
            )
            mail_result["erp_result"] = erp_result
        
        # 자동 회신 (첫 번째 처리된 Excel 첨부)
        if auto_reply and processed_excels:
            first_excel = processed_excels[0]
            reply_body = reply_template if reply_template.strip() else (
                "Dear Mr. Gu,\n\n"
                "We have reviewed the packing list and added the HS codes accordingly.\n"
                "Please find the attached file for your reference.\n\n"
                "Best regards,\n"
                "LINEUP SYSTEM CO., LTD."
            )
            reply_sent = send_reply_email(
                to_address=TARGET_SENDER,
                subject=mail_info["subject"],
                body=reply_body,
                attachment_data=first_excel["output_data"],
                attachment_filename=first_excel["filename"],
                in_reply_to=msg_id,
            )
            mail_result["reply_sent"] = reply_sent
        
        mail_result["oem_items"] = all_oem_items
        mail_result["erp_lines_count"] = len(all_erp_lines)
        mail_result["exchange_rate"] = exchange_rate
        mail_result["status"] = "completed"
        
        # DB 로그 저장
        if db_conn:
            try:
                db_conn.execute("""
                    INSERT OR REPLACE INTO mail_processing_log 
                    (message_id, subject, sender, received_at, attachment_count,
                     status, hs_code_count, reply_sent, processed_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    msg_id,
                    mail_info["subject"],
                    TARGET_SENDER,
                    str(mail_info.get("date_kst", "")),
                    len(mail_info["attachments"]),
                    "completed",
                    sum(r["stats"]["hs_filled"] for r in processed_excels if r.get("success")),
                    mail_result["reply_sent"],
                    datetime.now(KST).isoformat(),
                ))
                db_conn.commit()
            except Exception as e:
                logger.error(f"[DB] 로그 저장 실패: {e}")
        
        pipeline_results.append(mail_result)
    
    return {
        "exchange_rate": exchange_rate,
        "base_rate": rate_info.get("base_rate", exchange_rate),
        "spread": rate_info.get("spread", 0),
        "rate_source": rate_info.get("source", ""),
        "total_emails": len(emails),
        "new_processed": len(pipeline_results),
        "already_processed": len(processed_ids),
        "results": pipeline_results,
    }
