"""
선적 메일 파싱 서비스
- IMAP으로 메일서버 접속
- "Final shipping", "shipping list" 등 키워드 메일 검색
- BOR로 시작하는 엑셀 첨부파일 다운로드 → 모델명 파싱
- 오더리스트의 BOR 번호와 매칭 → 해당 품목들 선적 확인
- 선적일(메일 수신일) + 입고예정일(+8일) 저장
"""

import imaplib
import email
from email.header import decode_header
import logging
import io
import os
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))

logger = logging.getLogger(__name__)
KST = timezone(timedelta(hours=9))


def _decode_header_value(value):
    """메일 헤더 디코딩"""
    if not value:
        return ""
    decoded = decode_header(value)
    parts = []
    for part, charset in decoded:
        if isinstance(part, bytes):
            parts.append(part.decode(charset or "utf-8", errors="replace"))
        else:
            parts.append(str(part))
    return " ".join(parts)


def _parse_email_date(date_str):
    """메일 날짜 파싱 → KST datetime"""
    if not date_str:
        return None
    try:
        from email.utils import parsedate_to_datetime
        dt = parsedate_to_datetime(date_str)
        return dt.astimezone(KST)
    except Exception:
        return None


def scan_shipping_emails(
    imap_server: str,
    imap_user: str,
    imap_password: str,
    imap_port: int = 993,
    days_back: int = 90,
    search_folder: str = "INBOX",
) -> list:
    """
    IMAP 메일서버에서 선적 관련 메일을 검색하고 BOR 첨부파일을 파싱

    Returns: [{
        "bor_number": "BOR-2601001",
        "subject": "Final shipping list...",
        "email_date": "2026-03-15",
        "shipping_date": "2026-03-15",
        "arrival_date": "2026-03-23",
        "filename": "BOR-2601001.xlsx",
        "models": ["LS-1000H", "LS-1200HB", ...],
    }, ...]
    """
    results = []

    try:
        # IMAP 접속
        logger.info(f"[선적메일] IMAP 접속: {imap_server}:{imap_port}")
        mail = imaplib.IMAP4_SSL(imap_server, imap_port)
        mail.login(imap_user, imap_password)
        mail.select(search_folder, readonly=True)

        # 날짜 범위 설정
        since_date = (datetime.now(KST) - timedelta(days=days_back)).strftime("%d-%b-%Y")

        # 전체 메일 대상 (일부 IMAP 서버는 SUBJECT 검색 미지원)
        try:
            status, data = mail.search(None, f"(SINCE {since_date})")
        except Exception:
            status, data = mail.search(None, "ALL")
        all_uids = set(data[0].split()) if status == "OK" and data[0] else set()

        logger.info(f"[선적메일] 대상 메일 {len(all_uids)}건 (최근 {days_back}일)")

        for uid in all_uids:
            try:
                # Ecount IMAP 호환: 여러 FETCH 형식 시도
                msg_data = None
                for fetch_cmd in ["(RFC822)", "RFC822", "(BODY[])"]:
                    try:
                        status, msg_data = mail.fetch(uid, fetch_cmd)
                        if status == "OK" and msg_data and msg_data[0]:
                            break
                    except Exception:
                        continue
                if not msg_data or not msg_data[0]:
                    continue

                raw = msg_data[0][1] if isinstance(msg_data[0], tuple) else msg_data[0]
                if not isinstance(raw, bytes):
                    continue

                msg = email.message_from_bytes(raw)
                subject = _decode_header_value(msg.get("Subject", ""))
                date_str = msg.get("Date", "")
                email_dt = _parse_email_date(date_str)

                if not email_dt:
                    continue

                email_date = email_dt.strftime("%Y-%m-%d")
                arrival_date = (email_dt + timedelta(days=8)).strftime("%Y-%m-%d")

                # 첨부파일 검색 (BOR로 시작하는 엑셀)
                for part in msg.walk():
                    content_disposition = str(part.get("Content-Disposition", ""))
                    if "attachment" not in content_disposition:
                        continue

                    filename = _decode_header_value(part.get_filename() or "")
                    if not filename:
                        continue

                    # BOR로 시작하는 엑셀 파일만 (FTA 제외)
                    if not filename.upper().startswith("BOR"):
                        continue
                    if "FTA" in filename.upper():
                        continue
                    if not any(filename.lower().endswith(ext) for ext in [".xlsx", ".xls", ".csv"]):
                        continue

                    # BOR 번호 추출 (파일명에서)
                    bor_match = re.match(r'(BOR[-_]?\d+)', filename, re.IGNORECASE)
                    bor_number = bor_match.group(1).upper() if bor_match else filename.split(".")[0].upper()
                    # BOR-2601001 형식으로 통일
                    bor_number = bor_number.replace("_", "-")

                    # 엑셀 파싱 → 모델명 추출
                    file_data = part.get_payload(decode=True)
                    models = _parse_bor_excel(file_data, filename)

                    if models:
                        results.append({
                            "bor_number": bor_number,
                            "subject": subject[:200],
                            "email_date": email_date,
                            "shipping_date": email_date,
                            "arrival_date": arrival_date,
                            "filename": filename,
                            "models": models,
                            "model_count": len(models),
                        })
                        logger.info(f"[선적메일] {bor_number}: {len(models)}개 모델 (선적일: {email_date})")

            except Exception as e:
                logger.warning(f"[선적메일] 메일 파싱 실패 (uid={uid}): {e}")

        mail.logout()

    except imaplib.IMAP4.error as e:
        logger.error(f"[선적메일] IMAP 접속 실패: {e}")
        raise Exception(f"메일 접속 실패: {e}")
    except Exception as e:
        logger.error(f"[선적메일] 메일 스캔 오류: {e}", exc_info=True)
        raise

    # BOR 번호 기준 중복 제거 (최신 메일 우선)
    seen_bors = {}
    for r in sorted(results, key=lambda x: x["email_date"], reverse=True):
        bor = r["bor_number"]
        if bor not in seen_bors:
            seen_bors[bor] = r
    results = list(seen_bors.values())

    logger.info(f"[선적메일] 총 {len(results)}건 BOR 선적 정보 파싱 완료")
    return results


def _parse_bor_excel(file_data: bytes, filename: str) -> list:
    """BOR 엑셀 첨부파일에서 모델명 추출"""
    models = []
    try:
        if filename.lower().endswith(".xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
            sh = wb.active
            for r in range(1, sh.max_row + 1):
                for c in range(1, min(sh.max_column + 1, 10)):
                    val = str(sh.cell(r, c).value or "").strip()
                    # LS- 또는 LSP- 등 랜스타 모델명 패턴
                    if val and re.match(r'^(LS[- P]|LSN-|HT-)', val, re.IGNORECASE):
                        # 모델명만 추출 (쉼표 이전)
                        model = val.split(",")[0].strip()
                        if len(model) >= 4 and model not in models:
                            models.append(model)
        elif filename.lower().endswith(".xls"):
            import xlrd
            wb = xlrd.open_workbook(file_contents=file_data)
            sh = wb.sheet_by_index(0)
            for r in range(sh.nrows):
                for c in range(min(sh.ncols, 10)):
                    val = str(sh.cell_value(r, c)).strip()
                    if val and re.match(r'^(LS[- P]|LSN-|HT-)', val, re.IGNORECASE):
                        model = val.split(",")[0].strip()
                        if len(model) >= 4 and model not in models:
                            models.append(model)
    except Exception as e:
        logger.warning(f"[선적메일] 엑셀 파싱 실패 ({filename}): {e}")

    return models


# ─── DB 저장/조회 ─────────────────────────────────────────

def save_shipping_info(conn, shipping_data: list):
    """선적 정보를 DB에 저장"""
    now = datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S")
    saved = 0
    for item in shipping_data:
        bor = item["bor_number"]
        models_json = ",".join(item["models"])
        conn.execute("""
            INSERT INTO shipping_mail_info
                (bor_number, subject, email_date, shipping_date, arrival_date,
                 filename, models, model_count, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(bor_number) DO UPDATE SET
                subject=excluded.subject, email_date=excluded.email_date,
                shipping_date=excluded.shipping_date, arrival_date=excluded.arrival_date,
                filename=excluded.filename, models=excluded.models,
                model_count=excluded.model_count, updated_at=excluded.updated_at
        """, (bor, item["subject"], item["email_date"], item["shipping_date"],
              item["arrival_date"], item["filename"], models_json,
              item["model_count"], now))
        saved += 1
    conn.commit()
    logger.info(f"[선적메일] {saved}건 저장 완료")
    return saved


def get_shipping_info_map(conn) -> dict:
    """
    선적 정보를 모델명 기준으로 매핑
    BOR: 전체 모델이 같은 선적일
    NAM: 모델별로 다른 선적일 가능 (개별 레코드 우선)
    입고예정일이 오늘 이전인 제품은 재고 스냅샷으로 실제 입고 여부 2차 검증
    Returns: {"LS-1000H": {"bor": "...", "shipping_date": "...", "arrival_date": "...", "status": "shipping|arrived|delayed"}}
    """
    today = datetime.now(KST).strftime("%Y-%m-%d")
    today_fmt = today.replace("-", "")  # YYYYMMDD

    rows = conn.execute("""
        SELECT bor_number, shipping_date, arrival_date, models
        FROM shipping_mail_info
        ORDER BY email_date DESC
    """).fetchall()

    model_map = {}
    for r in rows:
        bor = r[0]
        ship_date = r[1] or ""
        arr_date = r[2] or ""
        models_str = r[3] or ""

        # 입고예정일이 지났으면 → 재고 스냅샷으로 실제 입고 여부 확인
        if arr_date and arr_date < today:
            # 해당 모델들의 입고 검증은 개별 모델 처리 시 수행
            pass

        # 개별 모델 레코드 (NAM: "LAN-PI20260304_LS-BDCH" 형태)
        if "_" in bor and models_str and "," not in models_str:
            m = models_str.strip().upper()
            if m and m not in model_map and ship_date:
                entry = _build_shipping_entry(conn, m, bor.split("_")[0], ship_date, arr_date, today, today_fmt)
                if entry:
                    model_map[m] = entry
            continue

        # BOR/PI 단위 레코드 → 소속 모델에 일괄 적용
        for m in models_str.split(","):
            m = m.strip().upper()
            if m and m not in model_map:
                entry = _build_shipping_entry(conn, m, bor, ship_date, arr_date, today, today_fmt)
                if entry:
                    model_map[m] = entry
    return model_map


def _build_shipping_entry(conn, model: str, bor: str, ship_date: str, arr_date: str,
                          today: str, today_fmt: str) -> dict:
    """
    선적 정보 엔트리 생성 + 2차 검증 (입고예정일 경과 시 재고 증가 확인)

    - 입고예정일 미도래 → status="shipping" (선적 중)
    - 입고예정일 경과 + 재고 증가 확인 → None (이미 입고 → 표시 안함)
    - 입고예정일 경과 + 재고 미증가 → status="delayed" (입고 지연)
    """
    if not arr_date:
        return {"bor_number": bor, "shipping_date": ship_date, "arrival_date": arr_date, "status": "shipping"}

    if arr_date >= today:
        # 아직 입고예정일 전
        return {"bor_number": bor, "shipping_date": ship_date, "arrival_date": arr_date, "status": "shipping"}

    # 입고예정일이 지남 → 재고 스냅샷으로 실제 입고 여부 확인
    # 입고예정일 전후 7일간의 재고 변화를 확인 (재고가 늘었으면 입고됨)
    try:
        arr_dt = datetime.strptime(arr_date, "%Y-%m-%d")
        check_before = (arr_dt - timedelta(days=3)).strftime("%Y%m%d")
        check_after = (arr_dt + timedelta(days=7)).strftime("%Y%m%d")

        # 해당 모델의 품목코드 찾기 (planning_targets에서)
        target_row = conn.execute(
            "SELECT prod_cd FROM inventory_planning_targets WHERE UPPER(model_name) = ?",
            (model,)
        ).fetchone()

        if target_row:
            prod_cd = target_row[0]
            # 입고예정일 전후 스냅샷 조회
            snapshots = conn.execute("""
                SELECT snapshot_date, bal_qty FROM inventory_snapshots
                WHERE prod_cd = ? AND snapshot_date BETWEEN ? AND ?
                ORDER BY snapshot_date ASC
            """, (prod_cd, check_before, check_after)).fetchall()

            if len(snapshots) >= 2:
                min_qty = min(s[1] for s in snapshots)
                max_qty = max(s[1] for s in snapshots)
                # 재고가 의미있게 증가했으면 입고 확인 (10개 이상 또는 20% 이상 증가)
                if max_qty > min_qty and (max_qty - min_qty >= 10 or
                                          (min_qty > 0 and (max_qty - min_qty) / min_qty >= 0.2)):
                    return None  # 입고 확인됨 → 표시 안함
    except Exception:
        pass

    # 입고예정일 지났는데 재고 증가 미확인 → 입고 지연
    return {"bor_number": bor, "shipping_date": ship_date, "arrival_date": arr_date, "status": "delayed"}


def get_all_shipping_info(conn) -> list:
    """전체 선적 정보 조회"""
    rows = conn.execute("""
        SELECT bor_number, subject, email_date, shipping_date, arrival_date,
               filename, models, model_count, updated_at
        FROM shipping_mail_info
        ORDER BY email_date DESC
    """).fetchall()
    return [dict(r) for r in rows]


# ═══════════════════════════════════════════════════════════
#  NAM 거래처 메일 파싱 (네이버 IMAP → 163.com 발신자)
# ═══════════════════════════════════════════════════════════

def scan_nam_shipping_emails(
    imap_server: str,
    imap_user: str,
    imap_password: str,
    sender_filter: str = "13428934642@163.com",
    imap_port: int = 993,
    days_back: int = 180,
) -> list:
    """
    네이버 메일에서 NAM 거래처(163.com) 발신 메일 검색
    첨부 엑셀의 첫 시트: B열=주문일, E열=모델명, K열=선적일
    입고예정 = 선적일 + 8일

    Returns: [{
        "source": "NAM",
        "pi_number": "LAN-PI20260304",
        "email_date": "2026-03-04",
        "filename": "nam-lanstar-pi20260304-01.xlsx",
        "items": [{"model": "LS-BDCH", "order_date": "2026-03-04", "shipping_date": "2026-03-20", "arrival_date": "2026-03-28", "qty": 200}, ...]
    }, ...]
    """
    results = []

    try:
        logger.info(f"[NAM메일] IMAP 접속: {imap_server}:{imap_port} ({imap_user})")
        mail = imaplib.IMAP4_SSL(imap_server, imap_port)
        mail.login(imap_user, imap_password)

        since_date = (datetime.now(KST) - timedelta(days=days_back)).strftime("%d-%b-%Y")

        # 모든 폴더에서 검색 (네이버는 하위 폴더에 분류됨)
        # IMAP SEARCH FROM이 일부 폴더에서 작동 안 하므로 직접 fetch + 헤더 확인
        status, folder_list = mail.list()
        folders_to_search = []
        for f in (folder_list or []):
            try:
                decoded = f.decode("utf-8", errors="replace")
                fname = decoded.rsplit('"', 2)[-2] if '"' in decoded else "INBOX"
                folders_to_search.append(fname)
            except Exception:
                pass
        if not folders_to_search:
            folders_to_search = ["INBOX"]

        all_uids_by_folder = []
        for folder in folders_to_search:
            try:
                st, _ = mail.select('"' + folder + '"', readonly=True)
                if st != "OK":
                    continue

                since_date = (datetime.now(KST) - timedelta(days=days_back)).strftime("%d-%b-%Y")
                # 먼저 SEARCH 시도
                try:
                    status, data = mail.search(None, f'(SINCE {since_date} FROM "{sender_filter}")')
                    uids = data[0].split() if status == "OK" and data[0] else []
                except Exception:
                    uids = []

                # SEARCH 결과 없으면 → 전체 fetch 후 FROM 헤더 직접 확인
                if not uids:
                    try:
                        status, data = mail.search(None, f"(SINCE {since_date})")
                        all_folder_uids = data[0].split() if status == "OK" and data[0] else []
                    except Exception:
                        status, data = mail.search(None, "ALL")
                        all_folder_uids = data[0].split() if status == "OK" and data[0] else []

                    for uid in all_folder_uids:
                        try:
                            st2, md2 = mail.fetch(uid, "(BODY.PEEK[HEADER.FIELDS (FROM)])")
                            from_hdr = md2[0][1].decode("utf-8", errors="replace").lower()
                            if sender_filter.lower() in from_hdr:
                                uids.append(uid)
                        except Exception:
                            pass

                if uids:
                    all_uids_by_folder.extend([(folder, uid) for uid in uids])
                    logger.info(f"[NAM메일] 폴더 '{folder}': {len(uids)}건")
            except Exception as e:
                logger.debug(f"[NAM메일] 폴더 '{folder}' 검색 실패: {e}")

        logger.info(f"[NAM메일] {sender_filter} 총 {len(all_uids_by_folder)}건 발견")

        for folder, uid in all_uids_by_folder:
            try:
                mail.select(f'"{folder}"', readonly=True)
                status, msg_data = mail.fetch(uid, "(RFC822)")
                if status != "OK":
                    continue

                msg = email.message_from_bytes(msg_data[0][1])
                subject = _decode_header_value(msg.get("Subject", ""))
                email_dt = _parse_email_date(msg.get("Date", ""))
                if not email_dt:
                    continue
                email_date = email_dt.strftime("%Y-%m-%d")

                # 엑셀 첨부파일 검색
                for part in msg.walk():
                    content_disposition = str(part.get("Content-Disposition", ""))
                    if "attachment" not in content_disposition:
                        continue

                    filename = _decode_header_value(part.get_filename() or "")
                    if not filename:
                        continue
                    if not any(filename.lower().endswith(ext) for ext in [".xlsx", ".xls"]):
                        continue

                    file_data = part.get_payload(decode=True)
                    items = _parse_nam_excel(file_data, filename)

                    # PI 번호 추출
                    pi_match = re.search(r'(LAN-PI\d+)', filename, re.IGNORECASE)
                    pi_number = pi_match.group(1).upper() if pi_match else filename.split(".")[0]

                    if items:
                        results.append({
                            "source": "NAM",
                            "pi_number": pi_number,
                            "subject": subject[:200],
                            "email_date": email_date,
                            "filename": filename,
                            "items": items,
                        })
                        logger.info(f"[NAM메일] {pi_number}: {len(items)}개 품목 파싱")

            except Exception as e:
                logger.warning(f"[NAM메일] 메일 파싱 실패 (uid={uid}): {e}")

        mail.logout()

    except Exception as e:
        logger.error(f"[NAM메일] 메일 스캔 오류: {e}", exc_info=True)
        raise

    return results


def _parse_nam_excel(file_data: bytes, filename: str) -> list:
    """
    NAM 거래처 PI 엑셀 파싱
    첫 시트: B열=주문일, E열=모델명, G열=수량, K열=선적일
    Stickers, color box, Battery 등 부자재 제외
    """
    items = []
    skip_keywords = {"stickers", "sticker", "color box", "logo", "battery", "box", "manual", "none", ""}

    try:
        if filename.lower().endswith(".xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
            sh = wb[wb.sheetnames[0]]  # 첫 시트

            for r in range(2, sh.max_row + 1):
                model = str(sh.cell(r, 5).value or "").strip()  # E열
                if not model or model.lower() in skip_keywords:
                    continue
                if model.startswith("型号") or model == "Model":
                    continue

                order_date_raw = sh.cell(r, 2).value  # B열
                shipping_date_raw = sh.cell(r, 11).value  # K열
                qty_raw = sh.cell(r, 7).value  # G열

                order_date = _parse_date_value(order_date_raw)
                shipping_date = _parse_date_value(shipping_date_raw)
                qty = int(float(qty_raw)) if qty_raw and str(qty_raw).replace(".", "").isdigit() else 0

                # 선적일이 없으면 상위 행에서 상속
                if not shipping_date and items:
                    shipping_date = items[-1].get("shipping_date", "")

                # 주문일이 없으면 상위 행에서 상속
                if not order_date and items:
                    order_date = items[-1].get("order_date", "")

                arrival_date = ""
                if shipping_date:
                    try:
                        sd = datetime.strptime(shipping_date, "%Y-%m-%d")
                        arrival_date = (sd + timedelta(days=8)).strftime("%Y-%m-%d")
                    except Exception:
                        pass

                items.append({
                    "model": model,
                    "order_date": order_date or "",
                    "shipping_date": shipping_date or "",
                    "arrival_date": arrival_date,
                    "qty": qty,
                })

    except Exception as e:
        logger.warning(f"[NAM메일] 엑셀 파싱 실패 ({filename}): {e}")

    return items


def _parse_date_value(val) -> str:
    """다양한 날짜 형식을 YYYY-MM-DD로 변환"""
    if not val:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    # "출고완료" 등 텍스트 포함 시 날짜 부분만 추출
    date_match = re.search(r'(\d{4}[-/]\d{1,2}[-/]\d{1,2})', s)
    if date_match:
        d = date_match.group(1).replace("/", "-")
        parts = d.split("-")
        return f"{parts[0]}-{int(parts[1]):02d}-{int(parts[2]):02d}"
    return ""


# ─── NAM 선적정보 DB 저장 ─────────────────────────────────

def save_nam_shipping_info(conn, scan_results: list):
    """NAM 거래처 선적 정보를 shipping_mail_info + orderlist_items 테이블에 저장"""
    now = datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S")
    saved = 0

    for result in scan_results:
        pi = result["pi_number"]
        items = result.get("items", [])
        models_list = [it["model"] for it in items if it.get("model")]
        models_csv = ",".join(models_list)

        # 1) shipping_mail_info에 PI 단위 저장
        conn.execute("""
            INSERT INTO shipping_mail_info
                (bor_number, subject, email_date, shipping_date, arrival_date,
                 filename, models, model_count, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(bor_number) DO UPDATE SET
                subject=excluded.subject, email_date=excluded.email_date,
                shipping_date=excluded.shipping_date, arrival_date=excluded.arrival_date,
                filename=excluded.filename, models=excluded.models,
                model_count=excluded.model_count, updated_at=excluded.updated_at
        """, (
            pi, result.get("subject", ""), result["email_date"],
            items[0]["shipping_date"] if items else "",
            items[0]["arrival_date"] if items else "",
            result["filename"], models_csv, len(models_list), now
        ))
        saved += 1

        # 2) shipping_mail_info에 개별 모델 레코드 저장 (모델별 선적일)
        for it in items:
            model = it["model"].strip().upper()
            if not model:
                continue
            key = f"{pi}_{model}"
            conn.execute("""
                INSERT INTO shipping_mail_info
                    (bor_number, subject, email_date, shipping_date, arrival_date,
                     filename, models, model_count, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(bor_number) DO UPDATE SET
                    shipping_date=excluded.shipping_date, arrival_date=excluded.arrival_date,
                    updated_at=excluded.updated_at
            """, (
                key, f"NAM:{pi}", result["email_date"],
                it.get("shipping_date", ""), it.get("arrival_date", ""),
                result["filename"], model, 1, now
            ))

        # 3) orderlist_items에 오더 내역 자동 등록 (기존 오더리스트와 동일 구조)
        year = result["email_date"][:4] if result["email_date"] else "2026"
        tab_name = f"NAM-{year}"

        # 해당 PI의 기존 데이터 삭제 후 재등록 (중복 방지)
        conn.execute(
            "DELETE FROM orderlist_items WHERE sheet_tab = ? AND order_no = ?",
            (tab_name, pi)
        )

        current_category = ""
        for idx, it in enumerate(items):
            model = it["model"].strip()
            if not model:
                continue

            conn.execute("""
                INSERT INTO orderlist_items
                    (sheet_tab, order_no, seller, order_date, category,
                     model_name, description, qty, unit, unit_price,
                     total_value, row_index, raw_row, synced_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                tab_name, pi, "NAM/PUSIMA",
                it.get("order_date", ""), current_category,
                model, "", it.get("qty", 0), "PCS",
                "", "", idx + 1, "", now
            ))

        # 동기화 로그
        conn.execute(
            "INSERT INTO orderlist_sync_log(sheet_tab, item_count, synced_at) VALUES(?,?,?)",
            (tab_name, len(items), now)
        )

    conn.commit()
    logger.info(f"[NAM메일] {saved}건 PI 저장 + orderlist_items 등록 완료")
    return saved


# ═══════════════════════════════════════════════════════════
#  구글시트 오더리스트 자동 기록 (서비스 계정)
# ═══════════════════════════════════════════════════════════

ORDERLIST_SHEET_ID = "1ej0cxyM3NHJKpF-KBXbZ16fH-lZcUrr3Z3eTwFVFSco"

def _get_sheets_credentials():
    """Google 서비스 계정으로 Sheets API 인증"""
    import json
    from google.oauth2.service_account import Credentials

    sa_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "")
    if not sa_json:
        return None

    try:
        sa_info = json.loads(sa_json)
        creds = Credentials.from_service_account_info(
            sa_info,
            scopes=["https://www.googleapis.com/auth/spreadsheets"]
        )
        return creds
    except Exception as e:
        logger.error(f"[구글시트] 서비스 계정 인증 실패: {e}")
        return None


def _sheets_api_request(method, url, body=None):
    """Google Sheets API 호출 (서비스 계정 인증)"""
    import httpx

    creds = _get_sheets_credentials()
    if not creds:
        raise Exception("GOOGLE_SERVICE_ACCOUNT_JSON 환경변수 없음")

    # 토큰 갱신
    from google.auth.transport.requests import Request
    creds.refresh(Request())

    headers = {
        "Authorization": f"Bearer {creds.token}",
        "Content-Type": "application/json",
    }

    if method == "GET":
        resp = httpx.get(url, headers=headers, timeout=30)
    elif method == "POST":
        resp = httpx.post(url, headers=headers, json=body, timeout=30)
    elif method == "PUT":
        resp = httpx.put(url, headers=headers, json=body, timeout=30)
    else:
        raise ValueError(f"Unsupported method: {method}")

    resp.raise_for_status()
    return resp.json()


def write_nam_orders_to_sheet(scan_results: list) -> dict:
    """
    NAM 거래처 오더를 구글시트 오더리스트에 기록
    기존 BOR 탭과 동일한 패턴으로 새 탭(NAM-2026) 또는 기존 탭에 추가
    """
    if not scan_results:
        return {"status": "no_data"}

    base_url = f"https://sheets.googleapis.com/v4/spreadsheets/{ORDERLIST_SHEET_ID}"
    year = datetime.now(KST).strftime("%Y")
    tab_name = f"NAM-{year}"

    try:
        # 1. 탭 존재 여부 확인
        sheet_meta = _sheets_api_request("GET", f"{base_url}?fields=sheets.properties")
        existing_tabs = [s["properties"]["title"] for s in sheet_meta.get("sheets", [])]

        if tab_name not in existing_tabs:
            # 새 탭 생성
            _sheets_api_request("POST", f"{base_url}:batchUpdate", {
                "requests": [{"addSheet": {"properties": {"title": tab_name}}}]
            })
            logger.info(f"[구글시트] '{tab_name}' 탭 생성")

        # 2. 기존 데이터 클리어
        _sheets_api_request("POST",
            f"{base_url}/values/'{tab_name}'!A1:Z1000:clear", {})

        # 3. 데이터 구성 (BOR 오더리스트와 유사한 패턴)
        rows = []
        for result in scan_results:
            pi = result["pi_number"]
            items = result.get("items", [])
            email_date = result.get("email_date", "")

            # 헤더 행
            rows.append(["Seller:", "", "No.:", pi])
            rows.append(["NAM/PUSIMA (深圳普思玛)", "", "Date:", email_date])
            rows.append(["Item", "Description", "Quantity", "Unit",
                        "Order Date", "Shipping Date", "Arrival Date"])

            # 품목 행
            for it in items:
                model = it.get("model", "")
                rows.append([
                    model,
                    "",
                    it.get("qty", 0),
                    "PCS",
                    it.get("order_date", ""),
                    it.get("shipping_date", ""),
                    it.get("arrival_date", ""),
                ])

            rows.append([])  # 빈 행 구분

        # 4. 시트에 쓰기
        _sheets_api_request("PUT",
            f"{base_url}/values/'{tab_name}'!A1?valueInputOption=USER_ENTERED",
            {"values": rows}
        )

        logger.info(f"[구글시트] '{tab_name}'에 {len(scan_results)}건 PI 기록 완료")
        return {"status": "ok", "tab": tab_name, "rows": len(rows)}

    except Exception as e:
        logger.error(f"[구글시트] 오더리스트 기록 실패: {e}")
        return {"status": "error", "error": str(e)}


# ═══════════════════════════════════════════════════════════
#  BOR 오더리스트 자동 최신화 (Ecount 메일 → 구글시트 덮어쓰기)
# ═══════════════════════════════════════════════════════════

def scan_bor_orderlist_emails(
    imap_server: str, imap_user: str, imap_password: str,
    imap_port: int = 993, days_back: int = 90,
    sender_filter: str = "guzhiyi@bor-cable.com",
) -> list:
    """
    kyu@lanstar.co.kr 메일에서 BOR 거래처(guzhiyi@bor-cable.com) 발신 메일 검색
    'rest' 키워드가 포함된 엑셀 첨부파일 = 최신 오더리스트
    가장 최근 메일의 첨부파일만 반환 (덮어쓰기용)
    """
    results = []

    try:
        logger.info(f"[BOR오더] IMAP 접속: {imap_server} ({imap_user})")
        mail = imaplib.IMAP4_SSL(imap_server, imap_port)
        mail.login(imap_user, imap_password)
        mail.select("INBOX", readonly=True)

        # Ecount 서버는 검색 미지원 → 전체 메일 가져와서 Python에서 필터
        status, data = mail.search(None, "ALL")
        all_uids = data[0].split() if status == "OK" and data[0] else []
        logger.info(f"[BOR오더] 전체 메일 {len(all_uids)}건, 발신자 필터: {sender_filter}")

        for uid in sorted(all_uids, reverse=True):  # 최신 먼저
            try:
                # Ecount IMAP 호환: 여러 FETCH 형식 시도
                msg_data = None
                for fetch_cmd in ["(RFC822)", "RFC822", "(BODY[])"]:
                    try:
                        status, msg_data = mail.fetch(uid, fetch_cmd)
                        if status == "OK" and msg_data and msg_data[0]:
                            break
                    except Exception:
                        continue
                if not msg_data or not msg_data[0]:
                    continue

                raw = msg_data[0][1] if isinstance(msg_data[0], tuple) else msg_data[0]
                if not isinstance(raw, bytes):
                    continue

                msg = email.message_from_bytes(raw)

                # Python에서 발신자 필터 (Ecount IMAP 검색 미지원)
                msg_from = _decode_header_value(msg.get("From", "")).lower()
                if sender_filter and sender_filter.lower() not in msg_from:
                    continue

                subject = _decode_header_value(msg.get("Subject", ""))
                email_dt = _parse_email_date(msg.get("Date", ""))
                if not email_dt:
                    continue

                for part in msg.walk():
                    content_disposition = str(part.get("Content-Disposition", ""))
                    if "attachment" not in content_disposition:
                        continue

                    filename = _decode_header_value(part.get_filename() or "")
                    if not filename:
                        continue

                    # "rest" 키워드가 있는 엑셀만 (FTA 제외)
                    fn_upper = filename.upper()
                    if "REST" not in fn_upper:
                        continue
                    if "FTA" in fn_upper:
                        continue
                    if not any(filename.lower().endswith(ext) for ext in [".xlsx", ".xls"]):
                        continue

                    file_data = part.get_payload(decode=True)

                    # BOR 번호 추출
                    bor_match = re.search(r'(BOR[-_]?\d+)', filename, re.IGNORECASE)
                    bor_number = bor_match.group(1).upper().replace("_", "-") if bor_match else ""

                    results.append({
                        "filename": filename,
                        "bor_number": bor_number,
                        "subject": subject[:200],
                        "email_date": email_dt.strftime("%Y-%m-%d"),
                        "file_data": file_data,
                    })
                    logger.info(f"[BOR오더] REST 파일 발견: {filename} ({email_dt.strftime('%Y-%m-%d')})")

            except Exception as e:
                logger.warning(f"[BOR오더] 메일 파싱 실패: {e}")

        mail.logout()

    except Exception as e:
        logger.error(f"[BOR오더] 메일 스캔 오류: {e}", exc_info=True)
        raise

    return results


def _parse_bor_rest_excel(file_data: bytes, filename: str) -> list:
    """BOR REST 엑셀을 구글시트에 쓸 수 있는 2D 배열로 변환"""
    rows = []
    try:
        if filename.lower().endswith(".xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
            sh = wb.active
            for r in range(1, sh.max_row + 1):
                row = []
                for c in range(1, sh.max_column + 1):
                    val = sh.cell(r, c).value
                    if val is None:
                        row.append("")
                    elif isinstance(val, datetime):
                        row.append(val.strftime("%Y-%m-%d"))
                    elif isinstance(val, (int, float)):
                        if val == int(val):
                            row.append(int(val))
                        else:
                            row.append(round(val, 4))
                    else:
                        row.append(str(val))
                rows.append(row)
        elif filename.lower().endswith(".xls"):
            import xlrd
            wb = xlrd.open_workbook(file_contents=file_data)
            sh = wb.sheet_by_index(0)
            for r in range(sh.nrows):
                row = []
                for c in range(sh.ncols):
                    val = sh.cell_value(r, c)
                    if isinstance(val, float) and val == int(val):
                        val = int(val)
                    row.append(val if val else "")
                rows.append(row)
    except Exception as e:
        logger.error(f"[BOR오더] 엑셀 파싱 실패 ({filename}): {e}")
    return rows


def sync_bor_orderlist_to_sheet(scan_results: list) -> dict:
    """
    BOR REST 엑셀 내용을 구글시트 오더리스트에 덮어쓰기
    파일명에서 BOR 번호 → 해당 탭에 덮어쓰기
    탭이 없으면 새로 생성
    """
    if not scan_results:
        return {"status": "no_data"}

    base_url = f"https://sheets.googleapis.com/v4/spreadsheets/{ORDERLIST_SHEET_ID}"
    synced_tabs = []

    try:
        # 기존 탭 목록
        sheet_meta = _sheets_api_request("GET", f"{base_url}?fields=sheets.properties")
        existing_tabs = [s["properties"]["title"] for s in sheet_meta.get("sheets", [])]

        for item in scan_results:
            filename = item["filename"]
            file_data = item["file_data"]

            # 탭 이름 결정: BOR 번호 기반 또는 연도
            bor = item.get("bor_number", "")
            # 기존 탭에서 BOR 번호로 시작하는 탭 찾기
            target_tab = None
            for tab in existing_tabs:
                if bor and bor in tab.upper():
                    target_tab = tab
                    break

            # 기존 연도별 탭에 덮어쓰기 (2026, 2025 등)
            if not target_tab:
                year = item["email_date"][:4]
                for tab in existing_tabs:
                    if tab.strip() == year:
                        target_tab = tab
                        break

            if not target_tab:
                target_tab = f"BOR-{item['email_date'][:4]}"
                if target_tab not in existing_tabs:
                    _sheets_api_request("POST", f"{base_url}:batchUpdate", {
                        "requests": [{"addSheet": {"properties": {"title": target_tab}}}]
                    })
                    existing_tabs.append(target_tab)

            # 엑셀 → 2D 배열
            rows = _parse_bor_rest_excel(file_data, filename)
            if not rows:
                continue

            # 기존 데이터 클리어 후 덮어쓰기
            try:
                _sheets_api_request("POST",
                    f"{base_url}/values/'{target_tab}'!A1:Z1000:clear", {})
            except Exception:
                pass

            _sheets_api_request("PUT",
                f"{base_url}/values/'{target_tab}'!A1?valueInputOption=USER_ENTERED",
                {"values": rows}
            )

            synced_tabs.append({"tab": target_tab, "rows": len(rows), "file": filename})
            logger.info(f"[BOR오더] '{target_tab}' 탭 덮어쓰기 완료 ({len(rows)}행)")

        return {"status": "ok", "tabs": synced_tabs}

    except Exception as e:
        logger.error(f"[BOR오더] 구글시트 동기화 실패: {e}")
        return {"status": "error", "error": str(e)}


# ─── 스캔 이력 저장/조회 ────────────────────────────────

def save_scan_log(conn, scan_type: str, result_summary: str, email_dates: str = ""):
    """스캔/최신화 실행 이력 저장"""
    now = datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S")
    conn.execute("""
        INSERT INTO shipping_scan_log (scan_type, executed_at, result_summary, email_dates)
        VALUES (?, ?, ?, ?)
    """, (scan_type, now, result_summary, email_dates))
    conn.commit()


def get_last_scan_info(conn) -> dict:
    """마지막 스캔/최신화 정보 조회"""
    result = {}
    for scan_type in ["shipping_scan", "orderlist_sync"]:
        row = conn.execute("""
            SELECT executed_at, result_summary, email_dates
            FROM shipping_scan_log
            WHERE scan_type = ?
            ORDER BY executed_at DESC LIMIT 1
        """, (scan_type,)).fetchone()
        if row:
            result[scan_type] = {
                "executed_at": row[0],
                "result_summary": row[1],
                "email_dates": row[2],
            }
    return result
