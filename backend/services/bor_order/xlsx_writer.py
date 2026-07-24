"""
중국오더(BOR) 모듈 - new order xlsx 생성

기존 발주 파일과 동일 구조:
  시트명 'new order', 헤더 NO / SPECIFICATION / MODEL NAME / QT'Y / REMARK
  SPECIFICATION은 '품목명 [품목코드]' 형식 (가능한 경우)
파일명: new order-YYYY-MMDD.xlsx
"""
import io
import logging

logger = logging.getLogger(__name__)

HEADERS = ["NO", "SPECIFICATION", "MODEL NAME", "QT'Y", "REMARK"]


def build_order_xlsx(lines: list, order_date: str) -> tuple:
    """발주 xlsx 생성 → (filename, bytes)

    lines: [{model, spec_text, qty_final, flag?, remark?}, ...] (발주 대상만)
    order_date: 'YYYY-MMDD' (또는 'YYYY-MM-DD' — 자동 변환)
    """
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    # 날짜 표기 정규화: 2026-07-24 → 2026-0724
    date_part = str(order_date).strip()
    if len(date_part) == 10 and date_part.count("-") == 2:
        y, m, d = date_part.split("-")
        date_part = f"{y}-{m}{d}"
    filename = f"new order-{date_part}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "new order"

    head_font = Font(name="Malgun Gothic", bold=True)
    body_font = Font(name="Malgun Gothic", size=10)
    head_fill = PatternFill("solid", fgColor="D9E1F2")
    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    # 헤더
    for j, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, j, h)
        cell.font = head_font
        cell.fill = head_fill
        cell.border = border
        cell.alignment = center

    # 품목 라인
    for i, ln in enumerate(lines, 1):
        row = i + 1
        spec = ln.get("spec_text") or ln.get("spec") or ""
        remark = ln.get("remark") or ""
        qty = ln.get("qty_final")
        if qty is None:
            qty = ln.get("qty", 0)
        qty = float(qty or 0)
        qty_out = int(qty) if qty == int(qty) else qty

        ws.cell(row, 1, i)
        ws.cell(row, 2, spec)
        ws.cell(row, 3, str(ln.get("model", "")).strip())
        ws.cell(row, 4, qty_out)
        ws.cell(row, 5, remark)
        for j in range(1, 6):
            cell = ws.cell(row, j)
            cell.font = body_font
            cell.border = border
        ws.cell(row, 1).alignment = center
        ws.cell(row, 4).number_format = "#,##0"

    # 컬럼 폭 (기존 발주 파일과 유사하게)
    for col, width in (("A", 6), ("B", 62), ("C", 24), ("D", 10), ("E", 14)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    data = output.read()
    logger.info(f"[BOR/xlsx] {filename} 생성 — {len(lines)}라인, {len(data):,} bytes")
    return filename, data
