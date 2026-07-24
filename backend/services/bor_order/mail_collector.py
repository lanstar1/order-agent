"""
중국오더(BOR) 모듈 - IMAP 메일 수집기

★ 이 서버(wmbox3.ecount.com)는 일반 SEARCH/FETCH를 거부 → 반드시 UID 명령 사용.
   M.uid('search', None, 'FROM', '"x@y.com"', 'SINCE', '01-Jan-2026')
   M.uid('fetch', uid_str, '(RFC822)')

수집 대상:
1. collect_restlist   : kngtiger@naver.com / guzhiyi@bor-cable.com 발신 메일 중
                        파일명에 'rest' 포함 xlsx → 가장 최신 메일 1건 파싱
                        → bor_restlist_lines 교체 저장 (이전 스냅샷 삭제)
2. collect_recent_orders:
   ① INBOX FROM kngtiger, 제목 'new order' 포함 + xlsx 첨부 → new order xlsx 파싱
   ② Sent 폴더 TO guzhiyi, 제목 'new order' → 본문 텍스트 정규식
      패턴: MODEL / N pcs (예: "LS-USB-AMAM-5M / 2,000pcs")
"""
import email
import html as html_lib
import imaplib
import io
import logging
import re
from datetime import datetime, timedelta, timezone
from email.header import decode_header
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from db.database import now_kst
from config import MAIL_IMAP_SERVER, MAIL_IMAP_PORT, MAIL_USER, MAIL_PASSWORD
from services.bor_order import db as bor_db
from services.bor_order.engine import normalize_model

logger = logging.getLogger(__name__)
KST = timezone(timedelta(hours=9))

TIGER_SENDER = "kngtiger@naver.com"      # 정정섭 (new order xlsx)
BOR_SENDER = "guzhiyi@bor-cable.com"     # BOR Mr. Gu (rest list 회신)
SENT_FOLDER = "Sent"                     # kyu 계정 보낸편지함 (수시 텍스트 오더)

# kyu 텍스트 오더 패턴: "LS-USB-AMAM-5M / 2,000pcs" (여러 줄 가능, 콤마 포함 숫자)
KYU_ORDER_RE = re.compile(
    r"(?:^|\n)\s*(LS[NP]?-\S+|ZOT\S*)\s*/\s*([\d,]+)\s*pcs",
    re.IGNORECASE,
)


# ─── 공통 유틸 ───────────────────────────────────────────
def _decode_header_value(value):
    if not value:
        return ""
    decoded = decode_header(value)
    parts = []
    for part, charset in decoded:
        if isinstance(part, bytes):
            parts.append(part.decode(charset or "utf-8", errors="replace"))
        else:
            parts.append(str(part))
    return " ".join(parts)


def _parse_email_date(date_str):
    if not date_str:
        return None
    try:
        from email.utils import parsedate_to_datetime
        dt = parsedate_to_datetime(date_str)
        return dt.astimezone(KST)
    except Exception:
        return None


def _html_to_text(html: str) -> str:
    """HTML 본문 → 텍스트 (kyu 발신 메일은 text/plain 파트가 없음)"""
    text = re.sub(r"(?is)<(script|style)[^>]*>.*?</\1>", " ", html)
    text = re.sub(r"(?i)<br\s*/?>", "\n", text)
    text = re.sub(r"(?i)</(p|div|tr|li|h[1-6])>", "\n", text)
    text = re.sub(r"<[^>]+>", " ", text)
    text = html_lib.unescape(text)
    # NBSP 등 정리
    text = text.replace("\xa0", " ")
    return "\n".join(line.strip() for line in text.splitlines())


def _get_mail_body_text(msg) -> str:
    """text/plain 우선, 없으면 text/html → 텍스트 변환"""
    plain, html = "", ""
    for part in msg.walk():
        if part.get_content_disposition() == "attachment":
            continue
        ctype = part.get_content_type()
        if ctype not in ("text/plain", "text/html"):
            continue
        try:
            charset = part.get_content_charset() or "utf-8"
            payload = part.get_payload(decode=True)
            if payload is None:
                continue
            decoded = payload.decode(charset, errors="replace")
        except Exception:
            continue
        if ctype == "text/plain" and not plain:
            plain = decoded
        elif ctype == "text/html" and not html:
            html = decoded
    if plain:
        return plain
    if html:
        return _html_to_text(html)
    return ""


def _to_float(val) -> float:
    """수량/단가 셀 안전 변환 — '600\\', '2,000' 등 문자열 오염 대응"""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group(0)) if m else 0.0


def _connect() -> imaplib.IMAP4_SSL:
    if not MAIL_USER or not MAIL_PASSWORD:
        raise RuntimeError("[BOR/메일] MAIL_USER/MAIL_PASSWORD 미설정")
    M = imaplib.IMAP4_SSL(MAIL_IMAP_SERVER, MAIL_IMAP_PORT)
    M.login(MAIL_USER, MAIL_PASSWORD)
    return M


def _uid_search(M, *criteria) -> list:
    """UID SEARCH → uid 문자열 리스트"""
    status, data = M.uid("search", None, *criteria)
    if status != "OK" or not data or not data[0]:
        return []
    return [u.decode() if isinstance(u, bytes) else str(u) for u in data[0].split()]


def _uid_fetch_message(M, uid: str):
    """UID FETCH RFC822 → email.message.Message (실패 시 None)"""
    status, msg_data = M.uid("fetch", uid, "(RFC822)")
    if status != "OK":
        return None
    for part in msg_data:
        if isinstance(part, tuple) and len(part) >= 2 and part[1]:
            return email.message_from_bytes(part[1])
    return None


def _iter_xlsx_attachments(msg):
    """(filename, bytes) 반복자 — xlsx 첨부만"""
    for part in msg.walk():
        if part.get_content_disposition() != "attachment":
            continue
        filename = _decode_header_value(part.get_filename() or "")
        if not filename.lower().endswith(".xlsx"):
            continue
        data = part.get_payload(decode=True)
        if data:
            yield filename, data


# ─── xlsx 파서: new order ────────────────────────────────
def parse_new_order_xlsx(file_data: bytes, source_file: str = "") -> list:
    """new order xlsx 파싱 → [{model, model_norm, qty, spec, remark}]

    구조: 헤더 1행(간혹 2행 — 1행이 회사명/빈행인 경우),
          A=NO, B=SPECIFICATION, C=MODEL NAME(MODEL/MODEL MANE 변형), D=QT'Y(QTY), E=REMARK
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
    # 첫 번째로 헤더가 발견되는 시트 사용 (보통 'new'/'NEW'/'NEW ORDER')
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 헤더 행 탐색 (1~5행): A='NO' && C에 MODEL 계열 표기
        header_row = 0
        for r in range(1, 6):
            a = str(ws.cell(r, 1).value or "").strip().upper()
            c = str(ws.cell(r, 3).value or "").strip().upper()
            if a == "NO" and ("MODEL" in c or "MANE" in c):
                header_row = r
                break
        if not header_row:
            continue

        items = []
        for r in range(header_row + 1, ws.max_row + 1):
            model_raw = str(ws.cell(r, 3).value or "").strip()
            if not model_raw:
                continue
            qty = _to_float(ws.cell(r, 4).value)
            if qty <= 0:
                continue
            items.append({
                "model": model_raw,
                "model_norm": normalize_model(model_raw),
                "qty": qty,
                "spec": str(ws.cell(r, 2).value or "").strip(),
                "remark": str(ws.cell(r, 5).value or "").strip(),
            })
        if items:
            return items
    logger.warning(f"[BOR/메일] new order 헤더 미발견: {source_file}")
    return []


# ─── xlsx 파서: marked Rest List ─────────────────────────
def _parse_po_date(val) -> str:
    """'July 14, 2026' / datetime → 'YYYY-MM-DD' (ISO, 문자열 비교 가능)"""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%B %d, %Y", "%b %d, %Y", "%Y-%m-%d", "%Y.%m.%d", "%d-%b-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def parse_rest_list_xlsx(file_data: bytes, source_file: str = "") -> list:
    """marked Rest List xlsx 파싱 → 원시 라인 리스트 (중복 미합산)

    구조: 연도 시트(예: '2026')에 PO 블록 세로 적층 ('Stock' 시트 무시)
      - A='Seller:' 행 → D=PO번호(BOR-XXXXXXX), 다음 행 D=PO날짜
      - 헤더행(Item/Description/Quantity...) 후 품목행:
        A=모델명+설명(콤마구분) 또는 A=모델명/B=설명,
        C=미선적잔량, F=단가, I=긴급마킹(M/VR), J=카톤입수, K=카톤수
      - 카테고리행: A 비고 B에 카테고리명 → 스킵
    Returns: [{po_number, po_date, model, model_norm, description,
               rest_qty, unit_price, urgency, sheet, row}]
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
    lines = []
    for sheet_name in wb.sheetnames:
        if not re.match(r"^\d{4}$", str(sheet_name).strip()):
            continue  # 'Stock' 등 연도 아닌 시트 무시
        ws = wb[sheet_name]
        po_number, po_date = "", ""

        for r in range(1, ws.max_row + 1):
            a_val = ws.cell(r, 1).value
            b_val = ws.cell(r, 2).value
            a_str = str(a_val or "").strip()

            # PO 블록 시작: A='Seller:' → D=PO번호, 다음 행 D=PO날짜
            if a_str.startswith("Seller"):
                po_number = str(ws.cell(r, 4).value or "").strip()
                po_date = _parse_po_date(ws.cell(r + 1, 4).value)
                continue
            # 헤더/카테고리/메타 행 스킵 (카테고리행은 A 비어있음)
            if not a_str:
                continue
            if a_str.upper().startswith(("ITEM", "TOTAL", "XIANGSHAN", "DATE", "SELLER")):
                continue
            qty = _to_float(ws.cell(r, 3).value)
            if qty <= 0:
                continue

            # A=모델+설명(콤마) 또는 A=모델 / B=설명
            b_str = str(b_val or "").strip()
            if b_str:
                model_raw, description = a_str, b_str
            elif "," in a_str:
                model_raw, description = a_str.split(",", 1)
                model_raw, description = model_raw.strip(), description.strip()
            else:
                model_raw, description = a_str, ""

            # 모델코드 형태 검증 — 공백 포함/대시 없는 텍스트는 모델 미상 품목
            # (예: 'F/UTP Cat.6 Patch Cord, ...' 커스텀 품목 → model='' + 전체를 설명으로)
            token = model_raw.strip()
            if " " in token or ("-" not in token and not token.upper().startswith("ZOT")):
                model_raw, description = "", a_str

            model_norm = normalize_model(model_raw)

            lines.append({
                "po_number": po_number,
                "po_date": po_date,
                "model": model_raw.strip(),
                "model_norm": model_norm,
                "description": description,
                "rest_qty": qty,
                "unit_price": _to_float(ws.cell(r, 6).value),
                "urgency": str(ws.cell(r, 9).value or "").strip(),
                "sheet": str(sheet_name),
                "row": r,
            })
    return lines


# ─── kyu 텍스트 오더 파서 ────────────────────────────────
def parse_kyu_order_text(text: str) -> list:
    """본문 텍스트에서 'MODEL / N pcs' 패턴 추출 → [(model, qty)]"""
    orders = []
    for m in KYU_ORDER_RE.finditer(text or ""):
        model = m.group(1).strip().rstrip(".,;:")
        qty = _to_float(m.group(2))
        if model and qty > 0:
            orders.append((model, qty))
    return orders


# ─── 수집 1: marked Rest List ────────────────────────────
def collect_restlist(days_back: int = 60) -> dict:
    """가장 최신 rest list 메일 1건 파싱 → bor_restlist_lines 교체 저장

    Returns: {snapshot_date, source_file, line_count, total_amount_usd, po_numbers}
    """
    bor_db.ensure_tables()
    since = (datetime.now(KST) - timedelta(days=days_back)).strftime("%d-%b-%Y")

    best = None  # (date_kst, uid, filename, data)
    M = None
    try:
        M = _connect()
        M.select("INBOX", readonly=True)
        for sender in (TIGER_SENDER, BOR_SENDER):
            uids = _uid_search(M, "FROM", f'"{sender}"', "SINCE", since)
            logger.info(f"[BOR/메일] rest 후보 — {sender}: {len(uids)}건")
            for uid in uids:
                try:
                    msg = _uid_fetch_message(M, uid)
                    if msg is None:
                        continue
                    date_kst = _parse_email_date(msg.get("Date", ""))
                    for filename, data in _iter_xlsx_attachments(msg):
                        if "rest" not in filename.lower():
                            continue
                        if best is None or (date_kst and best[0] and date_kst > best[0]) \
                                or (best[0] is None and date_kst):
                            best = (date_kst, uid, filename, data)
                except Exception as e:
                    logger.warning(f"[BOR/메일] rest 메일 처리 오류 (UID {uid}): {e}")
    finally:
        if M:
            try:
                M.close()
                M.logout()
            except Exception:
                pass

    if best is None:
        logger.info("[BOR/메일] rest list 메일 없음")
        return {"snapshot_date": "", "source_file": "", "line_count": 0,
                "total_amount_usd": 0, "po_numbers": []}

    date_kst, uid, filename, data = best
    raw_lines = parse_rest_list_xlsx(data, filename)
    snapshot_date = date_kst.strftime("%Y-%m-%d") if date_kst else datetime.now(KST).strftime("%Y-%m-%d")

    # 같은 (PO, 모델) 중복 라인 → 합산
    merged = {}
    for ln in raw_lines:
        key = (ln["po_number"], ln["model_norm"])
        if key in merged:
            merged[key]["rest_qty"] += ln["rest_qty"]
            if not merged[key]["urgency"] and ln["urgency"]:
                merged[key]["urgency"] = ln["urgency"]
        else:
            merged[key] = dict(ln)

    total_amount = sum(ln["rest_qty"] * ln["unit_price"] for ln in raw_lines)
    po_numbers = sorted({ln["po_number"] for ln in raw_lines if ln["po_number"]})

    # 교체 저장 (해당 스냅샷만 유지 — 이전 스냅샷 전체 삭제)
    # 다른 백그라운드 작업의 장시간 쓰기와 겹칠 수 있어 잠김 재시도 헬퍼 사용
    def _save(conn):
        conn.execute("DELETE FROM bor_restlist_lines")
        fetched_at = now_kst()
        for ln in merged.values():
            conn.execute(
                "INSERT INTO bor_restlist_lines "
                "(snapshot_date, source_uid, source_file, po_number, po_date, "
                " model, model_norm, description, rest_qty, unit_price, urgency, fetched_at) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                (
                    snapshot_date, uid, filename, ln["po_number"], ln["po_date"],
                    ln["model"], ln["model_norm"], ln["description"],
                    ln["rest_qty"], ln["unit_price"], ln["urgency"], fetched_at,
                ),
            )

    bor_db.write_with_retry(_save)

    logger.info(
        f"[BOR/메일] rest 스냅샷 저장 — {filename} ({snapshot_date}), "
        f"{len(merged)}라인 (원시 {len(raw_lines)}), USD {total_amount:,.0f}"
    )
    return {
        "snapshot_date": snapshot_date,
        "source_file": filename,
        "line_count": len(merged),
        "total_amount_usd": round(total_amount, 2),
        "po_numbers": po_numbers,
    }


# ─── 수집 2: 최근 오더 (양 채널) ─────────────────────────
def collect_recent_orders(days_back: int = 90) -> dict:
    """최근 오더 수집 → bor_recent_orders UPSERT (UNIQUE 충돌 무시)

    ① INBOX FROM kngtiger 제목 'new order' 포함(RE:/Fw: 포함 — 첨부 있으면 수용) → xlsx 파싱
    ② Sent 폴더 TO guzhiyi 제목 'new order' → 본문 정규식 (MODEL / N pcs)
    Returns: {tiger_orders, kyu_orders, lines}
    """
    bor_db.ensure_tables()
    since = (datetime.now(KST) - timedelta(days=days_back)).strftime("%d-%b-%Y")

    tiger_mails, kyu_mails = 0, 0
    rows = []  # (source, mail_uid, mail_date, order_date, subject, model, model_norm, qty, note)

    M = None
    try:
        M = _connect()

        # ── ① 정정섭 INBOX new order xlsx ──
        M.select("INBOX", readonly=True)
        uids = _uid_search(M, "FROM", f'"{TIGER_SENDER}"', "SINCE", since)
        logger.info(f"[BOR/메일] tiger 후보: {len(uids)}건")
        for uid in uids:
            try:
                msg = _uid_fetch_message(M, uid)
                if msg is None:
                    continue
                subject = _decode_header_value(msg.get("Subject", ""))
                if "new order" not in subject.lower():
                    continue
                date_kst = _parse_email_date(msg.get("Date", ""))
                mail_date = date_kst.strftime("%Y-%m-%d %H:%M:%S") if date_kst else ""
                order_date = date_kst.strftime("%Y-%m-%d") if date_kst else ""
                mail_items = 0
                for filename, data in _iter_xlsx_attachments(msg):
                    if "rest" in filename.lower():
                        continue  # 같은 메일의 rest list 첨부는 제외
                    for item in parse_new_order_xlsx(data, filename):
                        # note에 파일명 보존 (같은 오더 정정 재발행 dedupe에 사용)
                        note = filename + (f"; {item['remark']}" if item["remark"] else "")
                        rows.append((
                            "tiger_xlsx", uid, mail_date, order_date, subject,
                            item["model"], item["model_norm"], item["qty"], note,
                        ))
                        mail_items += 1
                if mail_items:
                    tiger_mails += 1
            except Exception as e:
                logger.warning(f"[BOR/메일] tiger 메일 처리 오류 (UID {uid}): {e}")

        # ── ② kyu Sent 텍스트 오더 ──
        M.select(SENT_FOLDER, readonly=True)
        uids = _uid_search(M, "TO", f'"{BOR_SENDER}"', "SINCE", since)
        logger.info(f"[BOR/메일] kyu Sent 후보: {len(uids)}건")
        for uid in uids:
            try:
                msg = _uid_fetch_message(M, uid)
                if msg is None:
                    continue
                subject = _decode_header_value(msg.get("Subject", ""))
                if "new order" not in subject.lower():
                    continue
                date_kst = _parse_email_date(msg.get("Date", ""))
                mail_date = date_kst.strftime("%Y-%m-%d %H:%M:%S") if date_kst else ""
                order_date = date_kst.strftime("%Y-%m-%d") if date_kst else ""
                body = _get_mail_body_text(msg)
                orders = parse_kyu_order_text(body)
                for model, qty in orders:
                    rows.append((
                        "kyu_text", uid, mail_date, order_date, subject,
                        model, normalize_model(model), qty, "",
                    ))
                if orders:
                    kyu_mails += 1
            except Exception as e:
                logger.warning(f"[BOR/메일] kyu 메일 처리 오류 (UID {uid}): {e}")
    finally:
        if M:
            try:
                M.close()
                M.logout()
            except Exception:
                pass

    # ── DB UPSERT (UNIQUE(source, mail_uid, model_norm, qty) 충돌 무시) ──
    def _save(conn):
        fetched_at = now_kst()
        for row in rows:
            conn.execute(
                "INSERT OR IGNORE INTO bor_recent_orders "
                "(source, mail_uid, mail_date, order_date, subject, "
                " model, model_norm, qty, note, fetched_at) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)",
                row + (fetched_at,),
            )

    bor_db.write_with_retry(_save)

    logger.info(
        f"[BOR/메일] 최근 오더 수집 — tiger {tiger_mails}건 메일 / "
        f"kyu {kyu_mails}건 메일 / 라인 {len(rows)}건"
    )
    return {"tiger_orders": tiger_mails, "kyu_orders": kyu_mails, "lines": len(rows)}
