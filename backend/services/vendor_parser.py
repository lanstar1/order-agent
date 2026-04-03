"""거래처 원장 엑셀 파서
패턴 A: 유니정보통신/파워네트/현대모아컴 (21컬럼, 거래장부내역)
패턴 B: 랜마스터 (13컬럼, 거래확인서)
"""
import os
import openpyxl
from dataclasses import dataclass, field, asdict


@dataclass
class VendorTransaction:
    """거래처 원장 거래 항목"""
    date: str = ""
    slip_no: str = ""
    tx_type: str = ""
    seq: str = ""
    product_category: str = ""
    product_name: str = ""
    model_name: str = ""
    qty: int = 0
    unit_price: float = 0
    amount: float = 0
    purchase_amount: float = 0
    sales_amount: float = 0
    balance: float = 0
    memo: str = ""

    def to_dict(self):
        return asdict(self)


def _safe_int(val) -> int:
    try:
        if val is None or val == "":
            return 0
        return int(float(str(val).replace(",", "")))
    except (ValueError, TypeError):
        return 0


def _safe_float(val) -> float:
    try:
        if val is None or val == "":
            return 0.0
        return float(str(val).replace(",", ""))
    except (ValueError, TypeError):
        return 0.0


def _safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def detect_pattern(ws) -> str:
    """시트의 양식 패턴 감지"""
    row1_val = _safe_str(ws.cell(1, 1).value)
    row2_val = _safe_str(ws.cell(2, 1).value)

    if "거래장부내역" in row1_val or "거래장부" in row1_val:
        return "A"
    if "거래확인서" in row1_val:
        return "B"
    if row2_val == "날짜":
        return "A"
    row10_val = _safe_str(ws.cell(10, 1).value)
    if row10_val == "날짜":
        return "B"
    return "unknown"


def parse_pattern_a(ws) -> list[VendorTransaction]:
    """패턴 A 파싱 (거래장부내역 — 21컬럼)"""
    transactions = []
    header_row = 2
    current_date = ""

    for row_idx in range(header_row + 1, ws.max_row + 1):
        tx_type = _safe_str(ws.cell(row_idx, 3).value)
        if tx_type in ("이전", "합계", "소계", ""):
            date_val = _safe_str(ws.cell(row_idx, 1).value)
            if date_val and tx_type == "":
                current_date = date_val
            continue
        if tx_type not in ("매입", "매출"):
            continue

        date_val = _safe_str(ws.cell(row_idx, 1).value)
        if date_val:
            current_date = date_val

        tx = VendorTransaction(
            date=current_date,
            slip_no=_safe_str(ws.cell(row_idx, 2).value),
            tx_type=tx_type,
            seq=_safe_str(ws.cell(row_idx, 4).value),
            product_category=_safe_str(ws.cell(row_idx, 5).value),
            product_name=_safe_str(ws.cell(row_idx, 6).value),
            model_name=_safe_str(ws.cell(row_idx, 7).value),
            qty=_safe_int(ws.cell(row_idx, 9).value),
            unit_price=_safe_float(ws.cell(row_idx, 10).value),
            amount=_safe_float(ws.cell(row_idx, 11).value),
            purchase_amount=_safe_float(ws.cell(row_idx, 12).value),
            sales_amount=_safe_float(ws.cell(row_idx, 13).value),
            balance=_safe_float(ws.cell(row_idx, 14).value),
            memo=_safe_str(ws.cell(row_idx, 20).value),
        )
        transactions.append(tx)

    return transactions


def parse_pattern_b(ws) -> list[VendorTransaction]:
    """패턴 B 파싱 (거래확인서 — 13컬럼)"""
    transactions = []
    header_row = 10
    current_date = ""

    for row_idx in range(header_row + 1, ws.max_row + 1):
        tx_type = _safe_str(ws.cell(row_idx, 3).value)
        if tx_type in ("이전", "합계", "소계", ""):
            date_val = _safe_str(ws.cell(row_idx, 1).value)
            if date_val:
                current_date = date_val
            continue
        if tx_type not in ("매입", "매출"):
            continue

        date_val = _safe_str(ws.cell(row_idx, 1).value)
        if date_val:
            current_date = date_val

        sales_amt = _safe_float(ws.cell(row_idx, 11).value)
        purchase_amt = _safe_float(ws.cell(row_idx, 12).value)

        tx = VendorTransaction(
            date=current_date,
            slip_no=_safe_str(ws.cell(row_idx, 2).value),
            tx_type=tx_type,
            seq=_safe_str(ws.cell(row_idx, 4).value),
            product_category=_safe_str(ws.cell(row_idx, 5).value),
            product_name=_safe_str(ws.cell(row_idx, 6).value),
            model_name="",
            qty=_safe_int(ws.cell(row_idx, 9).value),
            unit_price=_safe_float(ws.cell(row_idx, 10).value),
            amount=sales_amt if sales_amt else purchase_amt,
            purchase_amount=purchase_amt,
            sales_amount=sales_amt,
            balance=_safe_float(ws.cell(row_idx, 13).value),
            memo="",
        )
        transactions.append(tx)

    return transactions


def parse_vendor_ledger(file_path: str) -> dict:
    """거래처 원장 엑셀 파싱 (자동 패턴 감지)"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    vendor_name = ws.title
    if "라인업" in vendor_name:
        filename = os.path.basename(file_path)
        parts = filename.replace(".xlsx", "").split("_")
        if len(parts) > 1:
            vendor_name = parts[-1]

    pattern = detect_pattern(ws)

    if pattern == "A":
        transactions = parse_pattern_a(ws)
    elif pattern == "B":
        transactions = parse_pattern_b(ws)
    else:
        transactions = []

    wb.close()

    purchase_items = [t for t in transactions if t.tx_type == "매입"]
    sales_items = [t for t in transactions if t.tx_type == "매출"]

    return {
        "vendor_name": vendor_name,
        "pattern": pattern,
        "transactions": [t.to_dict() for t in transactions],
        "summary": {
            "total_purchases": len(purchase_items),
            "total_sales": len(sales_items),
            "total_purchase_amount": sum(t.amount for t in purchase_items),
            "total_sales_amount": sum(t.amount for t in sales_items),
        },
        "purchase_items": [t.to_dict() for t in purchase_items],
        "sales_items": [t.to_dict() for t in sales_items],
    }
