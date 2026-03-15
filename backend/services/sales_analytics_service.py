"""
판매현황 분석 서비스
- CSV 업로드 / 이카운트 ERP 자동 수집
- 8개 뷰 분석 + 4개 에이전트
"""
import io
import csv
import logging
import time
import os
import re
from datetime import datetime, timedelta, timezone
from typing import Optional

logger = logging.getLogger(__name__)

KST = timezone(timedelta(hours=9))


def _now_kst():
    return datetime.now(KST)


def _today_str():
    return _now_kst().strftime("%Y%m%d")


def _first_of_month():
    n = _now_kst()
    return n.replace(day=1).strftime("%Y%m%d")


class SalesAnalyticsService:
    def __init__(self):
        self._erp_session_id: Optional[str] = None
        self._erp_session_time: Optional[datetime] = None

    # ══════════════════════════════════════════════
    #  수집
    # ══════════════════════════════════════════════

    async def import_csv_bytes(self, content: bytes) -> dict:
        """CSV 바이트를 파싱하여 sales_records에 저장"""
        log_id = self._log_fetch("csv_upload", "running", "CSV 업로드 시작", 0)
        try:
            text = None
            for enc in ("utf-8-sig", "cp949", "euc-kr", "utf-8"):
                try:
                    text = content.decode(enc)
                    break
                except (UnicodeDecodeError, LookupError):
                    continue
            if text is None:
                raise ValueError("CSV 인코딩을 판별할 수 없습니다")

            rows = self._parse_csv_text(text)
            if not rows:
                self._update_log(log_id, "error", "파싱된 행이 없습니다", 0)
                return {"success": False, "error": "파싱된 행이 없습니다", "rows": 0}

            cnt = self._upsert_rows(rows)
            self._update_log(log_id, "success", f"{cnt}행 임포트 완료", cnt)
            return {"success": True, "rows": cnt}
        except Exception as e:
            self._update_log(log_id, "error", str(e), 0)
            logger.error(f"[SalesCSV] 임포트 오류: {e}", exc_info=True)
            return {"success": False, "error": str(e), "rows": 0}

    def _parse_csv_text(self, text: str) -> list:
        """이카운트 판매현황 CSV 파싱"""
        rows = []
        reader = csv.reader(io.StringIO(text))
        header = None
        for line in reader:
            if not line:
                continue
            # 헤더 감지
            if header is None:
                if any("품목코드" in c for c in line) or any("연/월/일" in c for c in line):
                    header = line
                continue
            # 합계행 제외: slip_no에 '-'가 없는 행
            date_cell = str(line[0]).strip() if line else ""
            if "-" not in date_cell:
                continue
            try:
                row = self._parse_csv_row(line)
                if row:
                    rows.append(row)
            except Exception as e:
                logger.debug(f"[SalesCSV] 행 파싱 스킵: {e}")
        return rows

    def _parse_csv_row(self, cols: list) -> Optional[dict]:
        """CSV 한 행 → dict"""
        if len(cols) < 10:
            return None
        date_raw = str(cols[0]).strip()
        # "20260101-1" → slip_date=20260101, slip_no=20260101-1
        parts = date_raw.split("-")
        if len(parts) < 2:
            return None
        slip_date = parts[0]
        slip_no = date_raw

        def _num(v):
            s = str(v).strip().replace(",", "").replace(" ", "")
            if not s or s == "-":
                return 0.0
            try:
                return float(s)
            except ValueError:
                return 0.0

        item_code = str(cols[1]).strip() if len(cols) > 1 else ""
        customer_name = str(cols[2]).strip() if len(cols) > 2 else ""
        item_name = str(cols[3]).strip() if len(cols) > 3 else ""
        model_name = str(cols[4]).strip() if len(cols) > 4 else ""
        quantity = _num(cols[5]) if len(cols) > 5 else 0
        unit_price = _num(cols[6]) if len(cols) > 6 else 0
        supply_amount = _num(cols[7]) if len(cols) > 7 else 0
        vat = _num(cols[8]) if len(cols) > 8 else 0
        total_amount = _num(cols[9]) if len(cols) > 9 else 0
        cost_price = _num(cols[10]) if len(cols) > 10 else 0
        warehouse = str(cols[11]).strip() if len(cols) > 11 else ""
        account_date = str(cols[12]).strip() if len(cols) > 12 else ""
        item_group = str(cols[13]).strip() if len(cols) > 13 else ""
        note = str(cols[14]).strip() if len(cols) > 14 else ""
        staff_name = str(cols[15]).strip() if len(cols) > 15 else ""
        customer_group = str(cols[16]).strip() if len(cols) > 16 else ""
        safety_stock = _num(cols[17]) if len(cols) > 17 else 0
        display_code = str(cols[18]).strip() if len(cols) > 18 else ""

        gross_profit = supply_amount - (cost_price * quantity)

        return {
            "slip_date": slip_date,
            "slip_no": slip_no,
            "item_code": item_code,
            "customer_name": customer_name,
            "item_name": item_name,
            "model_name": model_name,
            "quantity": quantity,
            "unit_price": unit_price,
            "supply_amount": supply_amount,
            "vat": vat,
            "total_amount": total_amount,
            "cost_price": cost_price,
            "warehouse": warehouse,
            "account_date": account_date,
            "item_group": item_group,
            "note": note,
            "staff_name": staff_name,
            "customer_group": customer_group,
            "safety_stock": safety_stock,
            "display_code": display_code,
            "gross_profit": gross_profit,
        }

    # ── ERP 자동 수집 ──

    async def auto_fetch_from_ecount(self) -> None:
        """ERP API 로그인 → 판매현황 엑셀 다운로드 시도"""
        log_id = self._log_fetch("auto_ecount", "running", "이카운트 자동 수집 시작", 0)
        try:
            session_id = await self._get_erp_session()
            if not session_id:
                self._update_log(log_id, "error", "ERP 세션 획득 실패", 0)
                return

            # requests 방식으로 엑셀 다운로드 시도
            csv_bytes = await self._download_sales_excel(session_id)
            if csv_bytes:
                result = await self.import_csv_bytes(csv_bytes)
                self._update_log(
                    log_id, "success" if result.get("success") else "error",
                    f"자동수집 완료: {result.get('rows', 0)}행", result.get("rows", 0)
                )
            else:
                # Playwright 폴백
                csv_bytes = await self._download_with_playwright(session_id)
                if csv_bytes:
                    result = await self.import_csv_bytes(csv_bytes)
                    self._update_log(
                        log_id, "success" if result.get("success") else "error",
                        f"Playwright 수집: {result.get('rows', 0)}행", result.get("rows", 0)
                    )
                else:
                    self._update_log(log_id, "error", "엑셀 다운로드 실패 (requests + Playwright)", 0)
        except Exception as e:
            self._update_log(log_id, "error", str(e), 0)
            logger.error(f"[SalesEcount] 자동 수집 오류: {e}", exc_info=True)

    async def _get_erp_session(self) -> Optional[str]:
        """ERP API 로그인 → SESSION_ID (12시간 캐싱, 최대 3회 재시도)"""
        # 캐시 확인
        if self._erp_session_id and self._erp_session_time:
            elapsed = (datetime.now() - self._erp_session_time).total_seconds()
            if elapsed < 43200:  # 12시간
                return self._erp_session_id

        from config import ERP_COM_CODE, ERP_USER_ID, ERP_API_KEY, ERP_ZONE
        if not ERP_API_KEY:
            logger.warning("[SalesEcount] ERP_API_KEY 미설정")
            return None

        import httpx
        login_url = f"https://oapi{ERP_ZONE.lower()}.ecount.com/OAPI/V2/OAPILogin"
        body = {
            "COM_CODE": ERP_COM_CODE,
            "USER_ID": ERP_USER_ID,
            "API_CERT_KEY": ERP_API_KEY,
            "LAN_TYPE": "ko-KR",
            "ZONE": ERP_ZONE,
        }

        for attempt in range(3):
            try:
                async with httpx.AsyncClient(timeout=15) as client:
                    r = await client.post(login_url, json=body)
                    data = r.json()
                    if str(data.get("Status")) == "200":
                        sid = data.get("Data", {}).get("Datas", {}).get("SESSION_ID")
                        if sid:
                            self._erp_session_id = sid
                            self._erp_session_time = datetime.now()
                            logger.info(f"[SalesEcount] ERP 세션 획득 성공 (시도 {attempt+1})")
                            return sid
                    logger.warning(f"[SalesEcount] 로그인 실패 (시도 {attempt+1}): {data}")
            except Exception as e:
                logger.warning(f"[SalesEcount] 로그인 예외 (시도 {attempt+1}): {e}")
            if attempt < 2:
                import asyncio
                await asyncio.sleep(2)

        logger.error("[SalesEcount] ERP 로그인 3회 실패 — 자동호출 중단")
        return None

    async def _download_sales_excel(self, session_id: str) -> Optional[bytes]:
        """requests 방식 엑셀 다운로드 시도"""
        try:
            import httpx
            from config import ERP_ZONE
            base = f"https://oapi{ERP_ZONE.lower()}.ecount.com"
            date_from = _first_of_month()
            date_to = _today_str()

            # 판매현황 리스트 API 시도
            url = f"{base}/OAPI/V2/Sale/GetSaleList"
            params = {
                "SESSION_ID": session_id,
                "SALE_FROM": date_from,
                "SALE_TO": date_to,
            }
            async with httpx.AsyncClient(timeout=30) as client:
                r = await client.post(url, json=params)
                if r.status_code == 200:
                    data = r.json()
                    if str(data.get("Status")) == "200" and data.get("Data"):
                        # API 응답을 CSV 바이트로 변환
                        return self._api_response_to_csv(data["Data"])
            logger.info("[SalesEcount] API 다운로드 실패 → Playwright 폴백")
            return None
        except Exception as e:
            logger.warning(f"[SalesEcount] 엑셀 다운로드 실패: {e}")
            return None

    def _api_response_to_csv(self, data) -> Optional[bytes]:
        """ERP API 응답을 CSV 바이트로 변환"""
        try:
            rows = data if isinstance(data, list) else data.get("Datas", [])
            if not rows:
                return None
            output = io.StringIO()
            writer = csv.writer(output)
            # 헤더
            writer.writerow([
                "연/월/일", "품목코드", "거래처명", "품명 및 규격", "모델명",
                "수량", "단가", "공급가액", "부가세", "합 계", "입고단가",
                "창고", "회계반영일자", "품목그룹1명", "비고사항",
                "전표출력(담당자)", "거래처그룹1명", "안전재고수량", "진열코드"
            ])
            for r in rows:
                writer.writerow([
                    r.get("SALE_DATE", ""), r.get("PROD_CD", ""),
                    r.get("CUST_NAME", ""), r.get("PROD_DES", ""),
                    r.get("MODEL_NAME", ""), r.get("QTY", 0),
                    r.get("PRICE", 0), r.get("SUPPLY_AMT", 0),
                    r.get("VAT", 0), r.get("TOTAL_AMT", 0),
                    r.get("COST_PRICE", 0), r.get("WH_NAME", ""),
                    r.get("ACC_DATE", ""), r.get("ITEM_GRP1", ""),
                    r.get("NOTE", ""), r.get("EMP_NAME", ""),
                    r.get("CUST_GRP1", ""), r.get("SAFE_STOCK", 0),
                    r.get("DISPLAY_CD", ""),
                ])
            return output.getvalue().encode("utf-8-sig")
        except Exception as e:
            logger.warning(f"[SalesEcount] API→CSV 변환 실패: {e}")
            return None

    async def _download_with_playwright(self, session_id: str) -> Optional[bytes]:
        """Playwright 폴백: SESSION_ID 쿠키 주입 방식"""
        try:
            from playwright.async_api import async_playwright
            from config import ERP_ZONE
        except ImportError:
            logger.warning("[SalesEcount] Playwright 미설치")
            return None

        try:
            date_from = _first_of_month()
            date_to = _today_str()
            target_url = (
                f"https://login{ERP_ZONE.lower()}.ecount.com/ec5/view/erp"
                f"?w_flag=1#menuType=MPGU_RT00000034&menuSeq=MPMU00003400040"
                f"&groupSeq=MPGU00003400003&prgId=E040207&depth=3"
            )

            async with async_playwright() as p:
                browser = await p.chromium.launch(headless=True)
                context = await browser.new_context()
                await context.add_cookies([{
                    "name": "SESSION_ID",
                    "value": session_id,
                    "domain": f"login{ERP_ZONE.lower()}.ecount.com",
                    "path": "/",
                }])
                page = await context.new_page()
                await page.goto(target_url, timeout=30000)
                await page.wait_for_timeout(5000)

                # 엑셀 다운로드 버튼 클릭 시도
                download_data = None
                try:
                    async with page.expect_download(timeout=15000) as download_info:
                        await page.click("button:has-text('엑셀'), a:has-text('엑셀')", timeout=5000)
                    download = await download_info.value
                    path = await download.path()
                    if path:
                        with open(path, "rb") as f:
                            download_data = f.read()
                except Exception as e:
                    logger.warning(f"[SalesEcount] Playwright 다운로드 실패: {e}")

                await browser.close()

                if download_data:
                    # xlsx → csv 변환
                    return self._xlsx_to_csv_bytes(download_data)
                return None
        except Exception as e:
            logger.error(f"[SalesEcount] Playwright 오류: {e}", exc_info=True)
            return None

    def _xlsx_to_csv_bytes(self, xlsx_data: bytes) -> Optional[bytes]:
        """xlsx 바이트 → CSV 바이트 변환"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_data), read_only=True)
            ws = wb[wb.sheetnames[0]]
            output = io.StringIO()
            writer = csv.writer(output)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)
            wb.close()
            return output.getvalue().encode("utf-8-sig")
        except Exception as e:
            logger.warning(f"[SalesEcount] xlsx→csv 변환 실패: {e}")
            return None

    async def get_fetch_status(self) -> dict:
        """최근 수집 상태"""
        from db.database import get_connection
        conn = get_connection()
        try:
            row = conn.execute(
                "SELECT * FROM sales_fetch_log ORDER BY id DESC LIMIT 1"
            ).fetchone()
            if row:
                return {
                    "id": row["id"], "fetch_type": row["fetch_type"],
                    "status": row["status"], "message": row["message"],
                    "rows_imported": row["rows_imported"],
                    "started_at": row["started_at"],
                    "finished_at": row["finished_at"],
                }
            return {"status": "none", "message": "수집 이력 없음"}
        finally:
            conn.close()

    async def get_scheduler_status(self) -> dict:
        """스케줄러 상태"""
        status = await self.get_fetch_status()
        from db.database import get_connection
        conn = get_connection()
        try:
            cnt = conn.execute("SELECT COUNT(*) as cnt FROM sales_records").fetchone()
            total = cnt["cnt"] if cnt else 0
        finally:
            conn.close()
        return {
            "total_records": total,
            "last_fetch": status,
        }

    # ══════════════════════════════════════════════
    #  공통 유틸
    # ══════════════════════════════════════════════

    def _date_filter(self, date_from, date_to):
        """날짜 필터 조건+파라미터 생성"""
        conditions = []
        params = []
        if date_from:
            conditions.append("slip_date >= ?")
            params.append(str(date_from).replace("-", ""))
        if date_to:
            conditions.append("slip_date <= ?")
            params.append(str(date_to).replace("-", ""))
        return conditions, params

    def _upsert_rows(self, rows: list) -> int:
        """slip_no 기준 DELETE→INSERT"""
        from db.database import get_connection
        conn = get_connection()
        try:
            cnt = 0
            for r in rows:
                if r.get("slip_no"):
                    conn.execute("DELETE FROM sales_records WHERE slip_no=?", (r["slip_no"],))
                conn.execute("""
                    INSERT INTO sales_records (
                        slip_date, slip_no, item_code, customer_name, item_name,
                        model_name, quantity, unit_price, supply_amount, vat,
                        total_amount, cost_price, warehouse, account_date, item_group,
                        note, staff_name, customer_group, safety_stock, display_code,
                        gross_profit
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    r["slip_date"], r["slip_no"], r["item_code"], r["customer_name"],
                    r["item_name"], r["model_name"], r["quantity"], r["unit_price"],
                    r["supply_amount"], r["vat"], r["total_amount"], r["cost_price"],
                    r["warehouse"], r["account_date"], r["item_group"], r["note"],
                    r["staff_name"], r["customer_group"], r["safety_stock"],
                    r["display_code"], r["gross_profit"],
                ))
                cnt += 1
            conn.commit()
            return cnt
        finally:
            conn.close()

    def _log_fetch(self, ftype, status, msg, rows) -> int:
        from db.database import get_connection
        conn = get_connection()
        try:
            cur = conn.execute(
                "INSERT INTO sales_fetch_log(fetch_type,status,message,rows_imported) VALUES(?,?,?,?)",
                (ftype, status, msg, rows)
            )
            conn.commit()
            return cur.lastrowid or 0
        finally:
            conn.close()

    def _update_log(self, log_id, status, msg, rows):
        from db.database import get_connection
        conn = get_connection()
        try:
            conn.execute(
                "UPDATE sales_fetch_log SET status=?, message=?, rows_imported=?, finished_at=CURRENT_TIMESTAMP WHERE id=?",
                (status, msg, rows, log_id)
            )
            conn.commit()
        finally:
            conn.close()

    # ══════════════════════════════════════════════
    #  뷰1: 전체 대시보드
    # ══════════════════════════════════════════════

    async def get_summary(self, date_from=None, date_to=None) -> dict:
        from db.database import get_connection
        conds, params = self._date_filter(date_from, date_to)
        where = ("WHERE " + " AND ".join(conds)) if conds else ""
        conn = get_connection()
        try:
            row = conn.execute(f"""
                SELECT
                    COALESCE(SUM(supply_amount),0) as total_supply,
                    COALESCE(SUM(gross_profit),0) as total_profit,
                    COUNT(*) as slip_count,
                    COUNT(DISTINCT customer_name) as customer_count
                FROM sales_records {where}
            """, params).fetchone()
            total_supply = row["total_supply"] or 0
            total_profit = row["total_profit"] or 0
            profit_rate = (total_profit / total_supply * 100) if total_supply else 0
            return {
                "total_supply": total_supply,
                "total_profit": total_profit,
                "profit_rate": round(profit_rate, 1),
                "slip_count": row["slip_count"] or 0,
                "customer_count": row["customer_count"] or 0,
            }
        finally:
            conn.close()

    async def get_monthly_trend(self, months=6, customer_name=None) -> list:
        from db.database import get_connection
        conn = get_connection()
        try:
            n = _now_kst()
            start = (n - timedelta(days=30 * months)).strftime("%Y%m01")
            conds = ["slip_date >= ?"]
            params = [start]
            if customer_name:
                conds.append("customer_name = ?")
                params.append(customer_name)
            where = "WHERE " + " AND ".join(conds)
            rows = conn.execute(f"""
                SELECT
                    SUBSTR(slip_date,1,6) as month,
                    COALESCE(SUM(supply_amount),0) as supply,
                    COALESCE(SUM(gross_profit),0) as profit,
                    COUNT(*) as cnt
                FROM sales_records {where}
                GROUP BY SUBSTR(slip_date,1,6)
                ORDER BY month
            """, params).fetchall()
            return [{"month": r["month"], "supply": r["supply"], "profit": r["profit"], "count": r["cnt"]} for r in rows]
        finally:
            conn.close()

    async def get_daily_trend(self, date_from=None, date_to=None, customer_name=None) -> list:
        from db.database import get_connection
        conds, params = self._date_filter(date_from, date_to)
        if customer_name:
            conds.append("customer_name = ?")
            params.append(customer_name)
        where = ("WHERE " + " AND ".join(conds)) if conds else ""
        conn = get_connection()
        try:
            rows = conn.execute(f"""
                SELECT slip_date,
                    COALESCE(SUM(supply_amount),0) as supply,
                    COALESCE(SUM(gross_profit),0) as profit,
                    COUNT(*) as cnt
                FROM sales_records {where}
                GROUP BY slip_date ORDER BY slip_date
            """, params).fetchall()
            return [{"date": r["slip_date"], "supply": r["supply"], "profit": r["profit"], "count": r["cnt"]} for r in rows]
        finally:
            conn.close()

    async def get_customer_ranking(self, date_from=None, date_to=None, limit=15, group=None) -> list:
        from db.database import get_connection
        conds, params = self._date_filter(date_from, date_to)
        if group:
            conds.append("customer_group = ?")
            params.append(group)
        where = ("WHERE " + " AND ".join(conds)) if conds else ""
        conn = get_connection()
        try:
            rows = conn.execute(f"""
                SELECT customer_name, customer_group,
                    COALESCE(SUM(supply_amount),0) as supply,
                    COALESCE(SUM(gross_profit),0) as profit,
                    COUNT(*) as slip_count,
                    COUNT(DISTINCT item_code) as item_count
                FROM sales_records {where}
                GROUP BY customer_name
                ORDER BY supply DESC
                LIMIT ?
            """, params + [limit]).fetchall()
            result = []
            for r in rows:
                s = r["supply"] or 0
                p = r["profit"] or 0
                result.append({
                    "customer_name": r["customer_name"],
                    "customer_group": r["customer_group"] or "",
                    "supply": s, "profit": p,
                    "profit_rate": round(p / s * 100, 1) if s else 0,
                    "slip_count": r["slip_count"],
                    "item_count": r["item_count"],
                })
            return result
        finally:
            conn.close()

    async def get_customer_list(self) -> list:
        """전체 거래처 목록 (자동완성용)"""
        from db.database import get_connection
        conn = get_connection()
        try:
            rows = conn.execute(
                "SELECT DISTINCT customer_name FROM sales_records WHERE customer_name != '' ORDER BY customer_name"
            ).fetchall()
            return [r["customer_name"] for r in rows]
        finally:
            conn.close()

    async def get_customer_detail(self, customer_name, date_from=None, date_to=None) -> dict:
        from db.database import get_connection
        conds, params = self._date_filter(date_from, date_to)
        conds.append("customer_name = ?")
        params.append(customer_name)
        where = "WHERE " + " AND ".join(conds)
        conn = get_connection()
        try:
            summary = conn.execute(f"""
                SELECT COALESCE(SUM(supply_amount),0) as supply,
                       COALESCE(SUM(gross_profit),0) as profit,
                       COUNT(*) as cnt, customer_group
                FROM sales_records {where}
                GROUP BY customer_group
            """, params).fetchone()
            # 월별 추이
            monthly = conn.execute(f"""
                SELECT SUBSTR(slip_date,1,6) as month,
                    COALESCE(SUM(supply_amount),0) as supply,
                    COALESCE(SUM(gross_profit),0) as profit
                FROM sales_records {where}
                GROUP BY SUBSTR(slip_date,1,6) ORDER BY month
            """, params).fetchall()
            # 품목별 TOP
            products = conn.execute(f"""
                SELECT item_code, model_name,
                    COALESCE(SUM(supply_amount),0) as supply,
                    COALESCE(SUM(quantity),0) as qty
                FROM sales_records {where}
                GROUP BY item_code ORDER BY supply DESC LIMIT 20
            """, params).fetchall()
            s = (summary["supply"] or 0) if summary else 0
            p = (summary["profit"] or 0) if summary else 0
            return {
                "customer_name": customer_name,
                "customer_group": (summary["customer_group"] or "") if summary else "",
                "supply": s, "profit": p,
                "profit_rate": round(p / s * 100, 1) if s else 0,
                "slip_count": (summary["cnt"] or 0) if summary else 0,
                "monthly": [{"month": r["month"], "supply": r["supply"], "profit": r["profit"]} for r in monthly],
                "products": [{"item_code": r["item_code"], "model_name": r["model_name"] or "", "supply": r["supply"], "qty": r["qty"]} for r in products],
            }
        finally:
            conn.close()

    async def get_customer_compare(self, names: list, date_from=None, date_to=None) -> list:
        results = []
        for name in names[:5]:
            detail = await self.get_customer_detail(name, date_from, date_to)
            results.append(detail)
        return results

    async def get_customer_groups(self) -> list:
        from db.database import get_connection
        conn = get_connection()
        try:
            rows = conn.execute(
                "SELECT DISTINCT customer_group FROM sales_records WHERE customer_group != '' ORDER BY customer_group"
            ).fetchall()
            return [r["customer_group"] for r in rows]
        finally:
            conn.close()

    async def get_product_ranking(self, date_from=None, date_to=None, limit=20,
                                   customer_name=None, item_group=None) -> list:
        from db.database import get_connection
        conds, params = self._date_filter(date_from, date_to)
        if customer_name:
            conds.append("customer_name = ?")
            params.append(customer_name)
        if item_group:
            conds.append("item_group = ?")
            params.append(item_group)
        where = ("WHERE " + " AND ".join(conds)) if conds else ""
        conn = get_connection()
        try:
            rows = conn.execute(f"""
                SELECT item_code, model_name, item_group,
                    COALESCE(SUM(quantity),0) as qty,
                    COALESCE(SUM(supply_amount),0) as supply,
                    COALESCE(SUM(gross_profit),0) as profit,
                    COUNT(*) as cnt
                FROM sales_records {where}
                GROUP BY item_code ORDER BY supply DESC LIMIT ?
            """, params + [limit]).fetchall()
            return [{
                "item_code": r["item_code"], "model_name": r["model_name"] or "",
                "item_group": r["item_group"] or "",
                "qty": r["qty"], "supply": r["supply"], "profit": r["profit"],
                "profit_rate": round(r["profit"] / r["supply"] * 100, 1) if r["supply"] else 0,
                "count": r["cnt"],
            } for r in rows]
        finally:
            conn.close()

    async def get_product_groups(self) -> list:
        from db.database import get_connection
        conn = get_connection()
        try:
            rows = conn.execute(
                "SELECT DISTINCT item_group FROM sales_records WHERE item_group != '' ORDER BY item_group"
            ).fetchall()
            return [r["item_group"] for r in rows]
        finally:
            conn.close()

    async def get_product_detail(self, item_code, date_from=None, date_to=None) -> dict:
        from db.database import get_connection
        conds, params = self._date_filter(date_from, date_to)
        conds.append("item_code = ?")
        params.append(item_code)
        where = "WHERE " + " AND ".join(conds)
        conn = get_connection()
        try:
            summary = conn.execute(f"""
                SELECT COALESCE(SUM(supply_amount),0) as supply,
                       COALESCE(SUM(gross_profit),0) as profit,
                       COALESCE(SUM(quantity),0) as qty,
                       COUNT(*) as cnt, model_name
                FROM sales_records {where}
            """, params).fetchone()
            customers = conn.execute(f"""
                SELECT customer_name, COALESCE(SUM(supply_amount),0) as supply,
                       COALESCE(SUM(quantity),0) as qty
                FROM sales_records {where}
                GROUP BY customer_name ORDER BY supply DESC LIMIT 20
            """, params).fetchall()
            monthly = conn.execute(f"""
                SELECT SUBSTR(slip_date,1,6) as month,
                    COALESCE(SUM(supply_amount),0) as supply,
                    COALESCE(SUM(quantity),0) as qty
                FROM sales_records {where}
                GROUP BY SUBSTR(slip_date,1,6) ORDER BY month
            """, params).fetchall()
            s = (summary["supply"] or 0) if summary else 0
            p = (summary["profit"] or 0) if summary else 0
            return {
                "item_code": item_code,
                "model_name": (summary["model_name"] or "") if summary else "",
                "supply": s, "profit": p, "qty": (summary["qty"] or 0) if summary else 0,
                "profit_rate": round(p / s * 100, 1) if s else 0,
                "slip_count": (summary["cnt"] or 0) if summary else 0,
                "customers": [{"customer_name": r["customer_name"], "supply": r["supply"], "qty": r["qty"]} for r in customers],
                "monthly": [{"month": r["month"], "supply": r["supply"], "qty": r["qty"]} for r in monthly],
            }
        finally:
            conn.close()

    # ══════════════════════════════════════════════
    #  뷰2: 이익률 분석
    # ══════════════════════════════════════════════

    async def get_profit_analysis(self, date_from=None, date_to=None, group=None) -> dict:
        from db.database import get_connection
        conds, params = self._date_filter(date_from, date_to)
        if group:
            conds.append("customer_group = ?")
            params.append(group)
        where = ("WHERE " + " AND ".join(conds)) if conds else ""
        conn = get_connection()
        try:
            rows = conn.execute(f"""
                SELECT supply_amount, gross_profit, customer_name, item_code, model_name,
                       slip_date, slip_no, customer_group, item_group, quantity, cost_price
                FROM sales_records {where}
            """, params).fetchall()

            buckets = [
                {"range": "적자(<0%)", "cnt": 0, "supply": 0, "profit": 0},
                {"range": "0-10%", "cnt": 0, "supply": 0, "profit": 0},
                {"range": "10-20%", "cnt": 0, "supply": 0, "profit": 0},
                {"range": "20-30%", "cnt": 0, "supply": 0, "profit": 0},
                {"range": "30-40%", "cnt": 0, "supply": 0, "profit": 0},
                {"range": "40%+", "cnt": 0, "supply": 0, "profit": 0},
            ]
            low_margin = []
            cgroup_map = {}
            igroup_map = {}
            disc3 = {"supply": 0, "profit": 0}
            disc5 = {"supply": 0, "profit": 0}
            normal = {"supply": 0, "profit": 0}

            for r in rows:
                s = r["supply_amount"] or 0
                p = r["gross_profit"] or 0
                rate = (p / s * 100) if s else 0

                # bucket
                if rate < 0:
                    buckets[0]["cnt"] += 1; buckets[0]["supply"] += s; buckets[0]["profit"] += p
                elif rate < 10:
                    buckets[1]["cnt"] += 1; buckets[1]["supply"] += s; buckets[1]["profit"] += p
                elif rate < 20:
                    buckets[2]["cnt"] += 1; buckets[2]["supply"] += s; buckets[2]["profit"] += p
                elif rate < 30:
                    buckets[3]["cnt"] += 1; buckets[3]["supply"] += s; buckets[3]["profit"] += p
                elif rate < 40:
                    buckets[4]["cnt"] += 1; buckets[4]["supply"] += s; buckets[4]["profit"] += p
                else:
                    buckets[5]["cnt"] += 1; buckets[5]["supply"] += s; buckets[5]["profit"] += p

                # 역마진
                if rate < 0 or (r["quantity"] or 0) < 0:
                    low_margin.append({
                        "customer_name": r["customer_name"], "item_code": r["item_code"],
                        "model_name": r["model_name"] or "", "slip_date": r["slip_date"],
                        "supply": s, "profit": p, "profit_rate": round(rate, 1),
                    })

                # 그룹별
                cg = r["customer_group"] or "미분류"
                if cg not in cgroup_map:
                    cgroup_map[cg] = {"supply": 0, "profit": 0}
                cgroup_map[cg]["supply"] += s
                cgroup_map[cg]["profit"] += p

                ig = r["item_group"] or "미분류"
                if ig not in igroup_map:
                    igroup_map[ig] = {"supply": 0, "profit": 0}
                igroup_map[ig]["supply"] += s
                igroup_map[ig]["profit"] += p

                # 할인품목 분류 (비고 or 품목명에 할인 키워드)
                note = (r.get("item_code") or "") + (r.get("model_name") or "")
                if "3%" in note or "3할" in note:
                    disc3["supply"] += s; disc3["profit"] += p
                elif "5%" in note or "5할" in note:
                    disc5["supply"] += s; disc5["profit"] += p
                else:
                    normal["supply"] += s; normal["profit"] += p

            def _gr(d):
                return {**d, "rate": round(d["profit"] / d["supply"] * 100, 1) if d["supply"] else 0}

            by_cg = [{"group": k, "supply": v["supply"], "profit": v["profit"],
                       "rate": round(v["profit"]/v["supply"]*100,1) if v["supply"] else 0}
                      for k, v in cgroup_map.items()]
            by_ig = [{"group": k, "supply": v["supply"], "profit": v["profit"],
                       "rate": round(v["profit"]/v["supply"]*100,1) if v["supply"] else 0}
                      for k, v in igroup_map.items()]

            return {
                "buckets": buckets,
                "low_margin_records": sorted(low_margin, key=lambda x: x["profit_rate"])[:100],
                "by_customer_group": sorted(by_cg, key=lambda x: x["supply"], reverse=True),
                "by_item_group": sorted(by_ig, key=lambda x: x["supply"], reverse=True),
                "discount_compare": {
                    "discount3": _gr(disc3),
                    "discount5": _gr(disc5),
                    "normal": _gr(normal),
                },
            }
        finally:
            conn.close()

    async def get_profit_heatmap(self, date_from=None, date_to=None) -> list:
        from db.database import get_connection
        conds, params = self._date_filter(date_from, date_to)
        where = ("WHERE " + " AND ".join(conds)) if conds else ""
        conn = get_connection()
        try:
            rows = conn.execute(f"""
                SELECT customer_name, item_code, model_name,
                    COALESCE(SUM(supply_amount),0) as supply,
                    COALESCE(SUM(gross_profit),0) as profit
                FROM sales_records {where}
                GROUP BY customer_name, item_code
                HAVING supply != 0
                ORDER BY (CAST(profit AS REAL) / CAST(supply AS REAL)) ASC
                LIMIT 200
            """, params).fetchall()
            return [{
                "customer_name": r["customer_name"], "item_code": r["item_code"],
                "model_name": r["model_name"] or "",
                "supply": r["supply"], "profit": r["profit"],
                "profit_rate": round(r["profit"] / r["supply"] * 100, 1) if r["supply"] else 0,
            } for r in rows]
        finally:
            conn.close()

    # ══════════════════════════════════════════════
    #  뷰3: 거래처 건강도
    # ══════════════════════════════════════════════

    async def get_customer_health(self, date_from=None, date_to=None) -> dict:
        from db.database import get_connection
        conn = get_connection()
        try:
            n = _now_kst()
            m0 = n.strftime("%Y%m")
            m1 = (n - timedelta(days=30)).strftime("%Y%m")
            m2 = (n - timedelta(days=60)).strftime("%Y%m")
            m3 = (n - timedelta(days=90)).strftime("%Y%m")

            # 각 월별 거래처와 매출
            def _month_data(ym):
                rows = conn.execute("""
                    SELECT customer_name, COALESCE(SUM(supply_amount),0) as supply
                    FROM sales_records WHERE SUBSTR(slip_date,1,6)=?
                    GROUP BY customer_name
                """, (ym,)).fetchall()
                return {r["customer_name"]: r["supply"] for r in rows}

            d0 = _month_data(m0)
            d1 = _month_data(m1)
            d2 = _month_data(m2)
            d3 = _month_data(m3)
            all_custs = set(d0) | set(d1) | set(d2) | set(d3)

            # 조회기간 내 첫 거래 확인
            first_trade = {}
            rows = conn.execute("""
                SELECT customer_name, MIN(slip_date) as first_date
                FROM sales_records GROUP BY customer_name
            """).fetchall()
            for r in rows:
                first_trade[r["customer_name"]] = r["first_date"]

            active, at_risk, churned, new_customers = [], [], [], []
            period_start = str(date_from).replace("-", "") if date_from else m3 + "01"

            for c in all_custs:
                in_m0 = c in d0
                in_m1 = c in d1
                in_m2 = c in d2

                # 신규: 첫 거래가 조회기간 내
                ft = first_trade.get(c, "")
                if ft >= period_start:
                    new_customers.append({"customer_name": c, "first_date": ft,
                                          "supply": d0.get(c, 0) + d1.get(c, 0)})
                    continue

                # 활성: 3개월 연속
                if in_m0 and in_m1 and in_m2:
                    active.append({"customer_name": c, "supply": d0.get(c, 0)})
                    continue

                # 이탈: 2개월 이상 미거래
                if not in_m0 and not in_m1:
                    churned.append({"customer_name": c, "last_supply": d2.get(c, 0) or d3.get(c, 0)})
                    continue

                # 위험: 전월 대비 -50% 이상
                prev = d1.get(c, 0)
                curr = d0.get(c, 0)
                if prev > 0 and curr < prev * 0.5:
                    change = round((curr - prev) / prev * 100, 1)
                    at_risk.append({"customer_name": c, "prev_supply": prev,
                                    "curr_supply": curr, "change_rate": change})
                elif in_m0:
                    active.append({"customer_name": c, "supply": d0.get(c, 0)})

            return {
                "active": sorted(active, key=lambda x: x["supply"], reverse=True),
                "at_risk": sorted(at_risk, key=lambda x: x["change_rate"]),
                "churned": churned[:50],
                "new_customers": new_customers,
                "summary": {
                    "active": len(active), "at_risk": len(at_risk),
                    "churned": len(churned), "new": len(new_customers),
                },
            }
        finally:
            conn.close()

    async def get_customer_growth(self, date_from=None, date_to=None) -> list:
        from db.database import get_connection
        conn = get_connection()
        try:
            n = _now_kst()
            m0 = n.strftime("%Y%m")
            m1 = (n - timedelta(days=30)).strftime("%Y%m")

            rows = conn.execute("""
                SELECT customer_name, SUBSTR(slip_date,1,6) as month,
                    COALESCE(SUM(supply_amount),0) as supply
                FROM sales_records
                WHERE SUBSTR(slip_date,1,6) IN (?,?)
                GROUP BY customer_name, SUBSTR(slip_date,1,6)
            """, (m0, m1)).fetchall()

            data = {}
            for r in rows:
                c = r["customer_name"]
                if c not in data:
                    data[c] = {"prev": 0, "curr": 0}
                if r["month"] == m1:
                    data[c]["prev"] = r["supply"]
                else:
                    data[c]["curr"] = r["supply"]

            result = []
            for c, d in data.items():
                growth = round((d["curr"] - d["prev"]) / d["prev"] * 100, 1) if d["prev"] else (100 if d["curr"] else 0)
                status = "growth" if growth > 0 else ("decline" if growth < 0 else "stable")
                result.append({
                    "customer_name": c, "prev_supply": d["prev"],
                    "curr_supply": d["curr"], "growth_rate": growth, "status": status,
                })
            return sorted(result, key=lambda x: x["growth_rate"], reverse=True)
        finally:
            conn.close()

    # ══════════════════════════════════════════════
    #  뷰4: 재고 위험도
    # ══════════════════════════════════════════════

    async def get_inventory_risk(self, months_back=1) -> list:
        from db.database import get_connection
        conn = get_connection()
        try:
            n = _now_kst()
            start = (n - timedelta(days=30 * months_back)).strftime("%Y%m%d")
            rows = conn.execute("""
                SELECT item_code, model_name,
                    MAX(safety_stock) as safety_stock,
                    COALESCE(SUM(CASE WHEN quantity>0 THEN quantity ELSE 0 END),0) as total_qty
                FROM sales_records WHERE slip_date >= ? AND safety_stock > 0
                GROUP BY item_code
            """, (start,)).fetchall()

            result = []
            for r in rows:
                ss = r["safety_stock"] or 0
                monthly_qty = (r["total_qty"] or 0) / max(months_back, 1)
                turnover = monthly_qty / ss if ss else 0
                days_to_stockout = round(ss / (monthly_qty / 30), 0) if monthly_qty > 0 else 999

                if turnover >= 1.5:
                    risk = "critical"
                elif turnover >= 1.0:
                    risk = "warning"
                else:
                    risk = "normal"

                result.append({
                    "item_code": r["item_code"],
                    "model_name": r["model_name"] or "",
                    "safety_stock": ss,
                    "monthly_qty": round(monthly_qty, 1),
                    "turnover_rate": round(turnover, 2),
                    "days_to_stockout": int(days_to_stockout),
                    "risk_level": risk,
                    "recommended_order": max(0, round(monthly_qty * 2 - ss)),
                })
            return sorted(result, key=lambda x: x["turnover_rate"], reverse=True)
        finally:
            conn.close()

    async def get_inventory_summary(self) -> dict:
        items = await self.get_inventory_risk(months_back=1)
        critical = sum(1 for i in items if i["risk_level"] == "critical")
        warning = sum(1 for i in items if i["risk_level"] == "warning")
        normal = sum(1 for i in items if i["risk_level"] == "normal")
        return {
            "critical": critical, "warning": warning, "normal": normal,
            "total": len(items),
        }

    # ══════════════════════════════════════════════
    #  뷰5: 담당자 성과
    # ══════════════════════════════════════════════

    async def get_staff_performance(self, date_from=None, date_to=None) -> list:
        from db.database import get_connection
        conds, params = self._date_filter(date_from, date_to)
        conds.append("staff_name != ''")
        where = "WHERE " + " AND ".join(conds)
        conn = get_connection()
        try:
            rows = conn.execute(f"""
                SELECT staff_name,
                    COALESCE(SUM(supply_amount),0) as supply,
                    COALESCE(SUM(gross_profit),0) as profit,
                    COUNT(*) as slip_count,
                    COUNT(DISTINCT customer_name) as cust_count,
                    COALESCE(SUM(CASE WHEN quantity<0 THEN 1 ELSE 0 END),0) as return_count
                FROM sales_records {where}
                GROUP BY staff_name ORDER BY supply DESC
            """, params).fetchall()

            total_supply = sum(r["supply"] or 0 for r in rows)
            result = []
            for r in rows:
                s = r["supply"] or 0
                p = r["profit"] or 0
                sc = r["slip_count"] or 0
                cc = r["cust_count"] or 0
                rc = r["return_count"] or 0
                result.append({
                    "staff_name": r["staff_name"],
                    "supply": s, "profit": p,
                    "profit_rate": round(p / s * 100, 1) if s else 0,
                    "slip_count": sc,
                    "customer_count": cc,
                    "return_count": rc,
                    "return_rate": round(rc / sc * 100, 1) if sc else 0,
                    "avg_per_customer": round(s / cc) if cc else 0,
                    "share_pct": round(s / total_supply * 100, 1) if total_supply else 0,
                })
            return result
        finally:
            conn.close()

    # ══════════════════════════════════════════════
    #  뷰6: 반품·이슈 관리
    # ══════════════════════════════════════════════

    async def get_returns_analysis(self, date_from=None, date_to=None) -> dict:
        from db.database import get_connection
        conds, params = self._date_filter(date_from, date_to)
        where = ("WHERE " + " AND ".join(conds)) if conds else ""
        conn = get_connection()
        try:
            # 전체 매출
            total_row = conn.execute(f"""
                SELECT COALESCE(SUM(supply_amount),0) as total_supply,
                       COUNT(*) as total_cnt
                FROM sales_records {where}
            """, params).fetchone()
            total_supply = total_row["total_supply"] or 0
            total_cnt = total_row["total_cnt"] or 0

            # 반품 (수량 < 0)
            rconds = list(conds) + ["quantity < 0"]
            rwhere = "WHERE " + " AND ".join(rconds)
            ret_rows = conn.execute(f"""
                SELECT slip_no, customer_name, item_code, model_name,
                       quantity, supply_amount, note, slip_date, staff_name
                FROM sales_records {rwhere}
                ORDER BY slip_date DESC
            """, params).fetchall()

            ret_cnt = len(ret_rows)
            ret_amt = sum(abs(r["supply_amount"] or 0) for r in ret_rows)

            # 거래처별 반품
            by_customer = {}
            for r in ret_rows:
                c = r["customer_name"]
                if c not in by_customer:
                    by_customer[c] = {"count": 0, "amount": 0}
                by_customer[c]["count"] += 1
                by_customer[c]["amount"] += abs(r["supply_amount"] or 0)

            by_cust_list = [{"customer_name": k, **v} for k, v in
                            sorted(by_customer.items(), key=lambda x: x[1]["count"], reverse=True)]

            # 품목별 반품
            by_item = {}
            for r in ret_rows:
                ic = r["item_code"]
                if ic not in by_item:
                    by_item[ic] = {"model_name": r["model_name"] or "", "count": 0, "amount": 0}
                by_item[ic]["count"] += 1
                by_item[ic]["amount"] += abs(r["supply_amount"] or 0)

            by_item_list = [{"item_code": k, **v} for k, v in
                            sorted(by_item.items(), key=lambda x: x[1]["count"], reverse=True)]

            # 월별 추이
            monthly = conn.execute(f"""
                SELECT SUBSTR(slip_date,1,6) as month,
                    COUNT(*) as cnt,
                    COALESCE(SUM(ABS(supply_amount)),0) as amt
                FROM sales_records {rwhere}
                GROUP BY SUBSTR(slip_date,1,6) ORDER BY month
            """, params).fetchall()

            return {
                "summary": {
                    "total_cnt": ret_cnt,
                    "total_amt": ret_amt,
                    "return_rate": round(ret_cnt / total_cnt * 100, 1) if total_cnt else 0,
                },
                "by_customer": by_cust_list[:20],
                "by_item": by_item_list[:20],
                "records": [{
                    "slip_no": r["slip_no"], "customer_name": r["customer_name"],
                    "item_code": r["item_code"], "model_name": r["model_name"] or "",
                    "quantity": r["quantity"], "supply_amount": r["supply_amount"],
                    "note": r["note"] or "", "slip_date": r["slip_date"],
                    "staff_name": r["staff_name"] or "",
                } for r in ret_rows[:100]],
                "monthly_trend": [{"month": r["month"], "count": r["cnt"], "amount": r["amt"]} for r in monthly],
            }
        finally:
            conn.close()

    # ══════════════════════════════════════════════
    #  뷰7: 단가 일관성
    # ══════════════════════════════════════════════

    async def get_price_inconsistency(self, date_from=None, date_to=None) -> list:
        from db.database import get_connection
        conds, params = self._date_filter(date_from, date_to)
        conds.append("unit_price > 0")
        where = "WHERE " + " AND ".join(conds)
        conn = get_connection()
        try:
            rows = conn.execute(f"""
                SELECT customer_name, item_code, model_name,
                    MIN(unit_price) as min_price, MAX(unit_price) as max_price,
                    COUNT(*) as cnt,
                    GROUP_CONCAT(DISTINCT CAST(unit_price AS TEXT)) as prices
                FROM sales_records {where}
                GROUP BY customer_name, item_code
                HAVING min_price != max_price
                ORDER BY (max_price - min_price) DESC
            """, params).fetchall()

            result = []
            for r in rows:
                mn = r["min_price"] or 0
                mx = r["max_price"] or 0
                var = round((mx - mn) / mn * 100, 1) if mn else 0
                prices = sorted(set(str(r["prices"] or "").split(",")))
                result.append({
                    "customer_name": r["customer_name"],
                    "item_code": r["item_code"],
                    "model_name": r["model_name"] or "",
                    "min_price": mn, "max_price": mx,
                    "variance_pct": var, "count": r["cnt"],
                    "prices": prices,
                })
            return result
        finally:
            conn.close()

    async def get_price_standards(self) -> list:
        from db.database import get_connection
        conn = get_connection()
        try:
            rows = conn.execute("SELECT * FROM sales_price_standards ORDER BY item_code").fetchall()
            return [{
                "id": r["id"], "item_code": r["item_code"],
                "customer_name": r["customer_name"] or "",
                "standard_price": r["standard_price"],
                "tolerance_pct": r["tolerance_pct"],
            } for r in rows]
        finally:
            conn.close()

    async def set_price_standard(self, item_code, customer_name, price, tolerance=10.0) -> dict:
        from db.database import get_connection
        conn = get_connection()
        try:
            # UPSERT
            conn.execute("DELETE FROM sales_price_standards WHERE item_code=? AND customer_name=?",
                         (item_code, customer_name or ""))
            conn.execute("""
                INSERT INTO sales_price_standards(item_code, customer_name, standard_price, tolerance_pct)
                VALUES(?,?,?,?)
            """, (item_code, customer_name or "", price, tolerance))
            conn.commit()
            return {"success": True}
        finally:
            conn.close()

    # ══════════════════════════════════════════════
    #  뷰8: 창고×채널 분석
    # ══════════════════════════════════════════════

    async def get_warehouse_channel(self, date_from=None, date_to=None) -> dict:
        from db.database import get_connection
        conds, params = self._date_filter(date_from, date_to)
        where = ("WHERE " + " AND ".join(conds)) if conds else ""
        conn = get_connection()
        try:
            # 창고별
            by_wh = conn.execute(f"""
                SELECT warehouse, COALESCE(SUM(supply_amount),0) as supply, COUNT(*) as cnt
                FROM sales_records {where}
                GROUP BY warehouse ORDER BY supply DESC
            """, params).fetchall()

            # 교차 매트릭스
            cross = conn.execute(f"""
                SELECT warehouse, customer_group, COALESCE(SUM(supply_amount),0) as supply
                FROM sales_records {where}
                GROUP BY warehouse, customer_group
                ORDER BY warehouse, supply DESC
            """, params).fetchall()

            # 배송 관련 (shipments 테이블)
            ship_data = {"logen": 0, "kyungdong": 0, "office": 0, "other": 0}
            shipping_by_cust = []
            try:
                ship_rows = conn.execute("""
                    SELECT COUNT(*) as cnt FROM shipments
                """).fetchone()
                if ship_rows and ship_rows["cnt"] > 0:
                    # 배송수단 분류 (goods_nm, memo 기반)
                    all_ships = conn.execute("SELECT * FROM shipments").fetchall()
                    for sr in all_ships:
                        ship_data["logen"] += 1  # shipments는 모두 로젠
                    # 거래처별 배송 건수
                    ship_cust = conn.execute("""
                        SELECT rcv_name, COUNT(*) as cnt
                        FROM shipments GROUP BY rcv_name ORDER BY cnt DESC LIMIT 10
                    """).fetchall()
                    shipping_by_cust = [{"customer_name": r["rcv_name"], "count": r["cnt"]} for r in ship_cust]
            except Exception:
                pass

            return {
                "by_warehouse": [{"warehouse": r["warehouse"] or "미지정", "supply": r["supply"], "slip_count": r["cnt"]} for r in by_wh],
                "cross_matrix": [{"warehouse": r["warehouse"] or "미지정", "customer_group": r["customer_group"] or "미분류", "supply": r["supply"]} for r in cross],
                "shipping_cost": shipping_by_cust,
                "delivery_breakdown": ship_data,
            }
        finally:
            conn.close()

    # ══════════════════════════════════════════════
    #  에이전트 1: 이탈위험 감지
    # ══════════════════════════════════════════════

    async def detect_churn_risk(self) -> list:
        from db.database import get_connection
        n = _now_kst()
        m0 = n.strftime("%Y%m")
        m1 = (n - timedelta(days=30)).strftime("%Y%m")
        conn = get_connection()
        try:
            rows = conn.execute("""
                SELECT customer_name, SUBSTR(slip_date,1,6) as month,
                    COALESCE(SUM(supply_amount),0) as supply
                FROM sales_records
                WHERE SUBSTR(slip_date,1,6) IN (?,?)
                GROUP BY customer_name, SUBSTR(slip_date,1,6)
            """, (m0, m1)).fetchall()

            data = {}
            for r in rows:
                c = r["customer_name"]
                if c not in data:
                    data[c] = {"prev": 0, "curr": 0}
                if r["month"] == m1:
                    data[c]["prev"] = r["supply"]
                else:
                    data[c]["curr"] = r["supply"]

            alerts = []
            for c, d in data.items():
                if d["prev"] <= 0:
                    continue
                change = (d["curr"] - d["prev"]) / d["prev"] * 100
                if change <= -50:
                    sev = "critical" if change <= -80 else "warning"
                    msg = f"[이탈위험] {c}: 전월 {d['prev']:,.0f} → 이번달 {d['curr']:,.0f} ({change:+.1f}%)"
                    conn.execute("""
                        INSERT INTO sales_alerts(alert_type, target_name, message, severity)
                        VALUES(?,?,?,?)
                    """, ("churn_risk", c, msg, sev))
                    alerts.append({"customer_name": c, "change": round(change, 1), "severity": sev, "message": msg})

            conn.commit()
            return alerts
        finally:
            conn.close()

    # ══════════════════════════════════════════════
    #  에이전트 2: 재고 발주 추천
    # ══════════════════════════════════════════════

    async def detect_stockout_risk(self) -> list:
        items = await self.get_inventory_risk(months_back=1)
        from db.database import get_connection
        conn = get_connection()
        try:
            alerts = []
            for it in items:
                if it["turnover_rate"] >= 1.0:
                    msg = (f"[재고위험] {it['item_code']} ({it['model_name']}): "
                           f"회전율 {it['turnover_rate']:.1f}x, 안전재고 {it['safety_stock']}, "
                           f"월판매 {it['monthly_qty']:.0f}, 추천발주 {it['recommended_order']:.0f}")
                    sev = "critical" if it["risk_level"] == "critical" else "warning"
                    conn.execute("""
                        INSERT INTO sales_alerts(alert_type, target_name, message, severity)
                        VALUES(?,?,?,?)
                    """, ("stockout_risk", it["item_code"], msg, sev))
                    alerts.append({**it, "message": msg, "severity": sev})
            conn.commit()
            return alerts
        finally:
            conn.close()

    # ══════════════════════════════════════════════
    #  에이전트 3: 단가 이탈 감지
    # ══════════════════════════════════════════════

    async def detect_price_anomalies(self) -> list:
        from db.database import get_connection
        conn = get_connection()
        try:
            standards = conn.execute("SELECT * FROM sales_price_standards").fetchall()
            if not standards:
                return []

            alerts = []
            n = _now_kst()
            recent = (n - timedelta(days=30)).strftime("%Y%m%d")

            for std in standards:
                ic = std["item_code"]
                cn = std["customer_name"] or ""
                sp = std["standard_price"]
                tol = std["tolerance_pct"] or 10.0

                conds = ["item_code = ?", "slip_date >= ?", "unit_price > 0"]
                params = [ic, recent]
                if cn:
                    conds.append("customer_name = ?")
                    params.append(cn)

                rows = conn.execute(f"""
                    SELECT slip_no, customer_name, unit_price, slip_date
                    FROM sales_records WHERE {' AND '.join(conds)}
                """, params).fetchall()

                for r in rows:
                    up = r["unit_price"]
                    dev = abs(up - sp) / sp * 100 if sp else 0
                    if dev > tol:
                        msg = (f"[단가이탈] {ic} / {r['customer_name']}: "
                               f"기준 {sp:,.0f} → 실제 {up:,.0f} (편차 {dev:.1f}%)")
                        conn.execute("""
                            INSERT INTO sales_alerts(alert_type, target_name, message, severity)
                            VALUES(?,?,?,?)
                        """, ("price_anomaly", ic, msg, "warning"))
                        alerts.append({"item_code": ic, "customer_name": r["customer_name"],
                                       "standard": sp, "actual": up, "deviation": round(dev, 1),
                                       "message": msg})

            conn.commit()
            return alerts
        finally:
            conn.close()

    # ══════════════════════════════════════════════
    #  알림 관리
    # ══════════════════════════════════════════════

    async def get_alerts(self, is_read=None, limit=50) -> list:
        from db.database import get_connection
        conn = get_connection()
        try:
            conds = []
            params = []
            if is_read is not None:
                conds.append("is_read = ?")
                params.append(1 if is_read else 0)
            where = ("WHERE " + " AND ".join(conds)) if conds else ""
            rows = conn.execute(f"""
                SELECT * FROM sales_alerts {where}
                ORDER BY created_at DESC LIMIT ?
            """, params + [limit]).fetchall()
            return [{
                "id": r["id"], "alert_type": r["alert_type"],
                "target_name": r["target_name"] or "",
                "message": r["message"] or "", "severity": r["severity"],
                "is_read": bool(r["is_read"]),
                "created_at": r["created_at"],
            } for r in rows]
        finally:
            conn.close()

    async def mark_alert_read(self, alert_id) -> dict:
        from db.database import get_connection
        conn = get_connection()
        try:
            conn.execute("UPDATE sales_alerts SET is_read=1 WHERE id=?", (alert_id,))
            conn.commit()
            return {"success": True}
        finally:
            conn.close()

    async def run_all_agents(self) -> dict:
        """3개 에이전트 한번에 실행"""
        # 이전 미읽음 알림 정리 (7일 이상 된 것)
        from db.database import get_connection
        conn = get_connection()
        try:
            conn.execute("DELETE FROM sales_alerts WHERE is_read=0 AND created_at < datetime('now','-7 days')")
            conn.commit()
        except Exception:
            pass
        finally:
            conn.close()

        churn = await self.detect_churn_risk()
        stockout = await self.detect_stockout_risk()
        price = await self.detect_price_anomalies()
        return {
            "churn_risk": len(churn),
            "stockout_risk": len(stockout),
            "price_anomalies": len(price),
            "total_alerts": len(churn) + len(stockout) + len(price),
        }

    # ══════════════════════════════════════════════
    #  에이전트 4: AI 거래처 분석
    # ══════════════════════════════════════════════

    async def get_ai_customer_analysis(self, customer_name: str) -> dict:
        """Claude API로 거래처 인사이트 생성"""
        from config import ANTHROPIC_API_KEY
        if not ANTHROPIC_API_KEY:
            return {"error": "ANTHROPIC_API_KEY 미설정"}

        # 데이터 수집
        detail = await self.get_customer_detail(customer_name)
        health = await self.get_customer_health()
        returns = await self.get_returns_analysis()

        # 해당 거래처 반품 건수
        ret_cnt = 0
        for rc in returns.get("by_customer", []):
            if rc["customer_name"] == customer_name:
                ret_cnt = rc["count"]
                break

        # AI 호출
        import anthropic
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

        data_json = {
            "customer_name": customer_name,
            "customer_group": detail.get("customer_group", ""),
            "total_supply": detail.get("supply", 0),
            "total_profit": detail.get("profit", 0),
            "profit_rate": detail.get("profit_rate", 0),
            "slip_count": detail.get("slip_count", 0),
            "monthly_trend": detail.get("monthly", []),
            "top_products": detail.get("products", [])[:10],
            "return_count": ret_cnt,
        }

        import json
        try:
            response = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=1024,
                system="당신은 랜스타(IT/전자 액세서리 제조사)의 영업 분석 전문가입니다. "
                       "주어진 거래처 데이터를 분석하여 한국어로 인사이트를 제공하세요. "
                       "간결하고 실용적인 제안을 해주세요.",
                messages=[{
                    "role": "user",
                    "content": f"""다음 거래처 데이터를 분석해주세요:

{json.dumps(data_json, ensure_ascii=False, indent=2)}

다음 JSON 형식으로 응답해주세요:
{{
  "summary": "거래처 현황 요약 (2~3줄)",
  "cross_sell": ["교차판매 추천 품목1", "품목2"],
  "risk_level": "low/medium/high",
  "action_items": ["액션 제안1", "액션 제안2"]
}}"""
                }],
            )

            result_text = response.content[0].text
            # JSON 파싱 시도
            try:
                # JSON 블록 추출
                json_match = re.search(r'\{[\s\S]*\}', result_text)
                if json_match:
                    return json.loads(json_match.group())
                return {"summary": result_text, "cross_sell": [], "risk_level": "medium", "action_items": []}
            except json.JSONDecodeError:
                return {"summary": result_text, "cross_sell": [], "risk_level": "medium", "action_items": []}

        except Exception as e:
            logger.error(f"[AI분석] Claude API 오류: {e}")
            return {"error": str(e)}
