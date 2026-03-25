"""
AICC 데이터 로더 — 서버 시작 시 1회 로드 후 메모리 유지
기존 shop/aicc 시스템의 데이터 구조 + 검색 로직 완전 이식
"""
import json, os, re
from typing import Dict, List, Optional, Set
import openpyxl

# Render에서는 cd backend로 시작하므로 ../data/aicc가 올바른 경로
_default_dir = "./data/aicc"
if not os.path.exists(_default_dir) and os.path.exists("../data/aicc"):
    _default_dir = "../data/aicc"
DATA_DIR = os.getenv("AICC_DATA_DIR", _default_dir)

# 드롭다운 제외 조건
EXCLUDE_KEYWORDS = {"주문제작", "취소", "반품불가", "제작케", "★", "제작/취소"}


class AICCDataLoader:
    def __init__(self):
        self.dropdown_models: List[dict] = []
        self.product_data: Dict[str, dict] = {}
        self.faq_list: List[dict] = []
        self.golden_answers: List[dict] = []
        self.wrong_answers_text: str = ""
        self.install_guide_text: str = ""
        self.compatibility_data: List[dict] = []
        self.error_data: List[dict] = []
        self.policy_as: str = ""
        self.policy_delivery: str = ""
        self.policy_return: str = ""
        self.price_restricted: Set[str] = set()
        self._erp_map: Dict[str, str] = {}  # model_name → erp_code
        self._driver_models: Set[str] = set()  # 드라이버가 있는 모델 목록

        # 기존 shop/aicc 시스템 데이터 (searchRelevantQna용)
        self.technical_qna: List[dict] = []     # lanstar_technical_qna.json
        self.unidentified_qna: List[dict] = []  # lanstar_unidentified_qna.json

        # 품목가격정보 (제품문의용)
        self.price_info: Dict[str, dict] = {}   # model_name → {품목명, 딜러가, 온라인노출가}
        self._price_model_map: Dict[str, str] = {}  # 모델명(괄호 제거) → 원본 키

        # 리뷰 데이터
        self.review_data: Dict[str, List[str]] = {}  # model_name → [리뷰텍스트, ...]

        # 유튜브 영상 데이터
        self.youtube_videos: List[dict] = []  # lanstar_videos.json

    def load_all(self):
        print("[AICC] 데이터 로딩 시작...")
        try:
            self._load_master()
            self._load_product_json()
            self._load_faq()
            self._load_golden()
            self._load_wrong()
            self._load_install()
            self._load_compat()
            self._load_errors()
            self._load_policies()
            self._load_price()
            self._load_driver_models()
            self._load_technical_qna()
            self._load_unidentified_qna()
            self._load_price_info()
            self._load_reviews()
            self._load_youtube_videos()
            total_qna = sum(len(p.get("qna", [])) for p in self.technical_qna)
            print(f"[AICC] 완료 — 드롭다운:{len(self.dropdown_models)} 제품:{len(self.product_data)} "
                  f"기술QnA:{total_qna}건({len(self.technical_qna)}모델) 미분류QnA:{len(self.unidentified_qna)}건")
        except Exception as e:
            print(f"[AICC] 로딩 오류: {e}")
            raise

    def _load_master(self):
        path = os.path.join(DATA_DIR, "product_master.xlsx")
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            erp_code, prod_name, model = row[0], row[1], row[2]
            if not model:
                continue
            m = str(model).strip()
            if any(kw in m for kw in EXCLUDE_KEYWORDS):
                continue
            item = {
                "erp_code": str(erp_code).strip() if erp_code else "",
                "product_name": str(prod_name).strip() if prod_name else "",
                "model_name": m,
            }
            self.dropdown_models.append(item)
            self._erp_map[m] = item["erp_code"]
        wb.close()

    def _load_product_json(self):
        path = os.path.join(DATA_DIR, "01_제품별_통합데이터.json")
        with open(path, "r", encoding="utf-8") as f:
            self.product_data = json.load(f)

    def _load_faq(self):
        path = os.path.join(DATA_DIR, "02_FAQ_Top150_카테고리별.xlsx")
        wb = openpyxl.load_workbook(path, read_only=True)
        for row in wb.active.iter_rows(min_row=2, values_only=True):
            if row[0]:
                self.faq_list.append({
                    "category": str(row[1] or ""),
                    "question": str(row[2] or ""),
                    "models": str(row[4] or ""),
                    "answer": str(row[5] or ""),
                })
        wb.close()

    def _load_golden(self):
        path = os.path.join(DATA_DIR, "03_모범답변_골든앤서.xlsx")
        wb = openpyxl.load_workbook(path, read_only=True)
        for row in wb.active.iter_rows(min_row=2, values_only=True):
            if row[0]:
                self.golden_answers.append({
                    "category": str(row[1] or ""),
                    "model": str(row[2] or ""),
                    "question": str(row[3] or ""),
                    "answer": str(row[4] or ""),
                    "keyword": str(row[5] or ""),
                    "warning": str(row[6] or "") if row[6] else "",
                })
        wb.close()

    def _load_wrong(self):
        path = os.path.join(DATA_DIR, "04_오답사례_주의목록.txt")
        with open(path, "r", encoding="utf-8") as f:
            text = f.read()
        self.wrong_answers_text = text[:2000]

    def _load_install(self):
        for fname in ["05_제품별_연결방법_설치가이드_정제.txt", "05_제품별_연결방법_설치가이드.txt"]:
            path = os.path.join(DATA_DIR, fname)
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    self.install_guide_text = f.read()
                return
        print("[AICC] 설치가이드 파일 없음 (스킵)")

    def _load_compat(self):
        for fname in ["06_호환성_매트릭스_정제.xlsx", "06_호환성_매트릭스.xlsx"]:
            path = os.path.join(DATA_DIR, fname)
            if os.path.exists(path):
                wb = openpyxl.load_workbook(path, read_only=True)
                for row in wb.active.iter_rows(min_row=2, values_only=True):
                    if row[1]:
                        self.compatibility_data.append({
                            "model": str(row[1]).strip(),
                            "question": str(row[3] or ""),
                            "answer": str(row[4] or ""),
                        })
                wb.close()
                return
        print("[AICC] 호환성 매트릭스 파일 없음 (스킵)")

    def _load_errors(self):
        path = os.path.join(DATA_DIR, "07_오류증상_대응표.xlsx")
        wb = openpyxl.load_workbook(path, read_only=True)
        for row in wb.active.iter_rows(min_row=2, values_only=True):
            if row[1]:
                self.error_data.append({
                    "model": str(row[1]).strip(),
                    "symptom_type": str(row[3] or ""),
                    "symptom": str(row[4] or ""),
                    "solution": str(row[5] or ""),
                })
        wb.close()

    def _load_policies(self):
        for attr, fname in [
            ("policy_as", "09_AS정책_전문.txt"),
            ("policy_delivery", "10_배송정책.txt"),
            ("policy_return", "11_교환반품_규정.txt"),
        ]:
            path = os.path.join(DATA_DIR, fname)
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    setattr(self, attr, f.read())

    def _load_price(self):
        path = os.path.join(DATA_DIR, "12_가격지도_적용품목.xlsx")
        wb = openpyxl.load_workbook(path, read_only=True)
        for row in wb.active.iter_rows(min_row=2, values_only=True):
            if row[1]:
                self.price_restricted.add(str(row[1]).strip())
        wb.close()

    def _load_driver_models(self):
        """08_기술자료실 파일에서 드라이버가 있는 모델 목록 파싱"""
        path = os.path.join(DATA_DIR, "08_기술자료실_파일목록_URL.txt")
        if not os.path.exists(path):
            print("[AICC] 기술자료실 파일 없음 (스킵)")
            return
        with open(path, "r", encoding="utf-8") as f:
            text = f.read()
        # [모델별 빠른 검색 URL 전체 목록] 섹션에서 모델명 추출
        # 패턴: 줄 시작 공백 + LS-모델명 (URL: 줄 바로 위에 있는 모델명)
        for match in re.finditer(r'^\s+(LS-[\w\-]+)\s*$', text, re.MULTILINE):
            model = match.group(1).strip()
            self._driver_models.add(model)
        print(f"[AICC] 드라이버 모델 로드: {len(self._driver_models)}개")

    def _load_price_info(self):
        """품목가격정보.json 로드 — 제품문의 시 가격 비교용"""
        path = os.path.join(DATA_DIR, "품목가격정보.json")
        if not os.path.exists(path):
            print("[AICC] 품목가격정보.json 없음 (스킵)")
            return
        with open(path, "r", encoding="utf-8") as f:
            raw = json.load(f)
        for full_key, val in raw.items():
            # 키: "LS-H21AOC-5M(AVO-HD선054)" → 모델명: "LS-H21AOC-5M"
            model_name = full_key.split("(")[0].strip()
            self.price_info[model_name] = {
                "품목명": val.get("품목명", ""),
                "딜러가": val.get("딜러가", 0),
                "온라인노출가": val.get("온라인노출가", 0),
                "원본키": full_key,
            }
            self._price_model_map[model_name.upper()] = model_name
        print(f"[AICC] 품목가격정보 로드: {len(self.price_info)}개")

    def _load_reviews(self):
        """review_merged.json 로드 — 고객 리뷰 데이터"""
        path = os.path.join(DATA_DIR, "review_merged.json")
        if not os.path.exists(path):
            print("[AICC] review_merged.json 없음 (스킵)")
            return
        with open(path, "r", encoding="utf-8") as f:
            raw = json.load(f)
        for item in raw:
            model_name = item.get("모델명", "").strip()
            if not model_name:
                continue
            reviews = []
            for r in item.get("리뷰", []):
                text = r.get("리뷰상세내용", "").strip()
                if text:
                    reviews.append(text)
            if reviews:
                self.review_data[model_name] = reviews
        print(f"[AICC] 리뷰 로드: {len(self.review_data)}개 모델")

    def _load_youtube_videos(self):
        """lanstar_videos.json 로드 — 유튜브 영상 데이터"""
        path = os.path.join(DATA_DIR, "lanstar_videos.json")
        if not os.path.exists(path):
            print("[AICC] lanstar_videos.json 없음 (스킵)")
            return
        with open(path, "r", encoding="utf-8") as f:
            raw = json.load(f)
        self.youtube_videos = raw.get("videos", [])
        print(f"[AICC] 유튜브 영상 로드: {len(self.youtube_videos)}개")

    def search_youtube_videos(self, query: str, model: str = "", max_results: int = 3) -> List[dict]:
        """
        고객 질문과 모델명으로 관련 유튜브 영상 검색.
        검색 대상: title, main_topic, sub_topics, keywords, summary
        점수: 모델명 매칭 +5, 키워드 매칭 +1/단어
        띄어쓰기 무시 매칭으로 '멀티허브'='멀티 허브' 등 처리
        """
        if not self.youtube_videos:
            return []

        upper_query = query.upper()
        words = [w for w in re.split(r'[\s,]+', upper_query) if len(w) >= 2]
        if not words and not model:
            return []

        # 모델명에서 LS- 제거한 검색어도 준비
        model_upper = model.upper() if model else ""
        model_short = model.replace("LS-", "").upper() if model else ""
        model_base = self._extract_model_base(model).upper() if model else ""

        results = []
        for video in self.youtube_videos:
            title = video.get("title", "")
            main_topic = video.get("main_topic", "")
            sub_topics = " ".join(video.get("sub_topics", []))
            keywords_list = video.get("keywords", [])
            keywords_text = " ".join(keywords_list)
            summary = video.get("summary", "")

            search_text = f"{title} {main_topic} {sub_topics} {keywords_text} {summary}".upper()
            # 띄어쓰기 제거 버전 (멀티허브=멀티 허브, 도킹스테이션=도킹 스테이션 등)
            search_text_nospace = search_text.replace(" ", "")

            score = 0

            # 모델명 매칭 (정확한 모델명이 영상에 포함된 경우)
            if model_upper and model_upper in search_text:
                score += 5
            elif model_short and model_short in search_text:
                score += 4
            elif model_base and model_base != model_upper and model_base in search_text:
                score += 3

            # 키워드 매칭 (원본 + 띄어쓰기 제거 양쪽에서 검색)
            for w in words:
                if w in search_text or w in search_text_nospace:
                    score += 1

            if score > 0:
                results.append({
                    "title": title,
                    "url": video.get("url", ""),
                    "summary": summary,
                    "keywords": keywords_list,
                    "main_topic": main_topic,
                    "duration": video.get("duration", ""),
                    "score": score,
                })

        # 점수 내림차순 정렬
        results.sort(key=lambda x: x["score"], reverse=True)
        return results[:max_results]

    @staticmethod
    def _extract_model_base(model: str) -> str:
        """
        모델명에서 길이 접미사를 제거하여 베이스 모델명 추출.
        길이 표기 패턴 (모두 같은 제품의 길이 변형):
          1) -숫자M / -숫자.숫자M  : LS-HF-30M → LS-HF
          2)  숫자M /  숫자.숫자M  : LS-7SD-BK1M → LS-7SD-BK
          3) -숫자  / -숫자.숫자   : LS-HF-30 → LS-HF
        길이 접미사가 없으면 원본 그대로 반환.
        """
        # 1) -숫자M / -숫자.숫자M (대시 + 숫자 + M)
        m = re.match(r'^(.+?)-\d+(?:\.\d+)?M$', model, re.IGNORECASE)
        if m:
            return m.group(1)
        # 2) 숫자M / 숫자.숫자M (대시 없이 문자 바로 뒤 숫자 + M)
        m = re.match(r'^(.+?[A-Za-z])\d+(?:\.\d+)?M$', model, re.IGNORECASE)
        if m:
            return m.group(1)
        # 3) -숫자 / -숫자.숫자 (대시 + 숫자, M 없음)
        m = re.match(r'^(.+?)-\d+(?:\.\d+)?$', model)
        if m:
            return m.group(1)
        return model

    def search_reviews(self, model: str, max_reviews: int = 5) -> List[str]:
        """
        모델명으로 리뷰 검색.
        1차: 정확한 모델명 매칭
        2차: 베이스 모델명 매칭 (길이 변형 통합, 예: LS-HF-30M → LS-HF 계열 전체)
        """
        # 1차: 정확 매칭
        reviews = self.review_data.get(model, [])
        if reviews:
            return reviews[:max_reviews]

        # 2차: 베이스 모델명 매칭 (길이 접미사 제거 후 비교)
        target_base = self._extract_model_base(model).upper()
        collected = []
        for review_model, review_list in self.review_data.items():
            review_base = self._extract_model_base(review_model).upper()
            if review_base == target_base:
                collected.extend(review_list)
        return collected[:max_reviews]

    def search_products_for_recommendation(self, query: str, max_results: int = 15) -> List[dict]:
        """
        고객 질문에서 키워드를 추출하여 제품 추천용 데이터 검색.
        01_제품별_통합데이터.json + 품목가격정보.json 교차 검색.
        """
        upper = query.upper()
        words = [w for w in re.split(r'[\s,]+', upper) if len(w) >= 2]
        if not words:
            return []

        results = []
        seen_models = set()

        # 1. product_data (01_제품별_통합데이터.json) 검색
        for model_name, product in self.product_data.items():
            feat = product.get("제품특징", {})
            cat = product.get("카테고리", "")

            # 검색 대상 텍스트 구성
            if isinstance(feat, dict):
                search_text = " ".join(str(v) for v in feat.values()) + " " + cat + " " + model_name
            elif isinstance(feat, list):
                search_text = " ".join(feat) + " " + cat + " " + model_name
            else:
                search_text = str(feat) + " " + cat + " " + model_name
            search_text = search_text.upper()

            score = 0
            for w in words:
                if w in search_text:
                    score += 1

            if score > 0:
                # 가격 정보 매칭
                price_data = self.price_info.get(model_name, {})
                product_name = ""
                if isinstance(feat, dict):
                    product_name = feat.get("제품명_full", "") or feat.get("용도", "")
                if not product_name and price_data:
                    product_name = price_data.get("품목명", "")

                results.append({
                    "model_name": model_name,
                    "category": cat,
                    "product_name": product_name,
                    "features": feat,
                    "price_tier": price_data.get("온라인노출가", 0),
                    "score": score,
                })
                seen_models.add(model_name.upper())

        # 2. price_info에만 있는 제품도 검색 (product_data에 없는 것)
        for model_name, pinfo in self.price_info.items():
            if model_name.upper() in seen_models:
                continue
            search_text = (pinfo.get("품목명", "") + " " + model_name).upper()
            score = 0
            for w in words:
                if w in search_text:
                    score += 1
            if score > 0:
                results.append({
                    "model_name": model_name,
                    "category": "",
                    "product_name": pinfo.get("품목명", ""),
                    "features": {},
                    "price_tier": pinfo.get("온라인노출가", 0),
                    "score": score,
                })

        # 점수 내림차순 정렬
        results.sort(key=lambda x: x["score"], reverse=True)
        return results[:max_results]

    def get_price_rank(self, model_name: str) -> Optional[int]:
        """해당 모델의 온라인노출가 반환 (상대 비교용)"""
        pinfo = self.price_info.get(model_name)
        if pinfo:
            return pinfo.get("온라인노출가", 0)
        return None

    def _load_technical_qna(self):
        """기존 shop/aicc의 lanstar_technical_qna.json 로드"""
        path = os.path.join(DATA_DIR, "lanstar_technical_qna.json")
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                self.technical_qna = json.load(f)
            print(f"[AICC] 기술QnA 로드: {len(self.technical_qna)}개 모델")
        else:
            print("[AICC] lanstar_technical_qna.json 없음 (스킵)")

    def _load_unidentified_qna(self):
        """기존 shop/aicc의 lanstar_unidentified_qna.json 로드"""
        path = os.path.join(DATA_DIR, "lanstar_unidentified_qna.json")
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                self.unidentified_qna = json.load(f)
            print(f"[AICC] 미분류QnA 로드: {len(self.unidentified_qna)}건")
        else:
            print("[AICC] lanstar_unidentified_qna.json 없음 (스킵)")

    # ── 조회 메서드 ──────────────────────────────────────────

    def search_models(self, query: str, limit: int = 15) -> List[dict]:
        q = query.upper().strip()
        if len(q) < 2:
            return []
        return [
            m for m in self.dropdown_models
            if q in m["model_name"].upper() or q in m["product_name"].upper()
        ][:limit]

    def get_product(self, model: str) -> dict:
        return self.product_data.get(model, {})

    def get_erp_code(self, model: str) -> Optional[str]:
        return self._erp_map.get(model)

    def get_install_section(self, model: str) -> str:
        pattern = rf'\[{re.escape(model)}\](.*?)(?=\n\[LS-|\Z)'
        m = re.search(pattern, self.install_guide_text, re.DOTALL)
        return m.group(1).strip()[:800] if m else ""

    def get_compat(self, model: str) -> List[dict]:
        return [d for d in self.compatibility_data if d["model"] == model][:4]

    def get_errors(self, model: str) -> List[dict]:
        return [d for d in self.error_data if d["model"] == model][:4]

    def get_golden_by_category(self, cat: str) -> List[dict]:
        return [g for g in self.golden_answers if g["category"] == cat][:3]

    def get_golden_by_model(self, model: str) -> List[dict]:
        return [g for g in self.golden_answers if g["model"] == model][:5]

    def get_faq_by_model(self, model: str) -> List[dict]:
        return [f for f in self.faq_list if model in f.get("models", "")][:5]

    def has_driver(self, model: str) -> bool:
        """해당 모델에 드라이버가 있는지 확인"""
        return model in self._driver_models

    def get_driver_url(self, model: str) -> str:
        search_word = model.replace("LS-", "").lower() if model.startswith("LS-") else model.lower()
        return f"https://www.lanstar.co.kr/board/list.php?bdId=lanstardownload&memNo=&noheader=&mypageFl=&searchField=subject&searchWord={search_word}"

    def get_product_url(self, model: str) -> str:
        return f"https://www.lanstar.co.kr/goods/goods_search.php?keyword={model}&recentCount=10"

    def is_price_restricted(self, model: str) -> bool:
        return model in self.price_restricted

    # ── 기존 shop/aicc searchRelevantQna 완전 이식 ──────────

    def _get_related_models(self, model: str) -> set:
        """같은 제품군의 관련 모델명 반환 (ADOOR↔ANDOOR 등)"""
        if not model:
            return set()
        upper = model.upper()
        related = {upper}

        # ADOOR ↔ ANDOOR 매핑 (같은 도어락 제품군)
        if "ANDOOR" in upper:
            base = upper.replace("ANDOOR", "ADOOR")
            related.add(base)
            # LS-ANDOOR-S → LS-ADOOR, LS-ADOOR-S 등
            related.add(re.sub(r'-[SB]$', '', base))
            related.add(re.sub(r'-[SB]$', '', upper))
        elif "ADOOR" in upper:
            base = upper.replace("ADOOR", "ANDOOR")
            related.add(base)
            related.add(re.sub(r'-[SB]$', '', base))
            related.add(re.sub(r'-[SB]$', '', upper))

        # 같은 제품의 변형 (LS-XXX-S, LS-XXX-B → LS-XXX 공통)
        base_no_suffix = re.sub(r'-[SB]$', '', upper)
        if base_no_suffix != upper:
            related.add(base_no_suffix)

        return related

    def search_relevant_qna(self, query: str, session_model: str, max_results: int = 10) -> List[dict]:
        """
        강화된 QnA 검색:
        1. lanstar_technical_qna.json의 모든 제품 QnA 검색
        2. lanstar_unidentified_qna.json (미분류 QnA) 검색
        3. 같은 모델 = +5점, 관련 모델(ADOOR↔ANDOOR) = +4점, 키워드 매칭 = +1점/단어
        4. [강화] 핵심 키워드(초기화, 등록, 설치 등) 매칭 시 가산점
        5. [강화] 같은 모델 QnA는 최소 보장 슬롯 확보
        """
        upper = query.upper()
        words = [w for w in re.split(r'[\s,]+', upper) if len(w) >= 2]
        results = []

        # 관련 모델 그룹 (ADOOR↔ANDOOR, -S↔-B 등)
        related_models = self._get_related_models(session_model)

        # [강화] 핵심 키워드 — 도어락/기술문의에서 자주 등장하는 단어에 가산점
        _BOOST_KEYWORDS = {
            "초기화": 3, "공장초기화": 4, "리셋": 3, "RESET": 3,
            "등록": 2, "지문": 2, "비밀번호": 2, "패스워드": 2,
            "설치": 2, "연결": 2, "인식": 2, "오류": 2, "안됨": 2, "안돼": 2,
            "충전": 2, "배터리": 2, "건전지": 2, "마스터키": 3, "비상키": 3,
            "음량": 2, "소리": 2, "볼륨": 2, "상시열림": 3,
        }

        # 1. technical_qna (모델별 QnA)
        for product in self.technical_qna:
            product_model = product.get("model", "")
            product_model_upper = product_model.upper()
            for qna in product.get("qna", []):
                q_text = qna.get("question", "")
                a_text = qna.get("answer", "")
                text = (q_text + " " + a_text + " " + product_model).upper()

                score = 0
                is_same_model = False
                # 정확히 같은 모델
                if session_model and product_model_upper == session_model.upper():
                    score += 5
                    is_same_model = True
                # 관련 모델 (ADOOR↔ANDOOR 등) — 같은 제품군이므로 높은 점수
                elif product_model_upper in related_models or any(
                    rm in product_model_upper or product_model_upper in rm
                    for rm in related_models
                ):
                    score += 4
                    is_same_model = True  # 관련 모델도 같은 제품군으로 취급

                # 기본 키워드 매칭
                for w in words:
                    if w in text:
                        score += 1

                # [강화] 핵심 키워드 가산점 — 질문 의도와 QnA 내용이 일치할 때
                for kw, boost in _BOOST_KEYWORDS.items():
                    kw_upper = kw.upper()
                    if kw_upper in upper and kw_upper in text:
                        score += boost

                if score > 0:
                    results.append({
                        "model": product_model,
                        "category": product.get("category", ""),
                        "question": q_text,
                        "answer": a_text,
                        "score": score,
                        "_same_model": is_same_model,
                    })

        # 2. unidentified_qna (미분류 QnA — 도어락 등 특수 제품 커버)
        for qna in self.unidentified_qna:
            q_text = qna.get("question", "")
            a_text = qna.get("answer", "")
            text = (q_text + " " + a_text).upper()

            score = 0
            is_same_model = False
            for w in words:
                if w in text:
                    score += 1
            # 미분류지만 모델명이 포함되어 있으면 가산
            if session_model and session_model.upper() in text:
                score += 3
                is_same_model = True

            # [강화] 핵심 키워드 가산점
            for kw, boost in _BOOST_KEYWORDS.items():
                kw_upper = kw.upper()
                if kw_upper in upper and kw_upper in text:
                    score += boost

            if score > 0:
                results.append({
                    "model": qna.get("original_product_name", ""),
                    "category": "",
                    "question": q_text,
                    "answer": a_text,
                    "score": score,
                    "_same_model": is_same_model,
                })

        # [강화] 같은 모델/관련 모델 QnA를 최소 보장 (최소 절반은 같은 모델)
        same_model_results = [r for r in results if r.get("_same_model")]
        other_results = [r for r in results if not r.get("_same_model")]
        same_model_results.sort(key=lambda x: x["score"], reverse=True)
        other_results.sort(key=lambda x: x["score"], reverse=True)

        # 같은 모델 QnA가 있으면 최소 절반 슬롯 보장
        min_same = min(len(same_model_results), max(max_results // 2, 3))
        combined = same_model_results[:min_same]
        remaining_slots = max_results - len(combined)
        # 나머지 슬롯은 점수순으로 채움 (같은 모델 남은 것 + 다른 모델)
        leftover = same_model_results[min_same:] + other_results
        leftover.sort(key=lambda x: x["score"], reverse=True)
        combined.extend(leftover[:remaining_slots])

        # 중복 제거 (질문 앞 50자)
        seen = set()
        unique = []
        for r in combined:
            key = r["question"][:50]
            if key not in seen:
                seen.add(key)
                # _same_model 내부 플래그 제거
                r.pop("_same_model", None)
                unique.append(r)

        return unique[:max_results]


data_loader = AICCDataLoader()
