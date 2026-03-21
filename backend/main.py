"""
Order Agent - FastAPI 메인 앱
"""
import asyncio
import logging
import time
from pathlib import Path
from fastapi import FastAPI, Request
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, Response
import sys
sys.path.insert(0, str(Path(__file__).parent))

from db.database import init_db
from api.routes.orders import router as orders_router
from api.routes.customers import router as customers_router
from api.routes.auth import router as auth_router
from api.routes.settings import router as settings_router
from api.routes.inventory import router as inventory_router
from api.routes.sale_orders import router as sale_orders_router
from api.routes.materials import router as materials_router
from api.routes.training import router as training_router
from api.routes.orderlist import router as orderlist_router
from api.routes.activity import router as activity_router
from api.routes.shipping import router as shipping_router
from api.routes.cs import router as cs_router

from api.routes.sales_analytics import router as sales_analytics_router
from api.routes.purchases import router as purchases_router
from api.routes.aicc import router as aicc_router
from api.routes.aicc_ws import customer_ws_handler, admin_ws_handler, admin_list_ws_handler
from fastapi import WebSocket

# Super Agent
try:
    from super_agent.routes.jobs import router as super_agent_router
    _HAS_SUPER_AGENT = True
except ImportError as _sa_err:
    _HAS_SUPER_AGENT = False
    logging.getLogger(__name__).warning(f"Super Agent 모듈 로드 실패: {_sa_err}")

# ─────────────────────────────────────────
#  로깅 설정
# ─────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s - %(message)s"
)
logger = logging.getLogger(__name__)

# ─────────────────────────────────────────
#  앱 초기화
# ─────────────────────────────────────────
app = FastAPI(
    title="AI 발주서 자동화 시스템",
    description="거래처 발주서를 AI로 자동 분석하여 ECOUNT ERP 판매 전표를 생성합니다.",
    version="0.2.0",
    docs_url="/docs" if __import__("config").DEBUG else None,
    redoc_url=None,
)

# CORS 설정 (환경변수에서 허용 도메인 지정)
from config import ALLOWED_ORIGINS, DEBUG
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE"],
    allow_headers=["Content-Type", "Authorization"],
)


# ─────────────────────────────────────────
#  보안 헤더 미들웨어
# ─────────────────────────────────────────
@app.middleware("http")
async def security_headers_middleware(request: Request, call_next):
    start_time = time.time()
    response = await call_next(request)

    # 보안 헤더 추가
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    if not DEBUG:
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"

    # 요청 처리 시간 로깅 (API 엔드포인트만)
    if request.url.path.startswith("/api/"):
        duration = (time.time() - start_time) * 1000
        if duration > 5000:  # 5초 이상이면 경고
            logger.warning(
                f"[Slow API] {request.method} {request.url.path} "
                f"- {duration:.0f}ms"
            )

    return response


# ─────────────────────────────────────────
#  활동 로그 미들웨어
# ─────────────────────────────────────────
# 기록할 API 액션 매핑 (경로 prefix → 액션 라벨)
_ACTIVITY_ACTIONS = {
    ("POST", "/api/auth/login"): "로그인",
    ("POST", "/api/orders/process"): "발주서 분석",
    ("POST", "/api/orders/process-image"): "발주서 이미지 분석",
    ("POST", "/api/orders/confirm"): "발주서 확정",
    ("POST", "/api/orders/submit-erp"): "발주서 ERP 전송",
    ("POST", "/api/sale-orders/process"): "견적서 분석",
    ("POST", "/api/sale-orders/process-image"): "견적서 이미지 분석",
    ("POST", "/api/sale-orders/confirm"): "견적서 확정",
    ("POST", "/api/sale-orders/submit-erp"): "견적서 ERP 전송",
    ("GET", "/api/inventory/search"): "재고 조회",
    ("POST", "/api/inventory/check"): "재고 확인",
    ("POST", "/api/materials/sync"): "자료 동기화",
    ("POST", "/api/orderlist/sync"): "오더리스트 동기화",
    ("POST", "/api/settings/models"): "AI 모델 변경",
    ("POST", "/api/training/upload"): "학습 데이터 업로드",
    ("POST", "/api/training/bulk/create-session"): "대량 학습 세션 생성",
    ("POST", "/api/shipping/register"): "택배 발송 등록",
    ("POST", "/api/shipping/register-bulk"): "택배 대량 등록",
    ("POST", "/api/shipping/upload-excel"): "택배 엑셀 업로드",
    ("POST", "/api/shipping/track"): "택배 화물추적",
    ("POST", "/api/shipping/sync"): "택배 운송장 동기화",
    ("POST", "/api/shipping/sync-excel"): "택배 엑셀 동기화",
    ("POST", "/api/shipping/auto-fetch"): "SmartLogen 자동 가져오기",
    ("POST", "/api/shipping/scheduler/run-now"): "택배 동기화 즉시 실행",
    ("GET", "/api/shipping/search"): "택배 검색",
    ("GET", "/api/shipping/daily"): "택배 일별 조회",
    ("POST", "/api/cs/tickets"): "CS 불량 접수",
    ("PUT", "/api/cs/tickets"): "CS 상태 변경",
    ("POST", "/api/cs/tickets"): "CS 테스트/파일",

    ("POST", "/api/sales/upload-csv"): "판매현황 CSV 업로드",
    ("POST", "/api/sales/fetch-ecount"): "판매현황 이카운트 수집",
    ("POST", "/api/sales/scheduler/run-now"): "판매현황 즉시 수집",
    ("POST", "/api/sales/agents/run"): "판매현황 에이전트 실행",
    ("GET", "/api/sales/agents/ai-analysis"): "판매현황 AI 거래처 분석",
    ("POST", "/api/purchases/process"): "구매입력 텍스트 분석",
    ("POST", "/api/purchases/process-image"): "구매입력 이미지 분석",
    ("POST", "/api/purchases/confirm"): "구매입력 확정",
    ("POST", "/api/purchases/submit-erp"): "구매입력 ERP 전송",
    ("POST", "/api/super-agent/jobs"): "Super Agent 작업 생성",
    ("GET", "/api/super-agent/jobs"): "Super Agent 이력 조회",
}

@app.middleware("http")
async def activity_log_middleware(request: Request, call_next):
    """주요 API 호출을 활동 로그에 기록"""
    response = await call_next(request)

    # 성공한 API 호출만 기록
    if response.status_code < 400:
        method = request.method
        path = request.url.path

        # 정확한 매칭 또는 prefix 매칭
        action_label = None
        for (m, p), label in _ACTIVITY_ACTIONS.items():
            if method == m and path.startswith(p):
                action_label = label
                break

        if action_label:
            # JWT에서 사용자 정보 추출 (미들웨어에서는 직접 파싱)
            try:
                from security import verify_token
                auth_header = request.headers.get("authorization", "")
                token = auth_header.replace("Bearer ", "") if auth_header.startswith("Bearer") else ""

                emp_cd = ""
                emp_name = ""

                if token:
                    payload = verify_token(token)
                    if payload:
                        emp_cd = payload.get("emp_cd", "")
                        emp_name = payload.get("name", "")

                # 로그인 성공 시 body에서 emp_cd 추출 (토큰 없음)
                if action_label == "로그인" and not emp_cd:
                    # 로그인 응답이 성공이면 request body에서 추출하기 어려우므로
                    # path 기반으로만 기록
                    pass

                if emp_cd:
                    ip = request.headers.get("x-forwarded-for", request.client.host if request.client else "")
                    detail = path
                    # query string 포함
                    if request.url.query:
                        detail += f"?{request.url.query}"

                    from services.activity_service import log_activity
                    log_activity(emp_cd, emp_name, action_label, detail, ip)
            except Exception as e:
                logger.debug(f"[ActivityLog] 미들웨어 기록 실패 (무시): {e}")

    return response


# ─────────────────────────────────────────
#  글로벌 예외 핸들러
# ─────────────────────────────────────────
@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    logger.error(f"[Unhandled] {request.method} {request.url.path}: {exc}", exc_info=True)
    if DEBUG:
        detail = str(exc)
    else:
        # 프로덕션: 에러 유형별 사용자 친화적 메시지
        exc_str = str(exc).lower()
        if "connection" in exc_str or "timeout" in exc_str:
            detail = "데이터베이스 연결에 실패했습니다. 잠시 후 다시 시도해주세요."
        elif "permission" in exc_str or "access" in exc_str:
            detail = "접근 권한이 없습니다."
        elif "not found" in exc_str:
            detail = "요청한 데이터를 찾을 수 없습니다."
        else:
            detail = "서버 처리 중 오류가 발생했습니다. 잠시 후 다시 시도해주세요."
    return JSONResponse(status_code=500, content={"detail": detail})


# 라우터 등록
app.include_router(orders_router)
app.include_router(customers_router)
app.include_router(auth_router)
app.include_router(settings_router)
app.include_router(inventory_router)
app.include_router(sale_orders_router)
app.include_router(materials_router)
app.include_router(training_router)
app.include_router(orderlist_router)
app.include_router(activity_router)
app.include_router(shipping_router)
app.include_router(cs_router)

app.include_router(sales_analytics_router)
app.include_router(purchases_router)
app.include_router(aicc_router, prefix="/api/aicc", tags=["AICC"])

# Super Agent 라우터
if _HAS_SUPER_AGENT:
    app.include_router(super_agent_router)

# AICC WebSocket 엔드포인트
@app.websocket("/ws/aicc/chat/{session_id}")
async def ws_chat(websocket: WebSocket, session_id: str):
    await customer_ws_handler(websocket, session_id)

@app.websocket("/ws/aicc/admin/{session_id}")
async def ws_admin(websocket: WebSocket, session_id: str):
    await admin_ws_handler(websocket, session_id)

@app.websocket("/ws/aicc/admin-list")
async def ws_admin_list(websocket: WebSocket):
    await admin_list_ws_handler(websocket)

# AI 대시보드 라우터
try:
    from api.routes.dashboard import router as dashboard_router
    app.include_router(dashboard_router)
except ImportError:
    pass

# 정적 파일 (프론트엔드) — JS 캐시 방지
FRONTEND_DIR = Path(__file__).parent.parent / "frontend"
STATIC_DIR   = FRONTEND_DIR / "static"

# JS 파일은 캐시 방지 헤더와 함께 직접 서빙
@app.get("/static/js/{filename}")
async def serve_js_no_cache(filename: str):
    file_path = STATIC_DIR / "js" / filename
    if not file_path.exists() or not file_path.is_file():
        return JSONResponse({"error": "not found"}, status_code=404)
    return FileResponse(
        file_path,
        media_type="application/javascript",
        headers={
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )

if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


# ─────────────────────────────────────────
#  시작 이벤트
# ─────────────────────────────────────────
@app.on_event("startup")
async def startup():
    logger.info("=== Order Agent 시작 (v0.2.0) ===")
    init_db()
    logger.info("데이터베이스 초기화 완료")

    # Super Agent DB 테이블 초기화
    try:
        from super_agent.db.sa_tables import init_super_agent_tables
        from db.database import get_connection
        sa_conn = get_connection()
        init_super_agent_tables(sa_conn)
        sa_conn.close()
        logger.info("Super Agent 테이블 초기화 완료")
    except Exception as e:
        logger.warning(f"Super Agent 테이블 초기화 실패 (서비스는 계속): {e}")

    # AICC 데이터 로드
    try:
        from services.aicc_data_loader import data_loader as aicc_loader
        aicc_loader.load_all()
        logger.info(f"AICC 데이터 로드 완료: 모델 {len(aicc_loader.dropdown_models)}개")
    except Exception as e:
        import traceback
        logger.warning(f"AICC 데이터 로딩 실패 (서비스는 계속): {e}\n{traceback.format_exc()}")

    # AICC: 이전 active 세션을 closed로 정리 (재배포 시 인메모리 초기화되므로)
    try:
        from db.database import get_connection as _get_conn
        _conn = _get_conn()
        _conn.execute("UPDATE aicc_sessions SET status='closed' WHERE status='active'")
        _conn.commit()
        _conn.close()
        logger.info("AICC 이전 active 세션 → closed 정리 완료")
    except Exception as e:
        logger.warning(f"AICC 세션 정리 실패: {e}")

    # AICC 제품 지식 DB 자동 로드 (JSON → DB, 매 배포시 최신 반영)
    try:
        from services import aicc_db as _aicc_db
        import json as _json
        knowledge_path = Path(__file__).parent.parent / "data" / "aicc" / "lanstar_product_knowledge.json"
        if knowledge_path.exists():
            with open(knowledge_path, encoding="utf-8") as f:
                products = _json.load(f)
            count = _aicc_db.bulk_upsert_product_knowledge(products)
            logger.info(f"AICC 제품 지식 DB 동기화 완료: {count}개 제품 (upsert)")
    except Exception as e:
        logger.warning(f"AICC 제품 지식 DB 로드 실패 (서비스는 계속): {e}")

    # products.csv 확인
    from config import PRODUCTS_CSV
    if PRODUCTS_CSV.exists():
        import csv
        with open(PRODUCTS_CSV, encoding="utf-8-sig") as f:
            row_count = sum(1 for _ in f) - 1  # 헤더 제외
        logger.info(f"products.csv 로드 완료: {PRODUCTS_CSV} ({row_count}개 품목)")
    else:
        logger.error(f"products.csv 파일을 찾을 수 없습니다: {PRODUCTS_CSV}")

    # DB에 저장된 모델 설정 복원
    try:
        from api.routes.settings import ensure_settings_table
        ensure_settings_table()
        from db.database import get_connection
        conn = get_connection()
        row = conn.execute("SELECT value FROM app_settings WHERE key='claude_model'").fetchone()
        conn.close()
        if row:
            import config as cfg
            cfg.CLAUDE_MODEL = row["value"]
            logger.info(f"Claude 모델 복원: {row['value']}")
    except Exception as e:
        logger.warning(f"모델 설정 복원 실패 (기본값 사용): {e}")

    # AI 메트릭 테이블 초기화
    try:
        from services.ai_metrics import ensure_metrics_tables
        ensure_metrics_tables()
        logger.info("AI 메트릭 테이블 초기화 완료")
    except Exception as e:
        logger.warning(f"AI 메트릭 테이블 초기화 실패: {e}")

    # 활동 로그 테이블 초기화
    try:
        from services.activity_service import ensure_activity_table
        ensure_activity_table()
        logger.info("활동 로그 테이블 초기화 완료")
    except Exception as e:
        logger.warning(f"활동 로그 테이블 초기화 실패: {e}")

    # 거래처 자동 동기화: customers 테이블이 비어있으면 ERP에서 자동 가져오기
    asyncio.create_task(_auto_sync_customers())

    # 자료관리: 서버 시작 시 마지막 동기화가 오래됐으면 자동 동기화
    asyncio.create_task(_auto_sync_on_startup())

    # SmartLogen 택배 자동 동기화 스케줄러 (매일 09:00 KST)
    try:
        from services.scheduler_service import start_scheduler, check_and_run_on_startup
        start_scheduler()
        asyncio.create_task(check_and_run_on_startup())
        logger.info("SmartLogen 자동 동기화 스케줄러 등록 완료 (1시간 간격)")
    except Exception as e:
        logger.warning(f"택배 스케줄러 시작 실패: {e}")

    # 판매현황 자동 수집 + 에이전트 스케줄러
    try:
        from services.sales_analytics_service import SalesAnalyticsService
        _sales_svc = SalesAnalyticsService()

        def _sales_sync_wrapper(coro_func):
            """비동기 함수를 동기 래핑 (APScheduler용)"""
            def wrapper():
                try:
                    loop = asyncio.new_event_loop()
                    asyncio.set_event_loop(loop)
                    try:
                        loop.run_until_complete(coro_func())
                    finally:
                        loop.close()
                except Exception as e:
                    logger.error(f"[판매현황 스케줄러] 오류: {e}", exc_info=True)
            return wrapper

        from services.scheduler_service import _scheduler_state
        scheduler = _scheduler_state.get("scheduler")
        if scheduler:
            from apscheduler.triggers.cron import CronTrigger
            # 매시간 정각 — 이카운트 자동 수집
            scheduler.add_job(
                _sales_sync_wrapper(_sales_svc.auto_fetch_from_ecount),
                CronTrigger(minute=0, timezone="Asia/Seoul"),
                id="sales_auto_fetch", replace_existing=True
            )
            # 매시간 5분 — 에이전트 3종 실행
            scheduler.add_job(
                _sales_sync_wrapper(_sales_svc.run_all_agents),
                CronTrigger(minute=5, timezone="Asia/Seoul"),
                id="sales_agents", replace_existing=True
            )
            logger.info("판매현황 스케줄러 등록 완료 (매시간 자동수집 + 에이전트)")
    except Exception as e:
        logger.warning(f"판매현황 스케줄러 시작 실패: {e}")


# ─────────────────────────────────────────
#  자료관리 자동 동기화 (서버 시작 시)
# ─────────────────────────────────────────
async def _auto_sync_customers():
    """서버 시작 시 거래처 테이블이 비어있으면 엑셀 파일에서 자동 임포트"""
    await asyncio.sleep(2)  # DB 초기화 대기
    try:
        from db.database import get_connection
        conn = get_connection()
        row = conn.execute("SELECT COUNT(*) as cnt FROM customers").fetchone()
        cnt = row["cnt"] if row else 0
        conn.close()

        if cnt == 0:
            # data/customer.xlsx에서 거래처 임포트
            import openpyxl
            xlsx_path = Path(__file__).parent.parent / "data" / "customer.xlsx"
            if not xlsx_path.exists():
                logger.warning(f"[자동동기화] 거래처 엑셀 파일 없음: {xlsx_path}")
                return

            logger.info(f"[자동동기화] 거래처 테이블이 비어있음 - 엑셀에서 임포트 시작: {xlsx_path}")
            wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
            ws = wb[wb.sheetnames[0]]

            customers = []
            for i, row_data in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:  # 헤더 스킵
                    continue
                code = str(row_data[0] or "").strip()
                name = str(row_data[1] or "").strip()
                if code and name:
                    customers.append((code, name))
            wb.close()

            if customers:
                conn = get_connection()
                for code, name in customers:
                    conn.execute(
                        "INSERT OR IGNORE INTO customers(cust_code, cust_name, alias) VALUES(?,?,?)",
                        (code, name, "")
                    )
                conn.commit()
                conn.close()
                logger.info(f"[자동동기화] 거래처 {len(customers)}개 엑셀에서 임포트 완료")
            else:
                logger.warning("[자동동기화] 엑셀 파일에서 거래처를 찾지 못했습니다")
        else:
            logger.info(f"[자동동기화] 거래처 {cnt}개 존재 - 임포트 불필요")
    except Exception as e:
        logger.error(f"[자동동기화] 거래처 임포트 오류: {e}", exc_info=True)


async def _auto_sync_on_startup():
    """
    서버 시작 시 마지막 동기화가 6시간 이상 지났으면 자동 동기화.
    Render 무료 플랜은 5분 후 서버가 꺼지므로, 오전 9시 스케줄러 방식은 작동하지 않음.
    대신 서버가 깨어날 때마다 동기화 필요 여부를 체크.
    """
    from datetime import datetime, timedelta

    # 서버 시작 직후 약간의 딜레이 (DB 초기화 완료 대기)
    await asyncio.sleep(5)

    try:
        from db.database import get_connection
        conn = get_connection()

        # 소스 유형별 가장 오래된 동기화 시간 확인
        row = conn.execute(
            "SELECT MIN(last_synced) as oldest FROM material_sources WHERE last_synced != '' AND is_active=1"
        ).fetchone()
        conn.close()

        oldest_sync = row["oldest"] if row and row["oldest"] else None
        need_sync = True

        if oldest_sync:
            try:
                last_dt = datetime.strptime(str(oldest_sync)[:19], "%Y-%m-%d %H:%M:%S")
                hours_ago = (datetime.now() - last_dt).total_seconds() / 3600
                logger.info(f"[자동동기화] 가장 오래된 동기화: {oldest_sync} ({hours_ago:.1f}시간 전)")
                if hours_ago < 6:
                    need_sync = False
                    logger.info("[자동동기화] 모든 소스 최근 동기화됨 → 스킵")
            except (ValueError, TypeError) as e:
                logger.warning(f"[자동동기화] 날짜 파싱 실패: {e}")

        if need_sync:
            logger.info("[자동동기화] 동기화 시작...")
            from services.materials_service import sync_all as sync_all_sources
            result = await sync_all_sources()
            sheets = result.get("sheets", {})
            drive = result.get("drive", {})
            logger.info(
                f"[자동동기화] 완료: "
                f"시트 {sheets.get('success_count',0)}/{sheets.get('total_sources',0)}개, "
                f"총 {sheets.get('total_rows',0)}행 / "
                f"Drive {drive.get('success_count',0)}/{drive.get('total_sources',0)}개, "
                f"총 {drive.get('total_files',0)}파일"
            )
            # Drive 동기화 실패 상세 로깅
            for d in drive.get("details", []):
                if not d.get("success"):
                    logger.warning(f"[자동동기화] Drive 실패: {d.get('name','?')} - {d.get('error','알 수 없음')}")

    except Exception as e:
        logger.error(f"[자동동기화] 오류: {e}", exc_info=True)

    # ── 오더리스트 자동동기화 ──
    try:
        from db.database import get_connection as _gc2
        conn2 = _gc2()
        ol_row = conn2.execute(
            "SELECT MAX(synced_at) as ts FROM orderlist_sync_log"
        ).fetchone()
        conn2.close()

        ol_last = ol_row["ts"] if ol_row and ol_row["ts"] else None
        ol_need_sync = True

        if ol_last:
            try:
                ol_dt = datetime.strptime(str(ol_last)[:19], "%Y-%m-%d %H:%M:%S")
                ol_hours = (datetime.now() - ol_dt).total_seconds() / 3600
                logger.info(f"[자동동기화] 오더리스트 마지막 동기화: {ol_last} ({ol_hours:.1f}시간 전)")
                if ol_hours < 6:
                    ol_need_sync = False
                    logger.info("[자동동기화] 오더리스트 최근 동기화됨 → 스킵")
            except (ValueError, TypeError) as e:
                logger.warning(f"[자동동기화] 오더리스트 날짜 파싱 실패: {e}")

        if ol_need_sync:
            logger.info("[자동동기화] 오더리스트 동기화 시작...")
            from services.orderlist_service import sync_orderlist
            ol_result = sync_orderlist()
            if ol_result.get("success"):
                logger.info(f"[자동동기화] 오더리스트 완료: {ol_result.get('total_items', 0)}건")
            else:
                logger.warning(f"[자동동기화] 오더리스트 실패: {ol_result.get('error', '알 수 없음')}")

    except Exception as e:
        logger.error(f"[자동동기화] 오더리스트 오류: {e}", exc_info=True)


# ─────────────────────────────────────────
#  프론트엔드 서빙
# ─────────────────────────────────────────
@app.get("/", include_in_schema=False)
async def serve_index():
    index_path = FRONTEND_DIR / "index.html"
    if index_path.exists():
        return FileResponse(str(index_path))
    return {"message": "Order Agent API is running. Visit /docs for API documentation."}

@app.get("/health")
async def health():
    """상세 헬스체크 (DB, 외부 API 상태 포함)"""
    from datetime import datetime
    checks = {"status": "ok", "version": "0.2.0", "timestamp": datetime.now().isoformat()}

    # DB 연결 확인
    try:
        from db.database import get_connection
        conn = get_connection()
        conn.execute("SELECT 1").fetchone()
        conn.close()
        checks["database"] = "ok"
    except Exception as e:
        checks["database"] = f"error: {e}"
        checks["status"] = "degraded"

    return checks


# ─────────────────────────────────────────
#  직접 실행 시
# ─────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn
    from config import HOST, PORT, DEBUG
    uvicorn.run("main:app", host=HOST, port=PORT, reload=DEBUG)
